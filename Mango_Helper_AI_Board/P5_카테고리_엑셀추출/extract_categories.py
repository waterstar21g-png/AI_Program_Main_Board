"""
P5_카테고리_엑셀추출 — 오픈마켓 카테고리 전체를 엑셀 분류표로 추출.

접근 URL (기본):
  https://tmg1898.cafe24.com/mall/admin/admin_category_set.php?tm=F&ps_ftid=790

동작 (스크린샷 순서 그대로):
  1) 위 화면에서 대상 마켓 행 (예: 옥션2.0 → `tr#mapping_category_AUC20`)
  2) **[전체카테고리]** 클릭
     `<a onclick="search_category('AUC20','openmarket_category_search_list_AUC20','allview');">`
  3) 목록 리스트박스가 채워지면
     `select#openmarket_category_search_list_AUC20`
  4) 옵션 전체(`e쿠폰/모바일상품권 > 교육/어학이용권 > 온라인교육/외국어` …)를 읽어
     **카테고리분류표(1~6단계)** 로 정리해 엑셀 저장

사용법:
  python extract_categories.py                      # 옥션2.0, 기본 출력폴더
  python extract_categories.py --market GMK20
  python extract_categories.py --out D:\\out\\분류표.xlsx
  python extract_categories.py --from-text 목록.txt  # 브라우저 없이 텍스트 → 엑셀
"""

from __future__ import annotations

import argparse
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
P2_DIR = ROOT / "P2"
if str(P2_DIR) not in sys.path:
    sys.path.insert(0, str(P2_DIR))

ProgressFn = Callable[[str], None]

HERE = Path(__file__).resolve().parent
STOP_FLAG_PATH = HERE / ".p5_stop"
OUTPUT_DIR = HERE / "output"

DEFAULT_URL = (
    "https://tmg1898.cafe24.com/mall/admin/admin_category_set.php?tm=F&ps_ftid=790"
)

# 마켓 코드 → 화면 표기 (admin_category_set.php 의 mapping_category_<코드>)
MARKETS: dict[str, str] = {
    "AUC20": "옥션2.0",       # tr#mapping_category_AUC20
    "11ST": "11번가",          # tr#mapping_category_11ST
    "GMK20": "G마켓2.0",       # tr#mapping_category_GMK20
    "SMART": "스마트스토어",   # tr#mapping_category_SMART
    "COUP": "쿠팡",            # tr#mapping_category_COUP
    "LTON": "롯데ON",          # tr#mapping_category_LTON
}
DEFAULT_MARKET = "AUC20"
ALL_MARKETS = "ALL"  # 전체 마켓 일괄 추출

# 카테고리분류표 양식
LEVELS = 6
LEVEL_HEADERS = [f"{i}단계" for i in range(1, LEVELS + 1)]
HEADERS = ["마켓", *LEVEL_HEADERS, "전체경로"]

PATH_SEPARATORS = (">", "&gt;", "》", "＞")

T_CLICK = 3_000
T_LIST = 15_000  # 전체카테고리 로딩 대기 (ajax)


@dataclass
class RunResult:
    ok: bool
    market: str = ""
    total: int = 0
    deepest: int = 0
    excel_path: str = ""
    errors: list[str] = field(default_factory=list)


def clear_stop_flag() -> None:
    try:
        STOP_FLAG_PATH.unlink(missing_ok=True)  # type: ignore[call-arg]
    except Exception:
        pass


def stop_requested() -> bool:
    return STOP_FLAG_PATH.is_file()


def _log(progress: ProgressFn | None, message: str, *, major: bool = False) -> None:
    line = f"##MAIN##{message}" if major else message
    print(line, flush=True)
    if progress:
        progress(line)


# ── 카테고리 경로 파싱 (순수 로직) ────────────────────────────────


def normalize_separator(text: str) -> str:
    out = str(text or "")
    for sep in PATH_SEPARATORS:
        out = out.replace(sep, ">")
    return out


def parse_category_path(text: str) -> list[str]:
    """`A > B > C` → ['A','B','C'] (빈 조각·공백 제거)."""
    parts = [p.strip() for p in normalize_separator(text).split(">")]
    return [p for p in parts if p]


def is_placeholder(text: str) -> bool:
    """`- 카테고리를 선택해주세요 -` 같은 안내 옵션."""
    t = "".join(str(text or "").split())
    if not t:
        return True
    if t.startswith("-") and t.endswith("-"):
        return True
    return any(k in t for k in ("선택해주세요", "선택하세요", "카테고리검색"))


def to_row(path: list[str], market: str) -> dict:
    """경로 → 분류표 한 행. 6단계보다 깊으면 나머지를 6단계에 합친다."""
    cells = list(path[: LEVELS - 1])
    rest = path[LEVELS - 1 :]
    last = " > ".join(rest) if rest else ""
    cells.append(last)
    cells += [""] * (LEVELS - len(cells))

    row = {"마켓": market, "전체경로": " > ".join(path)}
    for header, value in zip(LEVEL_HEADERS, cells):
        row[header] = value
    return row


def build_rows(options: Iterable[str], market: str) -> list[dict]:
    """옵션 텍스트 목록 → 분류표 행 목록 (중복·안내문구 제거, 순서 유지)."""
    rows: list[dict] = []
    seen: set[str] = set()
    for opt in options:
        if is_placeholder(opt):
            continue
        path = parse_category_path(opt)
        if not path:
            continue
        key = " > ".join(path)
        if key in seen:
            continue
        seen.add(key)
        rows.append(to_row(path, market))
    return rows


def deepest_level(options: Iterable[str]) -> int:
    depth = 0
    for opt in options:
        if is_placeholder(opt):
            continue
        depth = max(depth, len(parse_category_path(opt)))
    return depth


# ── 엑셀 저장 ────────────────────────────────────────────────────


def default_excel_path(market: str, when: datetime | None = None) -> Path:
    stamp = (when or datetime.now()).strftime("%Y%m%d_%H%M%S")
    code = (market or "").strip().upper()
    label = "전체마켓" if code == ALL_MARKETS else MARKETS.get(code, code)
    return OUTPUT_DIR / f"카테고리분류표_{label}_{stamp}.xlsx"


def write_excel(rows: list[dict], path: Path) -> Path:
    from openpyxl import Workbook  # noqa: WPS433
    from openpyxl.styles import Font, PatternFill

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "카테고리분류표"
    ws.append(HEADERS)

    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="F1EEEE")
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill

    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    widths = [10, 22, 22, 22, 22, 22, 26, 60]
    for idx, width in enumerate(widths[: len(HEADERS)], start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    wb.save(str(path))
    return path


# ── 화면 조작 ────────────────────────────────────────────────────


def all_view_selectors(market: str) -> tuple[str, ...]:
    """[전체카테고리] 버튼."""
    return (
        f"a[onclick*=\"search_category('{market}'\"][onclick*=\"allview\"]",
        f'xpath=//tr[@id="mapping_category_{market}"]'
        '//a[.//span[contains(normalize-space(.),"전체카테고리")]]',
        'xpath=//a[.//span[contains(normalize-space(.),"전체카테고리")]]',
    )


def list_select_id(market: str) -> str:
    return f"openmarket_category_search_list_{market}"


def list_select_ids(market: str) -> list[str]:
    """마켓별 목록 select — 화면에 따라 list_ / list2_ 중 채워지는 쪽이 다르다.

    11번가·롯데ON 처럼 둘 다 존재하고 보이는 쪽이 서로 다른 경우가 있어 모두 읽는다.
    """
    return [
        f"openmarket_category_search_list_{market}",
        f"openmarket_category_search_list2_{market}",
    ]


def click_all_categories(page, market: str, *, progress: ProgressFn | None = None) -> bool:
    for sel in all_view_selectors(market):
        try:
            loc = page.locator(sel).first
            if loc.count() == 0:
                continue
            loc.click(timeout=T_CLICK)
            _log(progress, f"  [전체카테고리] 클릭 · {MARKETS.get(market, market)}")
            return True
        except Exception:
            continue
    try:  # 최후: 함수 직접 호출
        page.evaluate(
            "(m) => search_category(m, 'openmarket_category_search_list_' + m, 'allview')",
            market,
        )
        _log(progress, "  [전체카테고리] search_category 직접 호출")
        return True
    except Exception as e:  # noqa: BLE001
        _log(progress, f"오류: 전체카테고리 클릭 실패 · {e}", major=True)
        return False


_OPTIONS_JS = """
(ids) => {
  const out = [];
  for (const id of ids) {
    const el = document.getElementById(id);
    if (!el) continue;
    const texts = Array.from(el.options).map(o => (o.textContent || '').trim());
    if (texts.length > out.length) out.length = 0, out.push(...texts);
  }
  return out;
}
"""


def read_option_texts(page, market: str, *, timeout_ms: int | None = None) -> list[str]:
    """리스트박스 옵션이 채워질 때까지 기다렸다가 전부 읽는다."""
    budget = T_LIST if timeout_ms is None else timeout_ms
    deadline = time.monotonic() + budget / 1000
    best: list[str] = []
    while True:
        try:
            texts = page.evaluate(_OPTIONS_JS, list_select_ids(market)) or []
        except Exception:
            texts = []
        real = [t for t in texts if not is_placeholder(t)]
        if len(real) > len(best):
            best = real
        if len(real) > 1 or time.monotonic() >= deadline or stop_requested():
            return best
        try:
            page.wait_for_timeout(300)
        except Exception:
            time.sleep(0.3)


def markets_to_run(market: str) -> list[str]:
    """`ALL` 이면 전체 마켓, 아니면 하나."""
    code = (market or DEFAULT_MARKET).strip().upper()
    if code == ALL_MARKETS:
        return list(MARKETS.keys())
    return [code]


def extract_one(page, market: str, *, progress: ProgressFn | None = None) -> list[str]:
    """한 마켓 — [전체카테고리] 클릭 → 목록 옵션 읽기."""
    label = MARKETS.get(market, market)
    _log(progress, f"{label} ({market}) — 전체카테고리 조회", major=True)
    if not click_all_categories(page, market, progress=progress):
        return []
    options = read_option_texts(page, market)
    _log(progress, f"  {label} 카테고리 {len(options)}건", major=True)
    return options


def run_extract(
    *,
    market: str = DEFAULT_MARKET,
    url: str = "",
    out_path: str | Path = "",
    progress: ProgressFn | None = None,
) -> RunResult:
    market = (market or DEFAULT_MARKET).strip().upper()
    codes = markets_to_run(market)
    result = RunResult(ok=False, market=market)
    clear_stop_flag()

    try:
        import collect as p2  # noqa: WPS433
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        result.errors.append(f"의존성 로드 실패: {e}")
        _log(progress, result.errors[0], major=True)
        return result

    target = (url or "").strip() or DEFAULT_URL
    _log(
        progress,
        "카테고리 추출 — " + ", ".join(f"{MARKETS.get(c, c)}({c})" for c in codes),
        major=True,
    )
    _log(progress, f"접근 URL={target}")

    rows: list[dict] = []
    per_market: dict[str, int] = {}
    deepest = 0
    try:
        with sync_playwright() as pw:
            _browser, page = p2.connect_browser(pw)
            page.goto(target, wait_until="domcontentloaded", timeout=60_000)
            _log(progress, "카테고리 매핑 화면 표시", major=True)

            for code in codes:
                if stop_requested():
                    _log(progress, "사용자 중단", major=True)
                    break
                options = extract_one(page, code, progress=progress)
                if not options:
                    result.errors.append(f"{MARKETS.get(code, code)} 목록 비어 있음")
                    continue
                market_rows = build_rows(options, MARKETS.get(code, code))
                rows.extend(market_rows)
                per_market[code] = len(market_rows)
                deepest = max(deepest, deepest_level(options))
    except Exception as e:  # noqa: BLE001
        result.errors.append(str(e))
        _log(progress, f"실행 오류: {e}", major=True)
        return result
    finally:
        clear_stop_flag()

    if not rows:
        if not result.errors:
            result.errors.append("카테고리 목록이 비어 있습니다 (전체카테고리 로딩 확인).")
        _log(progress, result.errors[0], major=True)
        return result

    result.total = len(rows)
    result.deepest = deepest

    path = Path(out_path) if out_path else default_excel_path(market)
    try:
        saved = write_excel(rows, path)
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"엑셀 저장 실패: {e}")
        _log(progress, result.errors[0], major=True)
        return result

    result.excel_path = str(saved)
    result.ok = True
    for code, cnt in per_market.items():
        _log(progress, f"  {MARKETS.get(code, code)}: {cnt}행")
    _log(progress, f"엑셀 저장 완료 · {result.total}행 · 최대 {result.deepest}단계", major=True)
    _log(progress, f"  {saved}", major=True)
    return result


def run_from_text(
    text_path: str | Path,
    *,
    market: str = DEFAULT_MARKET,
    out_path: str | Path = "",
    progress: ProgressFn | None = None,
) -> RunResult:
    """브라우저 없이 — 목록 텍스트(한 줄에 한 카테고리)를 엑셀로."""
    market = (market or DEFAULT_MARKET).strip().upper()
    result = RunResult(ok=False, market=market)
    try:
        lines = Path(text_path).read_text(encoding="utf-8").splitlines()
    except OSError as e:
        result.errors.append(f"목록 파일 읽기 실패: {e}")
        _log(progress, result.errors[0], major=True)
        return result

    rows = build_rows(lines, MARKETS.get(market, market))
    if not rows:
        result.errors.append("추출할 카테고리가 없습니다.")
        _log(progress, result.errors[0], major=True)
        return result

    path = Path(out_path) if out_path else default_excel_path(market)
    saved = write_excel(rows, path)
    result.total = len(rows)
    result.deepest = deepest_level(lines)
    result.excel_path = str(saved)
    result.ok = True
    _log(progress, f"엑셀 저장 완료 · {result.total}행 · 최대 {result.deepest}단계", major=True)
    _log(progress, f"  {saved}", major=True)
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="P5_카테고리_엑셀추출")
    parser.add_argument(
        "--market",
        default=DEFAULT_MARKET,
        help=(
            "마켓 코드 ("
            + " / ".join(f"{k}={v}" for k, v in MARKETS.items())
            + f" / {ALL_MARKETS}=전체)"
        ),
    )
    parser.add_argument("--url", default="", help=f"접근 URL (기본={DEFAULT_URL})")
    parser.add_argument("--out", default="", help="엑셀 저장 경로 (기본=output 폴더)")
    parser.add_argument("--from-text", default="", help="브라우저 없이 목록 텍스트 → 엑셀")
    args = parser.parse_args(argv)

    if args.from_text:
        result = run_from_text(args.from_text, market=args.market, out_path=args.out)
    else:
        result = run_extract(market=args.market, url=args.url, out_path=args.out)

    for e in result.errors:
        print(f"[오류] {e}", flush=True)
    return 0 if result.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
