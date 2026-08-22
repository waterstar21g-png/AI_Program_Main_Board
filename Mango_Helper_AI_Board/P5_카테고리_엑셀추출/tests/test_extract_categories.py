"""P5_카테고리_엑셀추출 단위테스트 — 브라우저 없이 파싱·엑셀·선택자 검증."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import extract_categories as ec  # noqa: E402

# 스크린샷 4의 실제 옵션 텍스트
SAMPLE = [
    "- 카테고리를 선택해주세요 -",
    "e쿠폰/모바일상품권 > 교육/어학이용권 > 온라인교육/외국어",
    "e쿠폰/모바일상품권 > 교육/어학이용권 > 자기개발/기타",
    "e쿠폰/모바일상품권 > 기타 상품권 > 구글/아이툰즈/게임",
    "e쿠폰/모바일상품권 > 도넛/아이스크림/분식 > 간식/분식",
    "e쿠폰/모바일상품권 > 미용/뷰티/스파 > 헤어(브랜드샵)",
]


def test_parse_category_path():
    assert ec.parse_category_path("A > B > C") == ["A", "B", "C"]
    assert ec.parse_category_path(" e쿠폰/모바일상품권 >  교육/어학이용권 ") == [
        "e쿠폰/모바일상품권",
        "교육/어학이용권",
    ]
    assert ec.parse_category_path("") == []


def test_parse_handles_other_separators():
    assert ec.parse_category_path("A &gt; B") == ["A", "B"]
    assert ec.parse_category_path("A ＞ B") == ["A", "B"]


def test_placeholder_detection():
    assert ec.is_placeholder("- 카테고리를 선택해주세요 -") is True
    assert ec.is_placeholder("") is True
    assert ec.is_placeholder("e쿠폰/모바일상품권 > 기타") is False


def test_to_row_fills_six_levels():
    row = ec.to_row(["A", "B", "C"], "옥션2.0")
    assert row["마켓"] == "옥션2.0"
    assert row["1단계"] == "A" and row["2단계"] == "B" and row["3단계"] == "C"
    assert row["4단계"] == "" and row["6단계"] == ""
    assert row["전체경로"] == "A > B > C"


def test_to_row_merges_deeper_than_six():
    row = ec.to_row(["1", "2", "3", "4", "5", "6", "7"], "옥션2.0")
    assert row["5단계"] == "5"
    assert row["6단계"] == "6 > 7"  # 6단계 양식 유지
    assert row["전체경로"].endswith("6 > 7")


def test_build_rows_skips_placeholder_and_duplicates():
    rows = ec.build_rows(SAMPLE + [SAMPLE[1]], "옥션2.0")
    assert len(rows) == 5  # 안내문구 1건 제외 · 중복 1건 제외
    assert rows[0]["1단계"] == "e쿠폰/모바일상품권"
    assert rows[0]["3단계"] == "온라인교육/외국어"


def test_headers_are_category_table_form():
    assert ec.HEADERS == [
        "마켓",
        "1단계",
        "2단계",
        "3단계",
        "4단계",
        "5단계",
        "6단계",
        "전체경로",
    ]


def test_deepest_level():
    assert ec.deepest_level(SAMPLE) == 3
    assert ec.deepest_level(["A > B > C > D > E > F > G"]) == 7


def test_write_excel_roundtrip(tmp_path):
    from openpyxl import load_workbook

    rows = ec.build_rows(SAMPLE, "옥션2.0")
    out = ec.write_excel(rows, tmp_path / "분류표.xlsx")
    assert out.is_file()

    ws = load_workbook(out).active
    assert [c.value for c in ws[1]] == ec.HEADERS
    assert ws.max_row == len(rows) + 1
    assert ws.cell(row=2, column=2).value == "e쿠폰/모바일상품권"
    assert ws.freeze_panes == "A2"


def test_run_from_text(tmp_path):
    src = tmp_path / "목록.txt"
    src.write_text("\n".join(SAMPLE), encoding="utf-8")
    out = tmp_path / "결과.xlsx"
    result = ec.run_from_text(src, market="AUC20", out_path=out)
    assert result.ok is True
    assert result.total == 5
    assert result.deepest == 3
    assert Path(result.excel_path) == out


def test_run_from_text_missing_file(tmp_path):
    result = ec.run_from_text(tmp_path / "없음.txt")
    assert result.ok is False
    assert "읽기 실패" in result.errors[0]


# ── 화면 선택자 (스크린샷 DOM) ────────────────────────────────────


def test_all_view_selectors_match_screenshot_dom():
    sels = ec.all_view_selectors("AUC20")
    joined = " ".join(sels)
    assert "search_category('AUC20'" in joined
    assert "allview" in joined
    assert "mapping_category_AUC20" in joined
    assert "전체카테고리" in joined


def test_list_select_id():
    assert ec.list_select_id("AUC20") == "openmarket_category_search_list_AUC20"


def test_default_url_is_category_set_page():
    assert ec.DEFAULT_URL == (
        "https://tmg1898.cafe24.com/mall/admin/admin_category_set.php?tm=F&ps_ftid=790"
    )


def test_markets_match_screenshots():
    """스크린샷 1~10 의 tr#mapping_category_<코드> 와 표기."""
    assert ec.MARKETS == {
        "AUC20": "옥션2.0",
        "11ST": "11번가",
        "GMK20": "G마켓2.0",
        "SMART": "스마트스토어",
        "COUP": "쿠팡",
        "LTON": "롯데ON",
    }
    assert ec.DEFAULT_MARKET == "AUC20"


def test_all_view_selector_per_market():
    for code in ec.MARKETS:
        joined = " ".join(ec.all_view_selectors(code))
        assert f"search_category('{code}'" in joined
        assert f"mapping_category_{code}" in joined
        assert "allview" in joined


def test_list_select_ids_cover_both_variants():
    """11번가·롯데ON 은 list_ / list2_ 중 보이는 쪽이 다르다."""
    ids = ec.list_select_ids("11ST")
    assert ids == [
        "openmarket_category_search_list_11ST",
        "openmarket_category_search_list2_11ST",
    ]


def test_markets_to_run_all():
    assert ec.markets_to_run("ALL") == list(ec.MARKETS.keys())
    assert ec.markets_to_run("coup") == ["COUP"]
    assert ec.markets_to_run("") == ["AUC20"]


def test_default_excel_path_for_all_markets():
    from datetime import datetime

    p = ec.default_excel_path("ALL", datetime(2026, 8, 22, 11, 50, 0))
    assert p.name == "카테고리분류표_전체마켓_20260822_115000.xlsx"


class DualSelectPage:
    """list_ 는 비어 있고 list2_ 에만 목록이 있는 화면 (11번가 형태)."""

    def __init__(self, options):
        self.options = list(options)
        self.script_ids = None

    def evaluate(self, script, *args):
        self.script_ids = args[0] if args else None
        # JS 동작 모사 — 두 id 중 더 긴 목록 채택
        return self.options

    def wait_for_timeout(self, ms):
        return None


def test_read_options_asks_both_select_ids():
    page = DualSelectPage(SAMPLE)
    texts = ec.read_option_texts(page, "LTON")
    assert len(texts) == 5
    assert page.script_ids == ec.list_select_ids("LTON")


class MarketLoopPage:
    """마켓별로 서로 다른 목록을 주는 화면 — ALL 추출 확인용."""

    def __init__(self):
        self.clicked: list[str] = []
        self.current = ""

    def goto(self, url, **kwargs):
        return None

    def locator(self, selector):
        for code in ec.MARKETS:
            if f"search_category('{code}'" in selector:
                self.current = code
                return _ClickOnce(self, code)
        return _Missing()

    def evaluate(self, script, *args):
        if "search_category" in script:
            return None
        return [f"{ec.MARKETS[self.current]}대분류 > 중분류"]

    def wait_for_timeout(self, ms):
        return None


class _ClickOnce:
    def __init__(self, page, code):
        self.page = page
        self.code = code

    @property
    def first(self):
        return self

    def count(self):
        return 1

    def click(self, timeout=None):
        self.page.clicked.append(self.code)


class _Missing:
    @property
    def first(self):
        return self

    def count(self):
        return 0


def test_extract_one_clicks_and_reads(monkeypatch):
    monkeypatch.setattr(ec, "T_LIST", 300)  # 1건만 오는 목록에서 대기 단축
    page = MarketLoopPage()
    logs: list[str] = []
    options = ec.extract_one(page, "SMART", progress=logs.append)
    assert page.clicked == ["SMART"]
    assert options == ["스마트스토어대분류 > 중분류"]
    assert any("스마트스토어" in l for l in logs)


class FakePage:
    """전체카테고리 클릭 후 두 번째 조회에서 목록이 채워지는 화면."""

    def __init__(self):
        self.reads = 0
        self.waits = 0

    def evaluate(self, script, *args):
        self.reads += 1
        if self.reads == 1:
            return ["- 카테고리를 선택해주세요 -"]
        return SAMPLE

    def wait_for_timeout(self, ms):
        self.waits += 1


def test_read_option_texts_waits_for_ajax_fill():
    page = FakePage()
    texts = ec.read_option_texts(page, "AUC20")
    assert len(texts) == 5  # 안내문구 제외
    assert page.waits >= 1


def test_default_excel_path_has_market_and_stamp():
    from datetime import datetime

    p = ec.default_excel_path("AUC20", datetime(2026, 8, 22, 11, 30, 0))
    assert p.name == "카테고리분류표_옥션2.0_20260822_113000.xlsx"
    assert p.parent == ec.OUTPUT_DIR


# ── 구현 제외 마켓 (요건 고정) ────────────────────────────────────


def test_excluded_markets_are_not_targets():
    """LFMall · 머스트잇 · 쇼피 · 큐텐(일본) · 플레이오토(EMP) 는 대상 아님."""
    assert set(ec.EXCLUDED_MARKETS) == {
        "LFMALL",
        "MUSTIT",
        "SHOPEE",
        "QOO10JP",
        "PLAYAUTO",
    }
    for code in ec.EXCLUDED_MARKETS:
        assert code not in ec.MARKETS


def test_all_run_skips_excluded_markets():
    codes = ec.markets_to_run("ALL")
    assert codes == ["AUC20", "11ST", "GMK20", "SMART", "COUP", "LTON"]
    assert not set(codes) & set(ec.EXCLUDED_MARKETS)


def test_excluded_market_requested_directly_is_dropped():
    assert ec.markets_to_run("SHOPEE") == []
