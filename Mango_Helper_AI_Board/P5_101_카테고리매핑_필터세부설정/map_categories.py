"""
P5_101_카테고리매핑_필터세부설정 — 필터별 마켓 카테고리 자동 매핑.

초기 1회
  1) 상품수집사이트 선택 (`select[name="site_id"]`)
  2) [선택조건으로 검색하기] (`onclick="search_filter('search')"`)
  3) 마켓별 카테고리 엑셀 읽어 메모리 보관 (P5 추출 결과 양식)

1단계 루프 — 체크된 행마다
  0) 행 정보 읽기 (체크박스 · 필터이름 · ftid)
  1) 필터이름(수정가능) 읽기
  2~3) 필터세부설정 열의 [설정수정] (`onclick="market_mapping_new('<ftid>')"`)
  4) 팝업 `admin_category_set.php?tm=F&ps_ftid=<ftid>`
  5) [AI 자동 매핑 시작하기] (`onclick="search_recommend_category_all(this)"`)
  6) 2단계 루프 — 마켓마다
       필터이름 ↔ 엑셀 카테고리 비교 → 최적 카테고리
       → 검색필드 입력 (`#openmarket_category_search_text_<코드>`)
       → [검색] (`search_category('<코드>','openmarket_category_search_list_<코드>','')`)
       → 결과 목록에서 일치 항목 선택 (`#openmarket_category_search_list_<코드>`)
  7) [검색필터 설정저장 (Alt+S)] (`onclick="config_save()"`)
  8) 모달 닫기 → 9) 다음 행

사용법:
  python map_categories.py --excel-dir D:\\카테고리엑셀
  python map_categories.py --site-id abcmart --excel AUC20=D:\\옥션.xlsx --excel 11ST=D:\\11번가.xlsx
  python map_categories.py --dry-run          # 화면 조작 없이 매칭 결과만 출력
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, Sequence

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
P2_DIR = ROOT / "P2"
P5_DIR = ROOT / "P5_카테고리_엑셀추출"
for _p in (P2_DIR, P5_DIR):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

ProgressFn = Callable[[str], None]

HERE = Path(__file__).resolve().parent
STOP_FLAG_PATH = HERE / ".map_stop"

DEFAULT_LIST_URL = (
    "https://tmg1898.cafe24.com/mall/admin/admin_group.php"
    "?pmode=filter_delete&uids=&pg=1&date_type=modify"
    "&start_yy=2026&start_mm=8&start_dd=22&end_yy=2026&end_mm=8&end_dd=22"
    "&site_id=&sales_yn=&ft_group=all&sch_field=title&sch_keyword="
    "&ft_num=10&ft_sort=modify_asc"
)
CATEGORY_PAGE = "admin_category_set.php"

# 매핑 대상 마켓 (P5 추출 대상과 동일)
MARKETS: dict[str, str] = {
    "AUC20": "옥션2.0",
    "11ST": "11번가",
    "GMK20": "G마켓2.0",
    "SMART": "스마트스토어",
    "COUP": "쿠팡",
    "LTON": "롯데ON",
}

# 화면 선택자 (스크린샷 기준)
SEARCH_FILTER_JS = "search_filter('search')"
SETTING_EDIT_JS = "market_mapping_new"
AI_MAPPING_JS = "search_recommend_category_all"
CONFIG_SAVE_JS = "config_save"

T_CLICK = 3_000
T_FIELD = 5_000
T_LIST = 8_000
GAP = 0.15

MIN_SCORE = 0.34  # 이 점수 미만이면 매칭 실패로 본다


@dataclass
class RowInfo:
    index: int
    ftid: str
    filter_name: str
    checked: bool = True


@dataclass
class MappedItem:
    market: str
    category: str
    score: float
    ok: bool = False
    reason: str = ""


@dataclass
class RunResult:
    ok: bool
    rows: int = 0
    mapped: int = 0
    failed: int = 0
    details: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def clear_stop_flag() -> None:
    try:
        STOP_FLAG_PATH.unlink(missing_ok=True)  # type: ignore[call-arg]
    except Exception:
        pass


def stop_requested() -> bool:
    return STOP_FLAG_PATH.is_file()


def _log(progress: ProgressFn | None, message: str, *, major: bool = False) -> None:
    line = f"##MAIN##{message}" if major else message
    print(line, flush=True)
    if progress:
        progress(line)


# ── 엑셀 카테고리 자료 (초기 1회 · 메모리 보관) ────────────────────


def market_from_filename(name: str) -> str:
    """파일명에서 마켓 코드 추정 (P5 출력: 카테고리분류표_옥션2.0_....xlsx)."""
    text = str(name)
    for code, label in MARKETS.items():
        if code.lower() in text.lower() or label in text:
            return code
    return ""


def load_categories(path: str | Path) -> list[str]:
    """엑셀에서 카테고리 전체경로 목록을 읽는다.

    P5 추출 양식(마켓·구분·1~6단계·전체경로) 우선, 없으면 단계 열을 이어 붙이거나
    첫 열을 그대로 쓴다.
    """
    from openpyxl import load_workbook  # noqa: WPS433

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip() for c in next(rows)]
    except StopIteration:
        return []

    full_idx = header.index("전체경로") if "전체경로" in header else -1
    level_idx = [i for i, h in enumerate(header) if re.fullmatch(r"\d단계", h)]

    out: list[str] = []
    seen: set[str] = set()
    for row in rows:
        if row is None:
            continue
        if full_idx >= 0 and full_idx < len(row) and row[full_idx]:
            path_text = str(row[full_idx]).strip()
        elif level_idx:
            parts = [str(row[i]).strip() for i in level_idx if i < len(row) and row[i]]
            path_text = " > ".join(parts)
        else:
            path_text = str(row[0]).strip() if row and row[0] else ""
        if not path_text or path_text in seen:
            continue
        seen.add(path_text)
        out.append(path_text)
    return out


def load_market_excels(
    paths: dict[str, str | Path], *, progress: ProgressFn | None = None
) -> dict[str, list[str]]:
    """마켓 코드 → 카테고리 목록."""
    data: dict[str, list[str]] = {}
    for code, path in paths.items():
        code = code.strip().upper()
        if not path:
            continue
        try:
            cats = load_categories(path)
        except Exception as e:  # noqa: BLE001
            _log(progress, f"엑셀 읽기 실패 · {code} · {e}", major=True)
            continue
        data[code] = cats
        _log(progress, f"  {MARKETS.get(code, code)} 카테고리 {len(cats)}건 로드", major=True)
    return data


def discover_market_excels(folder: str | Path) -> dict[str, str]:
    """폴더에서 마켓별 엑셀을 파일명으로 자동 매칭."""
    found: dict[str, str] = {}
    for p in sorted(Path(folder).glob("*.xlsx")):
        code = market_from_filename(p.name)
        if code and code not in found:
            found[code] = str(p)
    return found


# ── 필터이름 ↔ 카테고리 매칭 (순수 로직) ──────────────────────────

_SPLIT_RE = re.compile(r"[\s_\-/>·,()\[\]]+")


def tokenize(text: str) -> list[str]:
    """비교용 토큰 — 구분자·공백 제거, 소문자화."""
    raw = _SPLIT_RE.split(str(text or "").strip().lower())
    return [t for t in raw if t]


def leaf_of(path: str) -> str:
    parts = [p.strip() for p in str(path or "").split(">") if p.strip()]
    return parts[-1] if parts else ""


def similarity(filter_name: str, category_path: str) -> float:
    """필터이름과 카테고리 경로의 유사도 (0~1).

    마지막 단계(리프)에 가중치를 두고, 경로 전체 토큰 겹침도 함께 본다.
    """
    ftoks = set(tokenize(filter_name))
    if not ftoks:
        return 0.0
    path_toks = set(tokenize(category_path))
    leaf_toks = set(tokenize(leaf_of(category_path)))
    if not path_toks:
        return 0.0

    path_hit = len(ftoks & path_toks) / len(ftoks)
    leaf_hit = len(ftoks & leaf_toks) / len(ftoks) if leaf_toks else 0.0

    # 문자열 포함 보너스 (예: '남성비니' ↔ '비니')
    joined_f = "".join(sorted(ftoks))
    bonus = 0.0
    for tok in leaf_toks:
        if tok and (tok in joined_f or any(tok in f or f in tok for f in ftoks)):
            bonus = max(bonus, 0.25)
    return min(1.0, 0.5 * path_hit + 0.5 * leaf_hit + bonus)


def best_category(
    filter_name: str, categories: Sequence[str], *, min_score: float = MIN_SCORE
) -> tuple[str, float]:
    """필터이름에 가장 잘 맞는 카테고리 (없으면 ('', 점수))."""
    best = ""
    best_score = 0.0
    for cat in categories:
        score = similarity(filter_name, cat)
        if score > best_score or (score == best_score and cat and len(cat) < len(best)):
            best, best_score = cat, score
    if best_score < min_score:
        return "", best_score
    return best, best_score


def search_keyword_for(category_path: str) -> str:
    """카테고리 검색필드에 넣을 검색어 — 마지막 단계."""
    return leaf_of(category_path)


def pick_option(options: Sequence[str], category_path: str) -> str:
    """검색 결과 목록에서 고를 항목 — 완전일치 → 리프일치 → 최고 유사도."""
    target = str(category_path or "").strip()
    if not target:
        return ""
    norm = lambda s: "".join(str(s or "").split())  # noqa: E731
    for opt in options:
        if norm(opt) == norm(target):
            return opt
    leaf = leaf_of(target)
    for opt in options:
        if leaf and norm(leaf_of(opt)) == norm(leaf):
            return opt
    best, score = best_category(target, list(options), min_score=0.0)
    return best if score > 0 else ""


# ── 화면 조작 ────────────────────────────────────────────────────


def _first(page, selectors: Iterable[str]):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0:
                return loc
        except Exception:
            continue
    return None


def select_site(page, site_id: str, *, progress: ProgressFn | None = None) -> bool:
    """상품수집사이트 리스트박스 선택 (초기 1회)."""
    if not site_id:
        return True
    loc = _first(page, ('select[name="site_id"]',))
    if loc is None:
        _log(progress, "오류: 상품수집사이트 드롭다운 미검출", major=True)
        return False
    for how in ("label", "value"):
        try:
            if how == "label":
                loc.select_option(label=site_id, timeout=T_CLICK)
            else:
                loc.select_option(site_id, timeout=T_CLICK)
            _log(progress, f"상품수집사이트: {site_id}", major=True)
            return True
        except Exception:
            continue
    _log(progress, f"오류: 상품수집사이트 선택 실패 · {site_id}", major=True)
    return False


def click_search_filter(page, *, progress: ProgressFn | None = None) -> bool:
    """[선택조건으로 검색하기]."""
    loc = _first(
        page,
        (
            f'a[onclick*="{SEARCH_FILTER_JS}"]',
            'xpath=//a[.//span[contains(normalize-space(.),"선택조건으로 검색하기")]]',
            'xpath=//*[contains(normalize-space(.),"선택조건으로 검색하기")]',
        ),
    )
    if loc is None:
        _log(progress, "오류: [선택조건으로 검색하기] 미검출", major=True)
        return False
    try:
        loc.click(timeout=T_CLICK)
    except Exception as e:  # noqa: BLE001
        _log(progress, f"오류: 검색 클릭 실패 · {e}", major=True)
        return False
    _log(progress, "[선택조건으로 검색하기] 클릭", major=True)
    try:
        page.wait_for_load_state("domcontentloaded", timeout=T_FIELD)
    except Exception:
        pass
    time.sleep(GAP)
    return True


LIST_ROWS_JS = r"""
() => {
  const out = [];
  const table = document.querySelector('table#search_category') || document;
  const trs = Array.from(table.querySelectorAll('tr'));
  trs.forEach((tr, idx) => {
    const cb = tr.querySelector('input[type="checkbox"]');
    const edit = tr.querySelector('a[onclick*="market_mapping_new"]');
    if (!edit) return;
    const m = (edit.getAttribute('onclick') || '').match(/market_mapping_new\(\s*'?(\d+)'?/);
    const ftid = m ? m[1] : '';
    let name = '';
    const nameInput = Array.from(tr.querySelectorAll('input[type="text"], input:not([type])'))
      .map(i => (i.value || '').trim())
      .filter(v => v && !/^https?:/i.test(v) && !/^\d+$/.test(v));
    if (nameInput.length) name = nameInput[0];
    if (!name) {
      const td = tr.querySelector('td');
      name = td ? (td.innerText || '').trim() : '';
    }
    out.push({index: idx, ftid, filterName: name, checked: cb ? !!cb.checked : false});
  });
  return out;
}
"""


def list_rows(page) -> list[RowInfo]:
    try:
        data = page.evaluate(LIST_ROWS_JS) or []
    except Exception:
        data = []
    rows: list[RowInfo] = []
    for d in data:
        rows.append(
            RowInfo(
                index=int(d.get("index") or 0),
                ftid=str(d.get("ftid") or "").strip(),
                filter_name=str(d.get("filterName") or "").strip(),
                checked=bool(d.get("checked")),
            )
        )
    return rows


def build_mapping_url(ftid: str, *, list_url: str = DEFAULT_LIST_URL) -> str:
    """설정수정 팝업 URL (`admin_category_set.php?tm=F&ps_ftid=<ftid>`)."""
    from urllib.parse import urlencode, urlsplit, urlunsplit

    parts = urlsplit(list_url or DEFAULT_LIST_URL)
    base_dir = parts.path.rsplit("/", 1)[0] if "/" in parts.path else ""
    query = urlencode({"tm": "F", "ps_ftid": str(ftid)})
    return urlunsplit((parts.scheme, parts.netloc, f"{base_dir}/{CATEGORY_PAGE}", query, ""))


def open_setting_popup(page, row: RowInfo, *, list_url: str, progress: ProgressFn | None = None):
    """행의 [설정수정] → 팝업. 실패 시 팝업 URL 직접 오픈."""
    sel = f"a[onclick*=\"{SETTING_EDIT_JS}('{row.ftid}')\"]"
    try:
        with page.expect_popup(timeout=T_FIELD) as info:
            page.locator(sel).first.click(timeout=T_CLICK)
        popup = info.value
        _log(progress, f"  설정수정 팝업 (ftid={row.ftid})")
        return popup
    except Exception:
        pass

    url = build_mapping_url(row.ftid, list_url=list_url)
    try:
        popup = page.context.new_page()
        popup.goto(url, wait_until="domcontentloaded", timeout=30_000)
        _log(progress, f"  설정수정 직접 열기 (ftid={row.ftid})")
        return popup
    except Exception as e:  # noqa: BLE001
        _log(progress, f"오류: 설정수정 팝업 실패 · {e}", major=True)
        return None


def click_ai_mapping(popup, *, progress: ProgressFn | None = None) -> bool:
    """[AI 자동 매핑 시작하기]."""
    loc = _first(
        popup,
        (
            f'a[onclick*="{AI_MAPPING_JS}"]',
            'xpath=//a[.//span[contains(normalize-space(.),"AI 자동 매핑")]]',
        ),
    )
    if loc is None:
        _log(progress, "  경고: [AI 자동 매핑 시작하기] 미검출 — 수동 매핑만 진행")
        return False
    try:
        loc.click(timeout=T_CLICK)
        _log(progress, "  AI 자동 매핑 시작")
        time.sleep(0.5)
        return True
    except Exception:
        return False


def market_search_input(popup, market: str):
    return _first(
        popup,
        (
            f"#openmarket_category_search_text_{market}",
            f'input[id="openmarket_category_search_text_{market}"]',
        ),
    )


def click_market_search(popup, market: str) -> bool:
    loc = _first(
        popup,
        (
            f"a[onclick*=\"search_category('{market}','openmarket_category_search_list_{market}',''\"]",
            f'xpath=//tr[@id="mapping_category_{market}"]'
            '//a[.//span[normalize-space()="검색"]]',
        ),
    )
    if loc is None:
        return False
    try:
        loc.click(timeout=T_CLICK)
        return True
    except Exception:
        return False


# 결과 리스트박스는 마켓에 따라 list_ / list2_ 두 벌이고 보이는 쪽이 다르다
# (11번가·롯데ON). 보이는 select 을 우선 읽고, 선택도 그 id 에 한다.
RESULT_OPTIONS_JS = """
(ids) => {
  const texts = (el) => Array.from(el.options).map(o => (o.textContent || '').trim());
  const isVisible = (el) => {
    const st = window.getComputedStyle(el);
    if (st.display === 'none' || st.visibility === 'hidden') return false;
    return el.offsetParent !== null || st.display === 'inline-block';
  };
  const cands = ids.map(id => document.getElementById(id)).filter(Boolean);
  const pick = (list) => {
    let best = {texts: [], id: ''};
    for (const el of list) {
      const t = texts(el);
      if (t.length > best.texts.length) best = {texts: t, id: el.id || ''};
    }
    return best;
  };
  const visible = pick(cands.filter(isVisible));
  return visible.texts.length ? visible : pick(cands);
}
"""


def result_select_ids(market: str) -> list[str]:
    return [
        f"openmarket_category_search_list_{market}",
        f"openmarket_category_search_list2_{market}",
    ]


def read_result_options(
    popup, market: str, *, timeout_ms: int = T_LIST
) -> tuple[list[str], str]:
    """(옵션 목록, 사용한 select id) — 보이는 리스트박스 기준."""
    deadline = time.monotonic() + timeout_ms / 1000
    while True:
        try:
            info = popup.evaluate(RESULT_OPTIONS_JS, result_select_ids(market)) or {}
        except Exception:
            info = {}
        texts = list(info.get("texts") or []) if isinstance(info, dict) else list(info or [])
        sel_id = str(info.get("id") or "") if isinstance(info, dict) else ""
        real = [t for t in texts if t and not t.startswith("-")]
        if real or time.monotonic() >= deadline:
            return real, sel_id or result_select_ids(market)[0]
        try:
            popup.wait_for_timeout(200)
        except Exception:
            time.sleep(0.2)


def choose_option(popup, market: str, label: str, *, select_id: str = "") -> bool:
    ids = [select_id] if select_id else result_select_ids(market)
    for sid in ids:
        loc = _first(popup, (f"#{sid}",))
        if loc is None:
            continue
        try:
            loc.select_option(label=label, timeout=T_CLICK)
            return True
        except Exception:
            continue
    return False


def click_config_save(popup, *, progress: ProgressFn | None = None) -> bool:
    """[검색필터 설정저장 (Alt+S)]."""
    loc = _first(
        popup,
        (
            f'a[onclick*="{CONFIG_SAVE_JS}"]',
            'xpath=//a[.//span[contains(normalize-space(.),"검색필터 설정저장")]]',
        ),
    )
    if loc is None:
        _log(progress, "  오류: [검색필터 설정저장] 미검출", major=True)
        return False
    try:
        loc.click(timeout=T_CLICK)
        _log(progress, "  검색필터 설정저장")
        return True
    except Exception:
        try:
            popup.keyboard.press("Alt+s")
            _log(progress, "  검색필터 설정저장 (Alt+S)")
            return True
        except Exception:
            return False


def close_popup(popup) -> None:
    try:
        if not popup.is_closed():
            popup.close()
    except Exception:
        pass


def map_one_market(
    popup,
    market: str,
    filter_name: str,
    categories: Sequence[str],
    *,
    progress: ProgressFn | None = None,
) -> MappedItem:
    """한 마켓 매핑 — 최적 카테고리 → 검색어 입력 → 검색 → 목록 선택."""
    label = MARKETS.get(market, market)
    if not categories:
        return MappedItem(market, "", 0.0, False, "엑셀 자료 없음")

    category, score = best_category(filter_name, categories)
    if not category:
        _log(progress, f"  {label}: 매칭 실패 (최고점 {score:.2f})")
        return MappedItem(market, "", score, False, "유사 카테고리 없음")

    keyword = search_keyword_for(category)
    box = market_search_input(popup, market)
    if box is None:
        return MappedItem(market, category, score, False, "검색필드 미검출")
    try:
        box.fill(keyword, timeout=T_CLICK)
    except Exception as e:  # noqa: BLE001
        return MappedItem(market, category, score, False, f"검색어 입력 실패({e})")

    if not click_market_search(popup, market):
        return MappedItem(market, category, score, False, "검색 버튼 미검출")

    options, select_id = read_result_options(popup, market)
    if not options:
        return MappedItem(market, category, score, False, "검색 결과 없음")

    picked = pick_option(options, category)
    if not picked or not choose_option(popup, market, picked, select_id=select_id):
        return MappedItem(market, category, score, False, "목록 선택 실패")

    _log(progress, f"  {label}: {picked}  (점수 {score:.2f})")
    return MappedItem(market, picked, score, True, "")


def map_one_row(
    page,
    row: RowInfo,
    excels: dict[str, list[str]],
    *,
    list_url: str,
    markets: Sequence[str] | None = None,
    progress: ProgressFn | None = None,
) -> dict:
    """한 행 — 설정수정 팝업 → AI 매핑 → 마켓별 매핑 → 설정저장 → 닫기."""
    codes = list(markets or MARKETS.keys())
    detail: dict = {"ftid": row.ftid, "filter": row.filter_name, "items": []}

    _log(progress, f"필터 [{row.filter_name}] (ftid={row.ftid})", major=True)
    popup = open_setting_popup(page, row, list_url=list_url, progress=progress)
    if popup is None:
        detail["error"] = "팝업 실패"
        return detail

    try:
        popup.on("dialog", lambda d: d.accept())
    except Exception:
        pass

    try:
        click_ai_mapping(popup, progress=progress)
        for market in codes:
            if stop_requested():
                break
            item = map_one_market(
                popup, market, row.filter_name, excels.get(market, []), progress=progress
            )
            detail["items"].append(item.__dict__)
            if item.ok:
                click_config_save(popup, progress=progress)
                time.sleep(GAP)
        click_config_save(popup, progress=progress)
    finally:
        close_popup(popup)
    return detail


def run_mapping(
    *,
    site_id: str = "",
    excels: dict[str, list[str]] | None = None,
    excel_paths: dict[str, str] | None = None,
    list_url: str = "",
    markets: Sequence[str] | None = None,
    progress: ProgressFn | None = None,
) -> RunResult:
    result = RunResult(ok=False)
    clear_stop_flag()

    data = dict(excels or {})
    if not data and excel_paths:
        data = load_market_excels(excel_paths, progress=progress)
    if not data:
        result.errors.append("마켓별 카테고리 엑셀이 없습니다.")
        _log(progress, result.errors[0], major=True)
        return result

    try:
        import collect as p2  # noqa: WPS433
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        result.errors.append(f"의존성 로드 실패: {e}")
        _log(progress, result.errors[0], major=True)
        return result

    url = (list_url or "").strip() or DEFAULT_LIST_URL

    try:
        with sync_playwright() as pw:
            _browser, page = p2.connect_browser(pw)
            page.goto(url, wait_until="domcontentloaded", timeout=60_000)
            _log(progress, "검색필터 목록 화면", major=True)

            if not select_site(page, site_id, progress=progress):
                result.errors.append("상품수집사이트 선택 실패")
            click_search_filter(page, progress=progress)

            rows = [r for r in list_rows(page) if r.checked and r.ftid]
            result.rows = len(rows)
            if not rows:
                result.errors.append("체크된 행이 없습니다.")
                _log(progress, result.errors[0], major=True)
                return result
            _log(progress, f"체크된 행 {len(rows)}건 — 순차 매핑", major=True)

            for i, row in enumerate(rows, start=1):
                if stop_requested():
                    _log(progress, "사용자 중단", major=True)
                    break
                _log(progress, f"[{i}/{len(rows)}]", major=True)
                detail = map_one_row(
                    page,
                    row,
                    data,
                    list_url=url,
                    markets=markets,
                    progress=progress,
                )
                result.details.append(detail)
                ok_cnt = sum(1 for it in detail.get("items", []) if it.get("ok"))
                result.mapped += ok_cnt
                result.failed += len(detail.get("items", [])) - ok_cnt
                time.sleep(GAP)
    except Exception as e:  # noqa: BLE001
        result.errors.append(str(e))
        _log(progress, f"실행 오류: {e}", major=True)
        return result
    finally:
        clear_stop_flag()

    result.ok = result.mapped > 0
    _log(
        progress,
        f"완료 — 행 {result.rows} · 매핑 성공 {result.mapped} · 실패 {result.failed}",
        major=True,
    )
    return result


def run_dry(
    filter_names: Sequence[str],
    excels: dict[str, list[str]],
    *,
    progress: ProgressFn | None = None,
) -> list[dict]:
    """브라우저 없이 매칭 결과만 확인 (검증용)."""
    out: list[dict] = []
    for name in filter_names:
        row = {"filter": name, "items": []}
        for code, cats in excels.items():
            cat, score = best_category(name, cats)
            row["items"].append(
                {"market": code, "category": cat, "score": round(score, 3)}
            )
            _log(progress, f"{name} · {MARKETS.get(code, code)} → {cat or '(없음)'} ({score:.2f})")
        out.append(row)
    return out


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="P5_101_카테고리매핑_필터세부설정")
    parser.add_argument("--site-id", default="", help="상품수집사이트 (리스트박스 값/표기)")
    parser.add_argument("--list-url", default="", help=f"목록 URL (기본={DEFAULT_LIST_URL[:60]}…)")
    parser.add_argument("--excel-dir", default="", help="마켓별 엑셀 폴더 (파일명으로 자동 매칭)")
    parser.add_argument(
        "--excel",
        action="append",
        default=[],
        metavar="코드=경로",
        help="마켓별 엑셀 지정 (예: AUC20=D:\\옥션.xlsx)",
    )
    parser.add_argument("--markets", default="", help="대상 마켓 (쉼표, 기본=전체)")
    parser.add_argument("--dry-run", default="", help="매칭만 확인할 필터이름 (쉼표)")
    args = parser.parse_args(argv)

    paths: dict[str, str] = {}
    if args.excel_dir:
        paths.update(discover_market_excels(args.excel_dir))
    for item in args.excel:
        code, _, path = str(item).partition("=")
        if code and path:
            paths[code.strip().upper()] = path.strip()

    excels = load_market_excels(paths) if paths else {}

    if args.dry_run:
        run_dry([n.strip() for n in args.dry_run.split(",") if n.strip()], excels)
        return 0

    markets = [m.strip().upper() for m in args.markets.split(",") if m.strip()] or None
    result = run_mapping(
        site_id=args.site_id,
        excels=excels,
        list_url=args.list_url,
        markets=markets,
    )
    for e in result.errors:
        print(f"[오류] {e}", flush=True)
    return 0 if result.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
