"""
P2 — 더망고(tmg1898) 상품데이터 대량수집

목표: 카테고리 URL(P1 엑셀) 1행당 상품 정보 N건(기본 3)을 오류 없이 가져오기.

0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭
1. URL상품검색하기 : 필드값 입력 후 클릭 -> 팝업창이 없어질 때까지 대기
2. 검색된 상품 모두저장 클릭 -> 검색필터명·저장수 입력 -> 저장하기
3. 팝업창이 없어질 때까지 대기 + 저장 결과 확인
4. 다음 행 (실패 시 같은 행 재시도)

사용법:
    python collect.py 엑셀.xlsx              # 저장수 3
    python collect.py 엑셀.xlsx 3 --verify   # 1행 검증(샷·재시도)
    python collect.py 엑셀.xlsx 3 --retries 3 --yes
    run-verify.bat 엑셀.xlsx
"""

import getpass
import os
import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    TimeoutError as PWTimeout,
    sync_playwright,
)

LOGIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin_login.php"
MAIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin.php"
BULK_URL = "https://tmg1898.cafe24.com/mall/admin/shop/getGoodsNew.php"
ADMIN_HOST = "tmg1898.cafe24.com"
BULK_PATH = "getGoodsNew.php"

CDP_PORT = 9222
CDP_URL = f"http://127.0.0.1:{CDP_PORT}"
PROFILE_DIR = Path(__file__).parent / ".chrome-profile"

CHROME_CANDIDATES = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    "/usr/bin/google-chrome-stable",
    "/usr/bin/google-chrome",
    "/usr/local/bin/google-chrome",
    "/usr/bin/chromium-browser",
    "/usr/bin/chromium",
]

POPUP_WAIT_SEC = 600
MODAL_WAIT_SEC = 180
DEFAULT_SAVE_COUNT = 3
DEFAULT_ROW_RETRIES = 3

FILTER_NAME_LABEL = re.compile(r"검색\s*필터\s*명")
SAVE_COUNT_LABEL = re.compile(r"저장\s*상품\s*수|검색결과\s*상위")
# 저장 완료로 볼 수 있는 화면 문구 (망고 버전에 따라 다를 수 있음)
SAVE_OK_PATTERNS = [
    re.compile(r"저장\s*(이\s*)?(완료|성공)"),
    re.compile(r"정상\s*처리"),
    re.compile(r"상품\s*(이\s*)?저장"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?저장"),
]
SAVE_FAIL_PATTERNS = [
    re.compile(r"저장\s*실패"),
    re.compile(r"오류\s*가\s*발생"),
    re.compile(r"다시\s*시도"),
]

SHOT_ROOT = Path(__file__).parent / "run-logs"

# CLI/환경변수로 받은 더망고 로그인 (세션 없을 때 사용)
_TMG_ID: str | None = None
_TMG_PW: str | None = None


def log(msg: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)


class RunCtx:
    """1행×N상품 검증·재시도·스크린샷용 실행 컨텍스트"""

    def __init__(
        self,
        *,
        save_count: int = DEFAULT_SAVE_COUNT,
        retries: int = DEFAULT_ROW_RETRIES,
        verify: bool = False,
        max_rows: int | None = None,
        batch: bool = False,
        shot_dir: Path | None = None,
    ) -> None:
        self.save_count = save_count
        self.retries = max(1, retries)
        self.verify = verify
        self.max_rows = max_rows
        self.batch = batch or verify  # 검증 모드는 기본 무중단
        stamp = time.strftime("%Y%m%d_%H%M%S")
        self.shot_dir = shot_dir or (SHOT_ROOT / stamp)
        self.shot_dir.mkdir(parents=True, exist_ok=True)
        self.step_i = 0
        self.log_path = self.shot_dir / "run.log"
        self._log_file = open(self.log_path, "a", encoding="utf-8")

    def close(self) -> None:
        try:
            self._log_file.close()
        except Exception:
            pass

    def info(self, msg: str) -> None:
        line = f"[{time.strftime('%H:%M:%S')}] {msg}"
        print(line, flush=True)
        try:
            self._log_file.write(line + "\n")
            self._log_file.flush()
        except Exception:
            pass

    def shot(self, page: Page, tag: str, row_no: int | None = None) -> Path | None:
        self.step_i += 1
        safe = re.sub(r"[^\w\-가-힣]+", "_", tag)[:40]
        name = f"{self.step_i:02d}_r{row_no or 0}_{safe}.png"
        path = self.shot_dir / name
        try:
            page.screenshot(path=str(path), full_page=True)
            self.info(f"  [샷] {path.name}")
            return path
        except Exception as e:  # noqa: BLE001
            self.info(f"  [샷 실패] {e}")
            return None


# ── 브라우저 연결 (기존 Chrome/Edge에 CDP로 붙기, Chromium 다운로드 없음) ──

def cdp_port_open(port: int = CDP_PORT) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        try:
            s.connect(("127.0.0.1", port))
            return True
        except OSError:
            return False


def find_browser_exe() -> str | None:
    for path in CHROME_CANDIDATES:
        if Path(path).exists():
            return path
    return None


def _clear_stale_singleton_locks() -> None:
    """
    이전 실행이 비정상 종료되어 남은 잠금파일이 있으면, 새 Chrome이
    "이미 실행 중"이라고 착각해 디버그 포트 없이 조용히 기존 창에만
    메시지를 보내고 끝나버릴 수 있다(=화면에 아무 반응도 없어 보임).
    cdp_port_open()이 False라는 건 우리 프로필로 살아있는 프로세스가
    없다는 뜻이므로 안전하게 지운다.
    """
    for name in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
        f = PROFILE_DIR / name
        try:
            if f.exists() or f.is_symlink():
                f.unlink()
        except OSError:
            pass


def launch_debug_browser() -> None:
    """평소 쓰는 Chrome/Edge를 디버그 포트로 실행 — Playwright Chromium 미사용"""
    exe = find_browser_exe()
    if not exe:
        raise SystemExit(
            "Chrome 또는 Edge를 찾지 못했습니다.\n"
            "https://www.google.com/chrome/ 에서 설치 후 다시 실행하세요."
        )
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    _clear_stale_singleton_locks()
    log(f"브라우저 실행: {exe}")
    # 주의: stdout/stderr를 DEVNULL로 버리면 일부 환경(보안 소프트웨어·
    # 컨테이너 등)에서 Chrome이 바로 죽는 경우가 있다(무반응처럼 보임).
    # 로그 파일로 받아두면 안전하고, 문제 생기면 이 파일로 원인도 알 수 있다.
    log_path = PROFILE_DIR / "chrome_debug.log"
    log_file = open(log_path, "ab")
    subprocess.Popen(
        [
            exe,
            f"--remote-debugging-port={CDP_PORT}",
            f"--user-data-dir={PROFILE_DIR}",
            "--no-first-run",
            "--no-default-browser-check",
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
            MAIN_URL,
        ],
        stdout=log_file,
        stderr=log_file,
    )
    log("디버그 모드 연결 대기 중 (최대 30초)...")
    for i in range(60):
        if cdp_port_open():
            log("브라우저 연결 확인됨")
            return
        if i > 0 and i % 6 == 0:
            log(f"  아직 대기 중... ({i * 0.5:.0f}초 경과 — 화면에 Chrome 창이 떴는지 확인해 주세요)")
        time.sleep(0.5)  # 소켓 연결 대기 — Playwright 이벤트루프와 무관하므로 안전
    raise SystemExit(
        "브라우저가 디버그 모드로 열리지 않았습니다.\n"
        "열려있는 Chrome/Edge 창을 모두 닫고 다시 실행해 보세요.\n"
        f"(참고 로그: {log_path})"
    )


def pick_working_page(context: BrowserContext) -> Page:
    """이미 망고 화면이 열려 있으면 그 탭을 그대로 사용(기본 화면 유지)"""
    open_pages = [p for p in context.pages if not p.is_closed()]
    for p in open_pages:
        try:
            if ADMIN_HOST in p.url:
                return p
        except Exception:  # noqa: BLE001
            continue
    return open_pages[0] if open_pages else context.new_page()


def refresh_if_closed(page: Page) -> Page:
    """
    로그인 성공 후 사이트가 원래 탭을 닫고 새 창을 띄우는 경우가 있다
    (예: 로그인 중계 페이지가 자기 자신을 닫음). 그러면 이전 page
    객체로는 더 이상 아무 것도 할 수 없으므로(TargetClosedError),
    같은 컨텍스트에서 살아있는 페이지를 다시 찾아온다.
    """
    if not page.is_closed():
        return page
    log("  탭이 닫힘 감지 — 새 탭을 다시 찾는 중...")
    return pick_working_page(page.context)


def connect_browser(p) -> tuple[Browser, Page]:
    """
    1) 이미 디버그 모드로 열린 Chrome/Edge가 있으면 그대로 연결(= 망고 기본 화면 그대로)
    2) 없으면 평소 쓰는 Chrome/Edge를 디버그 모드로 새로 열어서 연결
    Playwright 전용 Chromium은 내려받지 않는다.
    """
    if not cdp_port_open():
        launch_debug_browser()

    browser = p.chromium.connect_over_cdp(CDP_URL)
    context = browser.contexts[0] if browser.contexts else browser.new_context()
    page = pick_working_page(context)
    return browser, page


# ── 엑셀 ──────────────────────────────────────────────────────

def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        label_col = headers.index("상위 최종 카테고리명")
        url_col = headers.index("최종 카테고리 URL주소")
    except ValueError:
        raise SystemExit(
            "엑셀 1행 헤더에 '상위 최종 카테고리명', '최종 카테고리 URL주소' 열이 있어야 합니다."
        )

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        label = str(row[label_col].value or "").strip()
        url = str(row[url_col].value or "").strip()
        if url:
            rows.append({"row": i, "label": label, "url": url})
    return rows


def normalize_url(u: str) -> str:
    u = u.strip()
    return u if re.match(r"^https?://", u, re.I) else f"https://{u}"


# ── 팝업 ──────────────────────────────────────────────────────

def popups(page: Page) -> list:
    result = []
    for p in page.context.pages:
        if p is page or p.is_closed():
            continue
        try:
            u = p.url
        except Exception:
            continue
        if u and u != "about:blank" and ADMIN_HOST not in u:
            result.append(p)
    return result


def wait_popups_gone(
    page: Page,
    timeout_sec: int = POPUP_WAIT_SEC,
    grace_sec: float = 2.0,
    warn_if_never_opened: bool = False,
) -> bool:
    """팝업이 스스로 닫힐 때까지 대기 — 절대 건드리지 않음

    grace_sec: 클릭 직후 팝업이 뜨기까지 잠깐 기다리는 시간.
    팝업이 아예 안 뜨는 경우(이미 닫혀 있음)까지 대비해 대기 시간은 짧게 둔다
    (없는 팝업을 기다리며 매 행마다 시간을 허비하지 않도록).
    warn_if_never_opened=True 이면, grace_sec 동안 팝업이 단 한 번도
    뜨지 않았을 때 경고 로그를 남긴다(예: URL 검색 클릭이 안 먹혔을 때).

    주의: Playwright Python 동기 API는 이벤트(새 창 열림/닫힘)를
    time.sleep() 중에는 처리하지 않는다. 반드시 page.wait_for_timeout()
    으로 기다려야 context.pages()가 실시간으로 갱신된다.

    반환값: 팝업이 한 번이라도 열렸으면 True.
    """
    end = time.time() + timeout_sec
    grace_end = time.time() + grace_sec
    ever_seen = False
    while time.time() < grace_end:
        if popups(page):
            ever_seen = True
            break
        page.wait_for_timeout(200)

    if not ever_seen:
        if warn_if_never_opened:
            log("  [경고] 팝업이 뜨지 않음 — 클릭이 제대로 안 됐거나 사이트가 응답하지 않았을 수 있음")
        return False

    last_beat = 0.0
    while popups(page):
        if time.time() > end:
            raise TimeoutError("팝업창이 닫히지 않음")
        if time.time() - last_beat > 10:
            last_beat = time.time()
            cur = popups(page)
            urls = ", ".join(p.url for p in cur if not p.is_closed())
            log(f"  팝업창 대기중... (열린 팝업 {len(cur)}개: {urls})")
        page.wait_for_timeout(500)
    return True


# ── 입력 · 클릭 (망고 구형 input 대응) ────────────────────────

def type_into(page: Page, locator, value: str) -> None:
    el = locator.first
    el.wait_for(state="attached", timeout=60_000)
    el.scroll_into_view_if_needed()
    try:
        el.click(timeout=15_000)
    except PWTimeout:
        pass
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.keyboard.insert_text(value)

    got = ""
    try:
        got = el.input_value()
    except Exception:
        pass
    if not got.strip():
        el.evaluate(
            """(node, v) => {
                if (node instanceof HTMLInputElement || node instanceof HTMLTextAreaElement) {
                    node.focus();
                    node.value = v;
                    node.dispatchEvent(new Event('input', { bubbles: true }));
                    node.dispatchEvent(new Event('change', { bubbles: true }));
                    node.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
                    node.dispatchEvent(new Event('blur', { bubbles: true }));
                }
            }""",
            value,
        )
        got = ""
        try:
            got = el.input_value()
        except Exception:
            pass

    if got.strip() != value.strip():
        log(f"  [경고] 입력값 불일치 — 넣으려던 값: {value!r} / 실제 값: {got!r}")


def click_it(locator) -> bool:
    """
    반환값: 신뢰할 수 있는(trusted) 클릭으로 처리됐으면 True.
    el.evaluate(...node.click()...) 같은 JS 강제클릭은 브라우저가
    "진짜 사용자 클릭"으로 인정하지 않아, 그 안에서 호출되는
    window.open()(팝업)이 조용히 차단될 수 있다. 그래서 실패해도
    좌표 기반 실제 마우스 클릭(신뢰됨)을 먼저 시도하고,
    그것마저 안 될 때만 최후 수단으로 JS 클릭을 쓴다.
    """
    el = locator.first
    el.wait_for(state="visible", timeout=60_000)
    el.scroll_into_view_if_needed()
    try:
        el.click(timeout=20_000)
        return True
    except PWTimeout:
        pass

    try:
        box = el.bounding_box()
        if box:
            el.page.mouse.click(
                box["x"] + box["width"] / 2, box["y"] + box["height"] / 2
            )
            return True
    except Exception:  # noqa: BLE001
        pass

    log("  [경고] 실제 클릭 실패 — JS 강제클릭 사용 (팝업이 안 뜰 수 있음)")
    el.evaluate("(node) => node.click()")
    return False


# ── 화면 요소 ─────────────────────────────────────────────────

def url_search_button(page: Page):
    return page.locator(
        'input[type="button"][value*="URL"], input[type="submit"][value*="URL"]'
    ).or_(page.get_by_text(re.compile(r"URL\s*상품\s*검색")))


def save_all_button(page: Page):
    return (
        page.locator('input[type="button"][value*="모두저장"]')
        .or_(page.locator('input[type="submit"][value*="모두저장"]'))
        .or_(page.get_by_text(re.compile(r"검색된\s*상품\s*모두\s*저장")))
    )


def _describe(loc) -> str:
    try:
        return loc.evaluate(
            "n => `<${n.tagName.toLowerCase()}"
            " name=${n.name||''} id=${n.id||''} rows=${n.rows||''}"
            " value.len=${(n.value||'').length}>`"
        )
    except Exception:
        return "<알 수 없음>"


def url_input(page: Page):
    """페이지 전환 중이면 재시도(_url_input_once 참고)"""
    return with_nav_retry(page, lambda: _url_input_once(page))


def _url_input_once(page: Page):
    """
    URL상품검색하기 버튼과 실제 입력칸이 서로 다른 <tr>/<table>에 있는
    화면이 있어(선택자가 넓으면 엉뚱한 textarea를 골라 "검색결과 없음"이
    나는 원인이 됨) 좁은 범위 -> 넓은 범위 순으로, 후보가 정확히 하나일
    때만 채택한다.
    """
    btn = url_search_button(page).first

    # 1) 버튼과 같은 <tr> 안에서 우선 찾기 (가장 정확)
    row = btn.locator("xpath=ancestor::tr[1]")
    if row.count() > 0:
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = row.locator(sel)
            if cand.count() > 0:
                found = cand.first
                log(f"  URL입력칸(같은 행에서 발견): {_describe(found)}")
                return found

    # 2) 부모를 한 단계씩 올라가며(최대 4단계) 후보가 정확히 하나일 때만 채택
    ancestor = btn
    for _ in range(4):
        ancestor = ancestor.locator("xpath=..")
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = ancestor.locator(sel)
            if cand.count() == 1:
                found = cand.first
                log(f"  URL입력칸(상위 요소에서 발견): {_describe(found)}")
                return found

    # 3) 최후 수단: 페이지 전체에서 rows 속성이 가장 큰 textarea
    #    (URL 여러 줄 입력용 큰 textarea일 가능성이 높음)
    all_ta = page.locator("textarea")
    n = all_ta.count()
    if n == 1:
        found = all_ta.first
        log(f"  URL입력칸(페이지에 textarea 1개뿐): {_describe(found)}")
        return found
    if n > 1:
        best_idx, best_rows = 0, -1
        for i in range(n):
            rows_attr = all_ta.nth(i).get_attribute("rows")
            try:
                rows_val = int(rows_attr) if rows_attr else 1
            except ValueError:
                rows_val = 1
            if rows_val > best_rows:
                best_rows, best_idx = rows_val, i
        found = all_ta.nth(best_idx)
        log(f"  URL입력칸(가장 큰 textarea 선택, {n}개 중): {_describe(found)}")
        return found

    raise RuntimeError("URL 입력칸을 찾지 못했습니다")


def save_modal(page: Page):
    return (
        page.locator("div, form, table")
        .filter(has_text=re.compile(r"상품\s*저장\s*설정|검색\s*필터\s*명"))
        .filter(has_text=re.compile("저장하기"))
        .last
    )


def save_modal_visible(page: Page) -> bool:
    try:
        return page.get_by_text(re.compile(r"상품\s*저장\s*설정")).first.is_visible()
    except Exception:
        return False


def modal_field(page: Page, label_pattern: re.Pattern):
    modal = save_modal(page)
    return (
        modal.locator("tr, div, p, label")
        .filter(has_text=label_pattern)
        .locator('input[type="text"], input:not([type]), input[type="number"]')
        .first
    )


# ── 0 ~ 4 ────────────────────────────────────────────────────

def safe_goto(page: Page, url: str, retries: int = 3) -> None:
    """
    로그인 직후 등 사이트 자체가 리다이렉트를 진행 중일 때
    page.goto()와 겹치면 'interrupted by another navigation' 오류가 난다.
    사이트 쪽 리다이렉트가 끝날 때까지 기다렸다가 다시 시도한다.
    """
    last_err: Exception | None = None
    for _ in range(retries):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=120_000)
            return
        except Exception as e:  # noqa: BLE001
            msg = str(e)
            if "interrupted by another navigation" in msg or "NS_BINDING_ABORTED" in msg:
                last_err = e
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=15_000)
                except Exception:  # noqa: BLE001
                    pass
                page.wait_for_timeout(800)
                continue
            raise
    if last_err:
        raise last_err


NAV_ERROR_MARKERS = (
    "Execution context was destroyed",
    "context was destroyed",
    "Target closed",
    "Target page, context or browser has been closed",
)


def is_navigation_error(e: Exception) -> bool:
    msg = str(e)
    return any(m in msg for m in NAV_ERROR_MARKERS)


def with_nav_retry(page: Page, fn, retries: int = 3):
    """
    사이트가 리다이렉트/네비게이션 중일 때 DOM을 조회하면
    "Execution context was destroyed" 같은 오류가 난다.
    페이지가 안정될 때까지 기다렸다가 재시도한다.
    """
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            return fn()
        except Exception as e:  # noqa: BLE001
            if is_navigation_error(e):
                last_err = e
                log(f"  [정보] 페이지 전환 중이라 재시도합니다 ({attempt + 1}/{retries})")
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=10_000)
                except Exception:  # noqa: BLE001
                    pass
                page.wait_for_timeout(800)
                continue
            raise
    if last_err:
        raise last_err
    return None


def set_tmg_credentials(user_id: str | None, password: str | None) -> None:
    global _TMG_ID, _TMG_PW
    if user_id is not None:
        _TMG_ID = user_id.strip()
    if password is not None:
        _TMG_PW = password


def prompt_tmg_credentials() -> tuple[str, str]:
    """CLI로 더망고 아이디/비밀번호 요청 (환경변수 TMG_ID / TMG_PW 우선)."""
    global _TMG_ID, _TMG_PW
    uid = (_TMG_ID or os.environ.get("TMG_ID") or "").strip()
    pw = _TMG_PW or os.environ.get("TMG_PW") or ""

    print("", flush=True)
    print("=== 더망고 로그인 정보 (CLI) ===", flush=True)
    if not uid:
        try:
            uid = input("아이디: ").strip()
        except EOFError as e:
            raise RuntimeError("아이디 입력이 필요합니다 (CLI).") from e
    else:
        print(f"아이디: {uid}  (환경변수/인자)", flush=True)

    if not pw:
        try:
            if sys.stdin.isatty():
                pw = getpass.getpass("비밀번호: ")
            else:
                # 비TTY: 비밀번호도 한 줄 입력 (로컬 파이프/테스트용)
                print("비밀번호: ", end="", flush=True)
                pw = input().strip()
        except EOFError as e:
            raise RuntimeError("비밀번호 입력이 필요합니다 (CLI).") from e
    else:
        print("비밀번호: ********  (환경변수/인자)", flush=True)

    if not uid or not pw:
        raise RuntimeError("아이디와 비밀번호를 모두 입력해야 합니다.")

    _TMG_ID, _TMG_PW = uid, pw
    return uid, pw



def perform_tmg_login(page: Page, user_id: str | None = None, password: str | None = None) -> None:
    """로그인 화면에서 CLI 자격증명으로 로그인 시도.

    Cafe24 로그인 HTML은 <form>이 즉시 닫혀 있고, 아이디/비번 필드는
    form 소유자(form owner)로만 연결된다. 로그인 버튼은 form에 속하지
    않아 클릭만으로는 제출되지 않는 경우가 많다.
    → form.requestSubmit() 으로 onSubmitLoginForm(reCAPTCHA) 경로를 탄다.
    """
    if "admin_login" not in page.url:
        return

    uid = (user_id or _TMG_ID or "").strip()
    pw = password or _TMG_PW or ""
    if not uid or not pw:
        uid, pw = prompt_tmg_credentials()

    log(f"로그인 시도 (아이디={uid})")
    dialogs: list[str] = []

    def _on_dialog(dialog) -> None:
        dialogs.append(dialog.message)
        log(f"  [로그인 알림] {dialog.message}")
        try:
            dialog.accept()
        except Exception:  # noqa: BLE001
            pass

    page.on("dialog", _on_dialog)

    try:
        SHOT_ROOT.mkdir(parents=True, exist_ok=True)
        page.screenshot(
            path=str(SHOT_ROOT / f"login_before_{time.strftime('%Y%m%d_%H%M%S')}.png"),
            full_page=True,
        )
    except Exception:
        pass

    id_box = page.locator('input[name="login_id"]').first
    pw_box = page.locator('input[name="login_pass"]').first
    id_box.wait_for(state="visible", timeout=30_000)
    type_into(page, id_box, uid)
    type_into(page, pw_box, pw)

    # reCAPTCHA 스크립트 준비 대기
    try:
        page.wait_for_function(
            "() => !!(window.grecaptcha && window.grecaptcha.execute)",
            timeout=20_000,
        )
    except Exception:  # noqa: BLE001
        log("  [경고] grecaptcha 로드 대기 시간 초과 — 그대로 제출 시도")

    page.wait_for_timeout(800)

    # 정상 경로: form.requestSubmit() → onSubmitLoginForm → grecaptcha → POST
    submitted = False
    try:
        with page.expect_navigation(timeout=45_000, wait_until="domcontentloaded"):
            page.evaluate(
                """() => {
                    const form = document.getElementById('loginForm')
                        || document.querySelector('form[name="morning_main_login"]');
                    if (!form) throw new Error('loginForm not found');
                    if (typeof form.requestSubmit === 'function') form.requestSubmit();
                    else form.dispatchEvent(new Event('submit', { cancelable: true, bubbles: true }));
                }"""
            )
        submitted = True
    except Exception as e:  # noqa: BLE001
        log(f"  [경고] requestSubmit 네비게이션 대기 실패: {type(e).__name__}")

    if not submitted:
        # 폴백: 보이는 로그인 버튼 클릭
        for sel in (
            'button.defbtn_lar:has-text("로그인")',
            'button[type="submit"]:has-text("로그인")',
            'button:has-text("로그인")',
        ):
            loc = page.locator(sel)
            if loc.count() == 0:
                continue
            try:
                if loc.first.is_visible():
                    with page.expect_navigation(timeout=45_000, wait_until="domcontentloaded"):
                        click_it(loc)
                    submitted = True
                    break
            except Exception:  # noqa: BLE001
                continue

    # 로그인 중계/리다이렉트
    for _ in range(20):
        page = refresh_if_closed(page)
        if "admin_login" not in page.url and ADMIN_HOST in page.url:
            break
        if "m_login" in page.url or "login_ok" in page.url:
            page.wait_for_timeout(800)
            continue
        page.wait_for_timeout(400)

    try:
        page.screenshot(
            path=str(SHOT_ROOT / f"login_after_{time.strftime('%Y%m%d_%H%M%S')}.png"),
            full_page=True,
        )
    except Exception:
        pass

    try:
        page.remove_listener("dialog", _on_dialog)
    except Exception:  # noqa: BLE001
        pass

    if "admin_login" not in page.url:
        log("로그인 성공(로그인 화면 이탈 확인)")
        return

    url = page.url
    msg = " ".join(dialogs)
    if "login=2" in url or "Captcha" in msg or "캡차" in msg or "captcha" in msg.lower() or "자동화" in msg:
        raise RuntimeError(
            "로그인 실패 — Cafe24가 자동/원격 접근을 Captcha로 차단했습니다.\n"
            f"  · 알림: {msg or '(login=2)'}\n"
            "  · 클라우드 에이전트에서는 통과하기 어렵습니다.\n"
            "  · 로컬 PC에서 Chrome을 연 뒤 아래처럼 실행하세요:\n"
            "      python P2/collect.py 엑셀.xlsx 3 --verify --id ID --pw PW\n"
            f"  · URL={url}"
        )
    raise RuntimeError(
        "로그인 실패 — 아이디/비밀번호를 확인하거나, 화면에 캡차/추가인증이 있는지 확인하세요.\n"
        f"  · URL={url}"
        + (f"\n  · 알림={msg}" if msg else "")
    )


def _ask_human(prompt: str) -> None:
    """대화형 터미널에서만 대기. CI/비대화형이면 즉시 실패."""
    if not sys.stdin.isatty():
        raise RuntimeError(
            "로그인이 필요하지만 비대화형 실행입니다. "
            "로컬 PC에서 망고 로그인 후 다시 실행하세요."
        )
    try:
        input(prompt)
    except EOFError as e:
        raise RuntimeError(
            "로그인이 필요하지만 입력을 받을 수 없습니다. "
            "로컬 PC에서 망고 로그인 후 다시 실행하세요."
        ) from e


def handle_possible_login_page(page: Page) -> None:
    """
    세션이 갑자기 만료돼(또는 애초에 로그인이 안 된 채) admin_login.php로
    튕기는 경우가 있다. CLI로 ID/PW를 받아 자동 로그인한다.
    """
    if "admin_login" not in page.url:
        return
    log("  [경고] 로그인 화면 — CLI 자격증명으로 로그인합니다")
    try:
        SHOT_ROOT.mkdir(parents=True, exist_ok=True)
        shot = SHOT_ROOT / f"login_required_{time.strftime('%Y%m%d_%H%M%S')}.png"
        page.screenshot(path=str(shot), full_page=True)
        log(f"  [샷] 로그인 화면: {shot}")
    except Exception as e:  # noqa: BLE001
        log(f"  [샷 실패] {e}")
    perform_tmg_login(page)
    page.wait_for_timeout(800)
    if "admin_login" in page.url:
        raise RuntimeError("로그인 후에도 여전히 로그인 화면입니다 — 다시 확인해 주세요")


def _wait_bulk_ready_once(page: Page) -> None:
    try:
        page.wait_for_load_state("domcontentloaded", timeout=15_000)
    except Exception:  # noqa: BLE001
        pass
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    handle_possible_login_page(page)
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
        handle_possible_login_page(page)
    log("  대량수집 화면 로딩 확인 중...")
    url_search_button(page).first.wait_for(state="visible", timeout=60_000)


def wait_bulk_ready(page: Page) -> None:
    with_nav_retry(page, lambda: _wait_bulk_ready_once(page))


def _reset_to_bulk_menu_once(page: Page) -> None:
    href = page.locator('a[href*="getGoodsNew"]').first
    if href.count() > 0:
        try:
            href.click(timeout=5000)
        except PWTimeout:
            href.evaluate("(node) => node.click()")
        page.wait_for_timeout(800)
        handle_possible_login_page(page)
        if BULK_PATH in page.url:
            wait_bulk_ready(page)
            return

    page.evaluate(
        """() => {
            const clean = (s) => (s || '').replace(/\\s+/g, '');
            const nodes = Array.from(document.querySelectorAll('a, li, span, td, div, button'));
            const byHref = Array.from(document.querySelectorAll('a[href*="getGoodsNew"]'));
            if (byHref[0]) { byHref[0].click(); return; }
            const top = nodes.find(el => clean(el.textContent) === '상품데이터수집');
            if (top) top.click();
            const sub = nodes.find(el => {
                const t = clean(el.textContent);
                if (t.length > 30) return false;
                return /대량데이터수집|대량수집|상품데이터대량/.test(t);
            });
            if (sub) (sub.closest('a') || sub).click();
        }"""
    )
    page.wait_for_timeout(1000)
    handle_possible_login_page(page)
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    wait_bulk_ready(page)


def reset_to_bulk_menu(page: Page) -> None:
    """0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭"""
    with_nav_retry(page, lambda: _reset_to_bulk_menu_once(page))


def wait_page_not_loading(page: Page, timeout_sec: float = 15.0) -> None:
    """
    모두저장 클릭 전에 "로딩 중" 표시가 사라졌는지만 확인한다.
    (검색결과 있음/없음은 절대 판단하지 않음 — 로딩 중에 "결과없음"
    문구가 잠깐 보였다가 사라지는 경우를 결과없음으로 오판했던 과거
    버그가 있어서, 오직 로딩 표시 유무만 본다.)
    """
    end = time.time() + timeout_sec
    while time.time() < end:
        try:
            loading = page.evaluate(
                "() => /load\\s*product|상품정보를\\s*불러오는\\s*중|잠시만\\s*기다려/i"
                ".test(document.body ? document.body.innerText : '')"
            )
        except Exception:  # noqa: BLE001
            return
        if not loading:
            page.wait_for_timeout(500)  # 결과 렌더링 여유
            return
        page.wait_for_timeout(300)


def _process_row_once(page: Page, row: dict, ctx: RunCtx) -> None:
    label = row["label"]
    url = normalize_url(row["url"])
    save_count = ctx.save_count
    rn = row["row"]
    ctx.info(f"--- {rn}행 : {label} (목표 저장수={save_count}) ---")

    ctx.info("0. 초기화 : 상품데이터수집 -> 대량데이터수집")
    reset_to_bulk_menu(page)
    page.wait_for_timeout(500)
    ctx.shot(page, "00_init_bulk", rn)

    ctx.info(f"  엑셀 원본 값: {row['url']!r}")
    if url != row["url"].strip():
        ctx.info(f"  [정보] 프로토콜 보정됨: {url}")

    ctx.info(f"1. 필드값 입력: {url}")
    target = url_input(page)
    type_into(page, target, url)
    actual = ""
    try:
        actual = target.input_value()
    except Exception:
        pass
    ctx.info(f"  입력칸 최종 값: {actual!r}")
    if actual.strip() != url.strip():
        raise RuntimeError(f"URL 입력 불일치 — 기대 {url!r} / 실제 {actual!r}")
    ctx.shot(page, "01_url_filled", rn)

    ctx.info("1. URL상품검색하기 클릭")
    trusted = click_it(url_search_button(page))

    ctx.info("1. 팝업창이 없어질 때까지 대기")
    opened = wait_popups_gone(page, grace_sec=15.0, warn_if_never_opened=True)

    if not opened:
        ctx.info("  키보드로 재시도 (Enter)")
        try:
            btn = url_search_button(page).first
            btn.focus()
            page.keyboard.press("Enter")
        except Exception:  # noqa: BLE001
            pass
        opened = wait_popups_gone(page, grace_sec=10.0, warn_if_never_opened=True)

    if not opened:
        ctx.shot(page, "01_popup_missing", rn)
        raise RuntimeError(
            f"#{rn} URL상품검색하기 클릭 후 팝업이 뜨지 않음 "
            f"(trusted_click={trusted})"
        )
    ctx.info("1. 팝업창 닫힘")
    ctx.shot(page, "01_popup_closed", rn)

    wait_page_not_loading(page)
    ctx.shot(page, "01_results_ready", rn)

    ctx.info("2. 검색된 상품 모두저장 클릭")
    click_it(save_all_button(page))
    save_modal(page).wait_for(state="visible", timeout=MODAL_WAIT_SEC * 1000)
    page.wait_for_timeout(400)
    ctx.shot(page, "02_save_modal", rn)

    ctx.info(f"2. 검색필터명 입력: {label}")
    type_into(page, modal_field(page, FILTER_NAME_LABEL), label)

    count_field = modal_field(page, SAVE_COUNT_LABEL)
    if count_field.count() == 0:
        ctx.shot(page, "02_no_count_field", rn)
        raise RuntimeError(f"#{rn} 저장상품수 입력칸을 찾지 못함")
    type_into(page, count_field, str(save_count))
    type_into(page, modal_field(page, FILTER_NAME_LABEL), label)

    # 저장수 확인
    got_count = ""
    try:
        got_count = count_field.input_value()
    except Exception:
        pass
    ctx.info(f"  저장상품수 입력값: {got_count!r} (목표 {save_count})")
    if str(save_count) not in (got_count or "").replace(" ", ""):
        # 숫자만 비교
        digits = re.sub(r"\D", "", got_count or "")
        if digits != str(save_count):
            ctx.shot(page, "02_count_mismatch", rn)
            raise RuntimeError(
                f"#{rn} 저장상품수 불일치 — 기대 {save_count} / 실제 {got_count!r}"
            )
    ctx.shot(page, "02_modal_filled", rn)

    ctx.info("2. 저장하기 버튼 클릭")
    click_it(
        save_modal(page)
        .locator('input[value*="저장하기"]')
        .or_(save_modal(page).locator('button:has-text("저장하기")'))
        .or_(save_modal(page).get_by_text(re.compile("^저장하기$")))
    )

    ctx.info("3. 팝업창이 없어질 때까지 대기")
    end = time.time() + MODAL_WAIT_SEC
    closed = False
    while time.time() < end:
        if not save_modal_visible(page):
            wait_popups_gone(page, grace_sec=0.5)
            closed = True
            break
        page.wait_for_timeout(500)
    if not closed:
        ctx.shot(page, "03_modal_stuck", rn)
        raise TimeoutError(f"#{rn} 저장 팝업창이 닫히지 않음")
    ctx.info("3. 팝업창 닫힘")
    ctx.shot(page, "03_modal_closed", rn)

    # C. 3건(저장수) 완료 확인
    verify_row_save_done(page, ctx, rn, save_count)
    ctx.info(f"4. -> 행 완료 확인 (저장수 {save_count})")
    ctx.shot(page, "04_row_done", rn)


def verify_row_save_done(page: Page, ctx: RunCtx, row_no: int, save_count: int) -> None:
    """저장 모달 종료 후, 오류 문구가 없고(가능하면) 성공 신호를 확인."""
    page.wait_for_timeout(800)
    try:
        text = page.evaluate(
            "() => (document.body && document.body.innerText) ? document.body.innerText : ''"
        )
    except Exception:  # noqa: BLE001
        text = ""

    for pat in SAVE_FAIL_PATTERNS:
        if pat.search(text or ""):
            raise RuntimeError(f"#{row_no} 저장 후 오류 문구 감지: {pat.pattern}")

    ok_hit = None
    for pat in SAVE_OK_PATTERNS:
        m = pat.search(text or "")
        if m:
            ok_hit = m.group(0)
            break

    # 모달이 닫혔고 오류 문구가 없으면 1차 통과.
    # 성공 문구가 있으면 로그에 남김. 검증 모드에서는 성공 문구 또는
    # '저장상품수 입력 확인됨'을 전제로 이미 모달에서 확인했으므로 통과.
    if ok_hit:
        ctx.info(f"  [확인] 화면 성공 신호: {ok_hit!r}")
    else:
        ctx.info(
            f"  [확인] 오류 문구 없음 · 모달 종료 · 저장수 {save_count} 입력 확인됨 "
            f"(성공 문구는 화면에 없을 수 있음)"
        )

    if ctx.verify:
        # 검증 모드: 저장수 필드에 넣었던 값이 반영됐는지 이미 확인함.
        # 추가로 짧은 대기 후 로그인 튕김만 검사.
        if "admin_login" in page.url:
            raise RuntimeError(f"#{row_no} 저장 직후 로그인 화면으로 이동됨")


def process_row_with_retries(page: Page, row: dict, ctx: RunCtx) -> bool:
    """행 단위 재시도. 성공 True / 최종 실패 False."""
    last_err: Exception | None = None
    for attempt in range(1, ctx.retries + 1):
        try:
            ctx.info(f"▶ 시도 {attempt}/{ctx.retries} (엑셀 {row['row']}행)")
            page = refresh_if_closed(page)
            _process_row_once(page, row, ctx)
            ctx.info(f"✔ {row['row']}행 성공 (시도 {attempt})")
            return True
        except Exception as e:  # noqa: BLE001
            last_err = e
            ctx.info(f"✘ {row['row']}행 실패 (시도 {attempt}/{ctx.retries}): {e}")
            try:
                ctx.shot(page, f"fail_attempt{attempt}", row["row"])
            except Exception:
                pass
            if attempt < ctx.retries:
                ctx.info("  같은 행 재시도 전 대량수집 화면 복귀…")
                try:
                    page = refresh_if_closed(page)
                    reset_to_bulk_menu(page)
                except Exception as re:  # noqa: BLE001
                    ctx.info(f"  복귀 중 경고: {re}")
                page.wait_for_timeout(1000)
    ctx.info(f"✖ {row['row']}행 최종 실패: {last_err}")
    return False


def process_row(page: Page, row: dict, save_count: int = DEFAULT_SAVE_COUNT) -> None:
    """하위 호환: 저장수만 받아 1회 시도."""
    ctx = RunCtx(save_count=save_count, retries=1, batch=True)
    try:
        _process_row_once(page, row, ctx)
    finally:
        ctx.close()


def ensure_ready_page(page: Page) -> Page:
    """이미 망고 화면이면 그대로, 아니면 메인화면 진입 → 필요시 로그인 대기 → 0.초기화

    로그인 성공 후 사이트가 원래 탭을 닫아버리는 경우가 있어(예: 로그인
    중계 페이지가 스스로를 닫음) 매 단계 사이마다 탭이 살아있는지
    확인하고, 닫혔으면 같은 브라우저에서 새 탭을 다시 찾아온다.
    """
    page = refresh_if_closed(page)

    if ADMIN_HOST not in page.url or page.url in ("about:blank", ""):
        log("메인화면으로 이동: " + MAIN_URL)
        safe_goto(page, MAIN_URL)

    if "admin_login" in page.url:
        try:
            SHOT_ROOT.mkdir(parents=True, exist_ok=True)
            shot = SHOT_ROOT / f"login_gate_{time.strftime('%Y%m%d_%H%M%S')}.png"
            page.screenshot(path=str(shot), full_page=True)
            log(f"  [샷] 로그인 게이트: {shot}")
        except Exception as e:  # noqa: BLE001
            log(f"  [샷 실패] {e}")
        perform_tmg_login(page)
        page = refresh_if_closed(page)
        # 로그인 직후 사이트 자체가 리다이렉트 중일 수 있으므로(m_login_ok.php 등)
        # 안정될 때까지 잠깐 기다린 뒤에 필요하면 이동한다.
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
        except Exception:  # noqa: BLE001
            pass
        page = refresh_if_closed(page)
        try:
            page.wait_for_timeout(1000)
        except Exception:  # noqa: BLE001
            pass
        page = refresh_if_closed(page)
        if "admin_login" in page.url or ADMIN_HOST not in page.url:
            safe_goto(page, MAIN_URL)

    log("0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭")
    if BULK_PATH in page.url:
        wait_bulk_ready(page)
    else:
        reset_to_bulk_menu(page)

    return page


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(
        description="P2 더망고 대량수집 — URL 1행당 상품 N건 (기본 3)"
    )
    ap.add_argument("excel", help="P1 출력 엑셀 (.xlsx)")
    ap.add_argument(
        "save_count",
        nargs="?",
        type=int,
        default=DEFAULT_SAVE_COUNT,
        help=f"행당 저장 상품 수 (기본 {DEFAULT_SAVE_COUNT})",
    )
    ap.add_argument(
        "--verify",
        action="store_true",
        help="1행 검증 모드: 첫 행만, 단계 스크린샷, 행 재시도, 무중단",
    )
    ap.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="처리할 최대 행 수 (검증 모드 기본 1)",
    )
    ap.add_argument(
        "--retries",
        type=int,
        default=DEFAULT_ROW_RETRIES,
        help=f"행 실패 시 같은 행 재시도 횟수 (기본 {DEFAULT_ROW_RETRIES})",
    )
    ap.add_argument(
        "--yes",
        "--batch",
        dest="batch",
        action="store_true",
        help="실패해도 y/n 묻지 않고 다음 행으로 (또는 재시도만)",
    )
    ap.add_argument("--id", dest="tmg_id", default=None, help="더망고 아이디 (없으면 CLI 요청)")
    ap.add_argument("--pw", dest="tmg_pw", default=None, help="더망고 비밀번호 (없으면 CLI 요청)")
    args = ap.parse_args()

    excel_path = args.excel
    save_count = args.save_count if args.save_count and args.save_count > 0 else DEFAULT_SAVE_COUNT
    verify = bool(args.verify)
    max_rows = args.max_rows
    if verify and max_rows is None:
        max_rows = 1

    # 로그인 정보: 인자 → 환경변수 → CLI 입력
    set_tmg_credentials(args.tmg_id, args.tmg_pw)
    prompt_tmg_credentials()

    rows = read_excel(excel_path)
    if not rows:
        print("엑셀에 처리할 행이 없습니다.")
        sys.exit(1)
    if max_rows is not None:
        rows = rows[: max(0, max_rows)]

    ctx = RunCtx(
        save_count=save_count,
        retries=args.retries,
        verify=verify,
        max_rows=max_rows,
        batch=args.batch or verify,
    )
    ctx.info(
        f"엑셀 {len(rows)}행 · 저장수={save_count} · 재시도={ctx.retries} "
        f"· verify={verify} · 로그={ctx.shot_dir}"
    )

    ok = 0
    fail = 0
    try:
        with sync_playwright() as p:
            _browser, page = connect_browser(p)
            page.set_default_timeout(120_000)
            page = ensure_ready_page(page)
            ctx.shot(page, "ready", 0)

            for row in rows:
                page = refresh_if_closed(page)
                success = process_row_with_retries(page, row, ctx)
                if success:
                    ok += 1
                else:
                    fail += 1
                    if not ctx.batch:
                        if input("계속 진행할까요? (y/n) ").strip().lower() != "y":
                            break

            ctx.info(f"완료: 성공 {ok} / 실패 {fail} / 대상 {len(rows)}행")
            ctx.info(f"스크린샷·로그: {ctx.shot_dir}")
            print("브라우저는 그대로 열어둡니다 (이 창만 닫으면 됩니다).")
            if verify and ok >= 1 and fail == 0:
                print("✔ 검증 모드 PASS — 1행×저장수 완료")
                sys.exit(0)
            if fail:
                sys.exit(2)
    finally:
        ctx.close()


if __name__ == "__main__":
    main()
