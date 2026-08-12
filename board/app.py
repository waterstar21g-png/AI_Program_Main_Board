"""
AI_Program_Main_Board — Python B안 보드
(P1 / P1_101 / P1_ZARA_DE / P2 / P3_필터_갱신)
아주 단순한 UI, 필요한 기능만.
"""

from __future__ import annotations

import importlib.util
import os
import subprocess
import sys
import threading
import time
import tkinter as tk
import webbrowser
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from types import ModuleType

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "board"))


def _load_crawl_module(mod_name: str, folder: str) -> ModuleType:
    """P1 / P1_ZARA_DE 각각 crawl.py 를 충돌 없이 로드."""
    path = ROOT / folder / "crawl.py"
    spec = importlib.util.spec_from_file_location(mod_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"crawl 모듈 로드 실패: {path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[mod_name] = mod
    spec.loader.exec_module(mod)
    return mod


def _load_py_module(mod_name: str, folder: str, filename: str) -> ModuleType:
    """폴더 내 임의 .py 모듈을 충돌 없이 로드."""
    path = ROOT / folder / filename
    spec = importlib.util.spec_from_file_location(mod_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"모듈 로드 실패: {path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[mod_name] = mod
    spec.loader.exec_module(mod)
    return mod


p1_crawl = _load_crawl_module("p1_crawl", "P1")
p1_zara_crawl = _load_crawl_module("p1_zara_de_crawl", "P1_ZARA_DE")
p1_101_extract = _load_py_module("p1_101_extract", "P1_101", "extract.py")
p3_update = _load_py_module("p3_update_filters", "P3_필터_갱신", "update_filters.py")

TOP_CELL_MAX_LEN = p1_crawl.TOP_CELL_MAX_LEN
TOP_GRID_COLS = p1_crawl.TOP_GRID_COLS
TOP_GRID_ROWS = p1_crawl.TOP_GRID_ROWS
crawl_site = p1_crawl.crawl_site
save_excel = p1_crawl.save_excel
zara_crawl_site = p1_zara_crawl.crawl_site
zara_save_excel = p1_zara_crawl.save_excel

from library import (  # noqa: E402
    add_paths,
    default_roots,
    entries_annotated,
    is_in_library,
    load,
    read_category_url_rows,
    remove_path,
    search_xlsx,
    set_selected,
)
from log_protocol import (  # noqa: E402
    META_FIELDS,
    META_INTERNAL_FIELDS,
    format_meta_line,
    parse_line,
    step_tag,
    strip_timestamp,
    sub_time_range,
)
from shot_viewer import latest_p3_shot_dir, latest_shot_dir, open_shot_viewer  # noqa: E402
from self_update import (  # noqa: E402
    latest_open_pr_url,
    launch_external_updater,
    local_version,
)

import re  # noqa: E402


def _read_version() -> str:
    """VERSION.txt(저장소 루트)를 단일 소스로 읽는다.

    ★2026-08-08: 여기 하드코딩된 문자열이 VERSION.txt와 따로 놀아서,
    실제 코드는 최신으로 갱신됐는데도 화면 제목의 버전 숫자만 옛날
    그대로 보이는 문제가 있었다 — 다시는 두 값이 따로 놀지 않도록
    VERSION.txt를 직접 읽어온다.
    """
    try:
        text = (ROOT / "VERSION.txt").read_text(encoding="utf-8")
        m = re.search(r"(?:버전|version)\s*([0-9]+(?:\.[0-9]+)+)", text, re.I)
        if not m:
            m = re.search(r"([0-9]+\.[0-9]+\.[0-9]+)", text)
        if m:
            return m.group(1)
    except OSError:
        pass
    return "?"


VERSION = _read_version()
APP_TITLE = "AI_Program_Main_Board"


class BoardApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{APP_TITLE}  v{VERSION}")
        self.geometry("960x720")
        self.minsize(820, 600)
        self.configure(bg="#1a4d5c")

        self._p1_result = None
        self._p1_zara_result = None
        self._p1_101_result = None
        self._p1_101_proc: subprocess.Popen | None = None
        self._p2_proc: subprocess.Popen | None = None
        self._p3_proc: subprocess.Popen | None = None
        self._last_shot_dir: Path | None = None
        self._build()
        self._show("p1")
        self._refresh_p2_list()
        self._refresh_p3_list()

    def _build(self) -> None:
        head = tk.Frame(self, bg="#164a59", pady=10)
        head.pack(fill="x")
        tk.Label(
            head,
            text=APP_TITLE,
            fg="white",
            bg="#164a59",
            font=("Malgun Gothic", 14, "bold"),
        ).pack()
        tk.Label(
            head,
            text="P1 · P1_101 · P1_ZARA_DE · P2 대량수집 · P3_필터_갱신",
            fg="#cbd5e1",
            bg="#164a59",
            font=("Malgun Gothic", 9),
        ).pack()

        body = tk.Frame(self, bg="#1a4d5c")
        body.pack(fill="both", expand=True, padx=8, pady=8)

        side = tk.Frame(body, bg="#d9d9d9", width=180)
        side.pack(side="left", fill="y", padx=(0, 8))
        side.pack_propagate(False)

        tk.Label(
            side,
            text=f"v{VERSION}\n프로그램",
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            pady=8,
        ).pack(fill="x")

        self.btn_p1 = tk.Button(
            side,
            text="P1\n카테고리 URL 추출",
            command=lambda: self._show("p1"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p1.pack(fill="x", padx=6, pady=6)

        self.btn_p1_101 = tk.Button(
            side,
            text="P1_101\n상품수 추출",
            command=lambda: self._show("p1_101"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p1_101.pack(fill="x", padx=6, pady=6)

        self.btn_p1_zara = tk.Button(
            side,
            text="P1_ZARA_DE\n독일자라 URL추출",
            command=lambda: self._show("p1_zara"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p1_zara.pack(fill="x", padx=6, pady=6)

        self.btn_p2 = tk.Button(
            side,
            text="P2\n더망고 대량수집",
            command=lambda: self._show("p2"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p2.pack(fill="x", padx=6, pady=6)

        self.btn_p3 = tk.Button(
            side,
            text="P3_필터_갱신\n수집건수 갱신",
            command=lambda: self._show("p3"),
            font=("Malgun Gothic", 9, "bold"),
            relief="groove",
            pady=10,
        )
        self.btn_p3.pack(fill="x", padx=6, pady=6)

        # ★요건: 좌측 하단 — 머지반영 업데이트 (아이콘/run.bat 대체, 버튼 하나)
        side_bottom = tk.Frame(side, bg="#d9d9d9")
        side_bottom.pack(side="bottom", fill="x", padx=6, pady=(4, 10))
        self.lbl_update_hint = tk.Label(
            side_bottom,
            text="종료 후 강제 버전갱신",
            bg="#d9d9d9",
            fg="#475569",
            font=("Malgun Gothic", 8),
        )
        self.lbl_update_hint.pack(fill="x", pady=(0, 4))
        self.btn_merge_update = tk.Button(
            side_bottom,
            text="머지반영\n업데이트",
            command=self._run_merge_update,
            bg="#0f766e",
            fg="white",
            activebackground="#0d9488",
            activeforeground="white",
            font=("Malgun Gothic", 9, "bold"),
            relief="raised",
            pady=12,
            cursor="hand2",
        )
        self.btn_merge_update.pack(fill="x")
        self.lbl_update_status = tk.Label(
            side_bottom,
            text=f"현재 v{VERSION}\n(또는 바탕화면 버전갱신)",
            bg="#d9d9d9",
            fg="#64748b",
            font=("Malgun Gothic", 7),
            wraplength=160,
            justify="left",
        )
        self.lbl_update_status.pack(fill="x", pady=(4, 0))

        self.main = tk.Frame(body, bg="#f1f5f9")
        self.main.pack(side="left", fill="both", expand=True)

        self.frame_p1 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p1_101 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p1_zara = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p2 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self.frame_p3 = tk.Frame(self.main, bg="#f1f5f9", padx=12, pady=10)
        self._build_p1(self.frame_p1)
        self._build_p1_101(self.frame_p1_101)
        self._build_p1_zara(self.frame_p1_zara)
        self._build_p2(self.frame_p2)
        self._build_p3(self.frame_p3)

    def _show(self, which: str) -> None:
        self.frame_p1.pack_forget()
        self.frame_p1_101.pack_forget()
        self.frame_p1_zara.pack_forget()
        self.frame_p2.pack_forget()
        self.frame_p3.pack_forget()
        self.btn_p1.configure(bg="#ececec")
        self.btn_p1_101.configure(bg="#ececec")
        self.btn_p1_zara.configure(bg="#ececec")
        self.btn_p2.configure(bg="#ececec")
        self.btn_p3.configure(bg="#ececec")
        if which == "p1":
            self.frame_p1.pack(fill="both", expand=True)
            self.btn_p1.configure(bg="#dbeafe")
        elif which == "p1_101":
            self.frame_p1_101.pack(fill="both", expand=True)
            self.btn_p1_101.configure(bg="#dbeafe")
        elif which == "p1_zara":
            self.frame_p1_zara.pack(fill="both", expand=True)
            self.btn_p1_zara.configure(bg="#dbeafe")
        elif which == "p3":
            self.frame_p3.pack(fill="both", expand=True)
            self.btn_p3.configure(bg="#dbeafe")
        else:
            self.frame_p2.pack(fill="both", expand=True)
            self.btn_p2.configure(bg="#dbeafe")

    # ── 좌측 하단: 머지반영 업데이트 ───────────────────
    def _run_merge_update(self) -> None:
        """보드를 종료한 뒤 외부 스크립트로 GitHub main 강제 반영·재시작.

        (실행 중 pull 하면 Windows 파일 잠금으로 버전이 안 바뀌는 문제 방지)
        """
        if getattr(self, "_merge_update_busy", False):
            messagebox.showinfo("안내", "이미 업데이트를 진행 중입니다.")
            return

        pr_url = ""
        try:
            pr_url = latest_open_pr_url()
        except Exception:
            pr_url = f"https://github.com/waterstar21g-png/AI_Program_Main_Board/pulls"

        cur = local_version(ROOT) or VERSION
        msg = (
            "보드를 종료한 뒤 GitHub main 을 강제 반영하고 재시작합니다.\n\n"
            f"현재 버전: v{cur}\n\n"
            "아직 PR 머지 전이면 아래 머지 URL에서 먼저 머지하세요.\n"
            f"{pr_url}\n\n"
            "계속할까요?"
        )
        if not messagebox.askyesno("머지반영 업데이트", msg, parent=self):
            return

        try:
            if pr_url.startswith("http"):
                webbrowser.open(pr_url)
        except Exception:
            pass

        self._merge_update_busy = True
        self.btn_merge_update.configure(state="disabled", text="종료 후 갱신…")
        self.lbl_update_status.configure(
            text="보드 종료 → 강제 버전갱신 → 재시작",
            fg="#0f172a",
        )

        # 실행 중 수집 중단
        try:
            if self._p1_101_proc and self._p1_101_proc.poll() is None:
                self._p1_101_stop_flag().write_text("stop\n", encoding="utf-8")
                self._p1_101_proc.terminate()
        except Exception:
            pass
        try:
            if self._p2_proc and self._p2_proc.poll() is None:
                self._stop_flag_path().write_text("stop\n", encoding="utf-8")
                self._p2_proc.terminate()
        except Exception:
            pass
        try:
            if self._p3_proc and self._p3_proc.poll() is None:
                self._p3_stop_flag().write_text("stop\n", encoding="utf-8")
                self._p3_proc.terminate()
        except Exception:
            pass

        ok, detail = launch_external_updater(ROOT, wait_pid=os.getpid())
        if not ok:
            self._merge_update_busy = False
            self.btn_merge_update.configure(state="normal", text="머지반영\n업데이트")
            self.lbl_update_status.configure(text="업데이터 실행 실패", fg="#b91c1c")
            messagebox.showerror(
                "업데이트 실패",
                "외부 버전갱신 실행에 실패했습니다.\n\n"
                f"{detail}\n\n"
                "바탕화면 'AI_보드_버전갱신' 아이콘을 사용하세요.\n"
                f"머지 URL:\n{pr_url}",
                parent=self,
            )
            return

        messagebox.showinfo(
            "버전갱신 시작",
            "보드를 종료합니다.\n\n"
            "이어서 자동으로:\n"
            "1) GitHub main 강제 반영\n"
            "2) 보드 재시작\n\n"
            f"(실패 시 바탕화면 'AI_보드_버전갱신' 아이콘)\n"
            f"머지 URL:\n{pr_url}",
            parent=self,
        )
        try:
            self.destroy()
        except Exception:
            pass
        sys.exit(0)

    # ── P1 ─────────────────────────────────────────────
    def _build_p1(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P1 — 사이트·상위 카테고리 → 엑셀 (P2 입력용)",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 8))

        form = tk.Frame(parent, bg="#ffffff", padx=10, pady=10, relief="solid", bd=1)
        form.pack(fill="x")

        self.var_site = tk.StringVar(value=p1_crawl.DEFAULT_SITE)
        self.var_url = tk.StringVar(value=p1_crawl.DEFAULT_URL)
        # ★요건: 기본 저장 폴더
        self.var_outdir = tk.StringVar(value=p1_crawl.DEFAULT_OUTDIR)

        self._row(form, "사이트명", self.var_site)
        self._row(form, "사이트 URL", self.var_url)

        # 상위 카테고리: 3행 × 10칸 입력 그리드 (칸당 한글 15자, 명1:명2=엑셀치환)
        tops_wrap = tk.Frame(form, bg="#ffffff")
        tops_wrap.pack(fill="x", pady=3)
        tk.Label(
            tops_wrap,
            text="상위 카테고리",
            width=16,
            anchor="nw",
            bg="#ffffff",
        ).pack(side="left", anchor="n", pady=2)
        tops_right = tk.Frame(tops_wrap, bg="#ffffff")
        tops_right.pack(side="left", fill="x", expand=True)
        tk.Label(
            tops_right,
            text=(
                f"{TOP_GRID_ROWS}행×{TOP_GRID_COLS}칸 · 칸당 {TOP_CELL_MAX_LEN}자 · "
                "명1:명2 입력 시 엑셀 상위명을 명2로 출력"
            ),
            bg="#ffffff",
            fg="#64748b",
            anchor="w",
            font=("Malgun Gothic", 8),
        ).pack(fill="x", pady=(0, 2))
        self._p1_top_vars: list[tk.StringVar] = []
        vcmd = (self.register(self._validate_p1_top_cell), "%P")
        grid = tk.Frame(tops_right, bg="#ffffff")
        grid.pack(fill="x")
        # ★요건: MEN:남성, WOMEN:여성, KIDS:키즈
        defaults = list(p1_crawl.DEFAULT_TOPS)
        idx = 0
        for r in range(TOP_GRID_ROWS):
            row_f = tk.Frame(grid, bg="#ffffff")
            row_f.pack(fill="x", pady=1)
            for c in range(TOP_GRID_COLS):
                var = tk.StringVar(value=defaults[idx] if idx < len(defaults) else "")
                self._p1_top_vars.append(var)
                tk.Entry(
                    row_f,
                    textvariable=var,
                    width=TOP_CELL_MAX_LEN + 1,
                    font=("Malgun Gothic", 9),
                    justify="center",
                    validate="key",
                    validatecommand=vcmd,
                ).pack(side="left", padx=1)
                idx += 1

        out_row = tk.Frame(form, bg="#ffffff")
        out_row.pack(fill="x", pady=3)
        tk.Label(out_row, text="저장 폴더", width=16, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(out_row, textvariable=self.var_outdir).pack(side="left", fill="x", expand=True)
        tk.Button(out_row, text="…", width=3, command=self._pick_outdir).pack(side="left", padx=4)

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=10)
        self.btn_crawl = tk.Button(
            actions,
            text="1. 수집 시작",
            command=self._run_p1,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=6,
        )
        self.btn_crawl.pack(side="left")
        self.btn_save = tk.Button(
            actions,
            text="2. 엑셀 저장",
            command=self._save_p1,
            state="disabled",
            padx=12,
            pady=6,
        )
        self.btn_save.pack(side="left", padx=8)
        tk.Button(actions, text="ABC 기본값", command=self._p1_defaults).pack(side="left")

        self.p1_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w", justify="left")
        self.p1_status.pack(fill="x", pady=4)

        self.p1_preview = tk.Text(parent, height=14, font=("Consolas", 9), wrap="none")
        self.p1_preview.pack(fill="both", expand=True)

    def _row(self, parent: tk.Frame, label: str, var: tk.StringVar) -> None:
        row = tk.Frame(parent, bg="#ffffff")
        row.pack(fill="x", pady=3)
        tk.Label(row, text=label, width=16, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(row, textvariable=var).pack(side="left", fill="x", expand=True)

    def _validate_p1_top_cell(self, new_value: str) -> bool:
        """상위 카테고리 칸 — 한글 포함 최대 TOP_CELL_MAX_LEN자."""
        return len(new_value) <= TOP_CELL_MAX_LEN

    def _validate_p1_zara_url_cell(self, new_value: str) -> bool:
        """P1_ZARA_DE 하위 카테고리 URL 칸."""
        return len(new_value) <= p1_zara_crawl.URL_CELL_MAX_LEN

    def _p1_top_values(self) -> list[str]:
        """그리드에서 비어 있지 않은 상위 카테고리 칸 값을 순서대로."""
        out: list[str] = []
        for var in getattr(self, "_p1_top_vars", []):
            s = (var.get() or "").strip()
            if s:
                out.append(s)
        return out

    def _pick_outdir(self) -> None:
        d = filedialog.askdirectory(initialdir=self.var_outdir.get() or str(Path.home()))
        if d:
            self.var_outdir.set(d)

    def _p1_defaults(self) -> None:
        self.var_site.set(p1_crawl.DEFAULT_SITE)
        self.var_url.set(p1_crawl.DEFAULT_URL)
        self.var_outdir.set(p1_crawl.DEFAULT_OUTDIR)
        defaults = list(p1_crawl.DEFAULT_TOPS)
        for i, var in enumerate(getattr(self, "_p1_top_vars", [])):
            var.set(defaults[i] if i < len(defaults) else "")

    def _run_p1(self) -> None:
        self.btn_crawl.configure(state="disabled")
        self.btn_save.configure(state="disabled")
        self.p1_status.configure(text="수집 중…")
        self.p1_preview.delete("1.0", "end")
        tops = self._p1_top_values()

        def work() -> None:
            result = crawl_site(self.var_site.get(), self.var_url.get(), tops)
            self.after(0, lambda: self._p1_done(result))

        threading.Thread(target=work, daemon=True).start()

    def _p1_done(self, result) -> None:
        self.btn_crawl.configure(state="normal")
        self._p1_result = result
        if not result.ok:
            self.p1_status.configure(text="실패: " + "; ".join(result.errors), fg="#b91c1c")
            return
        self.btn_save.configure(state="normal")
        msg = f"완료 · {result.platform} · {result.total}건"
        if result.warnings:
            msg += " · " + " / ".join(result.warnings)
        self.p1_status.configure(text=msg, fg="#15803d")
        lines = ["상위 | 중위 | 하위 | 최종 | 상위최종 | URL", "-" * 80]
        for r in result.rows[:80]:
            lines.append(
                f"{r.top} | {r.mid or '—'} | {r.low or '—'} | {r.final} | "
                f"{r.top_final_label} | {r.final_category_url} | "
                f"총{getattr(r, 'total_product_count', 0)} | "
                f"수집가능{getattr(r, 'collectible_count', 0)} | "
                f"검색{getattr(r, 'search_count', 0)} | "
                f"리뷰{getattr(r, 'review_count', 0)}"
            )
        if result.total > 80:
            lines.append(f"… 외 {result.total - 80}행 (엑셀에 전체 포함)")
        self.p1_preview.insert("1.0", "\n".join(lines))

    def _save_p1(self) -> None:
        if not self._p1_result or not self._p1_result.ok:
            return
        try:
            path = save_excel(self._p1_result.rows, self._p1_result.site_name, self.var_outdir.get())
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))
            return
        self.p1_status.configure(text=f"저장됨: {path}", fg="#15803d")
        # P2 디렉터리목록·카테고리URL목록에 바로 반영
        add_paths([str(path)])
        try:
            self.var_dir.set(str(Path(path).parent))
        except Exception:
            pass
        self._refresh_p2_list()
        self._load_category_url_list(str(path))
        if messagebox.askyesno("P2로 이동", f"엑셀 저장·카테고리URL목록에 반영했습니다.\n\n{path}\n\nP2 화면으로 갈까요?"):
            self._show("p2")

    # ── P1_101 상품수 추출 ─────────────────────────────
    def _build_p1_101(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P1_101 — 엑셀 URL → 팝업닫기 → 카드갯수 → 출력폴더 저장",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 8))

        form = tk.Frame(parent, bg="#ffffff", padx=10, pady=10, relief="solid", bd=1)
        form.pack(fill="x")

        # 메인보드에서 바로 실행 — 최근 선택 엑셀을 기본값으로
        initial_excel = ""
        try:
            last = str(load().get("last_selected") or "").strip()
            if last and Path(last).is_file():
                initial_excel = last
        except Exception:
            pass
        self.var_p1_101_excel = tk.StringVar(value=initial_excel)
        file_row = tk.Frame(form, bg="#ffffff")
        file_row.pack(fill="x", pady=3)
        tk.Label(
            file_row,
            text="엑셀자료 파일지정",
            width=16,
            anchor="w",
            bg="#ffffff",
        ).pack(side="left")
        tk.Entry(file_row, textvariable=self.var_p1_101_excel).pack(
            side="left", fill="x", expand=True
        )
        tk.Button(file_row, text="…", width=3, command=self._pick_p1_101_excel).pack(
            side="left", padx=4
        )

        tk.Label(
            form,
            text=(
                "메인보드 직접실행 · URL 열기 → 팝업 닫기 → "
                f"{p1_101_extract.POST_POPUP_WAIT_SEC:g}초 대기 → 상품수 수집 → "
                r"출력: P2_INPUT_건수집계 (파일명_상품수_v버전_시각.xlsx)"
            ),
            bg="#ffffff",
            fg="#64748b",
            anchor="w",
            font=("Malgun Gothic", 8),
            justify="left",
        ).pack(fill="x", pady=(6, 0))

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=10)
        self.btn_p1_101_run = tk.Button(
            actions,
            text="▶ 상품수 추출 실행",
            command=self._run_p1_101,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=6,
        )
        self.btn_p1_101_run.pack(side="left")
        self.btn_p1_101_stop = tk.Button(
            actions,
            text="중단",
            command=self._stop_p1_101,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=6,
            state="disabled",
        )
        self.btn_p1_101_stop.pack(side="left", padx=8)
        tk.Button(
            actions,
            text="로그 지우기",
            command=self._clear_p1_101_log,
        ).pack(side="left", padx=4)

        self.p1_101_status = tk.Label(
            parent,
            text="엑셀 지정 후 ▶ 실행 (별도 CLI 불필요 · 보드에서 직접 실행)",
            bg="#f1f5f9",
            anchor="w",
            justify="left",
            fg="#64748b",
        )
        self.p1_101_status.pack(fill="x", pady=4)

        log_frame = tk.LabelFrame(
            parent,
            text="실행 로그 (실시간)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        log_frame.pack(fill="both", expand=True, pady=(0, 6))
        self.p1_101_log = ttk.Treeview(
            log_frame,
            columns=("time", "step", "message"),
            show="headings",
            height=14,
        )
        self.p1_101_log.heading("time", text="시각")
        self.p1_101_log.heading("step", text="단계")
        self.p1_101_log.heading("message", text="내용")
        self.p1_101_log.column(
            "time", width=90, minwidth=70, stretch=False, anchor="center"
        )
        self.p1_101_log.column(
            "step", width=70, minwidth=50, stretch=False, anchor="center"
        )
        self.p1_101_log.column(
            "message", width=640, minwidth=200, stretch=True, anchor="w"
        )
        log_sb = tk.Scrollbar(
            log_frame, orient="vertical", command=self.p1_101_log.yview
        )
        self.p1_101_log.configure(yscrollcommand=log_sb.set)
        self.p1_101_log.pack(side="left", fill="both", expand=True)
        log_sb.pack(side="right", fill="y")
        self.p1_101_log.tag_configure("err", foreground="#b91c1c")
        self.p1_101_log.tag_configure("ok", foreground="#166534")

    def _pick_p1_101_excel(self) -> None:
        initial = self.var_p1_101_excel.get().strip()
        if initial and Path(initial).is_file():
            initial_dir = str(Path(initial).parent)
        else:
            roots = default_roots()
            initial_dir = roots[0] if roots else str(Path.home())
        path = filedialog.askopenfilename(
            title="엑셀자료 파일지정",
            initialdir=initial_dir,
            filetypes=[
                ("Excel", "*.xlsx *.xlsm"),
                ("모든 파일", "*.*"),
            ],
        )
        if path:
            self.var_p1_101_excel.set(path)
            try:
                add_paths([path])
                set_selected(path)
            except Exception:
                pass

    def _clear_p1_101_log(self) -> None:
        tv = getattr(self, "p1_101_log", None)
        if tv is not None:
            for item in tv.get_children():
                tv.delete(item)

    def _append_p1_101_log(self, step: str, message: str) -> None:
        tv = getattr(self, "p1_101_log", None)
        if tv is None:
            return
        ts = time.strftime("%H:%M:%S")
        tag = ()
        s = (step or "").strip()
        su = s.upper()
        if su in ("오류", "ERROR", "FAIL") or s == "오류":
            tag = ("err",)
        elif s in ("완료", "최종", "OK") or "완료" in (message or ""):
            tag = ("ok",)
        item = tv.insert("", "end", values=(ts, step, message), tags=tag)
        tv.see(item)

    def _p1_101_stop_flag(self) -> Path:
        return ROOT / "P1_101" / ".extract_stop"

    def _run_p1_101(self) -> None:
        """메인보드에서 P1_101/extract.py 를 서브프로세스로 직접 실행."""
        excel = self.var_p1_101_excel.get().strip()
        if not excel:
            messagebox.showinfo("안내", "엑셀자료 파일을 지정하세요.")
            return
        if not Path(excel).is_file():
            messagebox.showerror("오류", f"파일이 없습니다:\n{excel}")
            return
        if self._p1_101_proc and self._p1_101_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 상품수 추출이 진행 중입니다.")
            return

        extract_py = ROOT / "P1_101" / "extract.py"
        if not extract_py.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{extract_py}")
            return

        try:
            add_paths([excel])
            set_selected(excel)
        except Exception:
            pass

        # 중단 플래그 초기화
        try:
            self._p1_101_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except TypeError:
            fp = self._p1_101_stop_flag()
            if fp.exists():
                try:
                    fp.unlink()
                except OSError:
                    pass
        except OSError:
            pass

        args = [sys.executable, str(extract_py), excel]
        # Windows 로컬: 브라우저 창 표시(팝업 닫기 확인) / 그 외 headless
        if os.name == "nt":
            args.append("--headed")
        else:
            args.append("--headless")

        self._clear_p1_101_log()
        self.btn_p1_101_run.configure(state="disabled")
        self.btn_p1_101_stop.configure(state="normal")
        self.p1_101_status.configure(
            text=f"보드 직접실행 중… {Path(excel).name}",
            fg="#0f172a",
        )
        self._append_p1_101_log("실행", f"메인보드 → {extract_py.name} · {excel}")

        try:
            creationflags = 0
            if os.name == "nt":
                creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            self._p1_101_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P1_101"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            self.btn_p1_101_run.configure(state="normal")
            self.btn_p1_101_stop.configure(state="disabled")
            messagebox.showerror("실행 실패", str(e))
            self.p1_101_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p1_101_proc,
            args=(self._p1_101_proc, excel),
            daemon=True,
        ).start()

    def _stop_p1_101(self) -> None:
        proc = self._p1_101_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 상품수 추출이 없습니다.")
            return
        try:
            self._p1_101_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p1_101_status.configure(
                text=f"중단 플래그 기록 실패: {e}", fg="#b91c1c"
            )
            return
        self.p1_101_status.configure(text="중단 요청 중… (부분 UPDATE 저장)", fg="#b45309")
        self._append_p1_101_log("중단", "사용자 중단 요청")

        def force() -> None:
            for _ in range(24):
                if proc.poll() is not None:
                    return
                time.sleep(0.5)
            if proc.poll() is not None:
                return
            try:
                proc.terminate()
            except Exception:
                pass
            time.sleep(1.5)
            if proc.poll() is None:
                try:
                    proc.kill()
                except Exception:
                    pass

        threading.Thread(target=force, daemon=True).start()

    def _p1_101_log_ui(self, text: str) -> None:
        """extract.py stdout 한 줄 → 실행로그 그리드."""
        line = (text or "").strip()
        if not line:
            return
        m = re.match(r"^\[([^\]]+)\]\s*(.*)$", line)
        if m:
            step, msg = m.group(1).strip(), m.group(2).strip()
        else:
            step, msg = "로그", line
        self.after(0, lambda s=step, m=msg: self._append_p1_101_log(s, m))

    def _watch_p1_101_proc(self, proc: subprocess.Popen, path: str) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self._p1_101_log_ui(text)
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self._p1_101_log_ui(text)
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p1_101_status.configure(
                    text=f"로그 수신 오류: {e}", fg="#b91c1c"
                ),
            )
        code = proc.wait()
        self.after(0, lambda: self._on_p1_101_finished(path, code))

    def _on_p1_101_finished(self, path: str, code: int) -> None:
        self.btn_p1_101_run.configure(state="normal")
        self.btn_p1_101_stop.configure(state="disabled")
        self._p1_101_proc = None
        try:
            self._p1_101_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass
        if code == 0:
            self.p1_101_status.configure(
                text="완료 · 출력: P2_INPUT_건수집계 (파일명_상품수_v버전)",
                fg="#15803d",
            )
            self._append_p1_101_log(
                "완료",
                f"보드 직접실행 종료 · 입력 {Path(path).name} · 출력폴더 P2_INPUT_건수집계",
            )
        else:
            self.p1_101_status.configure(
                text=f"실패 (exit={code}) · {Path(path).name}",
                fg="#b91c1c",
            )
            self._append_p1_101_log("오류", f"종료 코드 {code}")

    # ── P1_ZARA_DE ─────────────────────────────────────
    def _build_p1_zara(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P1_ZARA_DE — 독일자라 카테고리 → 엑셀 (P2 입력용)",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 8))

        form = tk.Frame(parent, bg="#ffffff", padx=10, pady=10, relief="solid", bd=1)
        form.pack(fill="x")

        self.var_zara_site = tk.StringVar(value=p1_zara_crawl.DEFAULT_SITE)
        self.var_zara_url = tk.StringVar(value=p1_zara_crawl.DEFAULT_URL)
        self.var_zara_outdir = tk.StringVar(value=str(Path.home() / "Downloads"))

        self._row(form, "사이트명", self.var_zara_site)
        self._row(form, "사이트 URL", self.var_zara_url)

        # ★요건: 20행 × 3열 — 상위명, 중위명, 하위 URL (URL 칸=명 칸의 10배)
        z_rows = p1_zara_crawl.TOP_GRID_ROWS
        z_labels = p1_zara_crawl.COL_LABELS
        name_w = p1_zara_crawl.NAME_ENTRY_WIDTH
        url_w = p1_zara_crawl.URL_ENTRY_WIDTH
        col_widths = (name_w, name_w, url_w)

        tops_wrap = tk.Frame(form, bg="#ffffff")
        tops_wrap.pack(fill="x", pady=3)
        tk.Label(
            tops_wrap,
            text="카테고리 계층",
            width=16,
            anchor="nw",
            bg="#ffffff",
        ).pack(side="left", anchor="n", pady=2)
        tops_right = tk.Frame(tops_wrap, bg="#ffffff")
        tops_right.pack(side="left", fill="both", expand=True)

        # 가로·세로 스크롤 (20행)
        canvas = tk.Canvas(tops_right, bg="#ffffff", height=280, highlightthickness=0)
        h_sb = tk.Scrollbar(tops_right, orient="horizontal", command=canvas.xview)
        v_sb = tk.Scrollbar(tops_right, orient="vertical", command=canvas.yview)
        canvas.configure(xscrollcommand=h_sb.set, yscrollcommand=v_sb.set)
        h_sb.pack(side="bottom", fill="x")
        v_sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        grid = tk.Frame(canvas, bg="#ffffff")
        canvas.create_window((0, 0), window=grid, anchor="nw")

        def _on_grid_configure(_e=None) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        grid.bind("<Configure>", _on_grid_configure)

        self._p1_zara_grid_vars: list[list[tk.StringVar]] = []
        self._p1_zara_level_vars = []  # 하위 호환 비움
        vcmd_name = (self.register(self._validate_p1_top_cell), "%P")
        vcmd_url = (self.register(self._validate_p1_zara_url_cell), "%P")
        # 열 헤더: 상위 카테고리명 / 중위 카테고리명 / 하위 카테고리 URL
        hdr = tk.Frame(grid, bg="#ffffff")
        hdr.pack(fill="x")
        tk.Label(hdr, text="행", width=4, bg="#ffffff", fg="#94a3b8").pack(side="left")
        for label, w in zip(z_labels, col_widths):
            tk.Label(
                hdr,
                text=label,
                width=w,
                bg="#ffffff",
                fg="#64748b",
                font=("Malgun Gothic", 7),
                anchor="w",
            ).pack(side="left", padx=1)
        for row_i in range(z_rows):
            row_f = tk.Frame(grid, bg="#ffffff")
            row_f.pack(fill="x", pady=1)
            tk.Label(
                row_f,
                text=str(row_i + 1),
                width=4,
                anchor="center",
                bg="#ffffff",
                font=("Malgun Gothic", 8),
            ).pack(side="left")
            row_vars: list[tk.StringVar] = []
            for col_i, w in enumerate(col_widths):
                var = tk.StringVar(value="")
                row_vars.append(var)
                is_url = col_i == 2
                tk.Entry(
                    row_f,
                    textvariable=var,
                    width=w,
                    font=("Malgun Gothic", 8),
                    justify="left" if is_url else "center",
                    validate="key",
                    validatecommand=vcmd_url if is_url else vcmd_name,
                ).pack(side="left", padx=1)
            self._p1_zara_grid_vars.append(row_vars)
        self._p1_zara_top_vars = []

        out_row = tk.Frame(form, bg="#ffffff")
        out_row.pack(fill="x", pady=3)
        tk.Label(out_row, text="저장 폴더", width=16, anchor="w", bg="#ffffff").pack(
            side="left"
        )
        tk.Entry(out_row, textvariable=self.var_zara_outdir).pack(
            side="left", fill="x", expand=True
        )
        tk.Button(out_row, text="…", width=3, command=self._pick_zara_outdir).pack(
            side="left", padx=4
        )

        actions = tk.Frame(parent, bg="#f1f5f9")
        actions.pack(fill="x", pady=10)
        self.btn_zara_crawl = tk.Button(
            actions,
            text="1. 수집 시작",
            command=self._run_p1_zara,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=12,
            pady=6,
        )
        self.btn_zara_crawl.pack(side="left")
        self.btn_zara_save = tk.Button(
            actions,
            text="2. 엑셀 저장",
            command=self._save_p1_zara,
            state="disabled",
            padx=12,
            pady=6,
        )
        self.btn_zara_save.pack(side="left", padx=8)
        tk.Button(actions, text="독일자라 기본값", command=self._p1_zara_defaults).pack(
            side="left"
        )
        tk.Button(
            actions,
            text="로그 지우기",
            command=self._clear_p1_zara_log,
        ).pack(side="left", padx=6)
        tk.Button(
            actions,
            text="스크린샷 보기",
            command=self._show_p1_zara_shot,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=8,
            pady=4,
        ).pack(side="left", padx=6)

        self.p1_zara_status = tk.Label(
            parent, text="", bg="#f1f5f9", anchor="w", justify="left"
        )
        self.p1_zara_status.pack(fill="x", pady=4)

        # 1) 실행로그 그리드 — 2) 실시간 표시
        log_frame = tk.LabelFrame(
            parent,
            text="실행 로그 (실시간)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        log_frame.pack(fill="both", expand=True, pady=(0, 6))
        self.p1_zara_log = ttk.Treeview(
            log_frame,
            columns=("time", "step", "message"),
            show="headings",
            height=10,
        )
        self.p1_zara_log.heading("time", text="시각")
        self.p1_zara_log.heading("step", text="단계")
        self.p1_zara_log.heading("message", text="내용")
        self.p1_zara_log.column("time", width=90, minwidth=70, stretch=False, anchor="center")
        self.p1_zara_log.column("step", width=70, minwidth=50, stretch=False, anchor="center")
        self.p1_zara_log.column("message", width=640, minwidth=200, stretch=True, anchor="w")
        zlog_sb = tk.Scrollbar(log_frame, orient="vertical", command=self.p1_zara_log.yview)
        self.p1_zara_log.configure(yscrollcommand=zlog_sb.set)
        self.p1_zara_log.pack(side="left", fill="both", expand=True)
        zlog_sb.pack(side="right", fill="y")
        self.p1_zara_log.tag_configure("err", foreground="#b91c1c")
        self.p1_zara_log.tag_configure("ok", foreground="#166534")
        self.p1_zara_log.tag_configure("shot", foreground="#0f766e")

        # 3) 최종 스크린샷
        shot_frame = tk.LabelFrame(
            parent,
            text="최종 스크린샷",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        shot_frame.pack(fill="both", expand=False)
        self.p1_zara_shot_label = tk.Label(
            shot_frame,
            text="(수집 완료 후 최종 스크린샷이 여기에 표시됩니다)",
            bg="#f8fafc",
            fg="#64748b",
            anchor="center",
            height=8,
        )
        self.p1_zara_shot_label.pack(fill="both", expand=True)
        self._p1_zara_shot_photo: tk.PhotoImage | None = None
        self._p1_zara_shot_path: str = ""

    def _p1_zara_top_values(self) -> list[str]:
        """하위 호환 — 각 행의 상위 칸 값만 평탄 목록으로."""
        out: list[str] = []
        grid = getattr(self, "_p1_zara_grid_vars", None)
        if grid:
            for row in grid:
                if row:
                    s = (row[0].get() or "").strip()
                    if s:
                        out.append(s)
            return out
        for var in getattr(self, "_p1_zara_top_vars", []):
            s = (var.get() or "").strip()
            if s:
                out.append(s)
        return out

    def _p1_zara_grid_rows(self) -> list[tuple[str, ...]]:
        """20행 × 3열 원시 입력 (상위명·중위명·하위URL)."""
        grid = getattr(self, "_p1_zara_grid_vars", None)
        if not grid:
            return []
        return [tuple(var.get() for var in row) for row in grid]

    def _p1_zara_hierarchy_paths(self) -> list[tuple[str, str, str]]:
        """그리드 → (상위, 중위, 하위URL). crawl.expand_grid_rows_to_paths와 동일."""
        return p1_zara_crawl.expand_grid_rows_to_paths(self._p1_zara_grid_rows())

    def _pick_zara_outdir(self) -> None:
        d = filedialog.askdirectory(
            initialdir=self.var_zara_outdir.get() or str(Path.home())
        )
        if d:
            self.var_zara_outdir.set(d)

    def _p1_zara_defaults(self) -> None:
        self.var_zara_site.set(p1_zara_crawl.DEFAULT_SITE)
        self.var_zara_url.set(p1_zara_crawl.DEFAULT_URL)
        # 카테고리 칸은 비움 — 사용자가 입력으로 지정
        for row in getattr(self, "_p1_zara_grid_vars", []):
            for var in row:
                var.set("")
        for row in getattr(self, "_p1_zara_level_vars", []):
            for var in row:
                var.set("")
        for var in getattr(self, "_p1_zara_top_vars", []):
            var.set("")

    def _clear_p1_zara_log(self) -> None:
        tv = getattr(self, "p1_zara_log", None)
        if tv is not None:
            for item in tv.get_children():
                tv.delete(item)
        self._p1_zara_shot_photo = None
        self._p1_zara_shot_path = ""
        if getattr(self, "p1_zara_shot_label", None) is not None:
            self.p1_zara_shot_label.configure(
                image="",
                text="(수집 완료 후 최종 스크린샷이 여기에 표시됩니다)",
                fg="#64748b",
            )

    def _append_p1_zara_log(self, step: str, message: str) -> None:
        """크롤 스레드 → UI 스레드 실행로그 1행 추가."""
        tv = getattr(self, "p1_zara_log", None)
        if tv is None:
            return
        ts = time.strftime("%H:%M:%S")
        tag = ()
        s = (step or "").upper()
        if s in ("오류", "ERROR", "FAIL"):
            tag = ("err",)
        elif s in ("완료", "OK", "결과") and "완료" in (message or ""):
            tag = ("ok",)
        elif s in ("SHOT", "샷"):
            tag = ("shot",)
        item = tv.insert("", "end", values=(ts, step, message), tags=tag)
        tv.see(item)

    def _p1_zara_progress(self, step: str, message: str) -> None:
        self.after(0, lambda s=step, m=message: self._append_p1_zara_log(s, m))

    def _show_p1_zara_shot_image(self, path: str) -> None:
        """최종 스크린샷 PNG를 패널에 표시."""
        self._p1_zara_shot_path = path or ""
        label = getattr(self, "p1_zara_shot_label", None)
        if label is None:
            return
        p = Path(path) if path else None
        if p is None or not p.is_file():
            label.configure(
                image="",
                text="(최종 스크린샷 없음)",
                fg="#b91c1c",
            )
            self._p1_zara_shot_photo = None
            return
        try:
            img = tk.PhotoImage(file=str(p))
            # 패널에 맞게 축소
            w, h = img.width(), img.height()
            max_w, max_h = 720, 220
            factor = 1
            while (w // factor) > max_w or (h // factor) > max_h:
                factor += 1
                if factor > 20:
                    break
            if factor > 1:
                img = img.subsample(factor, factor)
            self._p1_zara_shot_photo = img
            label.configure(image=img, text="", compound="center")
        except tk.TclError:
            label.configure(
                image="",
                text=f"미리보기 불가 — 파일: {p}\n(스크린샷 보기 버튼으로 열기)",
                fg="#0f766e",
            )
            self._p1_zara_shot_photo = None

    def _show_p1_zara_shot(self) -> None:
        path = self._p1_zara_shot_path
        if not path or not Path(path).is_file():
            result = getattr(self, "_p1_zara_result", None)
            if result and getattr(result, "final_shot_path", ""):
                path = result.final_shot_path
        if not path or not Path(path).is_file():
            messagebox.showinfo("안내", "표시할 최종 스크린샷이 없습니다. 먼저 수집을 실행하세요.")
            return
        try:
            if os.name == "nt":
                os.startfile(path)  # type: ignore[attr-defined]
            else:
                webbrowser.open(Path(path).resolve().as_uri())
        except Exception as e:
            messagebox.showerror("열기 실패", str(e))

    def _run_p1_zara(self) -> None:
        self.btn_zara_crawl.configure(state="disabled")
        self.btn_zara_save.configure(state="disabled")
        self.p1_zara_status.configure(text="수집 중…", fg="#0f172a")
        self._clear_p1_zara_log()
        grid_rows = self._p1_zara_grid_rows()

        def work() -> None:
            result = zara_crawl_site(
                self.var_zara_site.get(),
                self.var_zara_url.get(),
                None,
                progress=self._p1_zara_progress,
                take_screenshot=True,
                run_root=ROOT / "P1_ZARA_DE",
                category_grid_rows=grid_rows,
            )
            self.after(0, lambda: self._p1_zara_done(result))

        threading.Thread(target=work, daemon=True).start()

    def _p1_zara_done(self, result) -> None:
        self.btn_zara_crawl.configure(state="normal")
        self._p1_zara_result = result
        if getattr(result, "final_shot_path", ""):
            self._show_p1_zara_shot_image(result.final_shot_path)
        if not result.ok:
            self.p1_zara_status.configure(
                text="실패: " + "; ".join(result.errors), fg="#b91c1c"
            )
            return
        self.btn_zara_save.configure(state="normal")
        msg = f"완료 · {result.platform} · {result.total}건"
        if result.warnings:
            msg += " · " + " / ".join(result.warnings)
        if getattr(result, "final_shot_path", ""):
            msg += f" · 샷: {Path(result.final_shot_path).name}"
        self.p1_zara_status.configure(text=msg, fg="#15803d")

    def _save_p1_zara(self) -> None:
        if not self._p1_zara_result or not self._p1_zara_result.ok:
            return
        try:
            path = zara_save_excel(
                self._p1_zara_result.rows,
                self._p1_zara_result.site_name,
                self.var_zara_outdir.get(),
            )
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))
            return
        self.p1_zara_status.configure(text=f"저장됨: {path}", fg="#15803d")
        add_paths([str(path)])
        try:
            self.var_dir.set(str(Path(path).parent))
        except Exception:
            pass
        self._refresh_p2_list()
        self._load_category_url_list(str(path))
        if messagebox.askyesno(
            "P2로 이동",
            f"엑셀 저장·카테고리URL목록에 반영했습니다.\n\n{path}\n\nP2 화면으로 갈까요?",
        ):
            self._show("p2")

    # ── P2 ─────────────────────────────────────────────
    def _build_p2(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P2 — 폴더의 엑셀 파일 선택 → 카테고리URL목록 확인 → 수집 실행",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        # 1. 디렉터리 파일 목록 (리스트박스 + 스크롤)
        search = tk.LabelFrame(
            parent, text="1. 디렉터리 파일 목록", bg="#ffffff", padx=8, pady=6
        )
        search.pack(fill="x")

        self.var_dir = tk.StringVar(value=(default_roots() or [str(Path.home())])[0])
        self.var_q = tk.StringVar(value="카테고리URL")

        r1 = tk.Frame(search, bg="#ffffff")
        r1.pack(fill="x", pady=2)
        tk.Label(r1, text="폴더", width=8, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r1, textvariable=self.var_dir).pack(side="left", fill="x", expand=True)
        tk.Button(r1, text="…", width=3, command=self._pick_search_dir).pack(side="left", padx=4)

        r2 = tk.Frame(search, bg="#ffffff")
        r2.pack(fill="x", pady=2)
        tk.Label(r2, text="필터", width=8, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r2, textvariable=self.var_q, width=20).pack(side="left")
        tk.Button(r2, text="파일 새로고침", command=self._search_xlsx, bg="#e2e8f0").pack(
            side="left", padx=6
        )
        tk.Button(r2, text="선택 파일 열기", command=self._add_found).pack(side="left")

        self._p2_file_list_height = 5
        self._p2_url_list_height = 8

        found_wrap = tk.Frame(search, bg="#ffffff")
        found_wrap.pack(fill="x", pady=4)
        self.found_list = tk.Listbox(
            found_wrap,
            height=self._p2_file_list_height,
            selectmode="browse",
            font=("Consolas", 9),
            exportselection=False,
        )
        found_sb = tk.Scrollbar(found_wrap, orient="vertical", command=self.found_list.yview)
        self.found_list.configure(yscrollcommand=found_sb.set)
        self.found_list.pack(side="left", fill="both", expand=True)
        found_sb.pack(side="right", fill="y")
        self.found_list.bind("<<ListboxSelect>>", self._on_found_select)
        self.found_list.bind("<Double-Button-1>", lambda _e: self._add_found())
        self.found_list.bind("<MouseWheel>", self._on_found_mousewheel)
        self.found_list.bind("<Button-4>", self._on_found_mousewheel)
        self.found_list.bind("<Button-5>", self._on_found_mousewheel)
        self._found_paths: list[str] = []

        # 실행 버튼 (좌) + 체크박스 MAIN / SUB / 1·2행 스크린샷 (최우측)
        actions = tk.LabelFrame(parent, text="실행", bg="#ffffff", padx=8, pady=6)
        actions.pack(fill="x", pady=(8, 0))

        self.var_show_main = tk.BooleanVar(value=True)
        self.var_show_sub = tk.BooleanVar(value=True)
        self.var_verify = tk.BooleanVar(value=True)
        btn_row = tk.Frame(actions, bg="#ffffff")
        btn_row.pack(fill="x")

        # ★최우측 순서: MAIN → SUB → 1·2행 스크린샷
        # (실행로그 안내 라벨은 삭제 — 요건: 스크린샷 2번째 LABEL 삭제)
        # pack 순서: right 먼저 → 항상 최우측 고정
        right_checks = tk.Frame(btn_row, bg="#ffffff")
        right_checks.pack(side="right")
        tk.Checkbutton(
            right_checks,
            text="MAIN",
            variable=self.var_show_main,
            command=self._toggle_log_panels,
            bg="#ffffff",
            font=("Malgun Gothic", 9),
        ).pack(side="left", padx=(0, 6))
        tk.Checkbutton(
            right_checks,
            text="SUB",
            variable=self.var_show_sub,
            command=self._toggle_log_panels,
            bg="#ffffff",
            font=("Malgun Gothic", 9),
        ).pack(side="left", padx=(0, 6))
        tk.Checkbutton(
            right_checks,
            text="1·2행 스크린샷",
            variable=self.var_verify,
            bg="#ffffff",
            font=("Malgun Gothic", 9),
        ).pack(side="left")

        left_btns = tk.Frame(btn_row, bg="#ffffff")
        left_btns.pack(side="left", fill="x", expand=True)
        tk.Button(
            left_btns,
            text="수집 시작",
            command=self._run_p2,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        ).pack(side="left")
        tk.Button(
            left_btns,
            text="수집 종료",
            command=self._stop_p2,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        ).pack(side="left", padx=6)
        tk.Button(left_btns, text="파일 목록에서 제거", command=self._remove_lib).pack(
            side="left", padx=6
        )
        tk.Button(left_btns, text="새로고침", command=self._refresh_p2_list).pack(side="left")
        tk.Button(left_btns, text="로그 지우기", command=self._clear_p2_log).pack(
            side="left", padx=6
        )
        tk.Button(
            left_btns,
            text="스크린샷 보기",
            command=self._show_shot_viewer,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=8,
            pady=4,
        ).pack(side="left", padx=6)

        # 2. 카테고리URL목록 — 엑셀 전체 행 + 진행중 행 적색
        lib = tk.LabelFrame(
            parent, text="카테고리URL목록", bg="#ffffff", padx=8, pady=4
        )
        lib.pack(fill="x", pady=(8, 0))

        lib_wrap = tk.Frame(lib, bg="#ffffff")
        lib_wrap.pack(fill="both", expand=True)
        self.lib_list = tk.Listbox(
            lib_wrap,
            height=self._p2_url_list_height,
            font=("Malgun Gothic", 10),
            exportselection=False,
            activestyle="none",
        )
        lib_sb = tk.Scrollbar(lib_wrap, orient="vertical", command=self.lib_list.yview)
        self.lib_list.configure(yscrollcommand=lib_sb.set)
        self.lib_list.pack(side="left", fill="both", expand=True)
        lib_sb.pack(side="right", fill="y")
        self.lib_list.bind("<MouseWheel>", self._on_lib_mousewheel)
        self.lib_list.bind("<Button-4>", self._on_lib_mousewheel)
        self.lib_list.bind("<Button-5>", self._on_lib_mousewheel)
        self._lib_paths: list[str] = []  # 하위호환(파일경로 1개 보관용)
        self._excel_rows: list[dict] = []
        self._current_excel_path: str = ""
        self._active_ordinal: int = 0  # 1-based, 0=없음

        self.p2_sel = tk.Label(lib, text="", bg="#ffffff", fg="#64748b", anchor="w")
        self.p2_sel.pack(fill="x", pady=(2, 0))

        # 3. 실행 로그 — main(13단계) / sub(단계별 추가정보·스크린샷) 두 그리드
        log_area = tk.Frame(parent, bg="#f1f5f9")
        log_area.pack(fill="both", expand=True, pady=(8, 0))

        style = ttk.Style(self)
        try:
            style.configure("P2Log.Treeview", rowheight=22, font=("Malgun Gothic", 9))
            style.configure("P2Log.Treeview.Heading", font=("Malgun Gothic", 9, "bold"))
        except tk.TclError:
            pass

        self._p2_log_area = log_area
        self.p2_main_frame = tk.LabelFrame(
            log_area,
            text="3-A. 실행 로그 MAIN (상단=엑셀정보 · 아래=1~13단계)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        self.p2_main_frame.pack(fill="both", expand=True)

        self.p2_main_log = ttk.Treeview(
            self.p2_main_frame,
            columns=("time", "step", "message"),
            show="headings",
            height=9,
            style="P2Log.Treeview",
        )
        self.p2_main_log.heading("time", text="시각")
        self.p2_main_log.heading("step", text="단계")
        self.p2_main_log.heading("message", text="내용 (1~13단계)")
        self.p2_main_log.column("time", width=130, minwidth=110, stretch=False, anchor="center")
        self.p2_main_log.column("step", width=44, minwidth=40, stretch=False, anchor="center")
        self.p2_main_log.column("message", width=620, minwidth=220, stretch=True, anchor="w")
        main_sb = tk.Scrollbar(
            self.p2_main_frame, orient="vertical", command=self.p2_main_log.yview
        )
        self.p2_main_log.configure(yscrollcommand=main_sb.set)
        self.p2_main_log.pack(side="left", fill="both", expand=True)
        main_sb.pack(side="right", fill="y")
        self.p2_main_log.bind("<<TreeviewSelect>>", self._on_main_log_select)
        self._setup_p2_log_tags()

        self.p2_sub_frame = tk.LabelFrame(
            log_area,
            text="3-B. 실행 로그 SUB (선택한 단계의 추가정보·스크린샷)",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        self.p2_sub_frame.pack(fill="both", expand=True, pady=(6, 0))

        self.p2_sub_log = ttk.Treeview(
            self.p2_sub_frame,
            columns=("time", "message"),
            show="headings",
            height=7,
            style="P2Log.Treeview",
        )
        self.p2_sub_log.heading("time", text="시각")
        self.p2_sub_log.heading("message", text="추가정보 · [샷]은 더블클릭으로 열기")
        self.p2_sub_log.column("time", width=130, minwidth=110, stretch=False, anchor="center")
        self.p2_sub_log.column("message", width=700, minwidth=220, stretch=True, anchor="w")
        sub_sb = tk.Scrollbar(
            self.p2_sub_frame, orient="vertical", command=self.p2_sub_log.yview
        )
        self.p2_sub_log.configure(yscrollcommand=sub_sb.set)
        self.p2_sub_log.pack(side="left", fill="both", expand=True)
        sub_sb.pack(side="right", fill="y")
        self.p2_sub_log.tag_configure("shot", foreground="#0f766e")
        self.p2_sub_log.bind("<Double-Button-1>", self._on_sub_log_double_click)

        # seq(단계 발생 고유번호) 기반 main↔sub 연결 데이터
        self._sub_by_seq: dict[int, list[tuple[str, str, str]]] = {}
        self._shot_path_by_seq: dict[tuple[int, int], str] = {}
        self._main_item_by_seq: dict[int, str] = {}
        self._main_ts_end: dict[int, str] = {}
        self._seq_by_main_item: dict[str, int] = {}
        self._meta_item_id: str | None = None
        self._meta_values: dict[str, str] = {f: "" for f in META_FIELDS}
        self._selected_seq: int | None = None
        self._latest_seq: int = 0
        self._follow_latest: bool = True

        self._setup_meta_rows()

        self.p2_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p2_status.pack(fill="x", pady=4)

    # ── P3_필터_갱신 (UI 구조 = P2와 유사 + 더망고 URL 입력) ──
    def _build_p3(self, parent: tk.Frame) -> None:
        tk.Label(
            parent,
            text="P3_필터_갱신 — 엑셀 선택 → 더망고 URL 입력 → 저장상품수 갱신",
            bg="#f1f5f9",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))

        search = tk.LabelFrame(
            parent, text="1. 디렉터리 파일 목록", bg="#ffffff", padx=8, pady=6
        )
        search.pack(fill="x")

        self.var_p3_dir = tk.StringVar(
            value=(default_roots() or [str(Path.home())])[0]
        )
        self.var_p3_q = tk.StringVar(value="카테고리URL")
        # ★요건: 더망고 URL 입력창에 망고 url 기본 세팅 + 변경 가능
        self.var_p3_mango_url = tk.StringVar(value=p3_update.load_mango_url_default())

        r1 = tk.Frame(search, bg="#ffffff")
        r1.pack(fill="x", pady=2)
        tk.Label(r1, text="폴더", width=10, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r1, textvariable=self.var_p3_dir).pack(
            side="left", fill="x", expand=True
        )
        tk.Button(r1, text="…", width=3, command=self._pick_p3_dir).pack(
            side="left", padx=4
        )

        r2 = tk.Frame(search, bg="#ffffff")
        r2.pack(fill="x", pady=2)
        tk.Label(r2, text="필터", width=10, anchor="w", bg="#ffffff").pack(side="left")
        tk.Entry(r2, textvariable=self.var_p3_q, width=20).pack(side="left")
        tk.Button(
            r2, text="파일 새로고침", command=self._search_p3_xlsx, bg="#e2e8f0"
        ).pack(side="left", padx=6)
        tk.Button(r2, text="선택 파일 열기", command=self._add_p3_found).pack(
            side="left"
        )

        r3 = tk.Frame(search, bg="#ffffff")
        r3.pack(fill="x", pady=(6, 2))
        tk.Label(
            r3, text="더망고 URL", width=10, anchor="w", bg="#ffffff",
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side="left")
        tk.Entry(r3, textvariable=self.var_p3_mango_url).pack(
            side="left", fill="x", expand=True
        )
        tk.Label(
            search,
            text="망고 URL이 기본으로 채워집니다. 필요하면 수정하세요 (변경값 기억)",
            bg="#ffffff",
            fg="#64748b",
            anchor="w",
            font=("Malgun Gothic", 8),
        ).pack(fill="x")

        self._p3_file_list_height = 5
        self._p3_url_list_height = 8

        found_wrap = tk.Frame(search, bg="#ffffff")
        found_wrap.pack(fill="x", pady=4)
        self.p3_found_list = tk.Listbox(
            found_wrap,
            height=self._p3_file_list_height,
            selectmode="browse",
            font=("Consolas", 9),
            exportselection=False,
        )
        found_sb = tk.Scrollbar(
            found_wrap, orient="vertical", command=self.p3_found_list.yview
        )
        self.p3_found_list.configure(yscrollcommand=found_sb.set)
        self.p3_found_list.pack(side="left", fill="both", expand=True)
        found_sb.pack(side="right", fill="y")
        self.p3_found_list.bind("<<ListboxSelect>>", self._on_p3_found_select)
        self.p3_found_list.bind("<Double-Button-1>", lambda _e: self._add_p3_found())
        self._p3_found_paths: list[str] = []

        actions = tk.LabelFrame(parent, text="실행", bg="#ffffff", padx=8, pady=6)
        actions.pack(fill="x", pady=(8, 0))
        btn_row = tk.Frame(actions, bg="#ffffff")
        btn_row.pack(fill="x")
        left_btns = tk.Frame(btn_row, bg="#ffffff")
        left_btns.pack(side="left", fill="x", expand=True)
        self.btn_p3_run = tk.Button(
            left_btns,
            text="작업시작",
            command=self._run_p3,
            bg="#2563eb",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        )
        self.btn_p3_run.pack(side="left")
        self.btn_p3_stop = tk.Button(
            left_btns,
            text="작업중단",
            command=self._stop_p3,
            bg="#b91c1c",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=10,
            pady=4,
        )
        self.btn_p3_stop.pack(side="left", padx=6)
        tk.Button(
            left_btns,
            text="스크린샷 보기",
            command=self._show_p3_shot_viewer,
            bg="#0f766e",
            fg="white",
            font=("Malgun Gothic", 9, "bold"),
            padx=8,
            pady=4,
        ).pack(side="left", padx=6)
        tk.Button(left_btns, text="새로고침", command=self._refresh_p3_list).pack(
            side="left", padx=6
        )
        tk.Button(left_btns, text="로그 지우기", command=self._clear_p3_log).pack(
            side="left", padx=6
        )

        lib = tk.LabelFrame(
            parent, text="카테고리URL목록 (엑셀)", bg="#ffffff", padx=8, pady=4
        )
        lib.pack(fill="x", pady=(8, 0))
        lib_wrap = tk.Frame(lib, bg="#ffffff")
        lib_wrap.pack(fill="both", expand=True)
        self.p3_lib_list = tk.Listbox(
            lib_wrap,
            height=self._p3_url_list_height,
            font=("Malgun Gothic", 10),
            exportselection=False,
            activestyle="none",
        )
        lib_sb = tk.Scrollbar(
            lib_wrap, orient="vertical", command=self.p3_lib_list.yview
        )
        self.p3_lib_list.configure(yscrollcommand=lib_sb.set)
        self.p3_lib_list.pack(side="left", fill="both", expand=True)
        lib_sb.pack(side="right", fill="y")
        self._p3_excel_rows: list[dict] = []
        self._p3_current_excel: str = ""
        self.p3_sel = tk.Label(lib, text="", bg="#ffffff", fg="#64748b", anchor="w")
        self.p3_sel.pack(fill="x", pady=(2, 0))

        log_frame = tk.LabelFrame(
            parent,
            text="실행 로그",
            bg="#ffffff",
            padx=6,
            pady=4,
        )
        log_frame.pack(fill="both", expand=True, pady=(8, 0))
        self.p3_log = ttk.Treeview(
            log_frame,
            columns=("time", "step", "message"),
            show="headings",
            height=10,
        )
        self.p3_log.heading("time", text="시각")
        self.p3_log.heading("step", text="단계")
        self.p3_log.heading("message", text="내용")
        self.p3_log.column("time", width=90, minwidth=70, stretch=False, anchor="center")
        self.p3_log.column("step", width=70, minwidth=50, stretch=False, anchor="center")
        self.p3_log.column(
            "message", width=640, minwidth=200, stretch=True, anchor="w"
        )
        log_sb = tk.Scrollbar(log_frame, orient="vertical", command=self.p3_log.yview)
        self.p3_log.configure(yscrollcommand=log_sb.set)
        self.p3_log.pack(side="left", fill="both", expand=True)
        log_sb.pack(side="right", fill="y")
        self.p3_log.tag_configure("err", foreground="#b91c1c")
        self.p3_log.tag_configure("ok", foreground="#166534")
        self.p3_log.tag_configure("shot", foreground="#0f766e")
        self.p3_log.bind("<Double-Button-1>", self._on_p3_log_double_click)
        self._p3_shot_paths: dict[str, str] = {}
        self._p3_last_shot_dir: Path | None = None

        self.p3_status = tk.Label(parent, text="", bg="#f1f5f9", anchor="w")
        self.p3_status.pack(fill="x", pady=4)

    def _pick_p3_dir(self) -> None:
        d = filedialog.askdirectory(
            initialdir=self.var_p3_dir.get() or str(Path.home())
        )
        if d:
            self.var_p3_dir.set(d)
            self._search_p3_xlsx()

    def _search_p3_xlsx(self) -> None:
        self.p3_found_list.delete(0, "end")
        self._p3_found_paths = []
        try:
            files = search_xlsx(
                self.var_p3_dir.get().strip(), self.var_p3_q.get().strip()
            )
        except Exception as e:
            messagebox.showerror("검색 실패", str(e))
            return
        for f in files:
            self._p3_found_paths.append(f["path"])
            self.p3_found_list.insert("end", f["name"])
        self.p3_status.configure(
            text=f"파일 {len(files)}개" if files else "해당 폴더에서 .xlsx 없음",
            fg="#0f172a",
        )

    def _on_p3_found_select(self, _event=None) -> None:
        sel = self.p3_found_list.curselection()
        if not sel:
            return
        path = self._p3_found_paths[sel[0]]
        self._load_p3_category_list(path)

    def _add_p3_found(self) -> None:
        sel = self.p3_found_list.curselection()
        if not sel:
            messagebox.showinfo("안내", "목록에서 엑셀 파일을 선택하세요.")
            return
        path = self._p3_found_paths[sel[0]]
        try:
            add_paths([path])
            set_selected(path)
        except Exception:
            pass
        self._load_p3_category_list(path)

    def _load_p3_category_list(self, path: str) -> None:
        """P2와 동일 엑셀 컬럼으로 카테고리URL목록 표시."""
        self.p3_lib_list.delete(0, "end")
        self._p3_excel_rows = []
        self._p3_current_excel = path
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            self.p3_sel.configure(text=f"엑셀 읽기 실패: {e}")
            return
        if not rows:
            self.p3_sel.configure(text="(빈 엑셀)")
            return
        headers = [str(h or "").strip() for h in rows[0]]
        try:
            label_i = headers.index("상위 최종 카테고리명")
            url_i = headers.index("최종 카테고리 URL주소")
        except ValueError:
            # P3: 검색필터 URL 헤더도 허용
            label_i = next(
                (i for i, h in enumerate(headers) if "카테고리명" in h or "필터" in h),
                0,
            )
            url_i = next(
                (i for i, h in enumerate(headers) if "URL" in h.upper()),
                None,
            )
            if url_i is None:
                self.p3_sel.configure(text="URL 열을 찾지 못했습니다")
                return
        for vals in rows[1:]:
            cells = list(vals) if vals else []
            url = str(cells[url_i] or "").strip() if url_i < len(cells) else ""
            if not url.lower().startswith("http"):
                continue
            label = (
                str(cells[label_i] or "").strip() if label_i < len(cells) else ""
            )
            item = {"label": label, "url": url}
            self._p3_excel_rows.append(item)
            self.p3_lib_list.insert(
                "end", f"{len(self._p3_excel_rows):03d}  {label}  |  {url}"
            )
        self.p3_sel.configure(
            text=f"선택: {Path(path).name} · {len(self._p3_excel_rows)}행"
        )

    def _refresh_p3_list(self) -> None:
        cur = self._p3_current_excel
        self._search_p3_xlsx()
        if cur and os.path.isfile(cur):
            self._load_p3_category_list(cur)

    def _clear_p3_log(self) -> None:
        tv = getattr(self, "p3_log", None)
        if tv is None:
            return
        for item in tv.get_children():
            tv.delete(item)
        self._p3_shot_paths = {}

    def _append_p3_log(self, step: str, message: str, *, shot_path: str = "") -> None:
        tv = getattr(self, "p3_log", None)
        if tv is None:
            return
        ts = time.strftime("%H:%M:%S")
        tag = ()
        if step in ("오류", "ERROR", "FAIL") or "실패" in (message or ""):
            tag = ("err",)
        elif step in ("완료", "OK") or "갱신OK" in (message or ""):
            tag = ("ok",)
        elif step == "샷" or "[샷]" in (message or "") or shot_path:
            tag = ("shot",)
        item = tv.insert("", "end", values=(ts, step, message), tags=tag)
        if shot_path:
            self._p3_shot_paths[item] = shot_path
            try:
                self._p3_last_shot_dir = Path(shot_path).parent
            except Exception:
                pass
        tv.see(item)

    def _on_p3_log_double_click(self, _event=None) -> None:
        """실행 로그 샷 행 더블클릭 → PNG 열기."""
        tv = getattr(self, "p3_log", None)
        if tv is None:
            return
        sel = tv.selection()
        if not sel:
            return
        item = sel[0]
        path = self._p3_shot_paths.get(item, "")
        if not path:
            vals = tv.item(item, "values") or ()
            msg = vals[2] if len(vals) >= 3 else ""
            m = re.search(r"((?:[A-Za-z]:)?[^\s]+\.png)\s*$", str(msg))
            if m:
                path = m.group(1)
        if not path or not os.path.isfile(path):
            return
        try:
            self._p3_last_shot_dir = Path(path).parent
        except Exception:
            pass
        try:
            if os.name == "nt":
                os.startfile(path)  # type: ignore[attr-defined]
            else:
                import subprocess as _sp

                _sp.Popen(["xdg-open", path])
        except Exception as e:  # noqa: BLE001
            messagebox.showerror("열기 실패", f"{path}\n{e}")

    def _show_p3_shot_viewer(self) -> None:
        """P3 실행 스크린샷 뷰어 (저장상품수 전·후 포함)."""
        folder = self._p3_last_shot_dir
        if folder is None or not Path(folder).is_dir():
            folder = latest_p3_shot_dir(ROOT)
        open_shot_viewer(
            self,
            shot_dir=folder,
            root=ROOT,
            prefer_p3=True,
            empty_hint=(
                "P3 스크린샷 폴더가 없습니다.\n"
                "작업시작 후 필터 일치 행에서 단계 샷이 생성됩니다.\n"
                "(3)저장상품수 갱신 전·후 샷 포함)"
            ),
        )

    def _p3_stop_flag(self) -> Path:
        return ROOT / "P3_필터_갱신" / ".filter_stop"

    def _run_p3(self) -> None:
        path = self._p3_current_excel
        if not path:
            sel = self.p3_found_list.curselection()
            if sel:
                path = self._p3_found_paths[sel[0]]
                self._load_p3_category_list(path)
        if not path or not os.path.isfile(path):
            messagebox.showinfo("안내", "엑셀 파일을 선택하세요.")
            return
        mango = self.var_p3_mango_url.get().strip()
        if not mango:
            messagebox.showinfo(
                "안내",
                "더망고 URL(검색필터·저장조건 화면)을 입력하세요.",
            )
            return
        if not mango.lower().startswith("http"):
            messagebox.showerror("오류", "더망고 URL은 http(s)로 시작해야 합니다.")
            return
        if self._p3_proc and self._p3_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 작업이 진행 중입니다.")
            return
        if not self._p3_excel_rows:
            messagebox.showerror("오류", "엑셀에 처리할 URL 행이 없습니다.")
            return

        update_py = ROOT / "P3_필터_갱신" / "update_filters.py"
        if not update_py.is_file():
            messagebox.showerror("오류", f"실행 파일 없음:\n{update_py}")
            return

        # 변경한 망고 URL을 다음 실행 기본값으로 저장
        try:
            p3_update.save_mango_url(mango)
        except Exception:
            pass

        try:
            self._p3_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass

        args = [
            sys.executable,
            str(update_py),
            path,
            "--mango-url",
            mango,
        ]
        try:
            add_paths([path])
            set_selected(path)
        except Exception:
            pass

        self._clear_p3_log()
        self.p3_status.configure(
            text=(
                f"작업시작: {Path(path).name} / 총 {len(self._p3_excel_rows)}행 — "
                "P2와 동일: Chrome(CDP)에서 더망고 로그인·팝업 확인"
            ),
            fg="#15803d",
        )
        self._append_p3_log("실행", f"엑셀={path}")
        self._append_p3_log("실행", f"더망고URL={mango}")
        self._append_p3_log(
            "실행",
            "P2 망고연동 재현: CDP Chrome · 확장설정 · 로그인대기 · 필터URL",
        )

        try:
            creationflags = 0
            if os.name == "nt":
                creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            self._p3_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P3_필터_갱신"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p3_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p3_proc,
            args=(self._p3_proc, path),
            daemon=True,
        ).start()

    def _stop_p3(self) -> None:
        proc = self._p3_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 작업이 없습니다.")
            return
        try:
            self._p3_stop_flag().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p3_status.configure(text=f"중단 플래그 실패: {e}", fg="#b91c1c")
            return
        self.p3_status.configure(text="작업중단 요청 중…", fg="#b45309")

    def _watch_p3_proc(self, proc: subprocess.Popen, path: str) -> None:
        assert proc.stdout is not None
        buf = b""
        try:
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = ""
                    for enc in ("utf-8", "cp949", "mbcs"):
                        try:
                            text = line.decode(enc).rstrip("\r")
                            break
                        except UnicodeDecodeError:
                            continue
                    if not text:
                        continue
                    # ##P3SHOT##<path>##<label>
                    if text.startswith("##P3SHOT##"):
                        parts = text.split("##")
                        # ['', 'P3SHOT', path, label...]
                        path = parts[2].strip() if len(parts) > 2 else ""
                        label = "##".join(parts[3:]).strip() if len(parts) > 3 else ""
                        msg = f"[샷] {label} -> {path}" if label else f"[샷] {path}"
                        self.after(
                            0,
                            lambda m=msg, p=path: self._append_p3_log(
                                "샷", m, shot_path=p
                            ),
                        )
                        continue
                    # [step] message
                    step, msg = "로그", text
                    if text.startswith("[") and "]" in text:
                        end = text.index("]")
                        step = text[1:end].strip() or "로그"
                        msg = text[end + 1 :].strip()
                    shot_path = ""
                    if step == "샷" or "[샷]" in msg:
                        m = re.search(
                            r"((?:[A-Za-z]:)?[^\s]+\.png)\s*$",
                            msg,
                        )
                        if m:
                            shot_path = m.group(1)
                    self.after(
                        0,
                        lambda s=step, m=msg, p=shot_path: self._append_p3_log(
                            s, m, shot_path=p
                        ),
                    )
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p3_status.configure(
                    text=f"로그 수신 오류: {e}", fg="#b91c1c"
                ),
            )
        code = proc.wait()
        self.after(0, lambda: self._on_p3_finished(path, code))

    def _on_p3_finished(self, path: str, code: int) -> None:
        self._p3_proc = None
        try:
            self._p3_stop_flag().unlink(missing_ok=True)  # type: ignore[call-arg]
        except Exception:
            pass
        if code == 0:
            self.p3_status.configure(
                text=f"완료 · {Path(path).name}",
                fg="#15803d",
            )
            self._append_p3_log("완료", "보드 직접실행 종료")
        else:
            self.p3_status.configure(
                text=f"종료 (exit={code}) · {Path(path).name}",
                fg="#b91c1c",
            )
            self._append_p3_log("오류", f"종료 코드 {code}")

    def _pick_search_dir(self) -> None:
        d = filedialog.askdirectory(initialdir=self.var_dir.get() or str(Path.home()))
        if d:
            self.var_dir.set(d)
            self._search_xlsx()

    def _search_xlsx(self) -> None:
        """디렉터리의 .xlsx 파일 목록을 리스트박스(+스크롤)에 표시."""
        self.found_list.delete(0, "end")
        self._found_paths = []
        try:
            files = search_xlsx(self.var_dir.get().strip(), self.var_q.get().strip())
        except Exception as e:
            messagebox.showerror("검색 실패", str(e))
            return
        for f in files:
            self._found_paths.append(f["path"])
            self.found_list.insert("end", f["name"])
        self.p2_status.configure(
            text=f"파일 {len(files)}개" if files else "해당 폴더에서 .xlsx 없음",
            fg="#0f172a",
        )
        # 마지막 선택 파일이 목록에 있으면 자동 선택·카테고리URL목록 로드
        data = load()
        last = str(data.get("last_selected") or "").strip()
        if last and last in self._found_paths:
            idx = self._found_paths.index(last)
            self.found_list.selection_clear(0, "end")
            self.found_list.selection_set(idx)
            self.found_list.see(idx)
            self._load_category_url_list(last)
        elif self._found_paths and not self._current_excel_path:
            self.found_list.selection_set(0)
            self.found_list.see(0)

    def _on_found_select(self, _evt=None) -> None:
        sel = self.found_list.curselection()
        if not sel:
            return
        path = self._found_paths[sel[0]]
        self._load_category_url_list(path)

    def _on_found_mousewheel(self, event) -> str:
        if getattr(event, "num", None) == 4 or getattr(event, "delta", 0) > 0:
            self.found_list.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5 or getattr(event, "delta", 0) < 0:
            self.found_list.yview_scroll(1, "units")
        return "break"

    def _add_found(self) -> None:
        """선택한 디렉터리 파일을 열고 카테고리URL목록에 엑셀 전체 행을 표시."""
        sel = list(self.found_list.curselection())
        if not sel:
            messagebox.showinfo("안내", "디렉터리 파일 목록에서 엑셀을 선택하세요.")
            return
        path = self._found_paths[sel[0]]
        add_paths([path])
        self._load_category_url_list(path)
        self.p2_status.configure(
            text=f"카테고리URL목록 로드: {Path(path).name} ({len(self._excel_rows)}행)",
            fg="#15803d",
        )

    def _format_category_row(self, row: dict, *, active: bool = False) -> str:
        mark = "▶ " if active else "   "
        return (
            f"{mark}{row['ordinal']:03d} | {row.get('label', '')} | {row.get('url', '')}"
        )

    def _load_category_url_list(self, path: str) -> None:
        """엑셀 전체 행을 카테고리URL목록 리스트박스에 표시."""
        self.lib_list.delete(0, "end")
        self._excel_rows = []
        self._active_ordinal = 0
        self._current_excel_path = ""
        self._lib_paths = []
        if not path or not os.path.isfile(path):
            self.p2_sel.configure(text="(파일 없음 — 위에서 엑셀을 선택하세요)")
            return
        try:
            rows = read_category_url_rows(path)
        except Exception as e:
            self.p2_sel.configure(text=f"(엑셀 읽기 실패: {e})")
            messagebox.showerror("엑셀 읽기 실패", str(e))
            return
        self._excel_rows = rows
        self._current_excel_path = path
        self._lib_paths = [path]
        for row in rows:
            self.lib_list.insert("end", self._format_category_row(row, active=False))
        set_selected(path)
        self.p2_sel.configure(text=f"{path}  ·  총 {len(rows)}행")
        if rows:
            self.lib_list.see(0)

    def _highlight_active_category_row(self, ordinal: int) -> None:
        """현재 작업 진행중인 행을 적색으로 표시."""
        try:
            ord_i = int(ordinal or 0)
        except (TypeError, ValueError):
            ord_i = 0
        if not self._excel_rows:
            return
        prev = self._active_ordinal
        self._active_ordinal = ord_i
        # 이전 활성 행 복원
        if 1 <= prev <= len(self._excel_rows):
            idx = prev - 1
            self.lib_list.delete(idx)
            self.lib_list.insert(
                idx, self._format_category_row(self._excel_rows[idx], active=False)
            )
            self.lib_list.itemconfig(idx, foreground="#0f172a", background="#ffffff")
        # 새 활성 행 적색
        if 1 <= ord_i <= len(self._excel_rows):
            idx = ord_i - 1
            self.lib_list.delete(idx)
            self.lib_list.insert(
                idx, self._format_category_row(self._excel_rows[idx], active=True)
            )
            self.lib_list.itemconfig(idx, foreground="#b91c1c", background="#fee2e2")
            self.lib_list.see(idx)

    def _refresh_p2_list(self) -> None:
        """파일 목록 새로고침 + 현재 엑셀 카테고리URL목록 재로드."""
        cur = self._current_excel_path
        self._search_xlsx()
        if cur and os.path.isfile(cur):
            self._load_category_url_list(cur)
            if self._active_ordinal:
                self._highlight_active_category_row(self._active_ordinal)
        elif not self._current_excel_path:
            # 보관 라이브러리 last_selected 우선
            data = load()
            last = str(data.get("last_selected") or "").strip()
            if last and os.path.isfile(last):
                self._load_category_url_list(last)

    def _remove_lib(self) -> None:
        path = self._current_excel_path
        if not path:
            sel = self.found_list.curselection()
            if sel:
                path = self._found_paths[sel[0]]
        if not path:
            messagebox.showinfo("안내", "제거할 파일이 없습니다.")
            return
        remove_path(path)
        self._current_excel_path = ""
        self._excel_rows = []
        self._active_ordinal = 0
        self.lib_list.delete(0, "end")
        self.p2_sel.configure(text="(비어 있음 — 위에서 엑셀을 선택하세요)")
        self._search_xlsx()

    def _on_lib_mousewheel(self, event) -> str:
        """카테고리URL목록 스크롤 (Windows/macOS/Linux)."""
        if getattr(event, "num", None) == 4 or getattr(event, "delta", 0) > 0:
            self.lib_list.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5 or getattr(event, "delta", 0) < 0:
            self.lib_list.yview_scroll(1, "units")
        return "break"

    def _clear_p2_log(self) -> None:
        for tv in (getattr(self, "p2_main_log", None), getattr(self, "p2_sub_log", None)):
            if tv is not None:
                for item in tv.get_children():
                    tv.delete(item)
        self._sub_by_seq = {}
        self._shot_path_by_seq = {}
        self._main_item_by_seq = {}
        self._main_ts_end = {}
        self._seq_by_main_item = {}
        self._meta_item_id = None
        self._meta_values = {f: "" for f in META_FIELDS}
        self._selected_seq = None
        self._latest_seq = 0
        self._follow_latest = True
        self._setup_meta_rows()

    def _setup_meta_rows(self) -> None:
        """main 상단 엑셀 진행 정보 — 5항목을 1줄(오렌지)로 표시."""
        tv = getattr(self, "p2_main_log", None)
        if tv is None:
            return
        self._meta_values = {f: "" for f in META_FIELDS}
        line = format_meta_line(self._meta_values)
        self._meta_item_id = tv.insert("", 0, values=("", "엑셀", line), tags=("meta",))

    def _update_meta_row(self, field: str, value: str) -> None:
        # ★요건: 순번 META 삭제 — 진행행 적색은 내부필드 '진행'으로만
        if field in META_INTERNAL_FIELDS:
            try:
                ord_i = int(str(value or "0").strip() or "0")
            except ValueError:
                ord_i = 0
            if ord_i > 0:
                self._highlight_active_category_row(ord_i)
            return
        if field not in META_FIELDS:
            return
        self._meta_values[field] = str(value or "").strip()
        if not self._meta_item_id:
            return
        line = format_meta_line(self._meta_values)
        self.p2_main_log.item(self._meta_item_id, values=("", "엑셀", line))

    def _setup_p2_log_tags(self) -> None:
        """main 실행로그 — 단계 성격별 색상 태그."""
        tv = self.p2_main_log
        tv.tag_configure("meta", foreground="#ea580c", background="#fff7ed")
        tv.tag_configure("normal", foreground="#0f172a")
        tv.tag_configure("login", foreground="#7c3aed", background="#f5f3ff")
        tv.tag_configure("init", foreground="#0f766e", background="#f0fdfa")
        tv.tag_configure("save", foreground="#5b21b6", background="#f3e8ff")
        tv.tag_configure("done", foreground="#166534", background="#dcfce7")

    def _toggle_log_panels(self) -> None:
        """MAIN / SUB 체크박스 — 실행로그 패널 표시/숨김 (MAIN 위 · SUB 아래)."""
        show_main = bool(self.var_show_main.get())
        show_sub = bool(self.var_show_sub.get())
        main_f = getattr(self, "p2_main_frame", None)
        sub_f = getattr(self, "p2_sub_frame", None)
        if main_f is not None:
            main_f.pack_forget()
        if sub_f is not None:
            sub_f.pack_forget()
        if show_main and main_f is not None:
            main_f.pack(fill="both", expand=True)
        if show_sub and sub_f is not None:
            sub_f.pack(
                fill="both",
                expand=True,
                pady=(6, 0) if show_main else (0, 0),
            )

    def _handle_collect_line(self, message: str) -> None:
        """collect.py stdout 한 줄 처리 — main/sub 프로토콜만 그리드에 반영.

        (요건: main엔 1~13단계만, 그 외 잡다한 로그는 화면에 출력하지 않음)
        """
        text = (message or "").rstrip()
        if not text:
            return
        t, text = strip_timestamp(text)
        parsed = parse_line(text)
        if parsed is None:
            return  # 마커 없는 줄은 화면에 출력하지 않음 — 요건 2

        kind = parsed[0]
        if kind == "meta":
            _, field, value = parsed
            self._update_meta_row(field, value)
        elif kind == "main":
            _, seq, n, msg = parsed
            self._insert_main_row(t, seq, n, msg)
        elif kind == "sub":
            _, seq, msg = parsed
            self._append_sub_entry(seq, t, "info", msg)
        elif kind == "subshot":
            _, seq, path, label = parsed
            self._capture_shot_dir_from_path(path)
            self._append_sub_entry(seq, t, "shot", f"[샷] {label} -> {Path(path).name}")
            self._shot_path_by_seq[(seq, len(self._sub_by_seq.get(seq, [])) - 1)] = path

    def _main_ts_for_seq(self, seq: int) -> str | None:
        """main 그리드에 기록된 시각 — sub와 동일하게 맞출 때 사용."""
        item = self._main_item_by_seq.get(seq)
        if not item:
            return None
        vals = self.p2_main_log.item(item, "values")
        return vals[0] if vals else None

    def _ts_for_sub(self, seq: int, t: str) -> str:
        """sub 시각 = 현단계 MAIN 진입 ~ 다음 MAIN 진입."""
        if "~" in (t or ""):
            return t
        start = self._main_ts_for_seq(seq)
        if not start:
            return t
        end = self._main_ts_end.get(seq, start)
        return sub_time_range(start, end)

    def _insert_main_row(self, t: str, seq: int, n: int, msg: str) -> None:
        if seq > 1:
            self._main_ts_end[seq - 1] = t
            if self._selected_seq == seq - 1:
                self._render_sub_grid(seq - 1)
        tag = step_tag(n)
        # step=0 → 엑셀 5필드 한 줄(오렌지). 표시는 sticky META와 동일하게 "엑셀"
        step_label: str | int = "엑셀" if n == 0 else n
        item = self.p2_main_log.insert(
            "", "end", values=(t, step_label, msg), tags=(tag,)
        )
        self._main_item_by_seq[seq] = item
        self._seq_by_main_item[item] = seq
        self._latest_seq = max(self._latest_seq, seq)
        self.p2_main_log.see(item)
        if self._follow_latest:
            self.p2_main_log.selection_set(item)
            self._selected_seq = seq
            self._render_sub_grid(seq)

    def _append_sub_entry(self, seq: int, t: str, kind: str, msg: str) -> None:
        display_t = self._ts_for_sub(seq, t)
        self._sub_by_seq.setdefault(seq, []).append((display_t, kind, msg))
        if self._selected_seq == seq:
            tag = ("shot",) if kind == "shot" else ()
            item = self.p2_sub_log.insert("", "end", values=(display_t, msg), tags=tag)
            self.p2_sub_log.see(item)

    def _render_sub_grid(self, seq: int) -> None:
        for item in self.p2_sub_log.get_children():
            self.p2_sub_log.delete(item)
        for t, kind, msg in self._sub_by_seq.get(seq, []):
            display_t = self._ts_for_sub(seq, t)
            tag = ("shot",) if kind == "shot" else ()
            self.p2_sub_log.insert("", "end", values=(display_t, msg), tags=tag)

    def _on_main_log_select(self, _evt=None) -> None:
        sel = self.p2_main_log.selection()
        if not sel:
            return
        seq = self._seq_by_main_item.get(sel[0])
        if seq is None:
            return
        if self._meta_item_id and sel[0] == self._meta_item_id:
            return
        self._selected_seq = seq
        self._follow_latest = seq == self._latest_seq
        self._render_sub_grid(seq)

    def _on_sub_log_double_click(self, _evt=None) -> None:
        sel = self.p2_sub_log.selection()
        if not sel or self._selected_seq is None:
            return
        idx = self.p2_sub_log.index(sel[0])
        path = self._shot_path_by_seq.get((self._selected_seq, idx))
        if not path:
            return
        p = Path(path)
        if not p.is_file():
            return
        try:
            if os.name == "nt":
                os.startfile(str(p))  # type: ignore[attr-defined]
            else:
                webbrowser.open(p.as_uri())
        except Exception:  # noqa: BLE001
            pass

    def _capture_shot_dir_from_path(self, path: str) -> None:
        try:
            p = Path(path)
            if p.parent.is_dir():
                self._last_shot_dir = p.parent
        except Exception:  # noqa: BLE001
            pass

    def _show_shot_viewer(self) -> None:
        folder = self._last_shot_dir
        if folder is None or not folder.is_dir():
            folder = latest_shot_dir(ROOT)
        open_shot_viewer(self, shot_dir=folder, root=ROOT)

    def _p2_log_ui(self, message: str) -> None:
        self.after(0, lambda: self._handle_collect_line(message))

    def _run_p2(self) -> None:
        path = self._current_excel_path
        if not path:
            sel = self.found_list.curselection()
            if sel:
                path = self._found_paths[sel[0]]
                self._load_category_url_list(path)
        if not path:
            messagebox.showinfo(
                "안내", "디렉터리 파일 목록에서 엑셀을 선택한 뒤 실행하세요."
            )
            return
        if not os.path.isfile(path):
            messagebox.showerror("오류", f"파일 없음:\n{path}")
            return
        # 라이브러리에 없으면 자동 등록 (선택 파일로 바로 실행 가능)
        if not is_in_library(path):
            add_paths([path])
        if self._p2_proc and self._p2_proc.poll() is None:
            messagebox.showwarning("실행 중", "이미 수집이 진행 중입니다.")
            return

        # 실행 직전 카테고리URL목록 최신화
        self._load_category_url_list(path)
        if not self._excel_rows:
            messagebox.showerror("오류", "엑셀에 처리할 카테고리URL 행이 없습니다.")
            return

        collect_py = ROOT / "P2" / "collect.py"
        stop_flag = ROOT / "P2" / ".collect_stop"
        try:
            stop_flag.unlink(missing_ok=True)  # type: ignore[call-arg]
        except TypeError:
            if stop_flag.exists():
                try:
                    stop_flag.unlink()
                except OSError:
                    pass
        except OSError:
            pass

        verify = bool(self.var_verify.get())
        args = [
            sys.executable,
            str(collect_py),
            path,
            "3",
            "--retries",
            "1",  # ★요건: 엑셀 각 행은 1번 시도로 끝냄(재시도 없음)
            "--yes",
            "--shot-first",
            "2",
        ]
        if verify:
            # ★스크린샷만 1·2행 — 처리 행 수는 엑셀 전체 (max_rows 강제 금지)
            args.append("--verify")

        set_selected(path)
        self._clear_p2_log()
        self._highlight_active_category_row(0)
        mode = "엑셀전체수집·1·2행샷" if verify else "엑셀전체수집"
        self.p2_status.configure(
            text=(
                f"수집 시작 ({mode}): {Path(path).name} "
                f"/ 총 {len(self._excel_rows)}행 — 브라우저에서 직접 로그인하세요"
            ),
            fg="#15803d",
        )

        try:
            # 보드 하단 로그로 stdout 수신 (별도 콘솔 창 없음)
            # Windows 기본 콘솔 코드페이지(CP949)와 UTF-8 혼용 대비:
            # 자식 Python은 UTF-8 강제 + 수신 시 utf-8/cp949 폴백 디코딩
            creationflags = 0
            if os.name == "nt":
                creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUTF8"] = "1"
            self._p2_proc = subprocess.Popen(
                args,
                cwd=str(ROOT / "P2"),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=False,
                bufsize=0,
                env=env,
                creationflags=creationflags,
            )
        except Exception as e:
            messagebox.showerror("실행 실패", str(e))
            self.p2_status.configure(text=f"실행 실패: {e}", fg="#b91c1c")
            return

        threading.Thread(
            target=self._watch_p2_proc,
            args=(self._p2_proc, path),
            daemon=True,
        ).start()

    def _stop_flag_path(self) -> Path:
        return ROOT / "P2" / ".collect_stop"

    def _stop_p2(self) -> None:
        """중도 수집 중단 — 화면 실행로그는 지우지 않고 보존."""
        proc = self._p2_proc
        if proc is None or proc.poll() is not None:
            messagebox.showinfo("안내", "실행 중인 수집이 없습니다.")
            return
        try:
            self._stop_flag_path().write_text("stop\n", encoding="utf-8")
        except OSError as e:
            self.p2_status.configure(text=f"중단 플래그 기록 실패: {e}", fg="#b91c1c")
        self.p2_status.configure(text="수집 종료 요청 중… (로그 보존)", fg="#b45309")
        threading.Thread(target=self._force_stop_p2, args=(proc,), daemon=True).start()

    def _force_stop_p2(self, proc: subprocess.Popen) -> None:
        """협조적 중단 후 응답 없으면 프로세스 종료."""
        for _ in range(24):  # ~12초
            if proc.poll() is not None:
                return
            time.sleep(0.5)
        if proc.poll() is not None:
            return
        self.after(
            0,
            lambda: self.p2_status.configure(
                text="협조 중단 지연 — 프로세스 강제 종료", fg="#b45309"
            ),
        )
        try:
            proc.terminate()
        except Exception:  # noqa: BLE001
            pass
        time.sleep(1.5)
        if proc.poll() is None:
            try:
                proc.kill()
            except Exception:  # noqa: BLE001
                pass

    @staticmethod
    def _decode_log_bytes(raw: bytes) -> str:
        """Windows CP949 / UTF-8 혼용 stdout을 깨지지 않게 디코딩."""
        if not raw:
            return ""
        data = raw.replace(b"\r\n", b"\n").replace(b"\r", b"\n")
        for enc in ("utf-8", "cp949", "mbcs"):
            try:
                return data.decode(enc)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    def _watch_p2_proc(self, proc: subprocess.Popen, path: str) -> None:
        try:
            assert proc.stdout is not None
            buf = b""
            while True:
                chunk = proc.stdout.read(256)
                if not chunk:
                    break
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    text = self._decode_log_bytes(line).rstrip()
                    if text:
                        self._p2_log_ui(text)
            if buf.strip():
                text = self._decode_log_bytes(buf).rstrip()
                if text:
                    self._p2_log_ui(text)
        except Exception as e:  # noqa: BLE001
            self.after(
                0,
                lambda: self.p2_status.configure(text=f"로그 수신 오류: {e}", fg="#b91c1c"),
            )
        code = proc.wait()
        if code == 0:
            self.after(0, lambda: self._on_p2_finished(True, path, code))
        elif code == 130:
            self.after(0, lambda: self._on_p2_finished(False, path, code, stopped=True))
        else:
            self.after(0, lambda: self._on_p2_finished(False, path, code))

    def _on_p2_finished(
        self,
        ok: bool,
        path: str,
        code: int = 0,
        *,
        stopped: bool = False,
    ) -> None:
        # 중단/완료 모두 실행로그는 그대로 둔다 (_clear_p2_log 호출 없음)
        if stopped:
            self.p2_status.configure(
                text="수집 종료(사용자 중단) — 실행로그 보존됨",
                fg="#b45309",
            )
            return
        if ok:
            self.p2_status.configure(text=f"수집 완료: {path}", fg="#15803d")
            folder = self._last_shot_dir or latest_shot_dir(ROOT)
            if folder and folder.is_dir() and any(folder.glob("*.png")):
                # 1행 전과정 샷이 있으면 바로 보여 줌
                if bool(self.var_verify.get()):
                    open_shot_viewer(self, shot_dir=folder, root=ROOT)
        else:
            self.p2_status.configure(text=f"수집 실패 (exit={code})", fg="#b91c1c")
            folder = self._last_shot_dir or latest_shot_dir(ROOT)
            if folder and folder.is_dir() and any(folder.glob("*.png")):
                if messagebox.askyesno(
                    "스크린샷",
                    f"실패했지만 단계 스크린샷이 있습니다.\n{folder}\n\n지금 볼까요?",
                    parent=self,
                ):
                    open_shot_viewer(self, shot_dir=folder, root=ROOT)


def main() -> None:
    app = BoardApp()
    app.mainloop()


if __name__ == "__main__":
    main()
