"""카테고리URL목록 엑셀 행 읽기 단위테스트."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from library import read_category_url_rows  # noqa: E402


def test_read_category_url_rows(tmp_path: Path):
    """★요건(2026-08-20): 카테고리URL목록에는 "최종 카테고리명"을 표시한다."""
    fp = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["최종 카테고리명", "최종 카테고리 URL주소", "기타"])
    ws.append(["목차", "https://example.com/toc", "skip-toc"])  # 총건수 제외
    ws.append(["MEN 스니커즈", "https://example.com/a", "x"])
    ws.append(["", "", "skip"])
    ws.append(["WOMEN", "https://example.com/b", "y"])
    wb.save(fp)

    rows = read_category_url_rows(str(fp))
    assert len(rows) == 2  # 목차·빈행 제외
    assert rows[0]["ordinal"] == 1
    assert rows[0]["label"] == "MEN 스니커즈"
    assert rows[0]["url"] == "https://example.com/a"
    assert rows[1]["ordinal"] == 2
    assert rows[1]["label"] == "WOMEN"
    assert all(r["label"] != "목차" for r in rows)


def test_final_category_name_wins_over_top_final(tmp_path: Path):
    """두 열이 다 있으면 "최종 카테고리명" 쪽을 쓴다."""
    fp = tmp_path / "both.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["상위 최종 카테고리명", "최종 카테고리명", "최종 카테고리 URL주소"])
    ws.append(["MEN", "MEN 스니커즈", "https://example.com/a"])
    wb.save(fp)

    rows = read_category_url_rows(str(fp))
    assert [r["label"] for r in rows] == ["MEN 스니커즈"]


def test_falls_back_to_top_final_for_old_excel(tmp_path: Path):
    """하위 호환: "최종 카테고리명"이 없는 옛 엑셀은 "상위 최종 카테고리명"."""
    fp = tmp_path / "old.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["상위 최종 카테고리명", "최종 카테고리 URL주소"])
    ws.append(["MEN 스니커즈", "https://example.com/a"])
    wb.save(fp)

    rows = read_category_url_rows(str(fp))
    assert [r["label"] for r in rows] == ["MEN 스니커즈"]


if __name__ == "__main__":
    import tempfile

    with tempfile.TemporaryDirectory() as d:
        test_read_category_url_rows(Path(d))
        test_final_category_name_wins_over_top_final(Path(d))
        test_falls_back_to_top_final_for_old_excel(Path(d))
    print("PASS read_category_url_rows")
