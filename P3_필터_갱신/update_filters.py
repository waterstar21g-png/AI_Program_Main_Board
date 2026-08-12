"""
P3_필터_갱신 — 더망고 검색필터(저장조건) 화면의 저장상품수 갱신.

0) 이미 열린 현재 망고 화면에서 진행 (로그인·진입 없음)
1) ★더망고 행의 URL을 기준값으로 엑셀에서 동일 URL을 찾음
   → 검색필터 동일 시 진행 (엑셀 중간공백→'_' 재비교 포함)

필터 일치 시:
  1) URL 클릭
  2) 첫 팝업창 닫기
  3) 스크롤 푸터까지 내리기
  4) 하단→상단 스크롤하며 상품수 카드 갯수 로그
  5) 엑셀 상품수와 비교값 출력
  6) URL 바로 오른쪽 「수집조건수정」만 클릭 — 2초 간격 최대 5회(팝업 열릴 때까지, href 금지)
  7) 팝업에서 저장상품수 수정 → 저장하기 → 확인

로그:
- 더망고 처음 10건: 더망고/엑셀 각각 1줄(검색필터·URL)
- ★망고: 이미 열려 있는 현재 망고 화면에서만 진행 (로그인·URL진입 없음)
- ★필터 매칭 시: 팝업/탭을 bring_to_front 로 실제 Chrome에 표시 + 단계샷
- 불일치 행: 검색필터·URL만

사용법:
  python update_filters.py 엑셀.xlsx --mango-url "https://..."
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
P2_DIR = ROOT / "P2"
P1_101_DIR = ROOT / "P1_101"
if str(P2_DIR) not in sys.path:
    sys.path.insert(0, str(P2_DIR))
if str(P1_101_DIR) not in sys.path:
    sys.path.insert(0, str(P1_101_DIR))

ProgressFn = Callable[[str, str], None]

STOP_FLAG_PATH = Path(__file__).resolve().parent / ".filter_stop"
P3_RUN_LOG_DIR = Path(__file__).resolve().parent / "run-logs"
P3_SHOT_MARK = "##P3SHOT##"  # ##P3SHOT##<path>##<label>
DETAIL_EXCEL_ROWS = 5  # 엑셀 1~5행 매칭 시 세부 로그
FIRST_COMPARE_LOG_N = 10  # 더망고 처음 10건: 더망고/엑셀 비교 2줄 로그

# ★요건: 보드 「더망고 URL」초기값 (검색필터·저장조건 화면)
DEFAULT_MANGO_URL = (
    "https://tmg1898.cafe24.com/mall/admin/shop/getGoodsCategory.php"
    "?pmode=filter_delete&uids=&pg=1&date_type=modify"
    "&start_yy=2026&start_mm=8&start_dd=12"
    "&end_yy=2026&end_mm=8&end_dd=12"
    "&site_id=zara_de&sales_yn=&sch_keyword="
    "&ft_num=all&ft_show=&ft_sort=modify_asc"
)
LAST_MANGO_URL_PATH = Path(__file__).resolve().parent / ".last_mango_url"


def load_mango_url_default() -> str:
    """★고정 초기값 — 검색필터(저장조건) getGoodsCategory.php URL."""
    return DEFAULT_MANGO_URL


def save_mango_url(url: str) -> None:
    """더망고 URL 기억 파일에 저장 (보드 표시용). 비어 있으면 초기값 저장."""
    u = (url or "").strip() or DEFAULT_MANGO_URL
    if not u.lower().startswith("http"):
        u = DEFAULT_MANGO_URL
    try:
        LAST_MANGO_URL_PATH.write_text(u + "\n", encoding="utf-8")
    except Exception:
        pass

URL_HEADERS = (
    "검색필터 URL",
    "최종 카테고리 URL주소",
    "최종 카테고리 URL",
    "카테고리 URL",
    "URL주소",
    "URL",
    "url",
)
FILTER_HEADERS = (
    "검색필터",
    "검색필터명",
    "상위 최종 카테고리명",
    "최종 카테고리명",
    "카테고리명",
)
COLLECTIBLE_HEADERS = ("상품수집가능개수", "총상품수")


@dataclass
class ExcelRow:
    excel_row: int  # 1-based sheet row
    url: str
    filter_name: str
    collectible: int


@dataclass
class RunResult:
    ok: bool
    total_demango: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0
    errors: list[str] = field(default_factory=list)


def _log(progress: ProgressFn | None, step: str, message: str) -> None:
    print(f"[{step}] {message}", flush=True)
    if progress:
        progress(step, message)


# ★P2와 동일: 실제 Chrome(CDP) 창을 OS 앞으로 가져와 동작을 보여 줌
STEP_VIEW_DWELL_SEC = 1.2


def describe_page_state(page) -> str:
    """화면 요약 한 줄 — URL · 제목 · 탭수."""
    url = ""
    title = ""
    tabs = 0
    try:
        url = (page.url or "").strip()
    except Exception:
        url = "(url읽기실패)"
    try:
        title = (page.title() or "").strip()
    except Exception:
        title = ""
    try:
        tabs = len(list(page.context.pages))
    except Exception:
        tabs = 0
    parts = [f"url={url[:140]}"]
    if title:
        parts.append(f"title={title[:80]}")
    if tabs:
        parts.append(f"tabs={tabs}")
    return " · ".join(parts)


def _activate_chrome_window(page) -> None:
    """최소화/뒤로 간 Chrome 창을 복원·앞으로 (CDP Browser.setWindowBounds)."""
    if page is None:
        return
    try:
        session = page.context.new_cdp_session(page)
    except Exception:
        return
    try:
        # targetId 없이 현재 페이지 기준 windowId 조회
        info = session.send("Browser.getWindowForTarget")
        wid = info.get("windowId") if isinstance(info, dict) else None
        if wid is None:
            return
        session.send(
            "Browser.setWindowBounds",
            {"windowId": wid, "bounds": {"windowState": "normal"}},
        )
    except Exception:
        pass
    finally:
        try:
            session.detach()
        except Exception:
            pass
    # Windows: 작업표시줄에 숨은 Chrome을 강제 앞으로
    if os.name == "nt":
        try:
            import ctypes

            user32 = ctypes.windll.user32  # type: ignore[attr-defined]
            SW_RESTORE = 9
            hwnd = user32.GetForegroundWindow()
            # Chrome 창 찾기 (클래스 Chrome_WidgetWin_1)
            found = ctypes.c_void_p(0)

            @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
            def _enum(h, _l):  # noqa: ANN001
                length = user32.GetWindowTextLengthW(h)
                if length <= 0:
                    return True
                buf = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(h, buf, length + 1)
                title = buf.value or ""
                if "Chrome" in title or "더망고" in title or "cafe24" in title.lower():
                    found.value = h
                    return False
                return True

            user32.EnumWindows(_enum, 0)
            if found.value:
                user32.ShowWindow(found.value, SW_RESTORE)
                user32.SetForegroundWindow(found.value)
            elif hwnd:
                user32.ShowWindow(hwnd, SW_RESTORE)
        except Exception:
            pass


def reveal_browser_page(
    page,
    progress: ProgressFn | None,
    *,
    step_no: str,
    action: str,
    dwell_s: float | None = None,
) -> None:
    """P2와 동일 — 망고 Chrome 창·팝업을 화면에 보이게 한다."""
    if page is None:
        return
    dwell = STEP_VIEW_DWELL_SEC if dwell_s is None else max(0.0, float(dwell_s))
    try:
        if hasattr(page, "is_closed") and page.is_closed():
            _log(progress, "화면", f"{step_no}) {action} — 창이 이미 닫힘")
            return
    except Exception:
        pass
    _activate_chrome_window(page)
    try:
        page.bring_to_front()
    except Exception:
        pass
    try:
        page.evaluate("() => { try { window.focus(); } catch (e) {} }")
    except Exception:
        pass
    state = ""
    try:
        state = describe_page_state(page)
    except Exception:
        state = ""
    msg = f"{step_no}) {action}"
    if state:
        msg += f" · {state}"
    _log(progress, "화면", msg + " ← 망고 Chrome 창 표시")
    if dwell > 0:
        time.sleep(dwell)


def attach_mango_browser_like_p2(p2, playwright, *, progress: ProgressFn | None = None):
    """★P2와 동일: connect_browser 로 실제 Chrome을 띄우거나 연결한 뒤 창을 앞으로.

    로그인 대기는 하지 않는다. 이후 검색필터 URL로 이동한다.
    """
    _log(
        progress,
        "준비",
        "P2와 동일 — 망고 Chrome(CDP) 연결/실행 · 화면에 창 표시",
    )
    browser, page = p2.connect_browser(playwright)
    if hasattr(p2, "refresh_if_closed"):
        page = p2.refresh_if_closed(page)
    try:
        page.set_default_timeout(120_000)
    except Exception:
        pass
    reveal_browser_page(
        page,
        progress,
        step_no="0",
        action="망고 Chrome 연동 창 표시",
        dwell_s=1.5,
    )
    try:
        cur = (page.url or "").strip()
    except Exception:
        cur = ""
    _log(progress, "준비", f"연결 직후 URL={cur[:160] or '(없음)'} → 검색필터 URL로 이동")
    return browser, page


# 하위호환 별칭
attach_current_mango_page = attach_mango_browser_like_p2


def clear_stop_flag() -> None:
    try:
        STOP_FLAG_PATH.unlink(missing_ok=True)  # type: ignore[call-arg]
    except TypeError:
        if STOP_FLAG_PATH.exists():
            try:
                STOP_FLAG_PATH.unlink()
            except OSError:
                pass
    except OSError:
        pass


def stop_requested() -> bool:
    return STOP_FLAG_PATH.is_file()


def normalize_url(url: str) -> str:
    """URL 비교용 정규화 (끝 슬래시·쿼리 일부 무시하지 않음, 스킴/호스트 소문자)."""
    s = (url or "").strip()
    if not s:
        return ""
    try:
        p = urlparse(s)
        path = unquote(p.path or "")
        if path.endswith("/") and len(path) > 1:
            path = path[:-1]
        netloc = (p.netloc or "").lower()
        scheme = (p.scheme or "https").lower()
        query = p.query or ""
        return f"{scheme}://{netloc}{path}" + (f"?{query}" if query else "")
    except Exception:
        return s.rstrip("/").lower()


def map_save_count(collectible: int) -> int:
    """상품수집가능개수 → 더망고 저장상품수."""
    n = max(0, int(collectible))
    if n <= 200:
        return n
    if n <= 500:
        return 300
    return 400


def _header_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    lowered = [(h or "").strip() for h in headers]
    for cand in candidates:
        for i, h in enumerate(lowered):
            if h == cand:
                return i
    for cand in candidates:
        c = cand.lower()
        for i, h in enumerate(lowered):
            if c in h.lower():
                return i
    return None


def _parse_int(raw) -> int:
    if raw is None:
        return 0
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return max(0, int(raw))
    s = re.sub(r"[^\d]", "", str(raw))
    return int(s) if s else 0


def read_excel_rows(path: Path) -> list[ExcelRow]:
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_vals = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("엑셀이 비어 있습니다.")
    headers = [str(h).strip() if h is not None else "" for h in header_vals]
    url_i = _header_index(headers, URL_HEADERS)
    if url_i is None:
        wb.close()
        raise ValueError(
            "URL 열 없음 — '검색필터 URL' 또는 '최종 카테고리 URL주소' 필요"
        )
    filter_i = _header_index(headers, FILTER_HEADERS)
    coll_i = _header_index(headers, COLLECTIBLE_HEADERS)
    out: list[ExcelRow] = []
    for offset, vals in enumerate(rows_iter, start=2):
        cells = list(vals) if vals else []
        url = str(cells[url_i] or "").strip() if url_i < len(cells) else ""
        if not url or not url.lower().startswith("http"):
            continue
        fname = ""
        if filter_i is not None and filter_i < len(cells):
            fname = str(cells[filter_i] or "").strip()
        coll = 0
        if coll_i is not None and coll_i < len(cells):
            coll = _parse_int(cells[coll_i])
        out.append(
            ExcelRow(
                excel_row=offset,
                url=url,
                filter_name=fname,
                collectible=coll,
            )
        )
    wb.close()
    return out


def excel_by_url(rows: list[ExcelRow]) -> dict[str, ExcelRow]:
    """엑셀 URL → 행. 조회 KEY는 더망고 URL(정규화) 기준."""
    m: dict[str, ExcelRow] = {}
    for r in rows:
        key = normalize_url(r.url)
        if key and key not in m:
            m[key] = r
    return m


def find_excel_by_demango_url(
    by_url: dict[str, ExcelRow], demango_url: str
) -> ExcelRow | None:
    """★요건: 더망고 URL을 기준값으로 엑셀에서 동일 URL 행을 찾는다."""
    key = normalize_url(demango_url)
    if not key:
        return None
    return by_url.get(key)


def log_first10_compare(
    progress: ProgressFn | None,
    *,
    ordinal: int,
    d_filter: str,
    d_url: str,
    ex: ExcelRow | None,
) -> None:
    """더망고 처음 10건 — 더망고/엑셀 검색필터·URL을 두 줄로 표시."""
    if ordinal > FIRST_COMPARE_LOG_N:
        return
    _log(
        progress,
        "비교",
        f"더망고 · 검색필터={d_filter} · URL={d_url}",
    )
    if ex is not None:
        _log(
            progress,
            "비교",
            f"엑셀 · 검색필터={ex.filter_name} · URL={ex.url}",
        )
    else:
        _log(
            progress,
            "비교",
            "엑셀 · 검색필터=- · URL=-",
        )


def filters_equal(excel_filter: str, demango_filter: str) -> bool:
    """검색필터 비교.

    1) 그대로 비교
    2) 불일치이고 엑셀 검색필터 값 *중간*에 공백이 있으면 공백→'_' 치환 후 재비교
    """
    a = (excel_filter or "").strip()
    b = (demango_filter or "").strip()
    if a == b:
        return True
    # 중간에 공백이 있는 경우만 (앞뒤 trim 이후에도 공백 존재)
    if " " in a:
        if a.replace(" ", "_") == b:
            return True
    return False


def filter_compare_note(excel_filter: str, demango_filter: str) -> str:
    """로그용 — 공백→_ 재비교로 맞은 경우 메모."""
    a = (excel_filter or "").strip()
    b = (demango_filter or "").strip()
    if a == b:
        return ""
    if " " in a and a.replace(" ", "_") == b:
        return "엑셀 중간공백→'_' 재비교 일치"
    return ""


def _detail(excel_row: int | None) -> bool:
    """엑셀 처음 5행(데이터 행 기준 sheet row 중 앞 5건) — excel_row 번호가 작을수록 앞행.

    데이터 시작이 보통 2행이므로 sheet row 2~6 을 세부 로그로 본다.
    """
    if excel_row is None:
        return False
    # header=1 → 데이터 1번째 행 = excel_row 2 … 5번째 = excel_row 6
    return 2 <= excel_row <= (1 + DETAIL_EXCEL_ROWS)


class Logger:
    def __init__(
        self,
        progress: ProgressFn | None,
        excel_row: int | None = None,
        *,
        force_verbose: bool = False,
    ):
        self.progress = progress
        self.excel_row = excel_row
        # ★필터 매칭 행은 전 단계 상세(화면상태·동작)를 그대로 출력
        self.verbose = bool(force_verbose) or _detail(excel_row)

    def step(self, step: str, detail: str, summary: str | None = None) -> None:
        if self.verbose:
            _log(self.progress, step, detail)
        else:
            _log(self.progress, step, summary if summary is not None else detail)


def navigate_mango_url(
    page,
    mango_url: str,
    *,
    progress: ProgressFn | None,
    p2=None,
):
    """검색필터(저장조건) URL로 이동 — 로그인 대기 없음. P2 safe_goto 사용."""
    url = (mango_url or "").strip() or DEFAULT_MANGO_URL
    _log(progress, "로직", f"0) 검색필터 URL 이동: {url}")
    for attempt in range(1, 3):
        try:
            if p2 is not None and hasattr(p2, "safe_goto"):
                p2.safe_goto(page, url)
                if hasattr(p2, "refresh_if_closed"):
                    page = p2.refresh_if_closed(page)
            else:
                page.goto(url, wait_until="domcontentloaded", timeout=90_000)
        except Exception as e:  # noqa: BLE001
            _log(
                progress,
                "경고",
                f"0) URL 이동 예외({attempt}/2): {str(e).split(chr(10))[0][:120]}",
            )
        time.sleep(0.6)
        try:
            cur = (page.url or "").strip()
        except Exception:
            cur = ""
        _log(progress, "준비", f"이동 후 화면 URL={cur[:180] or '(없음)'}")
        # admin.php 등에 남으면 재시도
        if "getGoodsCategory" in cur or "filter" in cur.lower():
            break
        if attempt < 2:
            _log(progress, "경고", "검색필터 화면 미도달 — URL 재이동")
    reveal_browser_page(
        page, progress, step_no="0", action="검색필터(저장조건) 화면 표시", dwell_s=0.5
    )
    try:
        body = page.locator("body").inner_text(timeout=3000) or ""
        if re.search(r"검색\s*필터|저장\s*조건|수집\s*조건|필터이름", body):
            _log(progress, "확인", "검색필터/저장조건 화면 문구 확인")
        else:
            _log(progress, "확인", "화면 문구 미검출 — 지정 URL로 계속 진행")
    except Exception:
        pass
    return page



# 더망고 목록 행 스캔 JS (스크린샷 컬럼 구조 기준) — 테스트에서도 재사용
LIST_DEMANGO_ROWS_JS = r"""() => {
  const out = [];
  const trs = Array.from(document.querySelectorAll('table tr, form tr, tr'));
  let filterCol = -1;
  let condCol = -1;
  let headerIdx = -1;

  const isBadFilterName = (v) => {
    if (!v) return true;
    const s = String(v).trim();
    if (!s) return true;
    if (/^https?:/i.test(s)) return true;
    if (/^\d+$/.test(s)) return true;
    if (/\.com(\/|$)/i.test(s)) return true;           // Zara.com/de 등 사이트열
    if (/\d+\s*개\s*\/\s*\d+\s*개/.test(s)) return true;
    if (/^\d{4}-\d{2}-\d{2}/.test(s)) return true;
    if (/URL\s*검색|수집조건|전체저장|상품확인|검색필터관리|별도관리/.test(s)) return true;
    if (s.length > 120) return true;
    return false;
  };

  const readInputValue = (inp) => {
    if (!inp) return '';
    return (inp.value || inp.getAttribute('value') || '').trim();
  };

  // 헤더: '필터이름(수정가능)' / '검색필터(저장조건)'
  for (let i = 0; i < trs.length; i++) {
    const cells = Array.from(trs[i].querySelectorAll('th, td'));
    if (cells.length < 2) continue;
    const labels = cells.map(c => (c.innerText || '').replace(/\s+/g, ''));
    const fi = labels.findIndex(t => t.includes('필터이름'));
    const ci = labels.findIndex(t =>
      t.includes('검색필터') && (t.includes('저장조건') || t.includes('저장'))
    );
    if (fi >= 0 || ci >= 0) {
      filterCol = fi;
      condCol = ci;
      headerIdx = i;
      break;
    }
  }

  for (let i = 0; i < trs.length; i++) {
    if (i === headerIdx) continue;
    const tr = trs[i];
    const t = (tr.innerText || '').replace(/\s+/g, ' ').trim();
    if (!t) continue;
    if (!/URL\s*검색|수집\s*조건\s*수정|https?:\/\//i.test(t)) continue;
    if (/필터이름\(수정가능\)/.test(t) && !/https?:\/\//i.test(t)) continue;

    const cells = Array.from(tr.querySelectorAll(':scope > td, :scope > th'));
    const cellsFallback = cells.length ? cells : Array.from(tr.querySelectorAll('td, th'));
    let filterName = '';
    let url = '';
    let editHref = '';
    let hasEdit = false;

    // 1) 검색필터 필드값 = '필터이름(수정가능)' 열의 <input> value
    const filterCell = (filterCol >= 0 && filterCol < cellsFallback.length)
      ? cellsFallback[filterCol] : null;
    if (filterCell) {
      const inp = filterCell.querySelector(
        'input[type="text"], input[type="search"], input:not([type])'
      );
      filterName = readInputValue(inp);
      if (isBadFilterName(filterName)) filterName = '';
    }
    if (!filterName) {
      // 폴백: 행 내 텍스트 input (사이트·숫자·URL 제외)
      const inputs = Array.from(tr.querySelectorAll('input')).filter(inp => {
        const ty = (inp.getAttribute('type') || 'text').toLowerCase();
        return ty === 'text' || ty === 'search' || ty === '';
      });
      for (const inp of inputs) {
        const v = readInputValue(inp);
        if (isBadFilterName(v)) continue;
        filterName = v;
        break;
      }
    }

    // 2) URL = '검색필터(저장조건)' 열의 'URL 검색:' 뒤 링크/텍스트
    const condCell = (condCol >= 0 && condCol < cellsFallback.length)
      ? cellsFallback[condCol] : tr;
    const scope = condCell || tr;
    const aHttp = scope.querySelector('a[href^="http"], a[href*="://"]');
    if (aHttp) {
      url = (aHttp.href || aHttp.getAttribute('href') || '').trim();
    }
    if (!url) {
      const raw = (scope.innerText || scope.textContent || '');
      const m = raw.match(/URL\s*검색\s*[:：]?\s*(https?:\/\/\S+)/i);
      if (m) url = m[1].replace(/[|｜].*$/, '').trim();
    }
    if (!url) {
      const m2 = t.match(/(https?:\/\/www\.zara\.com[^\s|]+)/i)
        || t.match(/(https?:\/\/[^\s|]+)/i);
      if (m2) url = m2[1].trim();
    }
    url = url.replace(/[\)\]\>,\;]+$/, '');

    // 3) 수집조건수정 버튼 → ★「수집개수 … 전체저장」바로 옆 버튼 우선
    const editNodes = Array.from(tr.querySelectorAll(
      'a, button, input[type="button"], input[type="submit"], input[value], span'
    ));
    const nearCollect = (el) => {
      let node = el;
      for (let d = 0; d < 5 && node; d++) {
        const parent = node.parentElement;
        if (!parent) break;
        let before = '';
        for (const child of Array.from(parent.childNodes)) {
          if (child === node || (child.contains && el && child.contains(el))) break;
          before += (child.innerText || child.textContent || '');
        }
        const compact = before.replace(/\s+/g, '');
        if (compact.includes('수집개수') && (
          compact.includes('전체저장') || /수집개수[:：]?\d+개/.test(compact)
        )) {
          return true;
        }
        const full = (parent.innerText || '').replace(/\s+/g, '');
        const iCnt = full.indexOf('수집개수');
        const iAll = full.indexOf('전체저장');
        const iBtn = full.indexOf('수집조건수정');
        if (iCnt >= 0 && iBtn > iCnt && (iAll < 0 || (iAll > iCnt && iAll < iBtn))) {
          return true;
        }
        node = parent;
      }
      return false;
    };
    const isEditLabel = (el) => {
      const label = (el.value || el.textContent || '').replace(/\s+/g, '');
      return label === '수집조건수정' || /^수집조건수정/.test(label);
    };
    const ranked = editNodes
      .filter(isEditLabel)
      .map((el, ord) => {
        const tag = (el.tagName || '').toUpperCase();
        let score = 0;
        if (nearCollect(el)) score += 100;
        if (tag === 'INPUT' || tag === 'BUTTON') score += 20;
        if (tag === 'A') score += 10;
        if (((el.value || el.textContent || '').replace(/\s+/g, '')) === '수집조건수정') {
          score += 15;
        }
        return { el, score, ord };
      })
      .sort((a, b) => b.score - a.score || a.ord - b.ord);
    for (const item of ranked) {
      const el = item.el;
      hasEdit = true;
      if (el.tagName === 'A') {
        editHref = el.href || el.getAttribute('href') || '';
      } else if (el.closest && el.closest('a')) {
        const a = el.closest('a');
        editHref = a.href || a.getAttribute('href') || '';
      }
      if (!editHref) {
        const oc = el.getAttribute('onclick') || '';
        const hrefInOc = oc.match(/location\.href\s*=\s*['"]([^'"]+)['"]/i)
          || oc.match(/window\.open\s*\(\s*['"]([^'"]+)['"]/i)
          || oc.match(/['"]([^'"]*admin_group_modify\.php[^'"]*)['"]/i);
        if (hrefInOc) editHref = hrefInOc[1];
        if (!editHref) {
          const fm = oc.match(/ps_fuid\s*=\s*(\d+)/)
            || oc.match(/modify_filter[^0-9]*(\d+)/i)
            || oc.match(/fuid["'\s:=]+(\d+)/i)
            || oc.match(/(?:go|open|modify|edit)\s*\(\s*(\d+)\s*\)/i);
          if (fm) {
            editHref = 'admin_group_modify.php?ps_mode=modify_filter&ps_fuid=' + fm[1];
          }
        }
      }
      break;
    }

    if (!url && !hasEdit && !filterName) continue;
    out.push({
      index: i,
      url,
      filterName,
      hasEdit,
      editHref,
      text: t.slice(0, 240),
    });
  }
  return out;
}"""


def list_demango_rows(page) -> list[dict]:
    """더망고 검색필터 목록 행 수집.

    스크린샷 순서·구조:
    1) 검색필터 필드값 = 열 '필터이름(수정가능)' 의 input.value (예: 여성헤어_헤어)
    2) URL = 열 '검색필터(저장조건)' 안의 'URL 검색:' 뒤 링크/텍스트
    3) 수집조건수정 = 같은 영역 버튼/링크 (href 또는 ps_fuid)
    """
    data = page.evaluate(LIST_DEMANGO_ROWS_JS)
    return list(data or [])


def _find_and_mark_edit_button(page, row_index: int, row_url: str = "") -> dict:
    """URL 바로 오른쪽·수집개수|전체저장 옆 「수집조건수정」을 data-p3-edit-target 마킹."""
    info = page.evaluate(
        """(args) => {
          const rowIndex = args.rowIndex;
          const urlHint = (args.urlHint || '').trim();
          const urlStem = urlHint.split('?')[0];
          document.querySelectorAll('[data-p3-edit-target]').forEach(el => {
            el.removeAttribute('data-p3-edit-target');
          });

          const isEditControl = (el) => {
            if (!el) return false;
            const tag = (el.tagName || '').toUpperCase();
            if (!(tag === 'A' || tag === 'BUTTON' || tag === 'INPUT')) return false;
            if (tag === 'INPUT') {
              const ty = (el.getAttribute('type') || 'button').toLowerCase();
              if (!(ty === 'button' || ty === 'submit' || ty === '')) return false;
            }
            const t = (el.value || el.textContent || '').replace(/\\s+/g, '');
            return t === '수집조건수정' || /^수집조건수정/.test(t);
          };

          const nearCollectCount = (el) => {
            let node = el;
            for (let depth = 0; depth < 5 && node; depth++) {
              const parent = node.parentElement;
              if (!parent) break;
              let before = '';
              for (const child of Array.from(parent.childNodes)) {
                if (child === node || (child.contains && child.contains(el))) break;
                before += (child.innerText || child.textContent || '');
              }
              const compact = before.replace(/\\s+/g, '');
              if (compact.includes('수집개수') && (
                compact.includes('전체저장') || /수집개수[:：]?\\d+개/.test(compact)
              )) {
                return true;
              }
              const full = (parent.innerText || '').replace(/\\s+/g, '');
              const iCnt = full.indexOf('수집개수');
              const iAll = full.indexOf('전체저장');
              const iBtn = full.indexOf('수집조건수정');
              if (iCnt >= 0 && iBtn > iCnt && (iAll < 0 || (iAll > iCnt && iAll < iBtn))) {
                return true;
              }
              node = parent;
            }
            return false;
          };

          const findUrlAnchor = (tr) => {
            const anchors = Array.from(tr.querySelectorAll('a[href]'));
            for (const a of anchors) {
              const h = a.href || a.getAttribute('href') || '';
              if (!h || h.indexOf('http') !== 0) continue;
              if (!urlHint) return a;
              if (h === urlHint || h.startsWith(urlStem) || urlHint.startsWith(h.split('?')[0])) {
                return a;
              }
            }
            return anchors.find(a => {
              const h = a.href || '';
              return h.indexOf('http') === 0 && !/수집조건수정/.test((a.textContent||''));
            }) || null;
          };

          const rightOfUrl = (el, urlA) => {
            if (!urlA || !el) return false;
            try {
              const ub = urlA.getBoundingClientRect();
              const eb = el.getBoundingClientRect();
              // URL 앵커의 오른쪽 (같은 행 근처)
              if (eb.left >= ub.right - 12) return true;
              if (eb.left > ub.left && eb.right > ub.right + 20) return true;
            } catch (e) {}
            return false;
          };

          const score = (el, urlA) => {
            let s = 0;
            if (rightOfUrl(el, urlA)) s += 120;
            if (nearCollectCount(el)) s += 100;
            const t = (el.value || el.textContent || '').replace(/\\s+/g, '');
            if (t === '수집조건수정') s += 20;
            const tag = (el.tagName || '').toUpperCase();
            if (tag === 'INPUT' || tag === 'BUTTON') s += 10;
            if (tag === 'A') s += 5;
            return s;
          };

          const rowMatchesUrl = (tr) => {
            if (!urlHint) return false;
            const anchors = Array.from(tr.querySelectorAll('a[href]'));
            for (const a of anchors) {
              const h = a.href || a.getAttribute('href') || '';
              if (h && (h === urlHint || h.startsWith(urlStem) || urlHint.startsWith(h.split('?')[0]))) {
                return true;
              }
            }
            const text = tr.innerText || '';
            return text.includes(urlStem) || text.includes(urlHint);
          };

          const trs = Array.from(document.querySelectorAll('table tr, form tr, tr'));
          let candidates = [];
          if (urlHint) candidates = trs.filter(rowMatchesUrl);
          if (!candidates.length && rowIndex >= 0 && rowIndex < trs.length) {
            candidates = [trs[rowIndex]];
          }
          if (!candidates.length) return { ok: false, reason: 'row-not-found' };

          for (const tr of candidates) {
            const urlA = findUrlAnchor(tr);
            const edits = Array.from(tr.querySelectorAll('a, button, input')).filter(isEditControl);
            if (!edits.length) continue;
            edits.sort((a, b) => score(b, urlA) - score(a, urlA));
            const right = edits.filter(e => rightOfUrl(e, urlA));
            const near = edits.filter(nearCollectCount);
            const pick = (right.length ? right : (near.length ? near : edits))[0];
            try { pick.scrollIntoView({ block: 'center', inline: 'nearest' }); } catch (e) {}
            pick.setAttribute('data-p3-edit-target', '1');
            if (urlA) urlA.setAttribute('data-p3-url-target', '1');
            const oc = pick.getAttribute('onclick') || '';
            return {
              ok: true,
              tag: pick.tagName,
              text: ((pick.value || pick.textContent || '') + '').replace(/\\s+/g, ' ').trim().slice(0, 40),
              nearCollect: nearCollectCount(pick),
              rightOfUrl: rightOfUrl(pick, urlA),
              score: score(pick, urlA),
              onclick: oc.slice(0, 160),
            };
          }
          return { ok: false, reason: 'button-not-found' };
        }""",
        {"rowIndex": int(row_index), "urlHint": (row_url or "").strip()},
    )
    return info if isinstance(info, dict) else {"ok": False, "reason": "evaluate-failed"}


def _p1_browse():
    """P1_101 팝업닫기·스크롤집계 재사용 (중단 플래그는 P3)."""
    import extract as p1_extract  # noqa: WPS433

    p1_extract.stop_requested = stop_requested
    return p1_extract


def dismiss_store_layers_only(page) -> int:
    """스토어(자라) 페이지의 레이어/쿠키만 닫기.

    ★다른 탭(더망고 목록)은 절대 page.close() 하지 않음.
    P1 dismiss_popups 는 다른 창을 닫아 더망고 핸들이 끊길 수 있음.
    """
    closed = 0
    try:
        page.on("dialog", lambda d: d.dismiss())
    except Exception:
        pass
    try:
        p1 = _p1_browse()
        selectors = getattr(p1, "POPUP_CLOSE_SELECTORS", ())
    except Exception:
        selectors = (
            'button[aria-label="Close"]',
            'button:has-text("닫기")',
            'button:has-text("Accept")',
            'button:has-text("동의")',
        )
    for sel in selectors:
        try:
            loc = page.locator(sel)
            count = loc.count()
        except Exception:
            continue
        for i in range(min(count, 3)):
            try:
                el = loc.nth(i)
                if el.is_visible(timeout=300):
                    el.click(timeout=800, force=True)
                    closed += 1
                    time.sleep(0.15)
            except Exception:
                continue
    try:
        page.keyboard.press("Escape")
        time.sleep(0.1)
    except Exception:
        pass
    return closed


def find_alive_mango_page(context, mango_url: str, prefer=None):
    """열려 있는 더망고 목록 탭을 다시 찾는다 (죽은 핸들 대체)."""
    pages: list = []
    try:
        pages = list(context.pages)
    except Exception:
        pages = []
    if prefer is not None and prefer not in pages:
        pages = [prefer] + pages

    mu = (mango_url or "").strip().lower()
    mu_stem = mu.split("?")[0]

    def _score(p) -> int:
        try:
            if p.is_closed():
                return -999
        except Exception:
            return -999
        try:
            u = (p.url or "").lower()
        except Exception:
            u = ""
        s = 0
        if prefer is not None and p is prefer:
            s += 30
        if mu and (mu in u or (mu_stem and mu_stem in u)):
            s += 120
        if any(k in u for k in ("demango", "admin_group", "filter", "mango")):
            s += 60
        if "zara.com" in u:
            s -= 100
        if u in ("", "about:blank"):
            s -= 20
        return s

    ranked = sorted(pages, key=_score, reverse=True)
    for p in ranked:
        if _score(p) > 0:
            return p
    return None


def page_is_usable(page) -> bool:
    """evaluate 가능한 살아 있는 페이지인지."""
    if page is None:
        return False
    try:
        if page.is_closed():
            return False
    except Exception:
        return False
    try:
        page.evaluate("() => 1", timeout=1500)
        return True
    except TypeError:
        # 구버전 playwright 는 evaluate timeout kw 없음
        try:
            page.evaluate("() => 1")
            return True
        except Exception:
            return False
    except Exception:
        return False


def resolve_demango_row_index_by_url(
    page,
    row_url: str,
    *,
    fallback_index: int | None = None,
    progress: ProgressFn | None = None,
) -> int | None:
    """더망고 목록에서 URL로 행 index를 다시 찾는다 (복귀 후 stale index 방지)."""
    url = (row_url or "").strip()
    if not url:
        return fallback_index
    try:
        rows = list_demango_rows(page)
    except Exception as e:  # noqa: BLE001
        _log(progress, "경고", f"행 재탐색 실패(목록스캔): {str(e).split(chr(10))[0][:100]}")
        return fallback_index

    target = normalize_url(url)
    for r in rows:
        ru = normalize_url((r.get("url") or "").strip())
        if ru and target and ru == target:
            idx = int(r.get("index") or 0)
            _log(
                progress,
                "로직",
                f"더망고 행 재탐색 OK · index={idx} · url={url[:100]}",
            )
            return idx
    # 느슨 매칭: stem
    stem = url.split("?")[0]
    for r in rows:
        ru = (r.get("url") or "").strip()
        if stem and stem in ru:
            idx = int(r.get("index") or 0)
            _log(
                progress,
                "로직",
                f"더망고 행 재탐색(느슨) OK · index={idx} · url={url[:100]}",
            )
            return idx

    _log(
        progress,
        "경고",
        f"더망고 행 재탐색 실패 → fallback index={fallback_index} · url={url[:100]}",
    )
    return fallback_index


def _find_and_mark_row_url(page, row_index: int, row_url: str = "") -> dict:
    """행의 URL 검색 링크를 data-p3-url-target 으로 마킹."""
    info = page.evaluate(
        """(args) => {
          const rowIndex = args.rowIndex;
          const urlHint = (args.urlHint || '').trim();
          const urlStem = urlHint.split('?')[0];
          document.querySelectorAll('[data-p3-url-target]').forEach(el => {
            el.removeAttribute('data-p3-url-target');
          });
          const rowMatchesUrl = (tr) => {
            if (!urlHint) return false;
            const text = tr.innerText || '';
            if (text.includes(urlStem) || text.includes(urlHint)) return true;
            return Array.from(tr.querySelectorAll('a[href]')).some(a => {
              const h = a.href || a.getAttribute('href') || '';
              return h === urlHint || h.startsWith(urlStem);
            });
          };
          const trs = Array.from(document.querySelectorAll('table tr, form tr, tr'));
          let candidates = urlHint ? trs.filter(rowMatchesUrl) : [];
          if (!candidates.length && rowIndex >= 0 && rowIndex < trs.length) {
            candidates = [trs[rowIndex]];
          }
          for (const tr of candidates) {
            const anchors = Array.from(tr.querySelectorAll('a[href]'));
            let pick = null;
            for (const a of anchors) {
              const h = a.href || a.getAttribute('href') || '';
              const label = (a.textContent || '').replace(/\\s+/g, '');
              if (/수집조건수정/.test(label)) continue;
              if (h.indexOf('http') !== 0) continue;
              if (!urlHint || h === urlHint || h.startsWith(urlStem) || urlHint.startsWith(h.split('?')[0])) {
                pick = a; break;
              }
            }
            if (!pick) {
              pick = anchors.find(a => {
                const h = a.href || '';
                return h.indexOf('http') === 0 && !/수집조건수정/.test((a.textContent||''));
              });
            }
            if (!pick) continue;
            try { pick.scrollIntoView({ block: 'center', inline: 'nearest' }); } catch (e) {}
            pick.setAttribute('data-p3-url-target', '1');
            return { ok: true, href: (pick.href || '').slice(0, 200) };
          }
          return { ok: false, reason: 'url-not-found' };
        }""",
        {"rowIndex": int(row_index), "urlHint": (row_url or "").strip()},
    )
    return info if isinstance(info, dict) else {"ok": False}


def click_demango_row_url(
    page,
    row_index: int,
    row_url: str = "",
    *,
    progress: ProgressFn | None = None,
):
    """1) 필터일치 행의 URL 클릭 → 스토어 페이지(팝업/새탭) 반환."""
    info = _find_and_mark_row_url(page, row_index, row_url)
    if not info.get("ok"):
        _log(progress, "오류", f"1) URL 링크 미검출 ({info})")
        return None
    _log(progress, "로직", f"1) URL 클릭 · {(info.get('href') or row_url)[:120]}")
    before = []
    try:
        before = list(page.context.pages)
    except Exception:
        before = [page]
    loc = page.locator('[data-p3-url-target="1"]').first
    store = None
    try:
        with page.expect_popup(timeout=10_000) as pop_info:
            loc.click(timeout=5_000, no_wait_after=True)
        store = pop_info.value
    except Exception:
        try:
            loc.click(timeout=5_000, force=True, no_wait_after=True)
        except Exception as e:
            _log(progress, "오류", f"1) URL 클릭 실패: {str(e).split(chr(10))[0][:120]}")
            return None
    time.sleep(0.6)
    if store is None:
        try:
            for p in page.context.pages:
                if p not in before:
                    store = p
                    break
        except Exception:
            pass
    if store is None:
        # 같은 탭 이동된 경우 — 목록 복귀는 호출측에서 mango 로
        _log(progress, "로직", "1) URL 클릭 후 새 창 없음 — 현재 탭을 스토어로 사용")
        store = page
    try:
        store.wait_for_load_state("domcontentloaded", timeout=30_000)
    except Exception:
        pass
    # ★P2와 동일: 실제 Chrome 팝업/탭을 앞으로
    reveal_browser_page(
        store,
        progress,
        step_no="1",
        action="URL클릭 → 스토어/팝업 창 표시",
        dwell_s=STEP_VIEW_DWELL_SEC,
    )
    return store


def browse_store_count_cards(
    store_page,
    *,
    excel_count: int,
    progress: ProgressFn | None = None,
    shot_dir: Path | None = None,
    row_no: int = 0,
) -> tuple[int, bool]:
    """2)~5) 첫팝업 닫기 → 푸터↓ → 상단↑ 카드수 → 엑셀 비교.

    ★팝업 닫기 시 다른 탭(더망고)을 닫지 않음.
    ★P2와 동일: 실제 Chrome 창·팝업을 bring_to_front 후 동작.
    """
    p1 = _p1_browse()
    # 스크롤 중 P1 이 dismiss_popups 로 타 탭을 닫지 못하게 교체
    orig_dismiss = p1.dismiss_popups
    p1.dismiss_popups = dismiss_store_layers_only  # type: ignore[assignment]
    try:
        # 2) 스토어 팝업/레이어를 Chrome 앞으로 보여 준 뒤 닫기 (P2: 실제 창 표시)
        reveal_browser_page(
            store_page,
            progress,
            step_no="2",
            action="스토어 첫 화면/팝업 노출",
            dwell_s=1.0,
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="02_before_dismiss",
            label="2)팝업닫기 전(브라우저표시)",
            row_no=row_no,
            progress=progress,
        )
        _log(progress, "동작", "2) 첫 팝업창 닫기 (스토어 레이어만 · 더망고탭 유지)")
        closed = dismiss_store_layers_only(store_page)
        time.sleep(0.8)
        closed += dismiss_store_layers_only(store_page)
        _log(progress, "동작", f"2) 팝업 닫기 완료 · closed={closed}")
        reveal_browser_page(
            store_page,
            progress,
            step_no="2",
            action=f"팝업닫기 후 본문 표시 · closed={closed}",
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="02_after_dismiss",
            label=f"2)팝업닫기 후 closed={closed}",
            row_no=row_no,
            progress=progress,
        )

        _log(progress, "동작", "3) 스크롤 푸터 영역까지 내리기")
        reveal_browser_page(
            store_page, progress, step_no="3", action="푸터 스크롤 시작", dwell_s=0.4
        )
        p1.scroll_down_to_footer(store_page, progress=progress)
        reveal_browser_page(
            store_page, progress, step_no="3", action="푸터까지 스크롤 완료"
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="03_footer",
            label="3)푸터까지 스크롤",
            row_no=row_no,
            progress=progress,
        )

        _log(progress, "동작", "4) 하단→상단 스크롤 · 상품수 카드 갯수 집계")
        reveal_browser_page(
            store_page, progress, step_no="4", action="상단 스크롤·카드집계 시작", dwell_s=0.4
        )
        card_n = int(p1.scroll_up_count_card_images(store_page, progress=progress) or 0)
        _log(progress, "동작", f"4) 상품수 카드 갯수={card_n}")
        reveal_browser_page(
            store_page,
            progress,
            step_no="4",
            action=f"카드갯수={card_n} 화면",
        )
        screenshot_step(
            store_page,
            shot_dir,
            step_tag="04_card_count",
            label=f"4)카드갯수={card_n}",
            row_no=row_no,
            progress=progress,
        )
    finally:
        p1.dismiss_popups = orig_dismiss  # type: ignore[assignment]

    excel_n = int(excel_count or 0)
    matched = card_n == excel_n
    _log(
        progress,
        "동작",
        f"5) 비교 · 카드상품수={card_n} · 엑셀상품수={excel_n} · 일치={'Y' if matched else 'N'}",
    )
    reveal_browser_page(
        store_page,
        progress,
        step_no="5",
        action=f"비교결과 카드={card_n} 엑셀={excel_n} {'Y' if matched else 'N'}",
    )
    screenshot_step(
        store_page,
        shot_dir,
        step_tag="05_compare",
        label=f"5)비교 카드={card_n} 엑셀={excel_n} {'Y' if matched else 'N'}",
        row_no=row_no,
        progress=progress,
    )
    return card_n, matched


def close_store_return_list(
    list_page,
    store_page,
    mango_url: str,
    *,
    progress: ProgressFn | None = None,
):
    """스토어 탭만 닫고, 살아 있는 더망고 목록 탭을 다시 찾아 반환."""
    ctx = None
    try:
        if list_page is not None:
            ctx = list_page.context
        elif store_page is not None:
            ctx = store_page.context
    except Exception:
        ctx = None

    # 스토어(자라) 탭만 닫기 — 더망고와 동일 핸들이면 닫지 않음
    try:
        if store_page is not None and store_page is not list_page:
            store_url = ""
            try:
                store_url = (store_page.url or "").lower()
            except Exception:
                store_url = ""
            # 더망고 URL 이면 닫지 않음
            mu = (mango_url or "").lower()
            is_mango = bool(mu) and (mu in store_url or store_url.startswith(mu.split("?")[0]))
            if (not is_mango) and ("zara.com" in store_url or "http" in store_url):
                if not store_page.is_closed():
                    store_page.close()
                    _log(progress, "로직", "스토어 탭 닫음 → 더망고 목록 재연결")
    except Exception as e:  # noqa: BLE001
        _log(progress, "경고", f"스토어 탭 닫기 예외: {str(e).split(chr(10))[0][:80]}")

    time.sleep(0.35)

    mango = None
    if ctx is not None:
        mango = find_alive_mango_page(ctx, mango_url, prefer=list_page)

    if mango is None or not page_is_usable(mango):
        # prefer 가 죽었어도 context 에서 재탐색
        if ctx is not None:
            mango = find_alive_mango_page(ctx, mango_url, prefer=None)
        if mango is None or not page_is_usable(mango):
            _log(progress, "오류", "더망고 목록 탭 재연결 실패 (usable page 없음)")
            return None

    try:
        mango.bring_to_front()
    except Exception:
        pass

    try:
        cur = mango.url or ""
        if (
            "modify_filter" in cur
            or "admin_group_modify" in cur
            or "zara.com" in cur.lower()
        ):
            _return_to_list(mango, mango_url)
    except Exception:
        try:
            _return_to_list(mango, mango_url)
        except Exception:
            pass

    if not page_is_usable(mango):
        _log(progress, "오류", "더망고 목록 탭 evaluate 불가")
        return None

    _log(
        progress,
        "로직",
        f"더망고 목록 탭 재연결 OK · url={(mango.url or '')[:120]}",
    )
    return mango


def _modify_ui_opened(page) -> bool:
    """수집조건수정 후 수정 팝업/페이지/iframe 이 실제로 열렸는지."""
    if page_shows_not_found(page):
        return False
    try:
        if wait_for_save_count_ready(page, timeout_ms=400):
            return True
    except Exception:
        pass
    try:
        target, kind = resolve_modify_target(page)
        if kind in ("page", "frame") and kind != "main":
            if wait_for_save_count_ready(target, timeout_ms=400):
                return True
        body = ""
        try:
            body = target.locator("body").inner_text(timeout=300) or ""
        except Exception:
            body = ""
        if "저장상품수" in body and (
            "검색필터 수정" in body or "검색결과" in body or "저장하기" in body
        ):
            return True
    except Exception:
        pass
    try:
        for p in page.context.pages:
            if page_shows_not_found(p):
                continue
            bu = p.url or ""
            if "modify_filter" in bu or "admin_group_modify" in bu:
                bt = ""
                try:
                    bt = p.locator("body").inner_text(timeout=300) or ""
                except Exception:
                    pass
                if "저장상품수" in bt or "검색필터 수정" in bt or (
                    "검색결과" in bt and "저장하기" in bt
                ):
                    return True
                # URL 은 맞는데 본문 로딩 중이면 잠깐 더
                try:
                    p.wait_for_load_state("domcontentloaded", timeout=2000)
                except Exception:
                    pass
                try:
                    bt = p.locator("body").inner_text(timeout=400) or ""
                except Exception:
                    bt = ""
                if page_shows_not_found(p):
                    continue
                if "저장상품수" in bt or "검색필터 수정" in bt:
                    return True
            else:
                try:
                    bt = p.locator("body").inner_text(timeout=250) or ""
                except Exception:
                    bt = ""
                if "저장상품수" in bt and ("검색결과" in bt or "저장하기" in bt):
                    return True
    except Exception:
        pass
    return False


# 6) 수집조건수정: 「전체저장」바로 우측에서 시작해 한글 1글자씩 우측 이동 · 최대 10회
EDIT_CLICK_MAX_TRIES = 10
EDIT_CLICK_CHAR_PAD_X = 2  # 전체저장 직후 여유(px)


def edit_click_char_steps(attempt: int) -> int:
    """시도 번호(1..N) → 전체저장 오른쪽에서 이동한 글자 수(0부터)."""
    return max(0, int(attempt) - 1)


def _find_allsave_anchor_geometry(page, row_index: int, row_url: str) -> dict | None:
    """행 안 '전체저장' 텍스트의 우측 끝 + 한글 1글자 폭을 구한다."""
    geo = page.evaluate(
        """(args) => {
          const rowIndex = args.rowIndex;
          const urlHint = (args.urlHint || '').trim();
          const urlStem = urlHint.split('?')[0];
          const needle = '전체저장';

          const rowMatchesUrl = (tr) => {
            if (!urlHint) return false;
            const text = tr.innerText || '';
            if (text.includes(urlStem) || text.includes(urlHint)) return true;
            return Array.from(tr.querySelectorAll('a[href]')).some(a => {
              const h = a.href || a.getAttribute('href') || '';
              return h === urlHint || h.startsWith(urlStem);
            });
          };

          const trs = Array.from(document.querySelectorAll('table tr, form tr, tr'));
          let candidates = urlHint ? trs.filter(rowMatchesUrl) : [];
          if (!candidates.length && rowIndex >= 0 && rowIndex < trs.length) {
            candidates = [trs[rowIndex]];
          }
          if (!candidates.length) return null;

          const findNeedleRect = (root) => {
            const walker = document.createTreeWalker(root, NodeFilter.SHOW_TEXT);
            let node;
            while ((node = walker.nextNode())) {
              const t = node.textContent || '';
              const i = t.indexOf(needle);
              if (i < 0) continue;
              const range = document.createRange();
              range.setStart(node, i);
              range.setEnd(node, i + needle.length);
              const rects = range.getClientRects();
              if (!rects || !rects.length) continue;
              const r = rects[rects.length - 1];
              // 한글 1글자 폭: '전' 한 글자 측정, 실패 시 전체/4
              let charW = Math.max(8, r.width / Math.max(1, needle.length));
              try {
                const r2 = document.createRange();
                r2.setStart(node, i);
                r2.setEnd(node, i + 1);
                const rr = r2.getClientRects();
                if (rr && rr.length && rr[0].width > 2) charW = rr[0].width;
              } catch (e) {}
              return {
                left: r.left,
                right: r.right,
                top: r.top,
                bottom: r.bottom,
                width: r.width,
                height: r.height,
                charW: charW,
                midY: r.top + r.height / 2,
              };
            }
            return null;
          };

          for (const tr of candidates) {
            // URL 오른쪽 영역 우선: 수집개수/전체저장이 있는 셀
            let rect = null;
            const cells = Array.from(tr.querySelectorAll('td, th'));
            for (const cell of cells) {
              const txt = (cell.innerText || '').replace(/\\s+/g, '');
              if (txt.includes('전체저장') || txt.includes('수집개수')) {
                rect = findNeedleRect(cell);
                if (rect) break;
              }
            }
            if (!rect) rect = findNeedleRect(tr);
            if (!rect) continue;
            try { tr.scrollIntoView({ block: 'center', inline: 'nearest' }); } catch (e) {}
            return rect;
          }
          return null;
        }""",
        {"rowIndex": int(row_index), "urlHint": (row_url or "").strip()},
    )
    return geo if isinstance(geo, dict) else None


def _edit_click_point_from_allsave(
    page,
    row_index: int,
    row_url: str,
    *,
    attempt: int,
) -> dict | None:
    """「전체저장」바로 우측 + (attempt-1)글자만큼 오른쪽 클릭 좌표."""
    # URL 마킹(스크롤/로그용)
    _find_and_mark_row_url(page, row_index, row_url)
    geo = _find_allsave_anchor_geometry(page, row_index, row_url)
    if not geo:
        return None
    char_w = float(geo.get("charW") or 14.0)
    if char_w < 6:
        char_w = 14.0
    steps = edit_click_char_steps(attempt)
    start_x = float(geo["right"]) + float(EDIT_CLICK_CHAR_PAD_X)
    x = start_x + steps * char_w + char_w * 0.35  # 글자 중심 부근
    y = float(geo.get("midY") or ((float(geo["top"]) + float(geo["bottom"])) / 2.0))
    try:
        vp = page.viewport_size or {}
        vw = float(vp.get("width") or 1280)
        vh = float(vp.get("height") or 800)
        x = min(max(2.0, x), vw - 2.0)
        y = min(max(2.0, y), vh - 2.0)
    except Exception:
        pass
    return {
        "x": x,
        "y": y,
        "char_w": char_w,
        "char_steps": steps,
        "start_x": start_x,
        "allsave_right": float(geo["right"]),
        "offset": int(round(x - start_x)),
    }


def click_edit_on_row(
    page,
    row_index: int,
    edit_href: str = "",  # 미사용 — href 재시도 금지(호출부 호환용)
    *,
    row_url: str = "",
    progress: ProgressFn | None = None,
    shot_dir: Path | None = None,
    row_no: int = 0,
    shot_count: int = 0,
    shot_interval_s: float = 0.0,
    max_tries: int = EDIT_CLICK_MAX_TRIES,
    try_interval_s: float = 2.0,
) -> bool:
    """6) 「전체저장」바로 우측부터 한글 1글자씩 우측 이동 클릭 — 2초×최대 10회.

    순서: URL 우측 → 수집개수 우측 → ★전체저장 바로 우측 시작 → 글자씩 이동.
    ★href / location 대체 절대 금지. 오로지 좌표 클릭만.
    """
    _ = edit_href
    tries = max(1, int(max_tries))
    gap = max(0.2, float(try_interval_s))

    for attempt in range(1, tries + 1):
        if stop_requested():
            return False

        info = _find_and_mark_edit_button(page, row_index, row_url)
        point = _edit_click_point_from_allsave(
            page, row_index, row_url, attempt=attempt
        )
        if point is None:
            _log(
                progress,
                "로직",
                f"6) '전체저장' 좌표 미검출 · 시도 {attempt}/{tries} — 버튼 폴백",
            )
            if not info.get("ok"):
                if attempt < tries:
                    time.sleep(gap)
                continue
            try:
                bbox = page.locator('[data-p3-edit-target="1"]').first.bounding_box()
            except Exception:
                bbox = None
            if not bbox:
                if attempt < tries:
                    time.sleep(gap)
                continue
            # 폴백: 버튼 왼쪽부터 글자폭(~14px)씩 이동
            char_w = 14.0
            steps = edit_click_char_steps(attempt)
            point = {
                "x": float(bbox["x"]) + 4 + steps * char_w,
                "y": float(bbox["y"]) + float(bbox["height"]) / 2.0,
                "char_w": char_w,
                "char_steps": steps,
                "offset": int(steps * char_w),
            }

        before_pages = []
        try:
            before_pages = list(page.context.pages)
        except Exception:
            before_pages = [page]

        x = float(point["x"])
        y = float(point["y"])
        steps = int(point.get("char_steps") or edit_click_char_steps(attempt))
        char_w = float(point.get("char_w") or 14.0)
        _log(
            progress,
            "로직",
            f"6) 수집조건수정 클릭 시도 {attempt}/{tries} · "
            f"전체저장우측 +{steps}글자(≈{char_w:.0f}px) → ({x:.0f},{y:.0f})",
        )

        popup = None
        try:
            with page.expect_popup(timeout=int(gap * 1000)) as pop_info:
                page.mouse.click(x, y)
            popup = pop_info.value
        except Exception:
            popup = None
            try:
                page.mouse.click(x, y)
            except Exception as e:
                _log(
                    progress,
                    "로직",
                    f"6) 좌표클릭 예외 시도{attempt}: {str(e).split(chr(10))[0][:100]}",
                )

        if popup is not None:
            try:
                popup.wait_for_load_state("domcontentloaded", timeout=12_000)
            except Exception:
                pass
            _log(
                progress,
                "로직",
                f"6) 팝업창 열림 · url={(popup.url or '')[:120]}",
            )
            # ★P2와 동일: 팝업 창을 앞으로
            reveal_browser_page(
                popup,
                progress,
                step_no="6",
                action="수집조건수정 팝업 표시",
                dwell_s=STEP_VIEW_DWELL_SEC,
            )

        deadline = time.time() + gap
        while time.time() < deadline:
            try:
                for p in page.context.pages:
                    if p not in before_pages:
                        try:
                            p.wait_for_load_state("domcontentloaded", timeout=2_000)
                        except Exception:
                            pass
                        try:
                            p.bring_to_front()
                        except Exception:
                            pass
            except Exception:
                pass
            if page_shows_not_found(page):
                _log(progress, "오류", "6) 팝업 not found — 중단 (href 재시도 없음)")
                return False
            if _modify_ui_opened(page):
                _log(
                    progress,
                    "로직",
                    f"6) 수집조건수정 팝업 확인 · 시도 {attempt}/{tries} · +{steps}글자",
                )
                if shot_dir is not None and shot_count > 0:
                    screenshot_after_edit_click_series(
                        page,
                        shot_dir,
                        row_no=row_no,
                        progress=progress,
                        prefer_page=popup,
                        count=shot_count,
                        interval_s=shot_interval_s,
                    )
                return True
            time.sleep(0.25)

        if page_shows_not_found(page):
            _log(progress, "오류", "6) 팝업 not found — 중단 (href 재시도 없음)")
            return False
        if _modify_ui_opened(page) or wait_modify_page(page, timeout_ms=800):
            _log(
                progress,
                "로직",
                f"6) 수집조건수정 팝업 확인 · 시도 {attempt}/{tries} · +{steps}글자",
            )
            if shot_dir is not None and shot_count > 0:
                screenshot_after_edit_click_series(
                    page,
                    shot_dir,
                    row_no=row_no,
                    progress=progress,
                    prefer_page=popup,
                    count=shot_count,
                    interval_s=shot_interval_s,
                )
            return True

        _log(
            progress,
            "로직",
            f"6) 팝업 미오픈 · 다음 시도 +{edit_click_char_steps(attempt + 1)}글자 우측",
        )

    _log(
        progress,
        "오류",
        f"6) 수집조건수정 클릭 {tries}회 실패 — 전체저장 우측 글자이동에도 팝업 미오픈",
    )
    return False


def page_shows_not_found(page) -> bool:
    """수정 팝업/페이지가 'not found' 등 잘못된 진입인지."""
    targets = [page]
    try:
        targets = list(page.context.pages) or [page]
    except Exception:
        targets = [page]
    for target in targets:
        try:
            title = ""
            try:
                title = target.title() or ""
            except Exception:
                title = ""
            body = ""
            try:
                body = target.locator("body").inner_text(timeout=400) or ""
            except Exception:
                body = ""
            blob = f"{title}\n{body}"
            # 정상 수정화면이면 not-found 로 보지 않음
            if "저장상품수" in blob and (
                "검색필터 수정" in blob or "검색결과" in blob or "저장하기" in blob
            ):
                continue
            if re.search(
                r"not\s*found|404\b|찾을\s*수\s*없|존재하지\s*않|페이지를\s*찾을|잘못된\s*접근",
                blob,
                re.I,
            ):
                return True
        except Exception:
            continue
    return False


def set_save_count(
    page,
    value: int,
    *,
    shot_dir: Path | None = None,
    progress: ProgressFn | None = None,
    row_no: int = 0,
) -> bool:
    """저장상품수 입력칸: 현재 숫자(스크린샷의 '3')가 있는 칸을 찾아 상품수값으로 대체.

    UI: 저장상품수 | 검색결과 상위 [ 3 ] 개 상품만 저장
    1) value가 '3'(또는 숫자)인 input 을 찾음
    2) 그 칸의 값을 상품수(target)로 덮어씀
    ★갱신 전·후 스크린샷은 성공/실패와 무관하게 항상 로그에 남김
    """
    target = str(int(value))

    work, kind = resolve_modify_target(page)
    wait_for_save_count_ready(work, timeout_ms=8_000)
    shot_page = page if kind == "frame" else work

    # ★요건: 3) 저장상품수 갱신 전 스크린샷 (항상)
    if shot_dir is not None:
        screenshot_step(
            shot_page,
            shot_dir,
            step_tag="03_save_count_before",
            label=f"3)저장상품수 갱신 전 →목표={target}",
            row_no=row_no,
            progress=progress,
        )

    # ★요건: 숫자 "3" 이 들어있는 칸 우선 탐색
    loc = find_save_count_locator(work, prefer_value="3")

    before_val = ""
    if loc is not None:
        try:
            before_val = (loc.input_value(timeout=500) or "").strip()
        except Exception:
            before_val = ""
        _log(
            progress,
            "로직",
            f"3) 저장상품수 칸 발견 현재값={before_val or '?'} → 상품수값={target}",
        )
        if shot_dir is not None:
            screenshot_save_count_grid(
                shot_page,
                loc,
                shot_dir,
                tag="before",
                row_no=row_no,
                note=f"현재값={before_val or '?'}",
                progress=progress,
            )

    def _replace_value(el) -> bool:
        """기존 숫자(3 등)를 지우고 상품수값으로 대체 입력."""
        try:
            el.scroll_into_view_if_needed(timeout=1500)
        except Exception:
            pass
        try:
            el.click(timeout=1500, force=True)
        except Exception:
            try:
                el.focus()
            except Exception:
                pass
        # 전체 선택 후 삭제 → 새 값 입력 (자동완성 대비 Escape)
        for key in ("Control+a", "Meta+a"):
            try:
                el.press(key)
                break
            except Exception:
                continue
        try:
            el.press("Backspace")
        except Exception:
            pass
        try:
            el.fill("")
        except Exception:
            pass
        try:
            el.type(target, delay=30)
        except Exception:
            try:
                el.fill(target)
            except Exception:
                return False
        # 브라우저 자동완성(30/3/35) 닫기
        try:
            el.press("Escape")
        except Exception:
            pass
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        time.sleep(0.15)
        try:
            got = (el.input_value(timeout=800) or "").strip()
            if got == target:
                return True
        except Exception:
            pass
        # JS 강제 대체
        try:
            ok = bool(
                el.evaluate(
                    """(el, want) => {
                      el.focus();
                      el.value = '';
                      el.value = String(want);
                      el.dispatchEvent(new Event('input', {bubbles:true}));
                      el.dispatchEvent(new Event('change', {bubbles:true}));
                      el.blur();
                      return (el.value || '').trim() === String(want);
                    }""",
                    target,
                )
            )
            return ok
        except Exception:
            return False

    filled = False
    if loc is not None and _replace_value(loc):
        filled = True
    else:
        # JS: value==='3' 인 칸을 우선 찾아 상품수값으로 대체
        try:
            filled = bool(
                work.evaluate(
                    """(n) => {
                      const want = String(n);
                      const isNumInput = (inp) => {
                        if (!inp) return false;
                        const ty = (inp.getAttribute('type') || 'text').toLowerCase();
                        if (!(ty === 'text' || ty === 'number' || ty === '')) return false;
                        if (inp.disabled || inp.readOnly) return false;
                        return true;
                      };
                      const setVal = (inp) => {
                        inp.focus();
                        try { inp.select(); } catch (e) {}
                        inp.value = '';
                        inp.value = want;
                        inp.dispatchEvent(new Event('input', { bubbles: true }));
                        inp.dispatchEvent(new Event('change', { bubbles: true }));
                        try { inp.blur(); } catch (e) {}
                        return (inp.value || '').trim() === want;
                      };
                      const inSaveRow = (inp) => {
                        const tr = inp.closest('tr');
                        const scope = tr || inp.closest('td,div,li') || inp.parentElement;
                        const t = ((scope && scope.innerText) || '').replace(/\\s+/g, '');
                        return t.includes('저장상품수')
                          || (t.includes('검색결과') && t.includes('상위'))
                          || t.includes('개상품만저장')
                          || (t.includes('상위') && t.includes('개'));
                      };
                      const all = Array.from(document.querySelectorAll(
                        'input[type="text"], input[type="number"], input:not([type])'
                      )).filter(isNumInput);

                      // 1) 정확히 value==='3' 이고 저장상품수 문맥
                      let pick = all.find(i => (i.value || '').trim() === '3' && inSaveRow(i));
                      // 2) value==='3' (문맥 느슨)
                      if (!pick) pick = all.find(i => (i.value || '').trim() === '3');
                      // 3) 저장상품수 행의 숫자 value
                      if (!pick) pick = all.find(i => inSaveRow(i) && /^\\d+$/.test((i.value || '').trim()));
                      // 4) 저장상품수 행 첫 숫자 input
                      if (!pick) pick = all.find(i => inSaveRow(i));
                      if (pick) return setVal(pick);
                      return false;
                    }""",
                    int(value),
                )
            )
        except Exception:
            filled = False
        if filled:
            loc = find_save_count_locator(work, prefer_value=target)

    after_val = ""
    if filled:
        after_val = target
        if loc is not None:
            try:
                after_val = (loc.input_value(timeout=500) or "").strip() or target
            except Exception:
                after_val = target
            if shot_dir is not None:
                screenshot_save_count_grid(
                    shot_page,
                    loc,
                    shot_dir,
                    tag="after",
                    row_no=row_no,
                    note=f"{before_val or '3'}→{after_val}",
                    progress=progress,
                )
        _log(
            progress,
            "로직",
            f"3) 저장상품수 대체완료 {before_val or '3'} → {after_val or target}",
        )

    # ★요건: 3) 저장상품수 갱신 후 스크린샷 (항상 — 성공/실패 공통)
    if shot_dir is not None:
        status = "성공" if filled else "실패"
        note = f"{before_val or '?'}→{after_val or target}" if filled else "칸미검출/대체실패"
        screenshot_step(
            shot_page,
            shot_dir,
            step_tag="03_save_count_after",
            label=f"3)저장상품수 갱신 후 ({status}) {note}",
            row_no=row_no,
            progress=progress,
        )

    return filled


def new_shot_dir() -> Path:
    """P3 실행별 스크린샷 폴더."""
    ts = time.strftime("%Y%m%d_%H%M%S")
    d = P3_RUN_LOG_DIR / ts
    d.mkdir(parents=True, exist_ok=True)
    return d


def find_save_count_locator(page, prefer_value: str = "3"):
    """저장상품수 숫자 입력칸 — 우선 현재값이 prefer_value(기본 '3')인 input.

    스크린샷: 검색결과 상위 [ 3 ] 개 상품만 저장
    """
    prefer = (prefer_value or "3").strip()

    # 0) JS로 value===prefer (문맥: 저장상품수/상위/개) 인 요소 핸들
    try:
        handle = page.evaluate_handle(
            """(prefer) => {
              const isNumInput = (inp) => {
                if (!inp) return false;
                const ty = (inp.getAttribute('type') || 'text').toLowerCase();
                if (!(ty === 'text' || ty === 'number' || ty === '')) return false;
                if (inp.disabled || inp.readOnly) return false;
                return true;
              };
              const inCtx = (inp) => {
                const tr = inp.closest('tr');
                const scope = tr || inp.closest('td,div') || inp.parentElement;
                const t = ((scope && scope.innerText) || '').replace(/\\s+/g, '');
                return t.includes('저장상품수')
                  || (t.includes('검색결과') && t.includes('상위'))
                  || (t.includes('상위') && t.includes('개'));
              };
              const all = Array.from(document.querySelectorAll(
                'input[type="text"], input[type="number"], input:not([type])'
              )).filter(isNumInput);
              return all.find(i => (i.value || '').trim() === String(prefer) && inCtx(i))
                || all.find(i => (i.value || '').trim() === String(prefer))
                || all.find(i => inCtx(i) && /^\\d+$/.test((i.value || '').trim()))
                || all.find(i => inCtx(i))
                || null;
            }""",
            prefer,
        )
        el = handle.as_element()
        if el is not None:
            # ElementHandle → Locator 대신 직접 쓰기 위해 wrapper
            # Playwright: page.locator로 재검색이 더 안정적
            pass
    except Exception:
        handle = None
        el = None

    # 1) value=prefer 정확 매칭 (저장상품수 행)
    try:
        loc = page.locator(
            "xpath=//tr[.//text()[contains(.,'저장상품수')] or "
            ".//*[contains(normalize-space(.),'저장상품수')]]"
            f"//input[(@type='text' or @type='number' or not(@type)) and @value='{prefer}']"
        ).first
        if loc.count() > 0 and loc.is_visible(timeout=400):
            return loc
    except Exception:
        pass

    # 2) '검색결과 상위' 셀 안 value=prefer
    try:
        loc = page.locator(
            "xpath=//td[contains(.,'검색결과') and contains(.,'상위') and contains(.,'개')]"
            f"//input[(@type='text' or @type='number' or not(@type)) and @value='{prefer}']"
        ).first
        if loc.count() > 0 and loc.is_visible(timeout=400):
            return loc
    except Exception:
        pass

    # 3) 화면에 보이는 input 중 value가 prefer 인 것 (런타임 value)
    try:
        cands = page.locator(
            'input[type="text"], input[type="number"], input:not([type])'
        )
        n = min(cands.count(), 40)
        for i in range(n):
            el = cands.nth(i)
            try:
                if not el.is_visible(timeout=150):
                    continue
                v = (el.input_value(timeout=200) or "").strip()
                if v != prefer:
                    continue
                # 부모 텍스트에 상위/개 있으면 채택
                ok_ctx = el.evaluate(
                    """(el) => {
                      const tr = el.closest('tr');
                      const t = ((tr && tr.innerText) || el.parentElement?.innerText || '')
                        .replace(/\\s+/g, '');
                      return t.includes('저장상품수')
                        || (t.includes('상위') && t.includes('개'))
                        || t.includes('검색결과');
                    }"""
                )
                if ok_ctx:
                    return el
            except Exception:
                continue
    except Exception:
        pass

    # 4) 기존 폴백: 저장상품수 행의 숫자 input
    selectors = (
        "xpath=//td[contains(.,'검색결과') and contains(.,'상위') and contains(.,'개')]"
        "//input[@type='text' or @type='number' or not(@type)]",
        "xpath=//tr[.//th[contains(normalize-space(.),'저장상품수')] or "
        ".//td[normalize-space()='저장상품수'] or "
        ".//td[starts-with(normalize-space(.),'저장상품수')] or "
        ".//*[contains(normalize-space(.),'저장상품수')]]"
        "//input[@type='text' or @type='number' or not(@type)]",
        "xpath=//*[contains(normalize-space(.),'개 상품만 저장')]"
        "/preceding::input[@type='text' or @type='number' or not(@type)][1]",
    )
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() == 0:
                continue
            if not loc.is_visible(timeout=400):
                continue
            try:
                v = (loc.input_value(timeout=300) or "").strip()
                if v and not re.fullmatch(r"\d+", v):
                    if len(v) > 8 or "http" in v.lower():
                        continue
            except Exception:
                pass
            return loc
        except Exception:
            continue
    return None


def resolve_modify_target(page):
    """검색필터 수정(저장상품수) 화면이 열린 page/frame 을 찾는다.

    팝업 창·iframe 모두 탐색. (page, kind) 반환. 없으면 (page, 'main').
    """
    candidates = []
    try:
        for p in page.context.pages:
            candidates.append(("page", p))
    except Exception:
        candidates.append(("page", page))

    # 현재 page 를 앞에
    ordered = [("page", page)]
    for kind, p in candidates:
        if p is not page:
            ordered.append((kind, p))

    for kind, p in ordered:
        try:
            url = p.url or ""
            if "modify_filter" in url or "admin_group_modify" in url:
                return p, kind
        except Exception:
            pass
        try:
            body = p.locator("body").inner_text(timeout=400) or ""
            if "저장상품수" in body and (
                "검색필터 수정" in body or "검색결과" in body or "저장하기" in body
            ):
                return p, kind
        except Exception:
            pass
        # frames
        try:
            for fr in p.frames:
                try:
                    t = fr.inner_text("body", timeout=300) or ""
                except Exception:
                    continue
                if "저장상품수" in t and ("검색결과" in t or "저장하기" in t):
                    return fr, "frame"
        except Exception:
            pass
    return page, "main"


def wait_for_save_count_ready(target, *, timeout_ms: int = 8000) -> bool:
    """저장상품수 입력칸이 보일 때까지 대기."""
    end = time.time() + timeout_ms / 1000.0
    while time.time() < end:
        try:
            # Page or Frame both have locator
            if find_save_count_locator(target) is not None:
                return True
        except Exception:
            pass
        try:
            body = ""
            if hasattr(target, "locator"):
                body = target.locator("body").inner_text(timeout=300) or ""
            if "저장상품수" in body or "개 상품만 저장" in body:
                # 문구는 있는데 locator 실패 — 한 번 더 여유
                time.sleep(0.3)
                if find_save_count_locator(target) is not None:
                    return True
        except Exception:
            pass
        time.sleep(0.25)
    return find_save_count_locator(target) is not None


def _capture_png(page, path: Path, *, timeout_ms: int = 3000) -> bool:
    """뷰포트 PNG 캡처 — Playwright 실패 시 CDP 폴백(샷이 본작업을 오래 막지 않음)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        page.screenshot(
            path=str(path),
            timeout=timeout_ms,
            animations="disabled",
        )
        if path.is_file() and path.stat().st_size > 0:
            return True
    except Exception:
        pass
    # CDP 폴백
    try:
        session = page.context.new_cdp_session(page)
        result = session.send("Page.captureScreenshot", {"format": "png", "fromSurface": True})
        import base64

        data = base64.b64decode(result.get("data") or "")
        if data:
            path.write_bytes(data)
            return path.is_file() and path.stat().st_size > 0
    except Exception:
        pass
    return False


def screenshot_step(
    page,
    shot_dir: Path | None,
    *,
    step_tag: str,
    label: str,
    row_no: int = 0,
    progress: ProgressFn | None = None,
    full_page: bool = False,
) -> Path | None:
    """필터 일치 행의 단계 스크린샷 → 실행 로그에 출력."""
    if shot_dir is None:
        return None
    shot_dir.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^\w\-]+", "_", step_tag)[:48]
    name = f"r{row_no:03d}_{safe}.png"
    path = shot_dir / name
    time.sleep(0.05)
    ok = False
    if full_page:
        try:
            page.screenshot(
                path=str(path),
                full_page=True,
                timeout=4_000,
                animations="disabled",
            )
            ok = path.is_file() and path.stat().st_size > 0
        except Exception:
            ok = False
    if not ok:
        ok = _capture_png(page, path, timeout_ms=3_000)
    if not ok:
        _log(progress, "샷", f"[샷 실패] {label}: 캡처 불가(타임아웃/CDP)")
        return None

    _log(progress, "샷", f"{label} -> {path}")
    print(f"{P3_SHOT_MARK}{path}##{label}", flush=True)
    return path


def _page_for_edit_shot(page, prefer_page=None):
    """수집조건수정 클릭 후 샷 대상 — 팝업/수정화면 우선."""
    if prefer_page is not None:
        try:
            if not prefer_page.is_closed():
                return prefer_page
        except Exception:
            pass
    try:
        target, kind = resolve_modify_target(page)
        if kind in ("page", "frame") and target is not None:
            return target if kind == "page" else page
    except Exception:
        pass
    try:
        pages = list(page.context.pages)
        for p in reversed(pages):
            if p is page:
                continue
            try:
                if not p.is_closed():
                    return p
            except Exception:
                continue
    except Exception:
        pass
    return page


def screenshot_after_edit_click_series(
    page,
    shot_dir: Path | None,
    *,
    row_no: int = 0,
    progress: ProgressFn | None = None,
    count: int = 3,
    interval_s: float = 3.0,
    prefer_page=None,
) -> list[Path]:
    """수집조건수정 클릭 직후 — 3초 간격 스크린샷 3장 (실행로그·뷰어 출력)."""
    out: list[Path] = []
    if shot_dir is None:
        return out
    n = max(1, int(count))
    gap = max(0.0, float(interval_s))
    _log(
        progress,
        "로직",
        f"2) 수집조건수정 클릭 직후 스크린샷 {n}장 ({gap:g}초 간격)",
    )
    for i in range(1, n + 1):
        elapsed = int(round((i - 1) * gap))
        target = _page_for_edit_shot(page, prefer_page=prefer_page)
        path = screenshot_step(
            target,
            shot_dir,
            step_tag=f"02_after_edit_{i}of{n}",
            label=f"2)수집조건수정 클릭후 샷 {i}/{n} (+{elapsed}s)",
            row_no=row_no,
            progress=progress,
        )
        if path is not None:
            out.append(path)
        if i < n and gap > 0:
            time.sleep(gap)
    return out


def screenshot_save_count_grid(
    page,
    loc,
    shot_dir: Path,
    *,
    tag: str,
    row_no: int = 0,
    note: str = "",
    progress: ProgressFn | None = None,
) -> Path | None:
    """판단한 저장상품수 입력그리드(행) 근접 스크린샷 → 로그 출력."""
    shot_dir.mkdir(parents=True, exist_ok=True)
    safe_tag = re.sub(r"[^\w\-]+", "_", tag)[:40]
    name = f"r{row_no:03d}_save_count_{safe_tag}.png"
    path = shot_dir / name
    label = f"저장상품수 입력그리드/{tag}"
    if note:
        label = f"{label} ({note})"

    ok = False
    if loc is not None:
        try:
            row = loc.locator("xpath=ancestor::tr[1]")
            if row.count() > 0:
                row.first.scroll_into_view_if_needed(timeout=1000)
                time.sleep(0.08)
                row.first.screenshot(path=str(path), timeout=3_000, animations="disabled")
                ok = path.is_file() and path.stat().st_size > 0
        except Exception:
            ok = False

        if not ok:
            try:
                loc.scroll_into_view_if_needed(timeout=1000)
                box = loc.bounding_box()
                if box:
                    pad_l, pad_r, pad_y = 160, 220, 28
                    clip = {
                        "x": max(0, box["x"] - pad_l),
                        "y": max(0, box["y"] - pad_y),
                        "width": max(80, box["width"] + pad_l + pad_r),
                        "height": max(40, box["height"] + pad_y * 2),
                    }
                    page.screenshot(
                        path=str(path),
                        clip=clip,
                        timeout=3_000,
                        animations="disabled",
                    )
                    ok = path.is_file() and path.stat().st_size > 0
            except Exception:
                ok = False

    if not ok:
        ok = _capture_png(page, path, timeout_ms=3_000)

    if not ok:
        _log(progress, "샷", f"[샷 실패] {label}: 캡처 불가")
        return None

    _log(progress, "샷", f"{label} -> {path}")
    print(f"{P3_SHOT_MARK}{path}##{label}", flush=True)
    return path


def click_save_button(page) -> bool:
    """검색필터 수정 화면 하단 '저장하기' (옆에 '닫기')."""
    selectors = (
        'input[type="submit"][value="저장하기"]',
        'input[type="button"][value="저장하기"]',
        'input[value="저장하기"]',
        'button:has-text("저장하기")',
        'a:has-text("저장하기")',
    )
    for sel in selectors:
        try:
            loc = page.locator(sel).last
            if loc.count() > 0 and loc.is_visible(timeout=500):
                loc.click(timeout=3000, force=True)
                return True
        except Exception:
            continue
    try:
        clicked = page.evaluate(
            """() => {
              const nodes = Array.from(document.querySelectorAll('a,button,input'));
              for (const el of nodes) {
                const t = (el.value || el.textContent || '').replace(/\\s+/g, '');
                if (t === '저장하기') { el.click(); return true; }
              }
              return false;
            }"""
        )
        return bool(clicked)
    except Exception:
        return False


def attach_native_dialog_handler(page) -> dict:
    """브라우저 alert/confirm 대비 — '수정되었습니다' 등 네이티브 다이얼로그 accept.

    저장하기 클릭 *전에* 등록해야 한다.
    """
    state: dict = {"seen": False, "message": "", "accepted": False}

    def _on_dialog(dialog) -> None:  # noqa: ANN001
        try:
            state["seen"] = True
            state["message"] = dialog.message or ""
            dialog.accept()
            state["accepted"] = True
        except Exception:
            try:
                dialog.dismiss()
            except Exception:
                pass

    try:
        page.on("dialog", _on_dialog)
    except Exception:
        pass
    return state


def is_modify_page_open(page) -> bool:
    """검색필터 수정 팝업/페이지가 열려 있는지."""
    try:
        url = page.url or ""
        if "modify_filter" in url or "admin_group_modify" in url:
            return True
    except Exception:
        pass
    try:
        body = page.locator("body").inner_text(timeout=400) or ""
        if "검색필터 수정" in body and "저장상품수" in body:
            return True
        if "저장상품수" in body and "검색결과" in body and "저장하기" in body:
            return True
    except Exception:
        pass
    return False


def wait_modify_page_closed(page, *, timeout_ms: int = 20000) -> bool:
    """저장하기 후 '검색필터 수정' 팝업/페이지 닫힘 확인."""
    end = time.time() + timeout_ms / 1000.0
    # 잠깐은 열려 있을 수 있음 — 닫힐 때까지 대기
    while time.time() < end:
        if not is_modify_page_open(page):
            return True
        # 수정되었습니다 팝업이 뜨면 수정화면은 사실상 닫힌 것으로 본다
        try:
            body = page.locator("body").inner_text(timeout=300) or ""
            if "수정되었습니다" in body or "수정 되었습니다" in body:
                return True
        except Exception:
            pass
        time.sleep(0.2)
    return not is_modify_page_open(page)


def click_modified_confirm(page, *, timeout_ms: int = 20000, dialog_state: dict | None = None) -> bool:
    """'수정되었습니다' 팝업에서 '확인' 버튼을 반드시 클릭.

    - 네이티브 alert: attach_native_dialog_handler 가 이미 accept
    - HTML 레이어/모달: '확인' 버튼 클릭
    """
    end = time.time() + timeout_ms / 1000.0

    # 1) 네이티브 다이얼로그가 이미 처리된 경우
    if dialog_state and dialog_state.get("accepted"):
        msg = str(dialog_state.get("message") or "")
        if (not msg) or ("수정" in msg) or ("완료" in msg) or ("저장" in msg):
            return True

    while time.time() < end:
        if dialog_state and dialog_state.get("accepted"):
            return True

        # 2) HTML 팝업: '수정되었습니다' 근처의 '확인'
        try:
            ok = page.evaluate(
                """() => {
                  const bodyText = (document.body && document.body.innerText) || '';
                  const hasMsg = /수정\\s*되었습니다|수정되었습니다/.test(bodyText);
                  const nodes = Array.from(document.querySelectorAll(
                    'button, a, input[type="button"], input[type="submit"], input[type="image"]'
                  ));
                  // 정확히 '확인' 우선
                  for (const el of nodes) {
                    const t = ((el.value || el.innerText || el.textContent || '') + '')
                      .replace(/\\s+/g, '');
                    if (t !== '확인') continue;
                    // 메시지 보이거나, alert 레이어 안이면 클릭
                    const scope = el.closest(
                      '.ui-dialog, .modal, .layer, .popup, .alert, [role="dialog"], form, body'
                    );
                    const scopeText = ((scope && scope.innerText) || bodyText);
                    if (hasMsg || /수정\\s*되었습니다|수정되었습니다/.test(scopeText)) {
                      el.click();
                      return true;
                    }
                  }
                  // 폴백: 화면에 수정되었습니다가 보이면 첫 '확인' 클릭
                  if (hasMsg) {
                    for (const el of nodes) {
                      const t = ((el.value || el.innerText || el.textContent || '') + '')
                        .replace(/\\s+/g, '');
                      if (t === '확인') { el.click(); return true; }
                    }
                  }
                  return false;
                }"""
            )
            if ok:
                time.sleep(0.3)
                return True
        except Exception:
            pass

        # 3) Playwright locator 폴백
        try:
            msg = page.locator("text=수정되었습니다").first
            if msg.count() > 0 and msg.is_visible(timeout=200):
                for sel in (
                    'button:has-text("확인")',
                    'input[type="button"][value="확인"]',
                    'input[value="확인"]',
                    'a:has-text("확인")',
                ):
                    loc = page.locator(sel).last
                    if loc.count() > 0 and loc.is_visible(timeout=200):
                        loc.click(timeout=2000, force=True)
                        time.sleep(0.3)
                        return True
        except Exception:
            pass

        time.sleep(0.25)

    # 마지막: 다이얼로그만 처리됐고 메시지가 비어있어도(일부 브라우저) 통과
    if dialog_state and dialog_state.get("accepted"):
        return True
    return False


def wait_modify_page(page, *, timeout_ms: int = 20000) -> bool:
    """수집조건수정 후 '검색필터 수정' / 저장상품수 화면 대기.

    팝업 창·iframe 포함. not found 화면이면 실패.
    """
    end = time.time() + timeout_ms / 1000.0
    saw_not_found = False
    while time.time() < end:
        try:
            if page_shows_not_found(page):
                saw_not_found = True
                # 잠깐 더 기다려 정상 화면으로 바뀌는지 확인
                time.sleep(0.35)
                if page_shows_not_found(page):
                    return False
            target, kind = resolve_modify_target(page)
            url = ""
            try:
                url = getattr(target, "url", None) or page.url or ""
            except Exception:
                url = page.url or ""
            if "modify_filter" in url or "admin_group_modify" in url:
                if not page_shows_not_found(page):
                    # URL만 맞고 본문이 not found 인 경우 제외
                    if wait_for_save_count_ready(target, timeout_ms=400):
                        return True
                    body = ""
                    try:
                        body = target.locator("body").inner_text(timeout=300) or ""
                    except Exception:
                        body = ""
                    if "저장상품수" in body or "검색필터 수정" in body:
                        return True
            if kind in ("page", "frame") and kind != "main":
                # resolve 가 저장상품수 화면을 찾음
                if wait_for_save_count_ready(target, timeout_ms=500):
                    return True
            body = ""
            try:
                body = page.locator("body").inner_text(timeout=400) or ""
            except Exception:
                pass
            if "저장상품수" in body and (
                "검색필터 수정" in body or "검색결과" in body
            ):
                return True
            # 다른 탭/팝업
            try:
                for p in page.context.pages:
                    if p is page:
                        continue
                    if page_shows_not_found(p):
                        saw_not_found = True
                        continue
                    bu = p.url or ""
                    if "modify_filter" in bu or "admin_group_modify" in bu:
                        if wait_for_save_count_ready(p, timeout_ms=400):
                            return True
                        bt = p.locator("body").inner_text(timeout=300) or ""
                        if "저장상품수" in bt or "검색필터 수정" in bt:
                            return True
                    bt = p.locator("body").inner_text(timeout=300) or ""
                    if "저장상품수" in bt:
                        return True
            except Exception:
                pass
        except Exception:
            pass
        time.sleep(0.25)
    if saw_not_found and page_shows_not_found(page):
        return False
    target, _kind = resolve_modify_target(page)
    if page_shows_not_found(page):
        return False
    return wait_for_save_count_ready(target, timeout_ms=800)


def _return_to_list(page, list_url: str) -> None:
    """저장 후 검색필터 목록으로 복귀."""
    url = (list_url or "").strip()
    if not url:
        return
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60_000)
        time.sleep(0.5)
    except Exception:
        try:
            page.go_back(wait_until="domcontentloaded", timeout=30_000)
            time.sleep(0.4)
        except Exception:
            pass


def run_update(
    excel_path: str | Path,
    mango_url: str,
    *,
    progress: ProgressFn | None = None,
) -> RunResult:
    path = Path(excel_path).expanduser().resolve()
    result = RunResult(ok=False)
    if not path.is_file():
        result.errors.append(f"파일 없음: {path}")
        return result
    # ★고정: 비어 있으면 지정된 검색필터 URL 초기값 사용
    mango = (mango_url or "").strip() or DEFAULT_MANGO_URL

    clear_stop_flag()
    try:
        rows = read_excel_rows(path)
    except Exception as e:  # noqa: BLE001
        result.errors.append(str(e))
        return result
    if not rows:
        result.errors.append("엑셀에 URL 행이 없습니다.")
        return result
    by_url = excel_by_url(rows)
    _log(
        progress,
        "준비",
        f"엑셀 {path.name} · URL {len(rows)}건 · 망고URL={mango[:120]}",
    )
    _log(
        progress,
        "준비",
        "P2와 동일 — 망고 Chrome 창을 화면에 띄운 뒤 검색필터 URL로 이동 (로그인대기 없음)",
    )

    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"Playwright 미설치: {e}")
        return result

    try:
        import collect as p2  # type: ignore
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"P2 collect 로드 실패: {e}")
        return result

    try:
        with sync_playwright() as p:
            # ★P2 connect_browser → 창 앞으로 → 검색필터 URL 이동
            try:
                browser, page = attach_mango_browser_like_p2(p2, p, progress=progress)
            except Exception as e:  # noqa: BLE001
                result.errors.append(str(e))
                return result
            page = navigate_mango_url(page, mango, progress=progress, p2=p2) or page
            reveal_browser_page(
                page,
                progress,
                step_no="0",
                action="검색필터 목록 화면 (연동 동작 시작)",
                dwell_s=1.5,
            )
            shot_dir = new_shot_dir()
            _log(progress, "준비", f"스크린샷 폴더: {shot_dir}")
            try:
                cur_after = (page.url or "").strip()
            except Exception:
                cur_after = ""
            _log(progress, "준비", f"목록스캔 전 URL={cur_after[:180] or '(없음)'}")

            demango_rows = list_demango_rows(page)
            result.total_demango = len(demango_rows)
            _log(progress, "준비", f"더망고 목록 {result.total_demango}행 검출")

            if result.total_demango == 0:
                result.errors.append(
                    "검색필터 행 0건 — "
                    f"목표URL={mango[:100]} · 현재URL={cur_after[:100] or '(없음)'}. "
                    "로그인된 망고 Chrome에서 다시 실행하세요."
                )
                return result

            for i, drow in enumerate(demango_rows, start=1):
                if stop_requested():
                    _log(progress, "중단", "사용자 중단 요청")
                    break

                d_url = (drow.get("url") or "").strip()
                d_filter = (drow.get("filterName") or "").strip()
                row_idx = int(drow.get("index") or 0)
                edit_href = (drow.get("editHref") or "").strip()
                # ★요건1: 더망고 URL을 기준값으로 엑셀에서 동일 URL 검색
                ex = find_excel_by_demango_url(by_url, d_url)

                # ★요건2: 더망고 처음 10건 — 더망고/엑셀 검색필터·URL 두 줄
                log_first10_compare(
                    progress,
                    ordinal=i,
                    d_filter=d_filter,
                    d_url=d_url,
                    ex=ex,
                )

                # 1) KEY 매칭 (기준=더망고 URL)
                if not ex:
                    result.skipped += 1
                    if i > FIRST_COMPARE_LOG_N:
                        _log(
                            progress,
                            "불일치",
                            f"검색필터={d_filter} · URL={d_url}",
                        )
                    continue

                # 검색필터 비교 (공백→_ 재비교 포함)
                if ex.filter_name and d_filter and not filters_equal(
                    ex.filter_name, d_filter
                ):
                    result.skipped += 1
                    if i > FIRST_COMPARE_LOG_N:
                        _log(
                            progress,
                            "불일치",
                            f"검색필터={ex.filter_name} · URL={ex.url}",
                        )
                    continue

                # ★필터 매칭 시 전 단계 상세 + 브라우저에 동작(팝업 포함) 실제 표시
                lg = Logger(progress, ex.excel_row, force_verbose=True)
                lg.step(
                    "행",
                    f"{i}/{result.total_demango} 더망고URL기준 매칭 · 엑셀행={ex.excel_row}",
                    f"{i}/{result.total_demango} URL매칭=Y",
                )
                lg.step(
                    "로직",
                    f"1) 더망고URL→엑셀 일치 · 필터일치 — "
                    f"엑셀필터={ex.filter_name!r} · 더망고필터={d_filter!r} · "
                    f"상품수집가능개수={ex.collectible}",
                    f"1) URL·필터일치 · 수집가능={ex.collectible}",
                )
                note = filter_compare_note(ex.filter_name, d_filter)
                if note:
                    lg.step("로직", f"1) {note}", note)

                reveal_browser_page(
                    page,
                    progress,
                    step_no="1",
                    action=f"필터일치 목록행 표시 · filter={d_filter}",
                )
                screenshot_step(
                    page,
                    shot_dir,
                    step_tag="01_matched_list",
                    label=f"1)필터일치 목록행 filter={d_filter}",
                    row_no=i,
                    progress=progress,
                )

                # 목록에 있는지 확인 (이전 저장 후 복귀)
                try:
                    cur = page.url or ""
                    if "modify_filter" in cur or "admin_group_modify" in cur:
                        _return_to_list(page, mango)
                except Exception:
                    pass

                list_page = page
                reveal_browser_page(
                    list_page,
                    progress,
                    step_no="1",
                    action="더망고 목록 창 앞으로",
                    dwell_s=0.5,
                )

                # 1) URL 클릭 → 스토어/팝업을 브라우저에 표시
                lg.step("로직", "1) URL 클릭 (브라우저·팝업 실제 표시)", "1) URL클릭")
                store = click_demango_row_url(
                    list_page, row_idx, d_url, progress=progress
                )
                if store is None:
                    result.failed += 1
                    _log(progress, "오류", f"행{i} URL 클릭 실패")
                    screenshot_step(
                        list_page,
                        shot_dir,
                        step_tag="01_url_click_fail",
                        label="1)URL클릭 실패",
                        row_no=i,
                        progress=progress,
                    )
                    _return_to_list(list_page, mango)
                    continue
                screenshot_step(
                    store,
                    shot_dir,
                    step_tag="01_store_opened",
                    label="1)URL클릭 후 스토어/팝업",
                    row_no=i,
                    progress=progress,
                )

                # 2)~5) 팝업닫기 → 푸터↓ → 상단↑ 카드수 → 엑셀 비교 (창 표시)
                try:
                    browse_store_count_cards(
                        store,
                        excel_count=ex.collectible,
                        progress=progress,
                        shot_dir=shot_dir,
                        row_no=i,
                    )
                except Exception as e:  # noqa: BLE001
                    _log(
                        progress,
                        "경고",
                        f"행{i} 스토어 스크롤/상품수 집계 예외: {str(e).split(chr(10))[0][:120]}",
                    )

                # 더망고 목록 탭 재연결 후 수집조건수정
                page = close_store_return_list(
                    list_page, store, mango, progress=progress
                )
                if page is None:
                    result.failed += 1
                    _log(progress, "오류", f"행{i} 더망고 목록 탭 재연결 실패")
                    continue
                reveal_browser_page(
                    page,
                    progress,
                    step_no="6",
                    action="더망고 목록 복귀·창 표시",
                    dwell_s=STEP_VIEW_DWELL_SEC,
                )

                # 복귀 후 URL로 행 index 재확정 (stale index 방지)
                row_idx2 = resolve_demango_row_index_by_url(
                    page,
                    d_url,
                    fallback_index=row_idx,
                    progress=progress,
                )
                if row_idx2 is None:
                    result.failed += 1
                    _log(progress, "오류", f"행{i} 더망고 목록에서 URL 행 재탐색 실패")
                    continue
                row_idx = int(row_idx2)

                # 6) 「전체저장」바로 우측부터 한글 1글자씩 우측 · 2초×최대 10회
                lg.step(
                    "로직",
                    "6) 수집조건수정 클릭 (전체저장우측 · 글자씩 · 2초×10회)",
                    "6) 조건수정 클릭",
                )
                if not page_is_usable(page):
                    result.failed += 1
                    _log(progress, "오류", f"행{i} 더망고 페이지 핸들 사용불가")
                    continue
                if not click_edit_on_row(
                    page,
                    row_idx,
                    edit_href,
                    row_url=d_url,
                    progress=progress,
                    shot_dir=shot_dir,
                    row_no=i,
                    max_tries=EDIT_CLICK_MAX_TRIES,
                    try_interval_s=2.0,
                ):
                    result.failed += 1
                    _log(progress, "오류", f"행{i} 수집조건수정 클릭 실패")
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="06_edit_fail",
                        label="6)수집조건수정 실패",
                        row_no=i,
                        progress=progress,
                    )
                    _return_to_list(page, mango)
                    continue
                if not wait_modify_page(page):
                    result.failed += 1
                    if page_shows_not_found(page):
                        _log(
                            progress,
                            "오류",
                            f"행{i} 수집조건수정 후 not found — 잘못된 버튼/링크 가능",
                        )
                    else:
                        _log(progress, "오류", f"행{i} 검색필터 수정 화면 미진입")
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="06_modify_missing",
                        label="6)검색필터수정 미진입/notfound",
                        row_no=i,
                        progress=progress,
                    )
                    _return_to_list(page, mango)
                    continue
                try:
                    mod_page, _kind = resolve_modify_target(page)
                except Exception:
                    mod_page = page
                reveal_browser_page(
                    mod_page if mod_page is not None else page,
                    progress,
                    step_no="6",
                    action="검색필터 수정 팝업/화면 표시",
                    dwell_s=STEP_VIEW_DWELL_SEC,
                )
                screenshot_step(
                    mod_page if mod_page is not None else page,
                    shot_dir,
                    step_tag="06_modify_opened",
                    label="6)검색필터 수정 화면",
                    row_no=i,
                    progress=progress,
                )

                # 7) 저장상품수 수정 → 저장하기 → 확인
                target = map_save_count(ex.collectible)
                lg.step(
                    "로직",
                    f"7) 저장상품수 갱신 — 수집가능={ex.collectible} → 저장상품수={target}",
                    f"7) 저장상품수={target}",
                )
                if not set_save_count(
                    page,
                    target,
                    shot_dir=shot_dir,
                    progress=progress,
                    row_no=i,
                ):
                    result.failed += 1
                    _log(progress, "오류", f"행{i} 저장상품수 입력칸 실패")
                    try:
                        page.keyboard.press("Escape")
                    except Exception:
                        pass
                    _return_to_list(page, mango)
                    continue

                # 저장하기 → 팝업 닫힘 → "수정되었습니다" 확인 클릭
                dialog_state = attach_native_dialog_handler(page)
                lg.step("로직", "7) 팝업 하단 저장하기 클릭", "7) 저장하기")
                if not click_save_button(page):
                    result.failed += 1
                    _log(progress, "오류", f"행{i} 저장하기 클릭 실패")
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="07_save_click_fail",
                        label="7)저장하기 클릭 실패",
                        row_no=i,
                        progress=progress,
                    )
                    _return_to_list(page, mango)
                    continue
                screenshot_step(
                    page,
                    shot_dir,
                    step_tag="07_after_save_click",
                    label="7)저장하기 클릭 후",
                    row_no=i,
                    progress=progress,
                )

                lg.step(
                    "로직",
                    "7) 검색필터 수정 팝업 닫힘 확인",
                    "7) 수정팝업 닫힘",
                )
                if not wait_modify_page_closed(page, timeout_ms=20_000):
                    _log(progress, "경고", f"행{i} 수정팝업 닫힘 대기 시간초과 — 확인 팝업 계속 시도")
                screenshot_step(
                    page,
                    shot_dir,
                    step_tag="07_modify_closed",
                    label="7)수정팝업 닫힘/확인대기",
                    row_no=i,
                    progress=progress,
                )

                lg.step(
                    "로직",
                    "7) '수정되었습니다' 확인 클릭",
                    "7) 수정완료 확인",
                )
                if not click_modified_confirm(
                    page, timeout_ms=20_000, dialog_state=dialog_state
                ):
                    result.failed += 1
                    _log(progress, "오류", f"행{i} 수정되었습니다 확인 클릭 실패")
                    screenshot_step(
                        page,
                        shot_dir,
                        step_tag="07_confirm_fail",
                        label="7)수정완료확인 실패",
                        row_no=i,
                        progress=progress,
                    )
                    _return_to_list(page, mango)
                    continue
                screenshot_step(
                    page,
                    shot_dir,
                    step_tag="07_confirmed",
                    label="7)수정완료 확인클릭",
                    row_no=i,
                    progress=progress,
                )

                result.updated += 1
                lg.step(
                    "완료",
                    f"행{i} 갱신완료 · 엑셀행{ex.excel_row} · 저장상품수={target}",
                    f"갱신OK · 저장상품수={target}",
                )
                _return_to_list(page, mango)
                screenshot_step(
                    page,
                    shot_dir,
                    step_tag="08_back_to_list",
                    label="8)목록복귀",
                    row_no=i,
                    progress=progress,
                )
                time.sleep(0.35)

                if stop_requested():
                    break

    except Exception as e:  # noqa: BLE001
        result.errors.append(f"실행 실패: {e}")
        _log(progress, "오류", str(e))
        return result

    clear_stop_flag()
    result.ok = result.updated > 0 and not result.errors
    _log(
        progress,
        "완료",
        f"갱신 {result.updated} · 건너뜀 {result.skipped} · 실패 {result.failed} "
        f"/ 더망고 {result.total_demango}행",
    )
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="P3_필터_갱신 — 저장상품수 갱신")
    parser.add_argument("excel", help="엑셀 파일 경로")
    parser.add_argument(
        "--mango-url",
        default=DEFAULT_MANGO_URL,
        help="더망고 검색필터 URL (기본=getGoodsCategory.php filter_delete)",
    )
    args = parser.parse_args(argv)
    result = run_update(args.excel, args.mango_url)
    if result.errors:
        for e in result.errors:
            print(f"[오류] {e}", flush=True)
    if result.updated > 0 or (result.ok and not result.errors):
        return 0
    if result.skipped > 0 and result.failed == 0 and not result.errors:
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
