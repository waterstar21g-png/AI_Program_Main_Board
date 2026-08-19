"""P2 — 수집사이트명/URL 검증(verify_selected_site) · 엑셀 라벨 컬럼(최종 카테고리명) 단위테스트.

★요건(2026-08-19):
1. P2 화면에 "수집사이트명"·"수집사이트URL" 입력값이 더망고 좌측 "대량수집
   사이트" 목록(#site_list)에서 선택(active)된 사이트와 다르면 상세한 오류
   메세지와 함께 실행을 중단한다.
2. 카테고리URL목록·수집필터명(검색필터명)에는 "최종 카테고리명"을 우선 사용하고,
   없으면 "상위 최종 카테고리명"으로 안전하게 대체한다(하위 호환).

선택자는 실제 더망고 화면을 개발자도구로 확인해 다음과 같이 확정했다:
    #site_list .sites.active .text-name   → 사이트명 (title 속성, 예: "MUSINSA.com")
    #site_list .sites.active a.icon-link   → 사이트URL (href 속성)
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import collect as C  # noqa: E402


class _FakeElement:
    """Playwright Locator 흉내 — count()/get_attribute()/inner_text()/locator()
    를 모두 하나의 객체로 제공한다 (실제 Locator.first 도 Locator 자신을 반환)."""

    def __init__(self, attrs: dict | None = None, text: str = "", present: bool = True):
        self._attrs = attrs or {}
        self._text = text
        self._present = present

    def count(self) -> int:
        return 1 if self._present else 0

    def get_attribute(self, name: str):
        return self._attrs.get(name)

    def inner_text(self) -> str:
        return self._text

    def locator(self, _selector: str) -> "_FakeElement":
        return _FakeElement(present=False)

    @property
    def first(self) -> "_FakeElement":
        return self


class _FakeLocatorResult:
    def __init__(self, elements: list):
        self._elements = elements

    @property
    def first(self):
        return self._elements[0] if self._elements else _FakeElement(present=False)

    def count(self) -> int:
        return len(self._elements)


class _FakeActiveRow(_FakeElement):
    """`#site_list .sites.active` 행 — 내부에 .text-name / a.icon-link 를 갖는다."""

    def __init__(self, name_title: str = "", name_text: str = "", url_href: str = ""):
        super().__init__(present=True)
        self._name_el = _FakeElement({"title": name_title}, text=name_text)
        self._link_el = _FakeElement({"href": url_href})

    def locator(self, selector: str) -> "_FakeElement":
        if selector == ".text-name":
            return self._name_el
        if selector == "a.icon-link":
            return self._link_el
        return _FakeElement(present=False)


class _FakePage:
    def __init__(self, active_row: "_FakeActiveRow | None" = None):
        self._active_row = active_row

    def locator(self, selector: str):
        if selector == "#site_list .sites.active":
            return _FakeLocatorResult([self._active_row] if self._active_row else [])
        return _FakeLocatorResult([])


class _FakeCtx:
    def __init__(self):
        self.messages: list[str] = []

    def info(self, msg: str) -> None:
        self.messages.append(msg)


def test_verify_selected_site_skips_when_both_empty():
    page = _FakePage(active_row=_FakeActiveRow(name_title="MUSINSA.com", url_href="https://www.musinsa.com"))
    C.verify_selected_site(page, "", "")  # 예외 없이 통과(하위 호환)


def test_verify_selected_site_passes_when_name_matches():
    page = _FakePage(active_row=_FakeActiveRow(name_title="MUSINSA.com", url_href="https://www.musinsa.com"))
    ctx = _FakeCtx()
    C.verify_selected_site(page, "MUSINSA.com", "", ctx=ctx)
    assert any("검증 OK" in m for m in ctx.messages)


def test_verify_selected_site_passes_when_url_matches_case_and_www_insensitive():
    page = _FakePage(active_row=_FakeActiveRow(name_title="MUSINSA.com", url_href="https://www.musinsa.com"))
    ctx = _FakeCtx()
    C.verify_selected_site(page, "", "MUSINSA.COM", ctx=ctx)
    assert any("검증 OK" in m for m in ctx.messages)


def test_verify_selected_site_raises_detailed_error_on_mismatch():
    page = _FakePage(active_row=_FakeActiveRow(name_title="MUSINSA.com", url_href="https://www.musinsa.com"))
    with pytest.raises(RuntimeError) as exc:
        C.verify_selected_site(page, "ABCmart.a-rt.com", "https://abcmart.a-rt.com")
    msg = str(exc.value)
    assert "수집사이트명 불일치" in msg
    assert "수집사이트URL 불일치" in msg
    assert "ABCmart.a-rt.com" in msg
    assert "MUSINSA.com" in msg
    assert "중단" in msg


def test_verify_selected_site_skips_when_active_row_not_found():
    """★안전장치: #site_list 자체를 못 찾으면 확정 불일치로 보지 않고 건너뛴다."""
    page = _FakePage(active_row=None)
    ctx = _FakeCtx()
    C.verify_selected_site(page, "MUSINSA.com", "https://www.musinsa.com", ctx=ctx)
    assert any("검증생략" in m for m in ctx.messages)


def test_read_selected_site_from_list_reads_title_and_href():
    page = _FakePage(active_row=_FakeActiveRow(name_title="MUSINSA.com", url_href="https://www.musinsa.com"))
    name, url = C.read_selected_site_from_list(page)
    assert name == "MUSINSA.com"
    assert url == "https://www.musinsa.com"


def test_read_selected_site_from_list_missing_returns_empty():
    page = _FakePage(active_row=None)
    name, url = C.read_selected_site_from_list(page)
    assert name == "" and url == ""


def test_norm_site_token():
    assert C._norm_site_token("https://www.MUSINSA.com/") == "musinsa.com"
    assert C._norm_site_token("MUSINSA.com") == "musinsa.com"
    assert C._norm_site_token("Zara.com/de") == "zara.com/de"


def test_read_excel_uses_final_category_name_column(tmp_path):
    """★요건: 엑셀 라벨 컬럼은 "최종 카테고리명"을 우선 사용한다."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["상위 카테고리명", "중위 카테고리명", "하위 카테고리명", "최종 카테고리명", "상위 최종 카테고리명", "최종 카테고리 URL주소"])
    ws.append(["MEN", "상의", "반팔티", "(주)스타컴퍼니-ABC마트MEN-상의-반팔티", "MEN 반팔티", "https://abcmart.a-rt.com/x"])
    p = tmp_path / "sample.xlsx"
    wb.save(p)

    rows = C.read_excel(str(p))
    assert len(rows) == 1
    assert rows[0]["label"] == "(주)스타컴퍼니-ABC마트MEN-상의-반팔티"
    assert rows[0]["url"] == "https://abcmart.a-rt.com/x"


def test_read_excel_falls_back_to_old_header_when_final_category_name_missing(tmp_path):
    """★안전장치: "최종 카테고리명"이 없는(옛 포맷) 엑셀도 수집이 멈추지 않고
    "상위 최종 카테고리명"으로 대체해 읽는다."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["상위 최종 카테고리명", "최종 카테고리 URL주소"])
    ws.append(["MEN 반팔티", "https://abcmart.a-rt.com/x"])
    p = tmp_path / "old_format.xlsx"
    wb.save(p)

    rows = C.read_excel(str(p))
    assert len(rows) == 1
    assert rows[0]["label"] == "MEN 반팔티"
    assert rows[0]["url"] == "https://abcmart.a-rt.com/x"


def test_read_excel_missing_both_label_headers_raises(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["엉뚱한헤더", "최종 카테고리 URL주소"])
    ws.append(["x", "https://abcmart.a-rt.com/x"])
    p = tmp_path / "bad_format.xlsx"
    wb.save(p)

    with pytest.raises(SystemExit):
        C.read_excel(str(p))
