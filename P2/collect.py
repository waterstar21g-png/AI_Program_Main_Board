"""
P2 — 더망고(tmg1898) 상품데이터 대량수집

목표: 카테고리 URL(P1 엑셀) 1행당 상품 정보 N건(기본 3)을 오류 없이 가져오기.

0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭
1. URL상품검색하기 : 필드값 입력 후 클릭 -> 팝업창이 없어질 때까지 대기
2. [버튼1] 검색된 상품 모두저장 클릭 → 검색필터명·저장수 입력
   → [버튼2] 모달 하단 저장하기 클릭 (모두저장과 다른 버튼, 서버 최종 갱신)
3. 저장 실행 팝업/알림 확인(수집건수) — 서버 반영 증거
4. 다음 행 (버튼2 성공 전에 진행 금지 / 실패 시 1회 재시도 후 다음 행)

사용법:
    python collect.py 엑셀.xlsx              # 저장수 3
    python collect.py 엑셀.xlsx 3 --verify   # 1행 검증(샷·재시도)
    python collect.py 엑셀.xlsx 3 --retries 3 --yes
    run-verify.bat 엑셀.xlsx
"""

import os
import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    TimeoutError as PWTimeout,
    sync_playwright,
)

# Windows cp949 콘솔/파이프에서 특수기호 출력 시 UnicodeEncodeError 방지
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass


def safe_print(msg: str = "", *, flush: bool = True) -> None:
    """stdout이 cp949여도 죽지 않게 출력."""
    text = "" if msg is None else str(msg)
    try:
        print(text, flush=flush)
        return
    except UnicodeEncodeError:
        pass
    enc = getattr(sys.stdout, "encoding", None) or "utf-8"
    try:
        data = (text + "\n").encode(enc, errors="replace")
        sys.stdout.buffer.write(data)
        if flush:
            sys.stdout.flush()
    except Exception:
        try:
            print(text.encode("ascii", errors="replace").decode("ascii"), flush=flush)
        except Exception:
            pass

LOGIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin_login.php"
MAIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin.php"
BULK_URL = "https://tmg1898.cafe24.com/mall/admin/shop/getGoodsNew.php"
ADMIN_HOST = "tmg1898.cafe24.com"
BULK_PATH = "getGoodsNew.php"

CDP_PORT = 9222
CDP_URL = f"http://127.0.0.1:{CDP_PORT}"
PROFILE_DIR = Path(__file__).parent / ".chrome-profile"

# 더망고 솔루션 Chrome 확장프로그램 (Web Store ID = 로컬 load-extension 동일)
MANGO_EXT_ID = "lgfjcapohoongednoojdaiedebgbcelp"
MANGO_EXT_DIR = Path(__file__).parent / "extensions" / "themango-solution"
MANGO_EXT_POPUP = f"chrome-extension://{MANGO_EXT_ID}/popup.html"
# 크롬 기동 시 확장 팝업에 반드시 넣을 값 (사용자 지정)
MANGO_SERVICE_URL = "https://tmg1898.cafe24.com"
MANGO_SERVICE_KEY = "y94Tmx9LbxxCJtk5uI9z0RjGWDtVW4"

CHROME_CANDIDATES = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    "/usr/bin/google-chrome-stable",
    "/usr/bin/google-chrome",
    "/usr/local/bin/google-chrome",
    "/usr/bin/chromium-browser",
    "/usr/bin/chromium",
]

POPUP_WAIT_SEC = 40  # 검색 팝업 닫힘 대기(초) — 초과 시 닫고 다음 단계로
MODAL_WAIT_SEC = 60
# 저장하기 클릭 후 서버 처리·팝업 생성에 주는 최소 시간(초)
# (이 전에 화면에 있던 '00건 수집' 문구로 즉시 통과·초기화 금지)
SAVE_POPUP_GRACE_SEC = 4.0
DEFAULT_SAVE_COUNT = 3
DEFAULT_ROW_RETRIES = 2
SEARCH_MAX_TRIES = 2  # URL 검색 재시도(행 안) — 적게 두고 다음 행으로 넘김
ROW_BUDGET_SEC = 180  # 한 입력 행에 쓸 수 있는 최대 시간(초) — 초과 시 다음 행

# 보드 "수집 종료" 버튼이 만드는 중단 플래그 (로그는 보드에 보존)
STOP_FLAG = Path(__file__).parent / ".collect_stop"

# 같은 행을 더 돌리지 않고 다음 입력 데이터로 넘길 확정 실패 문구
ROW_ADVANCE_FAIL_MARKERS = (
    "검색결과가 없습니다",
    "망고 자체 메세지",
    "더망고 자체 메세지",
    "행 제한시간 초과",
    "팝업창이 닫히지 않음",
    "저장 팝업창이 닫히지 않음",
    "검색결과 확인 실패",
    "팝업이 뜨지 않음",
    "0건이 수집",
    "수집건수 알림",
    "수집 알림",
    "저장하기 서버",
    "서버에 반영되지",
    "서버 최종 갱신",
    "팝업창 모달이 나타나지",
    "팝업 없이 초기화",
)


class CollectStopped(Exception):
    """사용자가 보드에서 수집 종료를 요청함."""


class RowBudgetExceeded(Exception):
    """한 입력 행 처리 시간 초과 — 다음 입력으로 진행."""


def clear_stop_flag() -> None:
    try:
        STOP_FLAG.unlink(missing_ok=True)  # type: ignore[call-arg]
    except TypeError:
        # py<3.8 호환 아님 — 3.10+ 환경
        if STOP_FLAG.exists():
            try:
                STOP_FLAG.unlink()
            except OSError:
                pass
    except OSError:
        pass


def stop_requested() -> bool:
    try:
        return STOP_FLAG.is_file()
    except OSError:
        return False


def check_stop(where: str = "") -> None:
    if stop_requested():
        detail = f" ({where})" if where else ""
        raise CollectStopped(f"사용자 수집 종료 요청{detail}")

FILTER_NAME_LABEL = re.compile(r"검색\s*필터\s*명")
# 화면 표기: 저장상품수 / 검색결과상위 / (사용자 호칭) 수집상품수
SAVE_COUNT_LABEL = re.compile(
    r"저장\s*상품\s*수|검색결과\s*상위|수집\s*상품\s*수"
)
# 저장 완료로 볼 수 있는 화면 문구 (망고 버전에 따라 다를 수 있음)
SAVE_OK_PATTERNS = [
    re.compile(r"저장\s*(이\s*)?(완료|성공)"),
    re.compile(r"정상\s*처리"),
    re.compile(r"상품\s*(이\s*)?저장"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?저장"),
]
SAVE_FAIL_PATTERNS = [
    re.compile(r"저장\s*실패"),
    re.compile(r"오류\s*가\s*발생"),
    re.compile(r"다시\s*시도"),
]

# 망고 자체 알림 — "3건이 수집되었다" / "00건이수집되었다" / 저장 건수 등
MANGO_COLLECT_ALERT_PATTERNS = [
    re.compile(r"(\d+)\s*건\s*(이\s*)?수집\s*되었"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?저장\s*되었"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?등록\s*되었"),
    re.compile(r"(\d+)\s*건\s*(이\s*)?수집\s*완료"),
    re.compile(r"(\d+)\s*건\s*(의\s*)?상품\s*(이\s*)?(수집|저장)"),
    re.compile(r"저장\s*된\s*상품\s*(\d+)\s*건"),
    re.compile(r"상품\s*(\d+)\s*건\s*(이\s*)?(수집|저장)"),
    re.compile(r"총\s*(\d+)\s*건\s*(이\s*)?(수집|저장)"),
]

# 더망고(자체 UI) — 검색 결과 없음 문구 (로딩 중 오판 금지, 로딩 종료 후에만 사용)
MANGO_NO_RESULT_PATTERNS = [
    re.compile(r"검색하신\s*검색에\s*대한\s*검색결과가\s*없습니다"),
    re.compile(r"검색결과가\s*없습니다"),
    re.compile(r"정확한\s*검색어인지\s*다시한번\s*확인"),
]

# 더망고 로딩 오버레이/문구 (빨간 "잠시만 기다려주세요" 등)
MANGO_LOADING_PATTERNS = [
    re.compile(r"load\s*product", re.I),
    re.compile(r"상품정보를\s*불러오는\s*중"),
    re.compile(r"잠시만\s*기다려"),
    re.compile(r"처리\s*중\s*입니다"),
    re.compile(r"검색\s*중"),
]

SHOT_ROOT = Path(__file__).parent / "run-logs"

# 현재 실행 컨텍스트(로그인 등 단계 샷을 같은 폴더에 모으기)
_ACTIVE_CTX: "RunCtx | None" = None

# 파일명 태그 → 한글 단계명 (갤러리/보드 표시용)
SHOT_STEP_LABELS: dict[str, str] = {
    "login_wait": "로그인 대기(창 표시)",
    "login_ok": "로그인 완료",
    "login_gate": "로그인 게이트",
    "login_required": "세션만료·재로그인",
    "ext_settings": "확장프로그램(더망고솔루션) 설정값 저장",
    "ext_settings_fail": "확장프로그램 설정 실패",
    "ready": "준비완료(대량수집 진입)",
    "00_overlays_clear": "다음행 전 — 팝업/모달 전부 닫힘 확인",
    "00_overlays_stuck": "다음행 전 — 팝업/모달 미종료(경고)",
    "00_init_bulk": "0. 초기화 — 대량데이터수집",
    "01_url_filled": "1. URL 입력 완료",
    "01_popup_missing": "1. 검색 팝업 미표시(오류)",
    "01_popup_opened": "1. 검색 팝업 열림",
    "01_popup_closed": "1. 검색 팝업 닫힘",
    "01_mango_no_results": "1. 망고 검색결과 없음(자체메세지)",
    "01_results_ready": "1. 검색 결과 준비",
    "02_save_modal": "2. 모두저장 모달",
    "02_no_count_field": "2. 저장수 필드 없음(오류)",
    "02_count_mismatch": "2. 저장수 불일치(오류)",
    "02_modal_filled": "2. 필터명·저장수 입력",
    "02_save_missing": "2. 저장하기 버튼 없음(오류)",
    "02_save_clicked": "2. 저장하기(서버제출) 클릭",
    "02_save_no_react": "2. 저장하기 클릭 무반응(재시도)",
    "02_save_failed": "2. 저장하기 서버제출 실패",
    "03_modal_stuck": "3. 저장 모달 미종료(오류)",
    "03_modal_closed": "3. 저장 모달 닫힘",
    "03_result_popup": "3. 저장 후 결과 팝업",
    "03_result_missing": "3. 저장 후 결과 팝업 없음(오류)",
    "03_collect_alert": "3. 망고 수집건수 알림 확인",
    "03_collect_alert_fail": "3. 망고 수집건수 알림 확인 실패",
    "04_row_done": "4. 행 완료",
}

# 사용자 수동 로그인 대기 제한 (초)
LOGIN_WAIT_SEC = 600


def log(msg: str) -> None:
    safe_print(f"[{time.strftime('%H:%M:%S')}] {msg}")


class RunCtx:
    """행 단위 수집·재시도·스크린샷용 실행 컨텍스트

    shot_first_n: 입력 데이터 앞 N건(기본 2=1·2행)만 단계별 스크린샷.
    """

    def __init__(
        self,
        *,
        save_count: int = DEFAULT_SAVE_COUNT,
        retries: int = DEFAULT_ROW_RETRIES,
        verify: bool = False,
        max_rows: int | None = None,
        batch: bool = False,
        shot_dir: Path | None = None,
        shot_first_n: int = 2,
    ) -> None:
        global _ACTIVE_CTX
        self.save_count = save_count
        self.retries = max(1, retries)
        self.verify = verify
        self.max_rows = max_rows
        self.batch = batch or verify  # 검증 모드는 기본 무중단
        self.shot_first_n = max(0, int(shot_first_n))
        stamp = time.strftime("%Y%m%d_%H%M%S")
        self.shot_dir = shot_dir or (SHOT_ROOT / stamp)
        self.shot_dir.mkdir(parents=True, exist_ok=True)
        self.step_i = 0
        self.shots: list[dict] = []
        self._gallery_written = False
        self.input_ordinal = 0  # 처리 중인 입력 순서(1부터)
        self.current_label = ""
        self.current_url = ""
        self.row_deadline: float | None = None  # 행당 제한시간(epoch)
        # 저장하기(서버 최종 갱신) 성공 여부 — True 되기 전 행 완료 금지
        self.server_save_ok: bool = False
        # 저장하기 후 팝업창 모달을 실제로 봤는지 — 없으면 초기화 금지
        self.save_popup_seen: bool = False
        self.log_path = self.shot_dir / "run.log"
        self._log_file = open(self.log_path, "a", encoding="utf-8")
        _ACTIVE_CTX = self
        self.info(f"[샷폴더] {self.shot_dir}")
        if self.shot_first_n > 0:
            self.info(f"[샷대상] 입력 데이터 1~{self.shot_first_n}행 단계별 스크린샷")

    def check_budget(self, where: str = "") -> None:
        """중단 요청 + 행 제한시간 검사."""
        check_stop(where)
        if self.row_deadline is not None and time.time() > self.row_deadline:
            detail = f" ({where})" if where else ""
            raise RowBudgetExceeded(
                f"행 제한시간 초과{detail} — {ROW_BUDGET_SEC}초 내 미완료, 다음 입력으로"
            )

    def begin_row(self, ordinal: int, row: dict) -> None:
        """행 처리 시작 — 로그에 카테고리명·URL 기록."""
        self.input_ordinal = ordinal
        self.current_label = str(row.get("label") or "").strip()
        self.current_url = str(row.get("url") or "").strip()
        self.row_deadline = time.time() + ROW_BUDGET_SEC
        self.server_save_ok = False
        self.save_popup_seen = False
        excel_row = row.get("row", "?")
        self.info(
            f"--- 입력#{ordinal} 엑셀{excel_row}행 | "
            f"상위 최종 카테고리명={self.current_label} | "
            f"최종 카테고리 URL주소={self.current_url} | "
            f"제한 {ROW_BUDGET_SEC}초 ---"
        )

    def wants_row_shot(self, row_no: int | None = None) -> bool:
        """공통(로그인 등 row_no=0) 또는 입력 1~N행만 샷."""
        if row_no in (None, 0):
            return True
        return 0 < self.input_ordinal <= self.shot_first_n

    def close(self) -> None:
        global _ACTIVE_CTX
        try:
            self.write_gallery()
        except Exception as e:  # noqa: BLE001
            try:
                self.info(f"  [갤러리 작성 실패] {e}")
            except Exception:
                pass
        try:
            self._log_file.close()
        except Exception:
            pass
        if _ACTIVE_CTX is self:
            _ACTIVE_CTX = None

    def info(self, msg: str) -> None:
        line = f"[{time.strftime('%H:%M:%S')}] {msg}"
        safe_print(line)
        try:
            self._log_file.write(line + "\n")
            self._log_file.flush()
        except Exception:
            pass

    @staticmethod
    def label_for_tag(tag: str) -> str:
        if tag in SHOT_STEP_LABELS:
            return SHOT_STEP_LABELS[tag]
        for key, label in SHOT_STEP_LABELS.items():
            if key in tag:
                return label
        if tag.startswith("fail_attempt"):
            return f"실패 시도 {tag.replace('fail_attempt', '')}"
        return tag

    def shot(self, page: Page, tag: str, row_no: int | None = None) -> Path | None:
        if not self.wants_row_shot(row_no):
            return None
        self.step_i += 1
        safe = re.sub(r"[^\w\-가-힣]+", "_", tag)[:40]
        ord_part = f"i{self.input_ordinal}" if self.input_ordinal else "i0"
        name = f"{self.step_i:02d}_{ord_part}_r{row_no or 0}_{safe}.png"
        path = self.shot_dir / name
        label = self.label_for_tag(tag)
        try:
            page.screenshot(path=str(path), full_page=True)
            self.shots.append(
                {
                    "step": self.step_i,
                    "ordinal": self.input_ordinal,
                    "row": row_no or 0,
                    "tag": tag,
                    "label": label,
                    "category": self.current_label,
                    "url": self.current_url,
                    "file": name,
                    "path": str(path),
                }
            )
            extra = ""
            if self.current_label or self.current_url:
                extra = f" | {self.current_label} | {self.current_url}"
            self.info(f"  [샷] {self.step_i:02d}. {label} -> {path.name}{extra}")
            return path
        except Exception as e:  # noqa: BLE001
            self.info(f"  [샷 실패] {label}: {e}")
            return None

    def write_gallery(self) -> Path | None:
        """1행 전과정 스크린샷 HTML 갤러리 + JSON 인덱스 작성."""
        if self._gallery_written:
            gallery = self.shot_dir / "index.html"
            return gallery if gallery.is_file() else None
        pngs = sorted(self.shot_dir.glob("*.png"))
        if not pngs and not self.shots:
            return None

        # shots 목록이 비어 있으면 파일명에서 재구성
        items = list(self.shots)
        if not items:
            for i, p in enumerate(pngs, start=1):
                stem = p.stem
                tag = stem
                m = re.match(r"^\d+_r\d+_(.+)$", stem)
                if m:
                    tag = m.group(1)
                items.append(
                    {
                        "step": i,
                        "row": 0,
                        "tag": tag,
                        "label": self.label_for_tag(tag),
                        "file": p.name,
                        "path": str(p),
                    }
                )

        import json

        idx_path = self.shot_dir / "shots.json"
        idx_path.write_text(
            json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        rows_html: list[str] = []
        for it in items:
            f = it["file"]
            label = it.get("label") or ""
            step = it.get("step", 0)
            cat = it.get("category") or ""
            url = it.get("url") or ""
            ord_n = it.get("ordinal") or 0
            meta_bits = [f]
            if ord_n:
                meta_bits.append(f"입력#{ord_n}")
            if cat:
                meta_bits.append(f"상위 최종 카테고리명={cat}")
            if url:
                meta_bits.append(f"최종 카테고리 URL주소={url}")
            meta = " | ".join(meta_bits)
            rows_html.append(
                f'<section class="shot">'
                f"<h2>{step:02d}. {label}</h2>"
                f'<p class="meta">{meta}</p>'
                f'<img src="{f}" alt="{label}"/>'
                f"</section>"
            )

        html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8"/>
<title>1·2행 전과정 스크린샷</title>
<style>
body {{ font-family: "Malgun Gothic", sans-serif; margin: 16px; background: #0f172a; color: #e2e8f0; }}
h1 {{ font-size: 20px; }}
.shot {{ margin: 24px 0; padding: 12px; background: #1e293b; border-radius: 8px; }}
.shot h2 {{ margin: 0 0 6px; font-size: 16px; color: #93c5fd; }}
.meta {{ margin: 0 0 10px; color: #94a3b8; font-size: 12px; word-break: break-all; }}
img {{ max-width: 100%; height: auto; border: 1px solid #334155; background: #fff; }}
</style>
</head>
<body>
<h1>1·2행 전과정 스크린샷 ({len(items)}장)</h1>
<p>폴더: {self.shot_dir}</p>
{''.join(rows_html)}
</body>
</html>
"""
        gallery = self.shot_dir / "index.html"
        gallery.write_text(html, encoding="utf-8")
        self._gallery_written = True
        self.info(f"[갤러리] {gallery} ({len(items)}장)")
        return gallery


def shot_now(page: Page, tag: str, row_no: int | None = 0) -> Path | None:
    """활성 RunCtx가 있으면 같은 샷폴더에, 없으면 run-logs 루트에 저장."""
    ctx = _ACTIVE_CTX
    if ctx is not None:
        return ctx.shot(page, tag, row_no)
    SHOT_ROOT.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^\w\-가-힣]+", "_", tag)[:40]
    path = SHOT_ROOT / f"{time.strftime('%Y%m%d_%H%M%S')}_{safe}.png"
    try:
        page.screenshot(path=str(path), full_page=True)
        log(f"  [샷] {path.name}")
        return path
    except Exception as e:  # noqa: BLE001
        log(f"  [샷 실패] {e}")
        return None


# ── 브라우저 연결 (기존 Chrome/Edge에 CDP로 붙기, Chromium 다운로드 없음) ──

def cdp_port_open(port: int = CDP_PORT) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        try:
            s.connect(("127.0.0.1", port))
            return True
        except OSError:
            return False


def find_browser_exe() -> str | None:
    for path in CHROME_CANDIDATES:
        if Path(path).exists():
            return path
    return None


def _clear_stale_singleton_locks() -> None:
    """
    이전 실행이 비정상 종료되어 남은 잠금파일이 있으면, 새 Chrome이
    "이미 실행 중"이라고 착각해 디버그 포트 없이 조용히 기존 창에만
    메시지를 보내고 끝나버릴 수 있다(=화면에 아무 반응도 없어 보임).
    cdp_port_open()이 False라는 건 우리 프로필로 살아있는 프로세스가
    없다는 뜻이므로 안전하게 지운다.
    """
    for name in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
        f = PROFILE_DIR / name
        try:
            if f.exists() or f.is_symlink():
                f.unlink()
        except OSError:
            pass


def launch_debug_browser() -> None:
    """평소 쓰는 Chrome/Edge를 디버그 포트로 실행 — Playwright Chromium 미사용"""
    exe = find_browser_exe()
    if not exe:
        raise SystemExit(
            "Chrome 또는 Edge를 찾지 못했습니다.\n"
            "https://www.google.com/chrome/ 에서 설치 후 다시 실행하세요."
        )
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    _clear_stale_singleton_locks()
    log(f"브라우저 실행: {exe}")
    # 주의: stdout/stderr를 DEVNULL로 버리면 일부 환경(보안 소프트웨어·
    # 컨테이너 등)에서 Chrome이 바로 죽는 경우가 있다(무반응처럼 보임).
    # 로그 파일로 받아두면 안전하고, 문제 생기면 이 파일로 원인도 알 수 있다.
    log_path = PROFILE_DIR / "chrome_debug.log"
    log_file = open(log_path, "ab")
    chrome_args = [
        exe,
        f"--remote-debugging-port={CDP_PORT}",
        f"--user-data-dir={PROFILE_DIR}",
        "--no-first-run",
        "--no-default-browser-check",
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--disable-gpu",
    ]
    # 더망고 솔루션 확장 로드 (전용 프로필에 설정값이 비어 있는 문제 방지)
    if MANGO_EXT_DIR.is_dir() and (MANGO_EXT_DIR / "manifest.json").is_file():
        ext_path = str(MANGO_EXT_DIR.resolve())
        chrome_args.append(f"--load-extension={ext_path}")
        chrome_args.append(f"--disable-extensions-except={ext_path}")
        log(f"더망고 솔루션 확장 로드: {ext_path}")
    else:
        log(f"[경고] 확장 폴더 없음 — {MANGO_EXT_DIR} (Web Store 설치분만 사용)")
    chrome_args.append(MAIN_URL)
    subprocess.Popen(
        chrome_args,
        stdout=log_file,
        stderr=log_file,
    )
    log("디버그 모드 연결 대기 중 (최대 30초)...")
    for i in range(60):
        if cdp_port_open():
            log("브라우저 연결 확인됨")
            return
        if i > 0 and i % 6 == 0:
            log(f"  아직 대기 중... ({i * 0.5:.0f}초 경과 — 화면에 Chrome 창이 떴는지 확인해 주세요)")
        time.sleep(0.5)  # 소켓 연결 대기 — Playwright 이벤트루프와 무관하므로 안전
    raise SystemExit(
        "브라우저가 디버그 모드로 열리지 않았습니다.\n"
        "열려있는 Chrome/Edge 창을 모두 닫고 다시 실행해 보세요.\n"
        f"(참고 로그: {log_path})"
    )


def pick_working_page(context: BrowserContext) -> Page:
    """이미 망고 화면이 열려 있으면 그 탭을 그대로 사용(기본 화면 유지)"""
    open_pages = [p for p in context.pages if not p.is_closed()]
    for p in open_pages:
        try:
            if ADMIN_HOST in p.url:
                return p
        except Exception:  # noqa: BLE001
            continue
    return open_pages[0] if open_pages else context.new_page()


def refresh_if_closed(page: Page) -> Page:
    """
    로그인 성공 후 사이트가 원래 탭을 닫고 새 창을 띄우는 경우가 있다
    (예: 로그인 중계 페이지가 자기 자신을 닫음). 그러면 이전 page
    객체로는 더 이상 아무 것도 할 수 없으므로(TargetClosedError),
    같은 컨텍스트에서 살아있는 페이지를 다시 찾아온다.
    """
    if not page.is_closed():
        return page
    log("  탭이 닫힘 감지 — 새 탭을 다시 찾는 중...")
    return pick_working_page(page.context)


def connect_browser(p) -> tuple[Browser, Page]:
    """
    1) 이미 디버그 모드로 열린 Chrome/Edge가 있으면 그대로 연결(= 망고 기본 화면 그대로)
    2) 없으면 평소 쓰는 Chrome/Edge를 디버그 모드로 새로 열어서 연결
    Playwright 전용 Chromium은 내려받지 않는다.
    """
    if not cdp_port_open():
        launch_debug_browser()

    browser = p.chromium.connect_over_cdp(CDP_URL)
    context = browser.contexts[0] if browser.contexts else browser.new_context()
    page = pick_working_page(context)
    return browser, page


def ensure_mango_extension_settings(
    context: BrowserContext,
    *,
    shot_ctx: "RunCtx | None" = None,
) -> None:
    """더망고 솔루션 확장에 서비스 URL·인증 KEY를 넣고 설정값 저장.

    Chrome 기동(전용 프로필) 직후 확장 팝업 값이 비어 있는 문제를 막는다.
    """
    want_url = MANGO_SERVICE_URL.strip()
    want_key = MANGO_SERVICE_KEY.strip()
    log(
        "더망고 솔루션 확장 설정 확인 — "
        f"URL={want_url} / KEY={want_key[:4]}…{want_key[-4:]}"
    )
    page = context.new_page()
    dialogs: list[str] = []

    def _on_dialog(dialog) -> None:
        try:
            dialogs.append(str(dialog.message or ""))
            dialog.accept()
        except Exception:  # noqa: BLE001
            pass

    page.on("dialog", _on_dialog)
    try:
        try:
            page.goto(MANGO_EXT_POPUP, wait_until="domcontentloaded", timeout=20_000)
        except Exception as e:  # noqa: BLE001
            if shot_ctx is not None:
                try:
                    shot_ctx.shot(page, "ext_settings_fail", 0)
                except Exception:  # noqa: BLE001
                    pass
            raise RuntimeError(
                "더망고 솔루션 확장프로그램 팝업을 열 수 없습니다.\n"
                f"  · 대상: {MANGO_EXT_POPUP}\n"
                "  · Chrome을 모두 닫고 P2를 다시 실행하세요 "
                "(전용 프로필에 확장이 로드됩니다).\n"
                f"  · 원인: {e}"
            ) from e

        page.wait_for_selector("#site_url", timeout=10_000)
        page.wait_for_selector("#site_key", timeout=5_000)
        # load_data.js 의 chrome.storage.local.get 반영 대기
        page.wait_for_timeout(700)

        cur_url = (page.input_value("#site_url") or "").strip()
        cur_key = (page.input_value("#site_key") or "").strip()
        if cur_url == want_url and cur_key == want_key:
            log("  확장 설정값 이미 올바름 — 저장 스킵")
            if shot_ctx is not None:
                shot_ctx.shot(page, "ext_settings", 0)
            return

        page.fill("#site_url", want_url)
        page.fill("#site_key", want_key)
        # 상품수집 ON 유지
        try:
            on = page.locator("#onoff")
            if on.count() and not on.is_checked():
                page.evaluate(
                    """() => {
                        const el = document.querySelector('#onoff');
                        if (!el) return;
                        if (window.jQuery) {
                            try { window.jQuery('#onoff').bootstrapToggle('on'); return; }
                            catch (e) {}
                        }
                        el.checked = true;
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                    }"""
                )
        except Exception:  # noqa: BLE001
            pass

        page.click("#sync_set")
        page.wait_for_timeout(400)
        # 저장 성공 문구 또는 storage 반영 확인
        saved_btn = False
        try:
            page.wait_for_function(
                """() => {
                    const b = document.querySelector('#sync_set');
                    return !!(b && (b.innerText || '').includes('저장되었습니다'));
                }""",
                timeout=5_000,
            )
            saved_btn = True
        except Exception:  # noqa: BLE001
            saved_btn = False

        stored = page.evaluate(
            """async () => {
                const local = await chrome.storage.local.get(
                    ['site_url', 'site_key', 'onoff']
                );
                return {
                    site_url: (local.site_url || '').trim(),
                    site_key: (local.site_key || '').trim(),
                    onoff: local.onoff || '',
                };
            }"""
        )
        got_url = str((stored or {}).get("site_url") or "").strip()
        got_key = str((stored or {}).get("site_key") or "").strip()
        if got_url != want_url or got_key != want_key:
            if shot_ctx is not None:
                try:
                    shot_ctx.shot(page, "ext_settings_fail", 0)
                except Exception:  # noqa: BLE001
                    pass
            raise RuntimeError(
                "더망고 솔루션 확장 설정값 저장 실패.\n"
                f"  · 기대 URL={want_url}\n"
                f"  · 실제 URL={got_url or '(비어있음)'}\n"
                f"  · KEY 일치={got_key == want_key}\n"
                f"  · 버튼저장문구={saved_btn}\n"
                f"  · 대화상자={dialogs[:2] if dialogs else '(없음)'}"
            )

        log("  확장 설정값 저장 완료 (서비스 URL + 인증 KEY)")
        if shot_ctx is not None:
            shot_ctx.shot(page, "ext_settings", 0)
    finally:
        try:
            page.close()
        except Exception:  # noqa: BLE001
            pass


# ── 엑셀 ──────────────────────────────────────────────────────

def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        label_col = headers.index("상위 최종 카테고리명")
        url_col = headers.index("최종 카테고리 URL주소")
    except ValueError:
        raise SystemExit(
            "엑셀 1행 헤더에 '상위 최종 카테고리명', '최종 카테고리 URL주소' 열이 있어야 합니다."
        )

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        label = str(row[label_col].value or "").strip()
        url = str(row[url_col].value or "").strip()
        if url:
            rows.append({"row": i, "label": label, "url": url})
    return rows


def normalize_url(u: str) -> str:
    u = u.strip()
    return u if re.match(r"^https?://", u, re.I) else f"https://{u}"


# ── 팝업 ──────────────────────────────────────────────────────

def _is_browser_internal_url(u: str) -> bool:
    low = (u or "").lower()
    return (
        low.startswith("chrome-extension://")
        or low.startswith("chrome://")
        or low.startswith("edge://")
        or low.startswith("devtools://")
        or low.startswith("about:")
    )


def popups(page: Page) -> list:
    """검색용 외부 팝업만 (확장/크롬 내부 탭 제외 — 무한대기 방지)."""
    result = []
    for p in page.context.pages:
        if p is page or p.is_closed():
            continue
        try:
            u = p.url
        except Exception:
            continue
        if not u or _is_browser_internal_url(u):
            continue
        if ADMIN_HOST in u:
            continue
        result.append(p)
    return result


def close_search_popups(page: Page) -> int:
    """남아 있는 검색 팝업을 닫고 닫은 개수를 반환."""
    closed = 0
    for p in list(popups(page)):
        try:
            if not p.is_closed():
                p.close()
                closed += 1
        except Exception:  # noqa: BLE001
            pass
    return closed


def wait_popup_open(page: Page, grace_sec: float = 15.0) -> list:
    """검색 팝업(새 창)이 열릴 때까지 대기. 열린 팝업 Page 리스트 반환."""
    grace_end = time.time() + max(0.5, float(grace_sec))
    while time.time() < grace_end:
        check_stop("팝업 열림 대기")
        cur = popups(page)
        if cur:
            return cur
        page.wait_for_timeout(200)
    return []


def wait_popups_close(page: Page, timeout_sec: int = POPUP_WAIT_SEC) -> None:
    """열린 검색 팝업이 모두 닫힐 때까지 대기.

    시간 초과 시 팝업을 닫고 다음 단계로 진행(1행에 무한대기하지 않음).
    """
    end = time.time() + max(5, int(timeout_sec))
    last_beat = 0.0
    while popups(page):
        check_stop("팝업 닫힘 대기")
        if time.time() > end:
            n = close_search_popups(page)
            log(
                f"  [경고] 팝업 닫힘 대기 {timeout_sec}초 초과 — "
                f"남은 팝업 {n}개 강제 닫고 다음 단계로 진행"
            )
            return
        if time.time() - last_beat > 10:
            last_beat = time.time()
            cur = popups(page)
            urls = ", ".join(p.url for p in cur if not p.is_closed())
            log(f"  팝업창 닫힘 대기중... (열린 팝업 {len(cur)}개: {urls})")
        page.wait_for_timeout(500)


def wait_popups_gone(
    page: Page,
    timeout_sec: int = POPUP_WAIT_SEC,
    grace_sec: float = 2.0,
    warn_if_never_opened: bool = False,
) -> bool:
    """팝업이 한 번 열린 뒤 스스로 닫힐 때까지 대기.

    반환값: 팝업이 한 번이라도 열렸으면 True.
    """
    opened = wait_popup_open(page, grace_sec=grace_sec)
    if not opened:
        if warn_if_never_opened:
            log(
                "  [경고] 팝업이 뜨지 않음 — 클릭이 제대로 안 됐거나 "
                "사이트가 응답하지 않았을 수 있음"
            )
        return False
    wait_popups_close(page, timeout_sec=timeout_sec)
    return True


def scroll_to_product_strip(page: Page) -> None:
    """수집 상품 썸네일/모두저장 버튼 영역이 보이도록 스크롤."""
    try:
        btn = save_all_button(page).first
        if btn.count() > 0 and btn.is_visible():
            btn.scroll_into_view_if_needed(timeout=5_000)
            page.wait_for_timeout(250)
    except Exception:  # noqa: BLE001
        pass
    try:
        page.evaluate(
            """() => {
                const re = /검색된\\s*상품\\s*모두\\s*저장|전체선택|선택상품저장/;
                const nodes = Array.from(document.querySelectorAll('input,button,a,td,div,span'));
                const hit = nodes.find(el => re.test((el.value || el.innerText || '').trim()));
                if (hit) hit.scrollIntoView({ block: 'center', inline: 'nearest' });
                const imgs = Array.from(document.querySelectorAll('img')).filter(
                    (i) => (i.naturalWidth || 0) >= 40 && (i.naturalHeight || 0) >= 40
                );
                if (imgs.length) {
                    imgs[Math.min(3, imgs.length - 1)].scrollIntoView({
                        block: 'center', inline: 'nearest'
                    });
                }
            }"""
        )
        page.wait_for_timeout(300)
    except Exception:  # noqa: BLE001
        pass


def wait_product_images(
    page: Page,
    *,
    min_count: int = 2,
    timeout_sec: float = 25.0,
) -> int:
    """상품 이미지(naturalWidth>=40)가 로드될 때까지 스크롤·대기. 로드 개수 반환."""
    end = time.time() + max(3.0, float(timeout_sec))
    best = 0
    while time.time() < end:
        check_stop("상품이미지 대기")
        try:
            n = page.evaluate(
                """() => {
                    let ok = 0;
                    for (const img of Array.from(document.querySelectorAll('img'))) {
                        const w = img.naturalWidth || 0;
                        const h = img.naturalHeight || 0;
                        const r = img.getBoundingClientRect();
                        if (w >= 40 && h >= 40 && r.width >= 20 && r.height >= 20) ok += 1;
                    }
                    return ok;
                }"""
            )
            best = max(best, int(n or 0))
        except Exception:  # noqa: BLE001
            pass
        if best >= min_count:
            scroll_to_product_strip(page)
            return best
        try:
            page.evaluate(
                "() => window.scrollBy(0, Math.max(280, Math.floor(window.innerHeight * 0.55)))"
            )
        except Exception:  # noqa: BLE001
            pass
        page.wait_for_timeout(350)
    scroll_to_product_strip(page)
    return best


def prepare_product_view_for_shot(page: Page, *, min_images: int = 2) -> int:
    """샷 직전에 상품 이미지가 보이도록 대기·스크롤. 로드된 이미지 수 반환."""
    wait_page_not_loading(page, timeout_sec=15.0)
    scroll_to_product_strip(page)
    n = wait_product_images(page, min_count=min_images, timeout_sec=25.0)
    scroll_to_product_strip(page)
    page.wait_for_timeout(400)
    return n


# ── 입력 · 클릭 (망고 구형 input 대응) ────────────────────────

def type_into(page: Page, locator, value: str) -> None:
    el = locator.first
    el.wait_for(state="attached", timeout=60_000)
    el.scroll_into_view_if_needed()
    try:
        el.click(timeout=15_000)
    except PWTimeout:
        pass
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.keyboard.insert_text(value)

    got = ""
    try:
        got = el.input_value()
    except Exception:
        pass
    if not got.strip():
        el.evaluate(
            """(node, v) => {
                if (node instanceof HTMLInputElement || node instanceof HTMLTextAreaElement) {
                    node.focus();
                    node.value = v;
                    node.dispatchEvent(new Event('input', { bubbles: true }));
                    node.dispatchEvent(new Event('change', { bubbles: true }));
                    node.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
                    node.dispatchEvent(new Event('blur', { bubbles: true }));
                }
            }""",
            value,
        )
        got = ""
        try:
            got = el.input_value()
        except Exception:
            pass

    if got.strip() != value.strip():
        log(f"  [경고] 입력값 불일치 — 넣으려던 값: {value!r} / 실제 값: {got!r}")


def click_it(locator) -> bool:
    """
    반환값: 신뢰할 수 있는(trusted) 클릭으로 처리됐으면 True.
    el.evaluate(...node.click()...) 같은 JS 강제클릭은 브라우저가
    "진짜 사용자 클릭"으로 인정하지 않아, 그 안에서 호출되는
    window.open()(팝업)이 조용히 차단될 수 있다. 그래서 실패해도
    좌표 기반 실제 마우스 클릭(신뢰됨)을 먼저 시도하고,
    그것마저 안 될 때만 최후 수단으로 JS 클릭을 쓴다.
    """
    el = locator.first
    el.wait_for(state="visible", timeout=60_000)
    el.scroll_into_view_if_needed()
    try:
        el.click(timeout=20_000)
        return True
    except PWTimeout:
        pass

    try:
        box = el.bounding_box()
        if box:
            el.page.mouse.click(
                box["x"] + box["width"] / 2, box["y"] + box["height"] / 2
            )
            return True
    except Exception:  # noqa: BLE001
        pass

    log("  [경고] 실제 클릭 실패 — JS 강제클릭 사용 (팝업이 안 뜰 수 있음)")
    el.evaluate("(node) => node.click()")
    return False


# ── 화면 요소 ─────────────────────────────────────────────────

def url_search_button(page: Page):
    return page.locator(
        'input[type="button"][value*="URL"], input[type="submit"][value*="URL"]'
    ).or_(page.get_by_text(re.compile(r"URL\s*상품\s*검색")))


def save_all_button(page: Page):
    """[버튼1] 검색결과 상단 — '검색된 상품 모두저장' (모달 하단 저장하기와 별개)."""
    return (
        page.locator('input[type="button"][value*="모두저장"]')
        .or_(page.locator('input[type="submit"][value*="모두저장"]'))
        .or_(page.get_by_text(re.compile(r"검색된\s*상품\s*모두\s*저장")))
    )


# 모달 하단 '저장하기'와 혼동하면 안 되는 결과목록 버튼 문구
_RESULT_LIST_SAVE_BTN = re.compile(
    r"모두\s*저장|선택상품\s*저장|검색된\s*상품"
)


def _describe(loc) -> str:
    try:
        return loc.evaluate(
            "n => `<${n.tagName.toLowerCase()}"
            " name=${n.name||''} id=${n.id||''} rows=${n.rows||''}"
            " value.len=${(n.value||'').length}>`"
        )
    except Exception:
        return "<알 수 없음>"


def url_input(page: Page):
    """페이지 전환 중이면 재시도(_url_input_once 참고)"""
    return with_nav_retry(page, lambda: _url_input_once(page))


def _url_input_once(page: Page):
    """
    URL상품검색하기 버튼과 실제 입력칸이 서로 다른 <tr>/<table>에 있는
    화면이 있어(선택자가 넓으면 엉뚱한 textarea를 골라 "검색결과 없음"이
    나는 원인이 됨) 좁은 범위 -> 넓은 범위 순으로, 후보가 정확히 하나일
    때만 채택한다.
    """
    btn = url_search_button(page).first

    # 1) 버튼과 같은 <tr> 안에서 우선 찾기 (가장 정확)
    row = btn.locator("xpath=ancestor::tr[1]")
    if row.count() > 0:
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = row.locator(sel)
            if cand.count() > 0:
                found = cand.first
                log(f"  URL입력칸(같은 행에서 발견): {_describe(found)}")
                return found

    # 2) 부모를 한 단계씩 올라가며(최대 4단계) 후보가 정확히 하나일 때만 채택
    ancestor = btn
    for _ in range(4):
        ancestor = ancestor.locator("xpath=..")
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = ancestor.locator(sel)
            if cand.count() == 1:
                found = cand.first
                log(f"  URL입력칸(상위 요소에서 발견): {_describe(found)}")
                return found

    # 3) 최후 수단: 페이지 전체에서 rows 속성이 가장 큰 textarea
    #    (URL 여러 줄 입력용 큰 textarea일 가능성이 높음)
    all_ta = page.locator("textarea")
    n = all_ta.count()
    if n == 1:
        found = all_ta.first
        log(f"  URL입력칸(페이지에 textarea 1개뿐): {_describe(found)}")
        return found
    if n > 1:
        best_idx, best_rows = 0, -1
        for i in range(n):
            rows_attr = all_ta.nth(i).get_attribute("rows")
            try:
                rows_val = int(rows_attr) if rows_attr else 1
            except ValueError:
                rows_val = 1
            if rows_val > best_rows:
                best_rows, best_idx = rows_val, i
        found = all_ta.nth(best_idx)
        log(f"  URL입력칸(가장 큰 textarea 선택, {n}개 중): {_describe(found)}")
        return found

    raise RuntimeError("URL 입력칸을 찾지 못했습니다")


def save_modal(page: Page):
    """상품저장설정 팝업 전체 — 하단 '저장하기'·'취소하기' 포함.

    안쪽 작은 div만 잡으면 하단 버튼이 범위 밖이 되어 클릭이 누락된다.
    '검색된 상품 모두저장' 결과목록 영역과 혼동하지 않는다.
    """
    full = (
        page.locator("div, form, table")
        .filter(has_text=re.compile(r"상품\s*저장\s*설정|검색\s*필터\s*명|적용\s*정책"))
        .filter(has_text=re.compile(r"저장하기"))
        .filter(has_text=re.compile(r"취소하기"))
    )
    try:
        if full.count() > 0:
            return full.last
    except Exception:  # noqa: BLE001
        pass
    return (
        page.locator("div, form, table")
        .filter(has_text=re.compile(r"상품\s*저장\s*설정|검색\s*필터\s*명"))
        .filter(has_text=re.compile(r"^[\s\S]*저장하기[\s\S]*$"))
        .filter(has_text=re.compile(r"취소하기"))
        .last
    )


def save_modal_visible(page: Page) -> bool:
    """상품저장설정 모달 열림 — 제목 또는 (저장하기+취소하기) 푸터."""
    try:
        if page.get_by_text(re.compile(r"상품\s*저장\s*설정")).first.is_visible():
            return True
    except Exception:  # noqa: BLE001
        pass
    try:
        # 스크린샷: 하단 파란 '저장하기' + '취소하기' 쌍
        has_save = page.get_by_text(re.compile(r"^저장하기$")).first.is_visible()
        has_cancel = page.get_by_text(re.compile(r"^취소하기$")).first.is_visible()
        if has_save and has_cancel:
            return True
    except Exception:  # noqa: BLE001
        pass
    return False


def save_submit_button(page: Page):
    """하위 호환: 상품저장설정 모달 하단 '저장하기' locator."""
    return resolve_save_submit_control(page)


def _first_visible(locator):
    """locator 후보 중 보이는 '저장하기'만. 모두저장/선택상품저장 제외."""
    try:
        n = locator.count()
    except Exception:  # noqa: BLE001
        return None
    for i in range(n):
        el = locator.nth(i)
        try:
            if not el.is_visible():
                continue
            val = (el.get_attribute("value") or "").strip()
            try:
                txt = re.sub(
                    r"\s+", " ", (el.inner_text(timeout=400) or "")
                ).strip()
            except Exception:  # noqa: BLE001
                txt = ""
            blob = f"{val} {txt}".strip()
            if _RESULT_LIST_SAVE_BTN.search(blob):
                continue
            # 버튼 라벨이 정확히 '저장하기' 인 경우만 (스크린샷 하단 파란 버튼)
            if val == "저장하기" or txt == "저장하기":
                return el
        except Exception:  # noqa: BLE001
            continue
    return None


def scroll_save_modal_to_footer(page: Page) -> None:
    """상품저장설정 모달 하단(저장하기·취소하기)이 보이도록 스크롤."""
    try:
        modal = save_modal(page)
        modal.evaluate(
            """(node) => {
                node.scrollTop = node.scrollHeight;
                let p = node.parentElement;
                for (let i = 0; i < 6 && p; i++) {
                    if (p.scrollHeight > p.clientHeight + 8) {
                        p.scrollTop = p.scrollHeight;
                    }
                    p = p.parentElement;
                }
                window.scrollTo(0, document.body.scrollHeight);
            }"""
        )
    except Exception:  # noqa: BLE001
        pass
    # 취소하기(푸터)가 보이면 그쪽을 기준으로 맞춤
    try:
        page.get_by_text(re.compile(r"^취소하기$")).last.scroll_into_view_if_needed(
            timeout=3_000
        )
    except Exception:  # noqa: BLE001
        pass
    try:
        page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
    except Exception:  # noqa: BLE001
        pass
    page.wait_for_timeout(250)


def dump_save_button_candidates(page: Page, ctx: "RunCtx | None" = None) -> str:
    """디버그: '저장하기' vs '모두저장' 후보 구분 나열."""
    try:
        info = page.evaluate(
            """() => {
                const out = [];
                const nodes = document.querySelectorAll(
                    'input, button, a, span, div, td, li'
                );
                for (const el of nodes) {
                    const val = (el.value || '').trim();
                    const txt = (el.innerText || el.textContent || '')
                        .replace(/\\s+/g, ' ').trim();
                    const blob = val + ' ' + txt;
                    if (!/저장/.test(blob)) continue;
                    const r = el.getBoundingClientRect();
                    if (r.width < 2 || r.height < 2) continue;
                    const kind =
                        (val === '저장하기' || txt === '저장하기')
                            ? 'MODAL_SAVE'
                            : /모두\\s*저장|선택상품/.test(blob)
                              ? 'LIST_SAVE'
                              : 'OTHER';
                    out.push({
                        kind,
                        tag: el.tagName,
                        type: el.getAttribute('type') || '',
                        value: val.slice(0, 40),
                        text: txt.slice(0, 40),
                        y: Math.round(r.top),
                    });
                    if (out.length >= 16) break;
                }
                return out;
            }"""
        )
    except Exception as e:  # noqa: BLE001
        info = [{"error": str(e)}]
    line = f"저장버튼구분={info!r}"
    if ctx is not None:
        ctx.info(f"  [진단] {line}")
    else:
        log(f"  [진단] {line}")
    return line


def _find_footer_save_by_cancel_pair(page: Page):
    """스크린샷 기준: 하단 [저장하기][취소하기] 쌍에서 저장하기만 반환.

    '검색된 상품 모두저장'과는 완전히 다른 버튼이다.
    """
    try:
        handle = page.evaluate_handle(
            """() => {
                const labelOf = (el) => {
                    const v = (el.value || '').trim();
                    const t = (el.innerText || el.textContent || '')
                        .replace(/\\s+/g, ' ').trim();
                    return { v, t, blob: (v + ' ' + t).trim() };
                };
                const visible = (el) => {
                    const r = el.getBoundingClientRect();
                    return r.width > 2 && r.height > 2;
                };
                const nodes = Array.from(document.querySelectorAll(
                    'input, button, a, span, div, td'
                ));
                // 1) 취소하기 기준 — 같은 부모 안의 저장하기 (푸터 쌍)
                const cancels = nodes.filter((el) => {
                    const { v, t } = labelOf(el);
                    return (v === '취소하기' || t === '취소하기') && visible(el);
                });
                for (const cancel of cancels) {
                    let root = cancel.parentElement;
                    for (let depth = 0; depth < 6 && root; depth++) {
                        const kids = Array.from(root.querySelectorAll(
                            'input, button, a, span, div'
                        ));
                        const save = kids.find((el) => {
                            if (el === cancel) return false;
                            const { v, t, blob } = labelOf(el);
                            if (v !== '저장하기' && t !== '저장하기') return false;
                            if (/모두\\s*저장|선택상품\\s*저장/.test(blob)) return false;
                            return visible(el);
                        });
                        if (save) return save;
                        root = root.parentElement;
                    }
                }
                // 2) 전역 exact 저장하기 (모두저장 제외)
                return nodes.find((el) => {
                    const { v, t, blob } = labelOf(el);
                    if (v !== '저장하기' && t !== '저장하기') return false;
                    if (/모두\\s*저장|선택상품\\s*저장/.test(blob)) return false;
                    return visible(el);
                }) || null;
            }"""
        )
        el = handle.as_element()
        if el is not None:
            return el
    except Exception:  # noqa: BLE001
        pass
    return None


def resolve_save_submit_control(page: Page):
    """[버튼2] 상품저장설정 하단 '저장하기' — [버튼1] 모두저장과 절대 혼동 금지.

    실화면: 파란 '저장하기' 옆에 '취소하기'. (밑줄 a/버튼/input)
    """
    scroll_save_modal_to_footer(page)

    # 0) 최우선: 취소하기 옆 푸터 쌍의 저장하기 (스크린샷 그대로)
    el = _find_footer_save_by_cancel_pair(page)
    if el is not None:
        return el

    modal = save_modal(page)

    # 1) 모달 안 — exact '저장하기' only
    modal_candidates = [
        modal.locator(
            'input[type="button"][value="저장하기"], '
            'input[type="submit"][value="저장하기"]'
        ),
        modal.locator("a, button").filter(has_text=re.compile(r"^저장하기$")),
        modal.locator("span, div, td").filter(has_text=re.compile(r"^저장하기$")),
        modal.get_by_text(re.compile(r"^저장하기$")),
    ]
    for loc in modal_candidates:
        found = _first_visible(loc)
        if found is not None:
            return found

    # 2) 페이지 — exact only, 모두저장 제외는 _first_visible에서 처리
    page_candidates = [
        page.locator(
            'input[type="button"][value="저장하기"], '
            'input[type="submit"][value="저장하기"]'
        ),
        page.locator("a, button").filter(has_text=re.compile(r"^저장하기$")),
        page.get_by_text(re.compile(r"^저장하기$")),
    ]
    for loc in page_candidates:
        found = _first_visible(loc)
        if found is not None:
            return found

    dump_save_button_candidates(page)
    raise RuntimeError(
        "상품저장설정 하단 '저장하기' 버튼을 찾지 못함 "
        "('검색된 상품 모두저장'과 다른 버튼). "
        "서버 최종 갱신을 할 수 없음"
    )


def trusted_click_save_submit(page: Page, el) -> bool:
    """하단 저장하기 클릭 — 마우스/Playwright 우선, 실패 시 force·JS.

    저장 실행 팝업이 window.open 이면 신뢰 클릭이 필요하지만,
    클릭 자체가 안 되면 서버 제출이 영원히 누락되므로 최후 JS도 허용.
    """
    try:
        el.scroll_into_view_if_needed(timeout=5_000)
    except Exception:  # noqa: BLE001
        pass
    page.wait_for_timeout(150)

    # 1) 일반 클릭
    try:
        el.click(timeout=15_000)
        return True
    except Exception:  # noqa: BLE001
        pass

    # 2) 좌표 마우스 클릭
    try:
        box = el.bounding_box()
        if box and box.get("width", 0) > 0 and box.get("height", 0) > 0:
            page.mouse.click(
                box["x"] + box["width"] / 2,
                box["y"] + box["height"] / 2,
            )
            return True
    except Exception:  # noqa: BLE001
        pass

    # 3) force 클릭
    try:
        el.click(timeout=8_000, force=True)
        return True
    except Exception:  # noqa: BLE001
        pass

    # 4) 최후 JS (미클릭으로 서버 미반영되는 것보다는 나음)
    try:
        el.evaluate("(node) => node.click()")
        log("  [경고] 저장하기 JS click 사용 — 팝업이 안 뜨면 재시도")
        return True
    except Exception:  # noqa: BLE001
        return False


def save_submit_reacted(
    page: Page,
    dialog_msgs: list[str] | None = None,
    *,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    timeout_sec: float = 15.0,
) -> bool:
    """저장하기 클릭 후 '저장 실행 팝업/알림'이 실제로 떴는지.

    ★ 상품저장설정 모달만 닫힌 것은 성공이 아님.
    ★ 클릭 전 화면의 '00건 수집' 등 잔여 문구는 성공이 아님.
    """
    end = time.time() + max(5.0, float(timeout_sec))
    while time.time() < end:
        has_signal, detail, _ = save_result_signal_present(
            page,
            dialog_msgs,
            before_popup_ids=before_popup_ids,
            baseline=baseline,
            dialog_from=dialog_from,
        )
        if has_signal:
            log(f"  [저장반응] {detail}")
            return True
        page.wait_for_timeout(350)
    return False


def try_dismiss_save_modal(page: Page) -> None:
    """저장 설정 모달이 남아 있으면 닫기/취소/Esc 로 해제 시도."""
    if not save_modal_visible(page):
        return
    try:
        modal = save_modal(page)
        closer = (
            modal.locator('button, a, input[type="button"], input[type="submit"]')
            .filter(has_text=re.compile(r"닫기|취소|취소하기|닫\s*기|×|X"))
            .first
        )
        if closer.count() > 0 and closer.is_visible():
            click_it(closer)
            page.wait_for_timeout(400)
            return
    except Exception:  # noqa: BLE001
        pass
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(300)
    except Exception:  # noqa: BLE001
        pass


def overlay_status(page: Page) -> dict:
    """열린 검색팝업·저장모달·로딩 여부."""
    n_pop = 0
    try:
        n_pop = len(popups(page))
    except Exception:  # noqa: BLE001
        n_pop = 0
    modal = False
    try:
        modal = bool(save_modal_visible(page))
    except Exception:  # noqa: BLE001
        modal = False
    loading = False
    try:
        loading = bool(is_mango_loading(page))
    except Exception:  # noqa: BLE001
        loading = False
    return {"popups": n_pop, "save_modal": modal, "loading": loading}


def finalize_row_overlays(page: Page, ctx: "RunCtx", row: dict) -> Page:
    """행 종료 직후 남은 검색팝업·저장모달을 정리(다음 행 진입 전 보조)."""
    page = refresh_if_closed(page)
    try:
        n = close_search_popups(page)
        if n:
            ctx.info(f"  [행종료정리] 검색 팝업 {n}개 닫음")
        if save_modal_visible(page):
            ctx.info("  [행종료정리] 저장 모달 닫기 시도")
            try_dismiss_save_modal(page)
            page.wait_for_timeout(400)
            if save_modal_visible(page):
                try_dismiss_save_modal(page)
    except Exception as e:  # noqa: BLE001
        ctx.info(f"  [행종료정리] 경고: {e}")
    return refresh_if_closed(page)


def ensure_overlays_closed_before_next(
    page: Page,
    ctx: "RunCtx",
    *,
    next_ordinal: int,
    next_row: dict,
    timeout_sec: float = 180.0,
) -> Page:
    """다음 입력(특히 2번) 상품수집 전 — 모든 모달/팝업이 닫힐 때까지 기다린 뒤 샷 보관.

    닫히지 않은 채 다음 행으로 넘어가지 않는다(반드시 종료 확인 후 진행).
    """
    rn = int(next_row.get("row") or next_ordinal)
    label = str(next_row.get("label") or "").strip()
    ctx.info(
        f"[다음행준비] 입력#{next_ordinal} 수집 전 — 팝업·모달 전부 종료될 때까지 대기 "
        f"(엑셀{rn}행 / {label})"
    )
    page = refresh_if_closed(page)
    try:
        page.bring_to_front()
    except Exception:  # noqa: BLE001
        pass

    end = time.time() + max(30.0, float(timeout_sec))
    last_log = 0.0
    clean_stable = 0
    while True:
        check_stop(f"입력#{next_ordinal} 전 모달 종료 대기")
        st = overlay_status(page)
        if st["popups"] == 0 and not st["save_modal"] and not st["loading"]:
            clean_stable += 1
            if clean_stable >= 3:  # ~1초 이상 안정적으로 닫힌 상태
                break
            page.wait_for_timeout(350)
            continue
        clean_stable = 0
        now = time.time()
        if now - last_log > 5:
            last_log = now
            ctx.info(
                f"  모달 종료 대기중... popups={st['popups']}, "
                f"save_modal={st['save_modal']}, loading={st['loading']}"
            )
        if st["popups"] > 0:
            n = close_search_popups(page)
            if n:
                ctx.info(f"  남은 검색 팝업 {n}개 닫음")
        if st["save_modal"]:
            try_dismiss_save_modal(page)
        if now > end:
            # 제한시간 후에도 강제 닫기 반복 — 그래도 안 닫히면 샷 후 오류
            for _ in range(5):
                close_search_popups(page)
                try_dismiss_save_modal(page)
                page.wait_for_timeout(500)
                st = overlay_status(page)
                if st["popups"] == 0 and not st["save_modal"]:
                    break
            st = overlay_status(page)
            if st["popups"] > 0 or st["save_modal"]:
                ctx.shot(page, "00_overlays_stuck", rn)
                raise RuntimeError(
                    f"입력#{next_ordinal} 수집 전 팝업/모달이 닫히지 않음 "
                    f"(popups={st['popups']}, save_modal={st['save_modal']}) "
                    "— 닫힌 뒤에만 다음 행 진행"
                )
            break
        page.wait_for_timeout(350)

    st = overlay_status(page)
    ctx.info(
        f"  닫힘 확인 완료: 검색팝업={st['popups']}개, "
        f"저장모달={'열림' if st['save_modal'] else '닫힘'}, "
        f"로딩={'중' if st['loading'] else '없음'}"
    )
    ctx.shot(page, "00_overlays_clear", rn)
    ctx.info(
        f"[다음행준비] 입력#{next_ordinal} — 모든 모달 종료 확인·스크린샷 보관 → 수집 시작"
    )
    return page


def modal_field(page: Page, label_pattern: re.Pattern):
    modal = save_modal(page)
    return (
        modal.locator("tr, div, p, label")
        .filter(has_text=label_pattern)
        .locator('input[type="text"], input:not([type]), input[type="number"]')
        .first
    )


# ── 0 ~ 4 ────────────────────────────────────────────────────

def safe_goto(page: Page, url: str, retries: int = 3) -> None:
    """
    로그인 직후 등 사이트 자체가 리다이렉트를 진행 중일 때
    page.goto()와 겹치면 'interrupted by another navigation' 오류가 난다.
    사이트 쪽 리다이렉트가 끝날 때까지 기다렸다가 다시 시도한다.
    """
    last_err: Exception | None = None
    for _ in range(retries):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=120_000)
            return
        except Exception as e:  # noqa: BLE001
            msg = str(e)
            if "interrupted by another navigation" in msg or "NS_BINDING_ABORTED" in msg:
                last_err = e
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=15_000)
                except Exception:  # noqa: BLE001
                    pass
                page.wait_for_timeout(800)
                continue
            raise
    if last_err:
        raise last_err


NAV_ERROR_MARKERS = (
    "Execution context was destroyed",
    "context was destroyed",
    "Target closed",
    "Target page, context or browser has been closed",
)


def is_navigation_error(e: Exception) -> bool:
    msg = str(e)
    return any(m in msg for m in NAV_ERROR_MARKERS)


def with_nav_retry(page: Page, fn, retries: int = 3):
    """
    사이트가 리다이렉트/네비게이션 중일 때 DOM을 조회하면
    "Execution context was destroyed" 같은 오류가 난다.
    페이지가 안정될 때까지 기다렸다가 재시도한다.
    """
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            return fn()
        except Exception as e:  # noqa: BLE001
            if is_navigation_error(e):
                last_err = e
                log(f"  [정보] 페이지 전환 중이라 재시도합니다 ({attempt + 1}/{retries})")
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=10_000)
                except Exception:  # noqa: BLE001
                    pass
                page.wait_for_timeout(800)
                continue
            raise
    if last_err:
        raise last_err
    return None


def _is_logged_in(page: Page) -> bool:
    """로그인 화면을 벗어나고 관리자 호스트에 있으면 로그인된 것으로 본다."""
    try:
        url = page.url or ""
    except Exception:  # noqa: BLE001
        return False
    if not url or url in ("about:blank", "chrome://newtab/"):
        return False
    if ADMIN_HOST not in url:
        return False
    if "admin_login" in url:
        return False
    if "m_login" in url:
        return False
    return True


def wait_for_user_login(page: Page, timeout_sec: int = LOGIN_WAIT_SEC) -> Page:
    """더망고 로그인 창만 띄우고, 사용자가 직접 로그인할 때까지 대기.

    자동 ID/PW 입력·글자 지연 입력·자동 제출은 하지 않는다.
    (Cafe24 보안/자동화 거절 회피 — 사용자 수동 로그인)
    """
    page = refresh_if_closed(page)
    if _is_logged_in(page):
        log("이미 로그인된 세션 — 사용자 로그인 대기 생략")
        return page

    if "admin_login" not in (page.url or ""):
        log("더망고 로그인창 열기: " + LOGIN_URL)
        safe_goto(page, LOGIN_URL)
        page = refresh_if_closed(page)

    try:
        page.bring_to_front()
    except Exception:  # noqa: BLE001
        pass

    safe_print("")
    safe_print("================================================")
    safe_print("  더망고 로그인창에서 직접 로그인하세요.")
    safe_print("  (프로그램이 ID/PW를 입력하지 않습니다)")
    safe_print(f"  로그인 완료 후 자동으로 계속됩니다. (최대 {timeout_sec}초)")
    safe_print("================================================")
    log("사용자 로그인 대기 중...")
    shot_now(page, "login_wait", 0)

    deadline = time.time() + max(30, int(timeout_sec))
    last_url = ""
    while time.time() < deadline:
        check_stop("사용자 로그인 대기")
        page = refresh_if_closed(page)
        try:
            cur = page.url or ""
        except Exception:  # noqa: BLE001
            cur = ""
        if cur != last_url:
            last_url = cur
            log(f"  [대기] URL={cur}")
        if _is_logged_in(page):
            log("사용자 로그인 확인 — 계속 진행")
            shot_now(page, "login_ok", 0)
            return page
        page.wait_for_timeout(1000)

    raise RuntimeError(
        "사용자 로그인 대기 시간 초과.\n"
        "  · Chrome에 열린 더망고 로그인창에서 직접 로그인한 뒤 다시 실행하세요.\n"
        f"  · 마지막 URL={last_url or '(unknown)'}"
    )


def perform_tmg_login(page: Page, user_id: str | None = None, password: str | None = None) -> Page:
    """하위 호환: 자동 입력 없음 — 사용자 수동 로그인 대기."""
    _ = user_id, password
    return wait_for_user_login(page)


def _ask_human(prompt: str) -> None:
    """대화형 터미널에서만 대기. CI/비대화형이면 즉시 실패."""
    if not sys.stdin.isatty():
        raise RuntimeError(
            "로그인이 필요하지만 비대화형 실행입니다. "
            "로컬 PC에서 망고 로그인 후 다시 실행하세요."
        )
    try:
        input(prompt)
    except EOFError as e:
        raise RuntimeError(
            "로그인이 필요하지만 입력을 받을 수 없습니다. "
            "로컬 PC에서 망고 로그인 후 다시 실행하세요."
        ) from e


def handle_possible_login_page(page: Page) -> None:
    """
    세션이 만료돼 admin_login.php로 튕기면 로그인창을 띄우고
    사용자가 직접 로그인할 때까지 기다린다. (자동 입력 없음)
    """
    if "admin_login" not in page.url:
        return
    log("  [경고] 로그인 화면 — 사용자 직접 로그인 대기")
    shot_now(page, "login_required", 0)
    wait_for_user_login(page)
    page = refresh_if_closed(page)
    if "admin_login" in page.url:
        raise RuntimeError("로그인 후에도 여전히 로그인 화면입니다 — 다시 확인해 주세요")


def _wait_bulk_ready_once(page: Page) -> None:
    try:
        page.wait_for_load_state("domcontentloaded", timeout=15_000)
    except Exception:  # noqa: BLE001
        pass
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    handle_possible_login_page(page)
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
        handle_possible_login_page(page)
    log("  대량수집 화면 로딩 확인 중...")
    url_search_button(page).first.wait_for(state="visible", timeout=60_000)


def wait_bulk_ready(page: Page) -> None:
    with_nav_retry(page, lambda: _wait_bulk_ready_once(page))


def _reset_to_bulk_menu_once(page: Page) -> None:
    href = page.locator('a[href*="getGoodsNew"]').first
    if href.count() > 0:
        try:
            href.click(timeout=5000)
        except PWTimeout:
            href.evaluate("(node) => node.click()")
        page.wait_for_timeout(800)
        handle_possible_login_page(page)
        if BULK_PATH in page.url:
            wait_bulk_ready(page)
            return

    page.evaluate(
        """() => {
            const clean = (s) => (s || '').replace(/\\s+/g, '');
            const nodes = Array.from(document.querySelectorAll('a, li, span, td, div, button'));
            const byHref = Array.from(document.querySelectorAll('a[href*="getGoodsNew"]'));
            if (byHref[0]) { byHref[0].click(); return; }
            const top = nodes.find(el => clean(el.textContent) === '상품데이터수집');
            if (top) top.click();
            const sub = nodes.find(el => {
                const t = clean(el.textContent);
                if (t.length > 30) return false;
                return /대량데이터수집|대량수집|상품데이터대량/.test(t);
            });
            if (sub) (sub.closest('a') || sub).click();
        }"""
    )
    page.wait_for_timeout(1000)
    handle_possible_login_page(page)
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    wait_bulk_ready(page)


def reset_to_bulk_menu(page: Page) -> None:
    """0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭"""
    with_nav_retry(page, lambda: _reset_to_bulk_menu_once(page))


def _page_body_text(page: Page) -> str:
    try:
        return page.evaluate(
            "() => (document.body && document.body.innerText) ? document.body.innerText : ''"
        ) or ""
    except Exception:  # noqa: BLE001
        return ""


def is_mango_loading(page: Page) -> bool:
    text = _page_body_text(page)
    return any(p.search(text) for p in MANGO_LOADING_PATTERNS)


def is_mango_no_results(page: Page) -> bool:
    """더망고 자체 '검색결과가 없습니다' 메세지 여부 (로딩 중이 아닐 때만 의미 있음)."""
    if is_mango_loading(page):
        return False
    text = _page_body_text(page)
    return any(p.search(text) for p in MANGO_NO_RESULT_PATTERNS)


def count_mango_result_products(page: Page) -> int:
    """대량수집 화면의 검색결과 상품(체크박스/썸네일) 대략 개수."""
    try:
        return int(
            page.evaluate(
                """() => {
                    // 결과 영역의 체크박스 + 일정 크기 이상 이미지
                    const boxes = Array.from(
                        document.querySelectorAll('input[type="checkbox"]')
                    ).filter((el) => {
                        const r = el.getBoundingClientRect();
                        return r.width > 0 && r.height > 0 && el.offsetParent !== null;
                    });
                    const imgs = Array.from(document.querySelectorAll('img')).filter((img) => {
                        const w = img.naturalWidth || 0;
                        const h = img.naturalHeight || 0;
                        const r = img.getBoundingClientRect();
                        return w >= 40 && h >= 40 && r.width >= 20 && r.height >= 20;
                    });
                    // 전체선택 등 UI 체크박스 1~2개는 제외 감안
                    const boxScore = Math.max(0, boxes.length - 1);
                    return Math.max(boxScore, imgs.length);
                }"""
            )
            or 0
        )
    except Exception:  # noqa: BLE001
        return 0


def wait_page_not_loading(page: Page, timeout_sec: float = 15.0) -> None:
    """
    "로딩 중" 표시가 사라질 때까지 확인한다.
    (검색결과 있음/없음은 여기서 판단하지 않음 — 로딩 중 결과없음 깜빡임 오판 방지)
    """
    end = time.time() + timeout_sec
    while time.time() < end:
        check_stop("로딩 대기")
        try:
            loading = is_mango_loading(page)
        except Exception:  # noqa: BLE001
            return
        if not loading:
            page.wait_for_timeout(500)  # 결과 렌더링 여유
            return
        page.wait_for_timeout(300)


def wait_mango_search_settle(
    page: Page,
    *,
    timeout_sec: float = 45.0,
) -> tuple[str, int]:
    """URL 검색 후 망고 화면이 안정될 때까지 대기.

    반환: (상태, 상품힌트수)
      - "products": 검색 상품이 보임
      - "no_results": 망고 자체 '검색결과가 없습니다' 메세지
      - "unknown": 로딩은 끝났으나 판별 어려움
    """
    end = time.time() + max(10.0, float(timeout_sec))
    # 1) 로딩 종료 대기
    while time.time() < end:
        check_stop("망고 검색 안정화")
        if not is_mango_loading(page):
            break
        page.wait_for_timeout(350)
    else:
        log("  [경고] 망고 로딩 대기 시간 초과")

    # 2) 결과 렌더 안정화 (로딩 직후 결과없음이 잠깐 뜨는 경우 대비)
    stable_need = 3
    stable = 0
    last_state = "unknown"
    last_count = 0
    while time.time() < end:
        check_stop("망고 검색 안정화")
        if is_mango_loading(page):
            stable = 0
            page.wait_for_timeout(300)
            continue
        no_res = is_mango_no_results(page)
        cnt = count_mango_result_products(page)
        if cnt >= 1 and not no_res:
            state = "products"
        elif no_res and cnt < 1:
            state = "no_results"
        elif cnt >= 1:
            state = "products"
        else:
            state = "unknown"
        if state == last_state and state != "unknown":
            stable += 1
        else:
            stable = 1 if state != "unknown" else 0
        last_state, last_count = state, cnt
        if stable >= stable_need:
            return state, last_count
        page.wait_for_timeout(400)

    return last_state, last_count


def _process_row_once(page: Page, row: dict, ctx: RunCtx) -> None:
    label = row["label"]
    url = normalize_url(row["url"])
    save_count = ctx.save_count
    rn = row["row"]
    ctx.info(
        f"처리 시작 | 상위 최종 카테고리명={label} | "
        f"최종 카테고리 URL주소={row['url']} | 목표 저장수={save_count}"
    )

    ctx.info("0. 초기화 : 상품데이터수집 -> 대량데이터수집")
    reset_to_bulk_menu(page)
    page.wait_for_timeout(500)
    ctx.shot(page, "00_init_bulk", rn)

    ctx.info(f"  엑셀 원본 URL: {row['url']}")
    if url != row["url"].strip():
        ctx.info(f"  [정보] 프로토콜 보정됨: {url}")

    # URL 검색: 망고 자체 '검색결과 없음'이면 소횟수 재시도 후 다음 행으로
    search_ok = False
    last_state = "unknown"
    last_count = 0
    popup_imgs = 0
    max_search = max(1, int(SEARCH_MAX_TRIES))
    for search_try in range(1, max_search + 1):
        ctx.check_budget(f"URL검색 {search_try}/{max_search}")
        ctx.info(
            f"1. URL 검색 시도 {search_try}/{max_search} | "
            f"상위 최종 카테고리명={label} | 최종 카테고리 URL주소={url}"
        )
        target = url_input(page)
        type_into(page, target, url)
        actual = ""
        try:
            actual = target.input_value()
        except Exception:
            pass
        ctx.info(f"  입력칸 최종 값: {actual!r}")
        if actual.strip() != url.strip():
            raise RuntimeError(f"URL 입력 불일치 — 기대 {url!r} / 실제 {actual!r}")
        if search_try == 1:
            ctx.shot(page, "01_url_filled", rn)

        ctx.info("1. URL상품검색하기 클릭")
        trusted = click_it(url_search_button(page))

        # 검색 팝업 "열림" 확인·샷 → 그 다음 "닫힘"
        ctx.info("1. 검색 팝업 열림 대기")
        opened_pages = wait_popup_open(page, grace_sec=15.0)
        if not opened_pages:
            ctx.info("  키보드로 재시도 (Enter)")
            try:
                btn = url_search_button(page).first
                btn.focus()
                page.keyboard.press("Enter")
            except Exception:  # noqa: BLE001
                pass
            opened_pages = wait_popup_open(page, grace_sec=10.0)

        if not opened_pages:
            ctx.shot(page, "01_popup_missing", rn)
            raise RuntimeError(
                f"#{rn} URL상품검색하기 클릭 후 팝업이 뜨지 않음 "
                f"(trusted_click={trusted})"
            )

        popup = opened_pages[0]
        try:
            popup.bring_to_front()
        except Exception:  # noqa: BLE001
            pass
        try:
            popup_imgs = prepare_product_view_for_shot(popup, min_images=2)
        except Exception as e:  # noqa: BLE001
            ctx.info(f"  [경고] 팝업 상품이미지 대기 실패: {e}")
            popup_imgs = 0
        ctx.info(f"1. 검색 팝업 열림 (상품이미지 약 {popup_imgs}개)")
        if search_try == 1:
            ctx.shot(popup, "01_popup_opened", rn)

        ctx.info("1. 검색 팝업 닫힘 대기")
        wait_popups_close(page)
        try:
            page.bring_to_front()
        except Exception:  # noqa: BLE001
            pass
        ctx.info("1. 검색 팝업 닫힘")
        if search_try == 1:
            ctx.shot(page, "01_popup_closed", rn)

        # 망고 로딩(빨간 잠시만 기다려주세요) 종료 후, 자체 무결과 메세지 판별
        ctx.info("1. 망고 검색결과 안정화 대기 (로딩 종료 후 판별)")
        last_state, last_count = wait_mango_search_settle(page, timeout_sec=45.0)
        if last_state == "no_results":
            ctx.info(
                "  [망고 자체메세지] 검색하신 검색에 대한 검색결과가 없습니다. "
                f"(팝업상품이미지약 {popup_imgs}개 / 망고결과힌트 {last_count})"
            )
            ctx.shot(page, "01_mango_no_results", rn)
            if search_try < max_search:
                ctx.info("  망고 무결과 — URL 검색 재시도")
                page.wait_for_timeout(800)
                continue
            raise RuntimeError(
                f"#{rn} 더망고 자체 메세지: 검색결과가 없습니다.\n"
                f"  · 상위 최종 카테고리명={label}\n"
                f"  · 최종 카테고리 URL주소={url}\n"
                f"  · ABC검색팝업 상품이미지 약 {popup_imgs}개였으나 "
                f"망고 수집결과에 상품이 없음\n"
                "  · URL/카테고리 접근·망고 세션을 확인하세요."
            )

        if last_state == "products" or last_count >= 1:
            search_ok = True
            break

        # unknown: 이미지 로드를 한번 더 기다려 보고 판단
        result_imgs = prepare_product_view_for_shot(page, min_images=2)
        if result_imgs >= 1 and not is_mango_no_results(page):
            last_state, last_count = "products", result_imgs
            search_ok = True
            break
        if is_mango_no_results(page):
            ctx.info("  [망고 자체메세지] 검색결과가 없습니다. (재확인)")
            ctx.shot(page, "01_mango_no_results", rn)
            if search_try < max_search:
                continue
            raise RuntimeError(
                f"#{rn} 더망고 자체 메세지: 검색결과가 없습니다.\n"
                f"  · 상위 최종 카테고리명={label}\n"
                f"  · 최종 카테고리 URL주소={url}"
            )
        ctx.info(
            f"  [경고] 검색결과 판별 불명 (state={last_state}, hint={last_count}, "
            f"imgs={result_imgs}) — 재시도"
        )
        if search_try < max_search:
            page.wait_for_timeout(800)
            continue

    if not search_ok and last_state != "products" and last_count < 1:
        raise RuntimeError(
            f"#{rn} 망고 검색결과 확인 실패 (state={last_state}, hint={last_count})"
        )

    # 08: 검색 결과 준비 — 하단 수집 상품 이미지가 보이도록 대기 후 샷
    result_imgs = prepare_product_view_for_shot(page, min_images=2)
    ctx.info(
        f"1. 검색 결과 준비 (하단 상품이미지 약 {result_imgs}개 / "
        f"망고상태={last_state}, hint={last_count})"
    )
    if is_mango_no_results(page):
        ctx.shot(page, "01_mango_no_results", rn)
        raise RuntimeError(
            f"#{rn} 더망고 자체 메세지: 검색결과가 없습니다.\n"
            f"  · 상위 최종 카테고리명={label}\n"
            f"  · 최종 카테고리 URL주소={url}"
        )
    if result_imgs < 1:
        ctx.info("  [경고] 검색 결과 상품이미지가 거의 보이지 않음 — 그대로 샷")
    ctx.shot(page, "01_results_ready", rn)

    # ── [버튼1] 검색결과 상단: 검색된 상품 모두저장 (모달 열기) ──
    ctx.info(
        "2-A. ★ [버튼1] '검색된 상품 모두저장' 클릭 "
        "(결과목록 상단 — 모달 하단 '저장하기'와 다른 버튼)"
    )
    scroll_to_product_strip(page)
    click_it(save_all_button(page))
    # 모달 열림: 제목 또는 하단 저장하기+취소하기
    end_modal = time.time() + MODAL_WAIT_SEC
    while time.time() < end_modal:
        if save_modal_visible(page):
            break
        page.wait_for_timeout(300)
    else:
        ctx.shot(page, "02_save_missing", rn)
        raise RuntimeError(
            f"#{rn} [버튼1] 모두저장 클릭 후 상품저장설정 모달이 열리지 않음"
        )
    try:
        modal_imgs = prepare_product_view_for_shot(page, min_images=2)
    except Exception as e:  # noqa: BLE001
        ctx.info(f"  [경고] 모달 화면 상품이미지 대기 실패: {e}")
        modal_imgs = 0
    page.wait_for_timeout(300)
    ctx.info(f"2-A. 상품저장설정 모달 열림 (상품이미지 약 {modal_imgs}개)")
    ctx.shot(page, "02_save_modal", rn)

    # 저장 단계 시간 확보 (검색 대기 후 예산이 거의 소진된 경우 대비)
    ctx.row_deadline = time.time() + max(120.0, float(MODAL_WAIT_SEC) + 60.0)

    # 검색필터명 → 저장상품수 → [버튼2] 하단 저장하기 (서버 최종 갱신)
    fill_save_modal_fields(page, ctx, rn, label, save_count)
    ctx.info(
        "2-B. ★ [버튼2] 모달 하단 '저장하기' 클릭 시작 "
        "([버튼1] 모두저장과 구분 — 서버 최종 갱신)"
    )
    run_save_submit_and_verify(page, ctx, rn, save_count)
    if not ctx.server_save_ok:
        raise RuntimeError(
            f"#{rn} [버튼2] 하단 저장하기 서버 최종 갱신 미확인 — "
            "필터정보·수집갯수·수집상품이 서버에 반영되지 않음"
        )


def fill_save_modal_fields(
    page: Page,
    ctx: RunCtx,
    rn: int,
    label: str,
    save_count: int,
) -> None:
    """상품저장설정 모달에 검색필터명·저장상품수를 입력하고 샷 보관."""
    ctx.info(f"2. 검색필터명 입력: {label}")
    type_into(page, modal_field(page, FILTER_NAME_LABEL), label)

    count_field = modal_field(page, SAVE_COUNT_LABEL)
    if count_field.count() == 0:
        ctx.shot(page, "02_no_count_field", rn)
        raise RuntimeError(f"#{rn} 저장상품수 입력칸을 찾지 못함")
    ctx.info(f"2. 저장상품수 입력: {save_count}")
    type_into(page, count_field, str(save_count))
    # 필터명이 지워지는 UI 대비 — 저장 직전 한 번 더
    type_into(page, modal_field(page, FILTER_NAME_LABEL), label)

    got_count = ""
    try:
        got_count = count_field.input_value()
    except Exception:
        pass
    ctx.info(f"  저장상품수 입력값: {got_count!r} (목표 {save_count})")
    if str(save_count) not in (got_count or "").replace(" ", ""):
        digits = re.sub(r"\D", "", got_count or "")
        if digits != str(save_count):
            ctx.shot(page, "02_count_mismatch", rn)
            raise RuntimeError(
                f"#{rn} 저장상품수 불일치 — 기대 {save_count} / 실제 {got_count!r}"
            )
    ctx.info("2. 검색필터명·저장상품수 입력 완료 → 저장하기 진행")
    ctx.shot(page, "02_modal_filled", rn)


def run_save_submit_and_verify(
    page: Page,
    ctx: RunCtx,
    rn: int,
    save_count: int,
) -> None:
    """★저장하기 = 서버 최종 갱신 (필터정보·수집갯수·수집상품).

    입력만 하고 이 버튼을 빼먹으면 서버에 아무것도 반영되지 않는다.
    클릭 → 서버 제출 반응 확인(최대 1회 재시도) → 결과 팝업/알림 → 수집건수 검증.
    재시도 후에도 동일 실패면 다음 행으로 진행한다.
    """
    ctx.server_save_ok = False
    ctx.save_popup_seen = False
    dialog_msgs: list[str] = []

    def _on_save_dialog(dialog) -> None:
        try:
            msg = str(dialog.message or "")
            dialog_msgs.append(msg)
            ctx.info(f"  [dialog] {msg!r}")
            dialog.accept()
        except Exception:  # noqa: BLE001
            try:
                dialog.dismiss()
            except Exception:  # noqa: BLE001
                pass

    page.on("dialog", _on_save_dialog)
    try:
        ctx.check_budget("저장하기(서버 최종 갱신) 전")
        if not save_modal_visible(page):
            ctx.shot(page, "02_save_missing", rn)
            raise RuntimeError(
                f"#{rn} 저장하기 서버 최종 갱신 실패 — "
                "상품저장설정 모달이 열려 있지 않음 "
                "(필터정보·수집갯수·수집상품이 서버에 반영되지 않음)"
            )

        ctx.info(
            "2-B. ★★★ [버튼2] 하단 파란 '저장하기' "
            "(옆=취소하기) = 서버 최종 갱신 — "
            "[버튼1] '검색된 상품 모두저장' 과 다름"
        )
        scroll_save_modal_to_footer(page)
        dump_save_button_candidates(page, ctx)
        ctx.shot(page, "02_modal_filled", rn)

        # 클릭 전 잔여 '00건 수집' 등 — 저장 팝업으로 오인 금지
        alert_baseline = collect_alert_baseline(page)
        dialog_from = len(dialog_msgs)
        if alert_baseline:
            ctx.info(
                "2-B. [주의] 저장하기 전 화면에 수집문구 잔여 "
                f"{sorted(alert_baseline)} — 저장 팝업으로 쓰지 않음"
            )

        before_popup_ids = {_popup_id(p) for p in popups(page)}
        clicked_ok = False
        # 최대 1회 재시도(총 2회). 동일 실패면 다음 행으로 진행.
        for attempt in range(1, 3):
            ctx.check_budget(f"[버튼2] 저장하기 서버제출 시도 {attempt}")
            if not save_modal_visible(page) and attempt > 1:
                ctx.info(
                    "  [경고] 재시도 시 상품저장설정 모달 없음 — "
                    "팝업 대기로 넘어감"
                )
                break
            scroll_save_modal_to_footer(page)
            try:
                btn = resolve_save_submit_control(page)
            except Exception as e:  # noqa: BLE001
                dump_save_button_candidates(page, ctx)
                ctx.shot(page, "02_save_missing", rn)
                raise RuntimeError(
                    f"#{rn} [버튼2] 하단 '저장하기' 없음 "
                    f"(모두저장과 다른 버튼). 다음 입력으로. 원인: {e}"
                ) from e

            try:
                desc = _describe(btn)
            except Exception:  # noqa: BLE001
                desc = "저장하기"
            ctx.info(
                f"2-B. ★ [버튼2] 하단 '저장하기' 클릭 "
                f"({attempt}/2, 재시도 최대 1회) | {desc}"
            )
            # 매 클릭 직전 기준 갱신
            alert_baseline = collect_alert_baseline(page)
            dialog_from = len(dialog_msgs)
            before_popup_ids = {_popup_id(p) for p in popups(page)}
            last_click_ok = trusted_click_save_submit(page, btn)
            ctx.info(
                f"  하단 저장하기 클릭 전송 "
                f"(ok={last_click_ok}, attempt={attempt}) "
                f"— 이후 최소 {SAVE_POPUP_GRACE_SEC:.0f}초 저장시간"
            )
            page.wait_for_timeout(500)
            ctx.shot(page, "02_save_clicked", rn)

            if not last_click_ok:
                ctx.info("  [경고] 하단 저장하기 클릭 실패 — 재시도")
                dump_save_button_candidates(page, ctx)
                ctx.shot(page, "02_save_no_react", rn)
                page.wait_for_timeout(500)
                continue

            clicked_ok = True
            # 클릭 직후: 잔여문구 무시 + 유예시간 포함 반응 확인
            if save_submit_reacted(
                page,
                dialog_msgs,
                before_popup_ids=before_popup_ids,
                baseline=alert_baseline,
                dialog_from=dialog_from,
                timeout_sec=max(8.0, SAVE_POPUP_GRACE_SEC + 2.0),
            ):
                ctx.info(
                    "2-B. ★ 저장하기 클릭 → 신규 저장 실행 팝업/알림 감지"
                )
                break

            ctx.info(
                "  [경고] 클릭 직후 신규 팝업 미감지 — "
                "잔여 '00건'/설정모달 닫힘만으로는 부족"
            )
            ctx.shot(page, "02_save_no_react", rn)
            if save_modal_visible(page) and attempt < 2:
                page.wait_for_timeout(600)
                continue
            break

        if not clicked_ok:
            dump_save_button_candidates(page, ctx)
            ctx.shot(page, "02_save_failed", rn)
            raise RuntimeError(
                f"#{rn} 하단 '저장하기' 클릭 실패 — "
                "팝업창 모달 없이 초기화할 수 없습니다. "
                "1회 재시도 후에도 동일. 다음 입력으로 진행."
            )

        # ★필수: 저장 시간 + 팝업창 모달 대기 — 없으면 초기화 금지
        ctx.info(
            "3. ★★★ 저장하기 후 저장시간·팝업창 모달을 반드시 기다림 "
            "(잔여 00건·모달닫힘으로 초기화 금지)"
        )
        wait_save_overlays_settle(
            page,
            ctx,
            rn,
            dialog_msgs=dialog_msgs,
            before_popup_ids=before_popup_ids,
            baseline=alert_baseline,
            dialog_from=dialog_from,
        )

        if not getattr(ctx, "save_popup_seen", False):
            ctx.shot(page, "03_result_missing", rn)
            raise RuntimeError(
                f"#{rn} 저장하기 후 팝업창 모달을 확인하지 못함 — "
                "팝업 없이 초기화할 수 없습니다."
            )

        # 최종: 망고 자체 메세지 — 저장된 상품 건수 = 서버 반영 확인
        verify_mango_collect_alert(
            page,
            ctx,
            rn,
            save_count,
            dialog_msgs=dialog_msgs,
            baseline=alert_baseline,
            dialog_from=dialog_from,
        )

        verify_row_save_done(page, ctx, rn, save_count)
        ctx.server_save_ok = True
        ctx.info(
            f"4. -> 서버 최종 갱신 완료 "
            f"(저장하기 OK + 팝업 확인 OK / 저장수 {save_count})"
        )
        ctx.shot(page, "04_row_done", rn)
    finally:
        try:
            page.remove_listener("dialog", _on_save_dialog)
        except Exception:  # noqa: BLE001
            pass


def _page_visible_text(page: Page) -> str:
    try:
        return (
            page.evaluate(
                "() => (document.body && document.body.innerText) "
                "? document.body.innerText : ''"
            )
            or ""
        )
    except Exception:  # noqa: BLE001
        return ""


def parse_mango_collect_count(text: str) -> tuple[int | None, str]:
    """망고 알림 문구에서 수집 건수 추출. (건수 또는 None, 매칭 원문)."""
    raw = text or ""
    for pat in MANGO_COLLECT_ALERT_PATTERNS:
        m = pat.search(raw)
        if not m:
            continue
        try:
            n = int(m.group(1))
        except (TypeError, ValueError):
            continue
        return n, m.group(0)
    return None, ""


def collect_alert_baseline(page: Page) -> set[str]:
    """저장하기 클릭 전 — 이미 화면에 있던 수집건수 문구 지문.

    검색 단계의 '00건이 수집되었다' 등이 저장 팝업으로 오인되는 것을 막는다.
    """
    fps: set[str] = set()
    try:
        text = _page_visible_text(page)
        n, hit = parse_mango_collect_count(text)
        if n is not None:
            fps.add(f"page:{n}:{hit}")
    except Exception:  # noqa: BLE001
        pass
    for p in popups(page):
        try:
            ptext = _page_visible_text(p)
            n, hit = parse_mango_collect_count(ptext)
            if n is not None:
                fps.add(f"popup:{n}:{hit}")
        except Exception:  # noqa: BLE001
            continue
    return fps


def find_mango_collect_alert(
    page: Page,
    dialog_msgs: list[str] | None = None,
    *,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
) -> tuple[int | None, str, str, Page | None]:
    """메인·팝업·JS dialog에서 수집건수 알림을 찾는다.

    baseline 이 있으면 클릭 전에 있던 문구(예: 00건 수집)는 무시한다.
    반환: (건수|None, 매칭문구, source, 해당 Page|None)
    """
    base = baseline or set()

    # 클릭 이후 새로 도착한 dialog 만
    for msg in list(dialog_msgs or [])[max(0, int(dialog_from)) :]:
        n, hit = parse_mango_collect_count(msg)
        if n is not None:
            fp = f"dialog:{n}:{hit or msg}"
            if fp not in base:
                return n, hit or msg, "dialog", page

    text = _page_visible_text(page)
    n, hit = parse_mango_collect_count(text)
    if n is not None:
        fp = f"page:{n}:{hit}"
        if fp not in base:
            return n, hit, "page", page

    for p in popups(page):
        ptext = _page_visible_text(p)
        n, hit = parse_mango_collect_count(ptext)
        if n is not None:
            fp = f"popup:{n}:{hit}"
            if fp not in base:
                return n, hit, "popup", p

    return None, "", "", None


def save_execution_layer_visible(
    page: Page,
    *,
    baseline: set[str] | None = None,
) -> bool:
    """저장하기 직후 뜨는 '저장 실행' 팝업/레이어.

    클릭 전 문구·상품저장설정 본문·'확인' 단독은 인정하지 않음.
    """
    base = baseline or set()
    try:
        n, hit, src, _ = find_mango_collect_alert(page, None, baseline=base)
        if n is not None and src:
            return True
    except Exception:  # noqa: BLE001
        pass
    try:
        layer = page.locator(
            '[role="dialog"], .ui-dialog, .modal, '
            'div[id*="layer"], div[id*="popup"], '
            'div[class*="layer"], div[class*="popup"], '
            'div[class*="alert"], div[class*="msg"]'
        ).filter(
            has_text=re.compile(
                r"(\d+)\s*건\s*(이\s*)?(수집|저장)\s*되었|"
                r"수집\s*완료|저장\s*완료|처리\s*완료|"
                r"상품\s*(이\s*)?(수집|저장)\s*되었"
            )
        )
        if layer.count() > 0 and layer.first.is_visible():
            txt = ""
            try:
                txt = layer.first.inner_text(timeout=500) or ""
            except Exception:  # noqa: BLE001
                txt = ""
            if re.search(r"상품\s*저장\s*설정", txt) and re.search(
                r"검색\s*필터\s*명|취소하기", txt
            ):
                return False
            n, hit = parse_mango_collect_count(txt)
            if n is not None and f"page:{n}:{hit}" in base:
                return False  # 검색 단계 잔여 '00건 수집' 등
            if n is not None or re.search(
                r"수집\s*완료|저장\s*완료|처리\s*완료", txt or ""
            ):
                return True
    except Exception:  # noqa: BLE001
        pass
    return False


def _popup_id(p) -> int:
    try:
        return id(p)
    except Exception:  # noqa: BLE001
        return 0


def save_result_signal_present(
    page: Page,
    dialog_msgs: list[str] | None = None,
    *,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
) -> tuple[bool, str, Page | None]:
    """저장하기 후 '저장 실행' 팝업/알림 모달 출현 여부.

    - 클릭 이후 새 브라우저 팝업
    - 클릭 전에 없던 수집건수 문구/레이어 (검색단계 '00건' 오인 금지)
    - 새 JS dialog
    ★ 상품저장설정 모달 닫힘만으로는 True 가 되지 않음
    """
    base = baseline or set()

    n, msg, src, hit_page = find_mango_collect_alert(
        page,
        dialog_msgs,
        baseline=base,
        dialog_from=dialog_from,
    )
    if n is not None:
        return True, f"{src}:{msg}", hit_page or page

    prev = before_popup_ids or set()
    for p in popups(page):
        if _popup_id(p) in prev:
            continue
        ptext = _page_visible_text(p)
        n2, hit2 = parse_mango_collect_count(ptext)
        if n2 is not None:
            fp = f"popup:{n2}:{hit2}"
            if fp in base:
                continue
            return True, f"new_popup:{hit2}", p
        return True, "new_popup_window", p

    if save_execution_layer_visible(page, baseline=base):
        return True, "inpage_save_layer", page

    return False, "", None


def wait_save_execution_popup(
    page: Page,
    ctx: RunCtx,
    rn: int,
    *,
    dialog_msgs: list[str] | None = None,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    timeout_sec: float | None = None,
    grace_sec: float | None = None,
) -> None:
    """★저장하기 클릭 후 — 저장 실행 팝업창 모달이 뜰 때까지 필수 대기.

    - 최소 grace_sec 동안은 서버 저장 시간을 줌 (즉시 초기화 금지)
    - 클릭 전 '00건 수집' 잔여 문구는 팝업으로 보지 않음
    - 팝업이 안 보이면 초기화/다음 단계로 절대 넘어가지 않음
    """
    wait_sec = float(timeout_sec or MODAL_WAIT_SEC)
    grace = max(2.0, float(grace_sec if grace_sec is not None else SAVE_POPUP_GRACE_SEC))
    base = baseline if baseline is not None else set()
    ctx.info(
        "3. ★★★ 저장하기 후 '팝업창 모달' 대기 중 "
        f"(최소 {grace:.0f}초 저장시간 + 최대 {int(wait_sec)}초) "
        "— 잔여 '00건' 문구로 통과·초기화 금지"
    )
    if base:
        ctx.info(f"  [기준] 클릭 전 무시할 수집문구={sorted(base)}")

    end = time.time() + wait_sec
    grace_until = time.time() + grace
    saw = False
    detail = ""
    target: Page | None = None
    logged_settings_closed = False
    logged_stale = False

    while time.time() < end:
        ctx.check_budget("저장하기 후 팝업창 모달 대기")
        if not save_modal_visible(page) and not logged_settings_closed:
            logged_settings_closed = True
            ctx.info(
                "  [대기] 상품저장설정 모달 닫힘 — "
                f"저장 처리·팝업을 최소 {grace:.0f}초 기다림"
            )

        # 유예 시간 중에도 '새 창 팝업'은 즉시 인정. 페이지 잔여문구는 무시.
        has_signal, detail, target = save_result_signal_present(
            page,
            dialog_msgs,
            before_popup_ids=before_popup_ids,
            baseline=base,
            dialog_from=dialog_from,
        )
        if has_signal:
            # 유예 전이라도 진짜 새 신호면 OK
            saw = True
            break

        # 잔여 00건이 화면에 있으면 안내 (오인 방지)
        if not logged_stale and base:
            stale_n, stale_hit = parse_mango_collect_count(_page_visible_text(page))
            if stale_n is not None and f"page:{stale_n}:{stale_hit}" in base:
                logged_stale = True
                ctx.info(
                    f"  [무시] 검색단계 잔여 문구 {stale_hit!r} "
                    "— 저장 팝업으로 보지 않음, 계속 대기"
                )

        if time.time() < grace_until:
            page.wait_for_timeout(350)
            continue

        page.wait_for_timeout(400)

    if not saw:
        ctx.shot(page, "03_result_missing", rn)
        raise TimeoutError(
            f"#{rn} 저장하기 후 팝업창 모달이 나타나지 않음. "
            "검색단계 '00건 수집' 잔여문구·설정모달 닫힘만으로 "
            "초기화/다음 행 진행 불가. "
            "저장하기 → (저장시간) → 팝업창 모달 열림 → 닫힘 → 건수확인 필수."
        )

    ctx.info(f"3. ★ 저장 실행 팝업창 모달 확인 — {detail}")
    shot_page = target or page
    try:
        shot_page.bring_to_front()
    except Exception:  # noqa: BLE001
        pass
    ctx.shot(shot_page, "03_result_popup", rn)
    ctx.save_popup_seen = True


def wait_save_overlays_settle(
    page: Page,
    ctx: RunCtx,
    rn: int,
    *,
    dialog_msgs: list[str] | None = None,
    before_popup_ids: set[int] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
) -> None:
    """저장 실행 팝업 출현(필수) 후 닫힘·정리 대기.

    팝업을 보기 전에는 반환하지 않는다. 모달 닫힘 ≠ 완료.
    """
    base = baseline if baseline is not None else set()
    wait_save_execution_popup(
        page,
        ctx,
        rn,
        dialog_msgs=dialog_msgs,
        before_popup_ids=before_popup_ids,
        baseline=base,
        dialog_from=dialog_from,
    )

    end = time.time() + min(30.0, float(MODAL_WAIT_SEC))
    while time.time() < end:
        ctx.check_budget("저장 팝업 정리 대기")
        alert_n, _, _, _ = find_mango_collect_alert(
            page, dialog_msgs, baseline=base, dialog_from=dialog_from
        )
        if alert_n is not None:
            ctx.info(
                f"3. 팝업/알림에 수집건수 문구 확인됨 → 검증 단계로 ({alert_n}건)"
            )
            return
        cur_new = [
            p
            for p in popups(page)
            if _popup_id(p) not in (before_popup_ids or set())
        ]
        if not cur_new and not save_modal_visible(page):
            page.wait_for_timeout(600)
            alert_n, _, _, _ = find_mango_collect_alert(
                page, dialog_msgs, baseline=base, dialog_from=dialog_from
            )
            if alert_n is not None:
                return
            if getattr(ctx, "save_popup_seen", False):
                ctx.info("3. 저장 팝업 닫힘 — 수집건수 알림 검증으로 진행")
                return
        page.wait_for_timeout(400)

    if getattr(ctx, "save_popup_seen", False):
        ctx.info("3. 팝업은 확인됨 — 수집건수 알림 검증으로 진행")
        return
    ctx.shot(page, "03_result_missing", rn)
    raise TimeoutError(
        f"#{rn} 저장하기 후 팝업창 모달/알림 확인 실패 — "
        "초기화로 진행할 수 없습니다."
    )


def dismiss_mango_alert_ui(page: Page) -> None:
    """화면 알림/레이어의 확인 버튼이 있으면 닫기."""
    try:
        btn = (
            page.locator("button, a, input[type='button'], input[type='submit']")
            .filter(has_text=re.compile(r"^(확인|닫기|OK|Yes)$", re.I))
            .first
        )
        if btn.count() > 0 and btn.is_visible():
            click_it(btn)
            page.wait_for_timeout(300)
    except Exception:  # noqa: BLE001
        pass
    # 팝업 창에도 확인 버튼이 있을 수 있음
    for p in list(popups(page)):
        try:
            btn = (
                p.locator("button, a, input[type='button'], input[type='submit']")
                .filter(has_text=re.compile(r"^(확인|닫기|OK|Yes)$", re.I))
                .first
            )
            if btn.count() > 0 and btn.is_visible():
                click_it(btn)
                p.wait_for_timeout(300)
        except Exception:  # noqa: BLE001
            pass
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(200)
    except Exception:  # noqa: BLE001
        pass


def verify_mango_collect_alert(
    page: Page,
    ctx: RunCtx,
    row_no: int,
    expect_count: int,
    *,
    dialog_msgs: list[str] | None = None,
    baseline: set[str] | None = None,
    dialog_from: int = 0,
    timeout_sec: float = 25.0,
) -> int:
    """결과 팝업/알림에서 망고 자체 'N건이 수집되었다'를 파악·검증.

    - 저장하기 클릭 이후 새로 뜬 알림만 사용 (검색단계 00건 잔여 무시)
    - 기대 건수(저장수)와 일치해야 통과
    - 0건이면 실패(다음 행으로 진행 가능하도록 확정 실패 문구 포함)
    """
    expect = max(0, int(expect_count))
    base = baseline or set()
    ctx.info(
        f"3. 망고 자체 알림 확인 — 'N건이 수집되었다' (기대 {expect}건) "
        f"| 잔여무시={len(base)}건"
    )
    end = time.time() + max(5.0, float(timeout_sec))
    found_n: int | None = None
    found_msg = ""
    source = ""
    shot_page: Page | None = page

    while time.time() < end:
        ctx.check_budget("망고 수집알림 대기")
        found_n, found_msg, source, hit_page = find_mango_collect_alert(
            page,
            dialog_msgs,
            baseline=base,
            dialog_from=dialog_from,
        )
        if found_n is not None:
            shot_page = hit_page or page
            break

        texts = [_page_visible_text(page)]
        for p in popups(page):
            texts.append(_page_visible_text(p))
        blob = "\n".join(texts)
        for pat in SAVE_FAIL_PATTERNS:
            if pat.search(blob or ""):
                ctx.shot(page, "03_collect_alert_fail", row_no)
                raise RuntimeError(
                    f"#{row_no} 망고 수집 알림 대신 오류 문구 감지: {pat.pattern}"
                )
        page.wait_for_timeout(400)

    if found_n is None:
        found_n, found_msg, source, hit_page = find_mango_collect_alert(
            page,
            dialog_msgs,
            baseline=base,
            dialog_from=dialog_from,
        )
        shot_page = hit_page or page
        if found_n is not None and not source:
            source = "final"

    if found_n is None:
        ctx.shot(page, "03_collect_alert_fail", row_no)
        raise RuntimeError(
            f"#{row_no} 망고 수집 알림을 찾지 못함 "
            f"(기대 문구 예: '{expect}건이 수집되었다'). "
            "저장하기 후 결과 팝업/알림 메세지를 확인하세요."
        )

    ctx.info(
        f"  [망고알림] source={source} | 문구={found_msg!r} | "
        f"수집건수={found_n} | 기대={expect}"
    )
    try:
        ctx.shot(shot_page or page, "03_collect_alert", row_no)
    except Exception:  # noqa: BLE001
        ctx.shot(page, "03_collect_alert", row_no)
    dismiss_mango_alert_ui(page)

    if found_n == 0:
        raise RuntimeError(
            f"#{row_no} 망고 자체 메세지: 0건이 수집되었다 "
            f"(알림={found_msg!r}). 다음 입력으로 진행."
        )
    if found_n != expect:
        raise RuntimeError(
            f"#{row_no} 망고 수집건수 알림 불일치 — "
            f"기대 {expect}건 / 알림 {found_n}건 (문구={found_msg!r})"
        )
    ctx.info(f"  [확인] 망고 알림 수집건수 OK — {found_n}건이 수집됨")
    return found_n


def verify_row_save_done(page: Page, ctx: RunCtx, row_no: int, save_count: int) -> None:
    """저장 모달·알림 확인 후, 오류 문구·로그인 튕김 등 부가 검사."""
    page.wait_for_timeout(400)
    text = _page_visible_text(page)

    for pat in SAVE_FAIL_PATTERNS:
        if pat.search(text or ""):
            raise RuntimeError(f"#{row_no} 저장 후 오류 문구 감지: {pat.pattern}")

    ok_hit = None
    for pat in SAVE_OK_PATTERNS:
        m = pat.search(text or "")
        if m:
            ok_hit = m.group(0)
            break
    if ok_hit:
        ctx.info(f"  [확인] 화면 부가 성공 신호: {ok_hit!r}")

    if ctx.verify:
        if "admin_login" in page.url:
            raise RuntimeError(f"#{row_no} 저장 직후 로그인 화면으로 이동됨")


def _is_advance_fail(err: BaseException) -> bool:
    """같은 행을 더 돌리지 않고 다음 입력으로 넘길 확정 실패인지."""
    text = str(err or "")
    return any(m in text for m in ROW_ADVANCE_FAIL_MARKERS)


def process_row_with_retries(page: Page, row: dict, ctx: RunCtx) -> bool:
    """행 단위 재시도. 성공 True / 최종 실패 False.

    망고 무결과·시간초과 등 확정 실패는 같은 행을 반복하지 않고
    즉시 끝내 다음 입력 데이터로 넘긴다. (1행 무한루프 방지)
    """
    last_err: Exception | None = None
    label = str(row.get("label") or "").strip()
    raw_url = str(row.get("url") or "").strip()
    success = False
    try:
        for attempt in range(1, ctx.retries + 1):
            try:
                ctx.check_budget(f"행 시도 전 엑셀{row['row']}행")
            except RowBudgetExceeded as e:
                last_err = e
                ctx.info(f"  [다음행] {e}")
                break
            try:
                ctx.info(
                    f"> 시도 {attempt}/{ctx.retries} (엑셀 {row['row']}행) | "
                    f"상위 최종 카테고리명={label} | 최종 카테고리 URL주소={raw_url}"
                )
                page = refresh_if_closed(page)
                _process_row_once(page, row, ctx)
                if not ctx.server_save_ok:
                    raise RuntimeError(
                        f"#{row['row']} 저장하기 서버 최종 갱신 미확인 — "
                        "필터정보·수집갯수·수집상품이 서버에 반영되지 않음. "
                        "저장하기 성공 전에 행을 완료할 수 없습니다."
                    )
                ctx.info(
                    f"[OK] 엑셀{row['row']}행 성공 (저장하기 서버갱신 OK, 시도 {attempt}) | "
                    f"상위 최종 카테고리명={label} | 최종 카테고리 URL주소={raw_url}"
                )
                success = True
                return True
            except CollectStopped:
                raise
            except RowBudgetExceeded as e:
                last_err = e
                ctx.info(
                    f"[FAIL] 엑셀{row['row']}행 제한시간 초과 — 다음 입력으로 | {e}"
                )
                try:
                    ctx.shot(page, "fail_budget", row["row"])
                except Exception:  # noqa: BLE001
                    pass
                break
            except Exception as e:  # noqa: BLE001
                last_err = e
                err_name = type(e).__name__
                ctx.info(
                    f"[FAIL] 엑셀{row['row']}행 실패 (시도 {attempt}/{ctx.retries}) | "
                    f"상위 최종 카테고리명={label} | 최종 카테고리 URL주소={raw_url} | "
                    f"{err_name}: {e}"
                )
                try:
                    page = refresh_if_closed(page)
                    ctx.shot(page, f"fail_attempt{attempt}", row["row"])
                except Exception:
                    pass
                # 확정 실패 → 같은 행 재시도 중단, 다음 입력으로
                if _is_advance_fail(e):
                    ctx.info(
                        "  [다음행] 확정 실패 — "
                        "같은 행 재시도 없이 다음 입력 데이터로 진행"
                    )
                    break
                if "TargetClosed" in err_name or "Target closed" in str(e):
                    ctx.info("  탭 닫힘 감지 — 작업 페이지 재연결 시도")
                    try:
                        page = refresh_if_closed(page)
                    except Exception as re:  # noqa: BLE001
                        ctx.info(f"  재연결 경고: {re}")
                if attempt < ctx.retries:
                    # 팝업 없이 성공 처리된 게 아님 — 실패 후 재시도용 초기화만
                    ctx.info(
                        "  같은 행 재시도 전 초기화 "
                        "(저장 팝업 미확인/실패 후 복귀 — 성공 경로 아님)"
                    )
                    try:
                        page = refresh_if_closed(page)
                        reset_to_bulk_menu(page)
                    except Exception as re:  # noqa: BLE001
                        ctx.info(f"  복귀 중 경고: {re}")
                    try:
                        page.wait_for_timeout(800)
                    except Exception:
                        page = refresh_if_closed(page)
        if not success:
            ctx.info(f"[FAIL] {row['row']}행 최종 실패: {last_err}")
        return success
    finally:
        # 성공/실패와 무관하게 남은 모달·팝업 정리 → 다음 행 대기 부담 감소
        try:
            page = finalize_row_overlays(page, ctx, row)
        except Exception as e:  # noqa: BLE001
            ctx.info(f"  [행종료정리] 실패: {e}")
        ctx.row_deadline = None


def process_row(page: Page, row: dict, save_count: int = DEFAULT_SAVE_COUNT) -> None:
    """하위 호환: 저장수만 받아 1회 시도."""
    ctx = RunCtx(save_count=save_count, retries=1, batch=True)
    try:
        _process_row_once(page, row, ctx)
    finally:
        ctx.close()


def ensure_ready_page(page: Page) -> Page:
    """이미 망고 화면이면 그대로, 아니면 메인화면 진입 → 필요시 로그인 대기 → 0.초기화

    로그인 성공 후 사이트가 원래 탭을 닫아버리는 경우가 있어(예: 로그인
    중계 페이지가 스스로를 닫음) 매 단계 사이마다 탭이 살아있는지
    확인하고, 닫혔으면 같은 브라우저에서 새 탭을 다시 찾아온다.
    """
    page = refresh_if_closed(page)

    if ADMIN_HOST not in page.url or page.url in ("about:blank", ""):
        log("메인화면으로 이동: " + MAIN_URL)
        safe_goto(page, MAIN_URL)

    # 세션 없으면 더망고 로그인창을 직접 연다 (메인 경유 리다이렉트만 기다리지 않음)
    need_login = "admin_login" in page.url
    if not need_login:
        # 메인에 들어갔는데 세션 쿠키가 없으면 곧 로그인으로 튕김 → 선제 확인
        try:
            if page.locator('input[name="login_id"]').count() > 0:
                need_login = True
        except Exception:  # noqa: BLE001
            pass
    if need_login or ADMIN_HOST not in page.url:
        if "admin_login" not in page.url:
            log("더망고 로그인창으로 이동: " + LOGIN_URL)
            safe_goto(page, LOGIN_URL)
            page = refresh_if_closed(page)
        shot_now(page, "login_gate", 0)
        page = wait_for_user_login(page)
        page = refresh_if_closed(page)
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
        except Exception:  # noqa: BLE001
            pass
        page = refresh_if_closed(page)
        try:
            page.wait_for_timeout(500)
        except Exception:  # noqa: BLE001
            pass
        page = refresh_if_closed(page)
        if "admin_login" in page.url or ADMIN_HOST not in page.url:
            safe_goto(page, MAIN_URL)

    log("0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭")
    if BULK_PATH in page.url:
        wait_bulk_ready(page)
    else:
        reset_to_bulk_menu(page)

    return page


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser(
        description="P2 더망고 대량수집 — URL 1행당 상품 N건 (기본 3)"
    )
    ap.add_argument("excel", help="P1 출력 엑셀 (.xlsx)")
    ap.add_argument(
        "save_count",
        nargs="?",
        type=int,
        default=DEFAULT_SAVE_COUNT,
        help=f"행당 저장 상품 수 (기본 {DEFAULT_SAVE_COUNT})",
    )
    ap.add_argument(
        "--verify",
        action="store_true",
        help="검증 모드: 앞 N행 처리(기본 2), 1·2행 단계 스크린샷, 재시도, 무중단",
    )
    ap.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="처리할 최대 행 수 (검증 모드 기본 2)",
    )
    ap.add_argument(
        "--shot-first",
        type=int,
        default=2,
        help="단계별 스크린샷을 남길 앞쪽 입력 행 수 (기본 2 = 1·2행)",
    )
    ap.add_argument(
        "--retries",
        type=int,
        default=DEFAULT_ROW_RETRIES,
        help=f"행 실패 시 같은 행 재시도 횟수 (기본 {DEFAULT_ROW_RETRIES})",
    )
    ap.add_argument(
        "--yes",
        "--batch",
        dest="batch",
        action="store_true",
        help="실패해도 y/n 묻지 않고 다음 행으로 (또는 재시도만)",
    )
    ap.add_argument(
        "--id",
        dest="tmg_id",
        default=None,
        help="(미사용) 자동 로그인 제거됨 — 브라우저에서 직접 로그인",
    )
    ap.add_argument(
        "--pw",
        dest="tmg_pw",
        default=None,
        help="(미사용) 자동 로그인 제거됨 — 브라우저에서 직접 로그인",
    )
    args = ap.parse_args()

    excel_path = args.excel
    save_count = args.save_count if args.save_count and args.save_count > 0 else DEFAULT_SAVE_COUNT
    verify = bool(args.verify)
    max_rows = args.max_rows
    shot_first = max(0, int(args.shot_first))
    if verify and max_rows is None:
        max_rows = max(2, shot_first)  # 검증 기본: 1·2행

    if args.tmg_id or args.tmg_pw:
        log("[안내] --id/--pw 는 더 이상 사용하지 않습니다. 브라우저에서 직접 로그인하세요.")

    all_rows = read_excel(excel_path)
    if not all_rows:
        safe_print("엑셀에 처리할 행이 없습니다.")
        sys.exit(1)

    ctx = RunCtx(
        save_count=save_count,
        retries=args.retries,
        verify=verify,
        max_rows=max_rows,
        batch=args.batch or verify,
        shot_first_n=shot_first,
    )

    # 모든 입력 데이터 카테고리명·URL을 실행 로그에 기록
    ctx.info(f"[입력목록] 파일={excel_path}")
    ctx.info(f"[입력목록] 총 {len(all_rows)}건 (상위 최종 카테고리명 / 최종 카테고리 URL주소)")
    for i, r in enumerate(all_rows, start=1):
        ctx.info(
            f"  입력#{i} 엑셀{r['row']}행 | "
            f"상위 최종 카테고리명={r['label']} | "
            f"최종 카테고리 URL주소={r['url']}"
        )

    rows = all_rows
    if max_rows is not None:
        rows = all_rows[: max(0, max_rows)]
    ctx.info(
        f"처리대상 {len(rows)}행 / 전체입력 {len(all_rows)}행 · 저장수={save_count} · "
        f"재시도={ctx.retries} · verify={verify} · 샷=입력1~{shot_first}행 · "
        f"로그={ctx.shot_dir}"
    )

    clear_stop_flag()
    ok = 0
    fail = 0
    stopped = False
    try:
        with sync_playwright() as p:
            _browser, page = connect_browser(p)
            page.set_default_timeout(120_000)
            # 망고 Chrome 기동 직후 — 더망고 솔루션 확장에 URL/KEY 필수 세팅
            ensure_mango_extension_settings(page.context, shot_ctx=ctx)
            page = refresh_if_closed(page)
            page = ensure_ready_page(page)
            ctx.shot(page, "ready", 0)

            for ordinal, row in enumerate(rows, start=1):
                success = False
                try:
                    check_stop(f"입력#{ordinal} 시작 전")
                    # 2번(및 이후) 수집 전: 팝업·모달 전부 닫힘 확인 + 스크린샷 보관
                    if ordinal >= 2:
                        page = ensure_overlays_closed_before_next(
                            page,
                            ctx,
                            next_ordinal=ordinal,
                            next_row=row,
                        )
                    ctx.begin_row(ordinal, row)
                    page = refresh_if_closed(page)
                    success = process_row_with_retries(page, row, ctx)
                except CollectStopped as e:
                    stopped = True
                    ctx.info(f"[중단] {e} — 로그·브라우저는 유지합니다")
                    break
                except RuntimeError as e:
                    if "수집 전 팝업/모달이 닫히지 않음" in str(e):
                        ctx.info(
                            f"[FAIL] 입력#{ordinal} 시작 전 팝업/모달 미정리 — "
                            f"해당 입력 건너뛰고 다음으로 | {e}"
                        )
                        fail += 1
                        continue
                    raise
                if success:
                    ok += 1
                else:
                    fail += 1
                    if not ctx.batch:
                        if input("계속 진행할까요? (y/n) ").strip().lower() != "y":
                            break
                ctx.info(
                    f"==== 입력#{ordinal} 종료 (성공={success}) "
                    f"| 엑셀{row['row']}행 ===="
                )
                if ordinal < len(rows) and not stopped:
                    nxt = rows[ordinal]  # next item (0-based index = ordinal)
                    ctx.info(
                        f"==== 다음 입력#{ordinal + 1} 로 진행 "
                        f"(엑셀{nxt['row']}행 / {nxt.get('label', '')}) ===="
                    )

            if stopped:
                ctx.info(
                    f"[중단완료] 성공 {ok} / 실패 {fail} / "
                    f"대상 {len(rows)}행 중 일부 — 화면 로그 보존"
                )
            else:
                ctx.info(
                    f"완료: 성공 {ok} / 실패 {fail} / "
                    f"대상 {len(rows)}행 / 입력전체 {len(all_rows)}건"
                )
            gallery = ctx.write_gallery()
            ctx.info(f"스크린샷·로그: {ctx.shot_dir}")
            if gallery:
                ctx.info(f"[갤러리] {gallery}")
            safe_print("브라우저는 그대로 열어둡니다 (이 창만 닫으면 됩니다).")
            if stopped:
                sys.exit(130)
            if verify and ok >= 1 and fail == 0:
                safe_print(
                    f"[OK] 검증 모드 PASS — {ok}행 완료 · "
                    f"입력 1~{shot_first}행 전과정 스크린샷 기록됨"
                )
                sys.exit(0)
            if fail:
                sys.exit(2)
    except CollectStopped as e:
        stopped = True
        ctx.info(f"[중단] {e} — 로그·브라우저는 유지합니다")
        try:
            ctx.write_gallery()
        except Exception:  # noqa: BLE001
            pass
        sys.exit(130)
    finally:
        clear_stop_flag()
        ctx.close()


if __name__ == "__main__":
    main()
