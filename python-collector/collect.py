"""
더망고(tmg1898) 상품데이터 대량수집 — 요건 0~4 그대로

0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭
1. URL상품검색하기 : 필드값 입력 후 클릭 -> 팝업창이 없어질 때까지 대기
2. 검색된 상품 모두저장 클릭 -> 팝업창에서 검색필터명 입력 -> 저장하기 버튼 클릭
3. 팝업창이 없어질 때까지 대기
4. -> 0. 초기화

사용법:
    python collect.py 엑셀파일.xlsx
    python collect.py 엑셀파일.xlsx 5     (저장수 5개, 기본 3)

엑셀 헤더(1행): 상위 최종 카테고리명 | 최종 카테고리 URL주소

Playwright의 별도 Chromium(다운로드본)을 쓰지 않는다. PC에 이미 설치된
Chrome/Edge — 평소 망고 화면을 여는 그 브라우저 — 를 디버그 모드로 열어
그대로 이어서 작업한다(CDP 연결). 이미 그 창이 열려 있고 로그인도 되어
있으면 로그인 화면을 거치지 않고 바로 메인화면에서 시작한다.
팝업창은 스크립트가 절대 열거나 닫지 않습니다 — 항상 "스스로 닫힐 때까지" 기다립니다.
"""

import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    TimeoutError as PWTimeout,
    sync_playwright,
)

LOGIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin_login.php"
MAIN_URL = "https://tmg1898.cafe24.com/mall/admin/admin.php"
BULK_URL = "https://tmg1898.cafe24.com/mall/admin/shop/getGoodsNew.php"
ADMIN_HOST = "tmg1898.cafe24.com"
BULK_PATH = "getGoodsNew.php"

CDP_PORT = 9222
CDP_URL = f"http://127.0.0.1:{CDP_PORT}"
PROFILE_DIR = Path(__file__).parent / ".chrome-profile"

CHROME_CANDIDATES = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
]

POPUP_WAIT_SEC = 600
MODAL_WAIT_SEC = 180
DEFAULT_SAVE_COUNT = 3

FILTER_NAME_LABEL = re.compile(r"검색\s*필터\s*명")
SAVE_COUNT_LABEL = re.compile(r"저장\s*상품\s*수|검색결과\s*상위")


def log(msg: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)


# ── 브라우저 연결 (기존 Chrome/Edge에 CDP로 붙기, Chromium 다운로드 없음) ──

def cdp_port_open(port: int = CDP_PORT) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        try:
            s.connect(("127.0.0.1", port))
            return True
        except OSError:
            return False


def find_browser_exe() -> str | None:
    for path in CHROME_CANDIDATES:
        if Path(path).exists():
            return path
    return None


def _clear_stale_singleton_locks() -> None:
    """
    이전 실행이 비정상 종료되어 남은 잠금파일이 있으면, 새 Chrome이
    "이미 실행 중"이라고 착각해 디버그 포트 없이 조용히 기존 창에만
    메시지를 보내고 끝나버릴 수 있다(=화면에 아무 반응도 없어 보임).
    cdp_port_open()이 False라는 건 우리 프로필로 살아있는 프로세스가
    없다는 뜻이므로 안전하게 지운다.
    """
    for name in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
        f = PROFILE_DIR / name
        try:
            if f.exists() or f.is_symlink():
                f.unlink()
        except OSError:
            pass


def launch_debug_browser() -> None:
    """평소 쓰는 Chrome/Edge를 디버그 포트로 실행 — Playwright Chromium 미사용"""
    exe = find_browser_exe()
    if not exe:
        raise SystemExit(
            "Chrome 또는 Edge를 찾지 못했습니다.\n"
            "https://www.google.com/chrome/ 에서 설치 후 다시 실행하세요."
        )
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    _clear_stale_singleton_locks()
    log(f"브라우저 실행: {exe}")
    # 주의: stdout/stderr를 DEVNULL로 버리면 일부 환경(보안 소프트웨어·
    # 컨테이너 등)에서 Chrome이 바로 죽는 경우가 있다(무반응처럼 보임).
    # 로그 파일로 받아두면 안전하고, 문제 생기면 이 파일로 원인도 알 수 있다.
    log_path = PROFILE_DIR / "chrome_debug.log"
    log_file = open(log_path, "ab")
    subprocess.Popen(
        [
            exe,
            f"--remote-debugging-port={CDP_PORT}",
            f"--user-data-dir={PROFILE_DIR}",
            "--no-first-run",
            "--no-default-browser-check",
            MAIN_URL,
        ],
        stdout=log_file,
        stderr=log_file,
    )
    log("디버그 모드 연결 대기 중 (최대 30초)...")
    for i in range(60):
        if cdp_port_open():
            log("브라우저 연결 확인됨")
            return
        if i > 0 and i % 6 == 0:
            log(f"  아직 대기 중... ({i * 0.5:.0f}초 경과 — 화면에 Chrome 창이 떴는지 확인해 주세요)")
        time.sleep(0.5)  # 소켓 연결 대기 — Playwright 이벤트루프와 무관하므로 안전
    raise SystemExit(
        "브라우저가 디버그 모드로 열리지 않았습니다.\n"
        "열려있는 Chrome/Edge 창을 모두 닫고 다시 실행해 보세요.\n"
        f"(참고 로그: {log_path})"
    )


def pick_working_page(context: BrowserContext) -> Page:
    """이미 망고 화면이 열려 있으면 그 탭을 그대로 사용(기본 화면 유지)"""
    for p in context.pages:
        if p.is_closed():
            continue
        try:
            if ADMIN_HOST in p.url:
                return p
        except Exception:
            continue
    return context.pages[0] if context.pages else context.new_page()


def connect_browser(p) -> tuple[Browser, Page]:
    """
    1) 이미 디버그 모드로 열린 Chrome/Edge가 있으면 그대로 연결(= 망고 기본 화면 그대로)
    2) 없으면 평소 쓰는 Chrome/Edge를 디버그 모드로 새로 열어서 연결
    Playwright 전용 Chromium은 내려받지 않는다.
    """
    if not cdp_port_open():
        launch_debug_browser()

    browser = p.chromium.connect_over_cdp(CDP_URL)
    context = browser.contexts[0] if browser.contexts else browser.new_context()
    page = pick_working_page(context)
    return browser, page


# ── 엑셀 ──────────────────────────────────────────────────────

def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        label_col = headers.index("상위 최종 카테고리명")
        url_col = headers.index("최종 카테고리 URL주소")
    except ValueError:
        raise SystemExit(
            "엑셀 1행 헤더에 '상위 최종 카테고리명', '최종 카테고리 URL주소' 열이 있어야 합니다."
        )

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        label = str(row[label_col].value or "").strip()
        url = str(row[url_col].value or "").strip()
        if url:
            rows.append({"row": i, "label": label, "url": url})
    return rows


def normalize_url(u: str) -> str:
    u = u.strip()
    return u if re.match(r"^https?://", u, re.I) else f"https://{u}"


# ── 팝업 ──────────────────────────────────────────────────────

def popups(page: Page) -> list:
    result = []
    for p in page.context.pages:
        if p is page or p.is_closed():
            continue
        try:
            u = p.url
        except Exception:
            continue
        if u and u != "about:blank" and ADMIN_HOST not in u:
            result.append(p)
    return result


def wait_popups_gone(
    page: Page, timeout_sec: int = POPUP_WAIT_SEC, grace_sec: float = 2.0
) -> None:
    """팝업이 스스로 닫힐 때까지 대기 — 절대 건드리지 않음

    grace_sec: 클릭 직후 팝업이 뜨기까지 잠깐 기다리는 시간.
    팝업이 아예 안 뜨는 경우(이미 닫혀 있음)까지 대비해 대기 시간은 짧게 둔다
    (없는 팝업을 기다리며 매 행마다 시간을 허비하지 않도록).

    주의: Playwright Python 동기 API는 이벤트(새 창 열림/닫힘)를
    time.sleep() 중에는 처리하지 않는다. 반드시 page.wait_for_timeout()
    으로 기다려야 context.pages()가 실시간으로 갱신된다.
    """
    end = time.time() + timeout_sec
    grace_end = time.time() + grace_sec
    while not popups(page) and time.time() < grace_end:
        page.wait_for_timeout(200)

    last_beat = 0.0
    while popups(page):
        if time.time() > end:
            raise TimeoutError("팝업창이 닫히지 않음")
        if time.time() - last_beat > 10:
            last_beat = time.time()
            log(f"  팝업창 대기중... (열린 팝업 {len(popups(page))}개)")
        page.wait_for_timeout(500)


# ── 입력 · 클릭 (망고 구형 input 대응) ────────────────────────

def type_into(page: Page, locator, value: str) -> None:
    el = locator.first
    el.wait_for(state="attached", timeout=60_000)
    el.scroll_into_view_if_needed()
    try:
        el.click(timeout=15_000)
    except PWTimeout:
        pass
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.keyboard.insert_text(value)

    got = ""
    try:
        got = el.input_value()
    except Exception:
        pass
    if not got.strip():
        el.evaluate(
            """(node, v) => {
                if (node instanceof HTMLInputElement || node instanceof HTMLTextAreaElement) {
                    node.focus();
                    node.value = v;
                    node.dispatchEvent(new Event('input', { bubbles: true }));
                    node.dispatchEvent(new Event('change', { bubbles: true }));
                    node.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true }));
                    node.dispatchEvent(new Event('blur', { bubbles: true }));
                }
            }""",
            value,
        )
        got = ""
        try:
            got = el.input_value()
        except Exception:
            pass

    if got.strip() != value.strip():
        log(f"  [경고] 입력값 불일치 — 넣으려던 값: {value!r} / 실제 값: {got!r}")


def click_it(locator) -> None:
    el = locator.first
    el.wait_for(state="visible", timeout=60_000)
    el.scroll_into_view_if_needed()
    try:
        el.click(timeout=20_000)
    except PWTimeout:
        el.evaluate("(node) => node.click()")


# ── 화면 요소 ─────────────────────────────────────────────────

def url_search_button(page: Page):
    return page.locator(
        'input[type="button"][value*="URL"], input[type="submit"][value*="URL"]'
    ).or_(page.get_by_text(re.compile(r"URL\s*상품\s*검색")))


def save_all_button(page: Page):
    return (
        page.locator('input[type="button"][value*="모두저장"]')
        .or_(page.locator('input[type="submit"][value*="모두저장"]'))
        .or_(page.get_by_text(re.compile(r"검색된\s*상품\s*모두\s*저장")))
    )


def _describe(loc) -> str:
    try:
        return loc.evaluate(
            "n => `<${n.tagName.toLowerCase()}"
            " name=${n.name||''} id=${n.id||''} rows=${n.rows||''}"
            " value.len=${(n.value||'').length}>`"
        )
    except Exception:
        return "<알 수 없음>"


def url_input(page: Page):
    """
    URL상품검색하기 버튼과 실제 입력칸이 서로 다른 <tr>/<table>에 있는
    화면이 있어(선택자가 넓으면 엉뚱한 textarea를 골라 "검색결과 없음"이
    나는 원인이 됨) 좁은 범위 -> 넓은 범위 순으로, 후보가 정확히 하나일
    때만 채택한다.
    """
    btn = url_search_button(page).first

    # 1) 버튼과 같은 <tr> 안에서 우선 찾기 (가장 정확)
    row = btn.locator("xpath=ancestor::tr[1]")
    if row.count() > 0:
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = row.locator(sel)
            if cand.count() > 0:
                found = cand.first
                log(f"  URL입력칸(같은 행에서 발견): {_describe(found)}")
                return found

    # 2) 부모를 한 단계씩 올라가며(최대 4단계) 후보가 정확히 하나일 때만 채택
    ancestor = btn
    for _ in range(4):
        ancestor = ancestor.locator("xpath=..")
        for sel in ("textarea", 'input[type="text"]:not([name*="login"]):not([readonly])'):
            cand = ancestor.locator(sel)
            if cand.count() == 1:
                found = cand.first
                log(f"  URL입력칸(상위 요소에서 발견): {_describe(found)}")
                return found

    # 3) 최후 수단: 페이지 전체에서 rows 속성이 가장 큰 textarea
    #    (URL 여러 줄 입력용 큰 textarea일 가능성이 높음)
    all_ta = page.locator("textarea")
    n = all_ta.count()
    if n == 1:
        found = all_ta.first
        log(f"  URL입력칸(페이지에 textarea 1개뿐): {_describe(found)}")
        return found
    if n > 1:
        best_idx, best_rows = 0, -1
        for i in range(n):
            rows_attr = all_ta.nth(i).get_attribute("rows")
            try:
                rows_val = int(rows_attr) if rows_attr else 1
            except ValueError:
                rows_val = 1
            if rows_val > best_rows:
                best_rows, best_idx = rows_val, i
        found = all_ta.nth(best_idx)
        log(f"  URL입력칸(가장 큰 textarea 선택, {n}개 중): {_describe(found)}")
        return found

    raise RuntimeError("URL 입력칸을 찾지 못했습니다")


def save_modal(page: Page):
    return (
        page.locator("div, form, table")
        .filter(has_text=re.compile(r"상품\s*저장\s*설정|검색\s*필터\s*명"))
        .filter(has_text=re.compile("저장하기"))
        .last
    )


def save_modal_visible(page: Page) -> bool:
    try:
        return page.get_by_text(re.compile(r"상품\s*저장\s*설정")).first.is_visible()
    except Exception:
        return False


def modal_field(page: Page, label_pattern: re.Pattern):
    modal = save_modal(page)
    return (
        modal.locator("tr, div, p, label")
        .filter(has_text=label_pattern)
        .locator('input[type="text"], input:not([type]), input[type="number"]')
        .first
    )


# ── 0 ~ 4 ────────────────────────────────────────────────────

def safe_goto(page: Page, url: str, retries: int = 3) -> None:
    """
    로그인 직후 등 사이트 자체가 리다이렉트를 진행 중일 때
    page.goto()와 겹치면 'interrupted by another navigation' 오류가 난다.
    사이트 쪽 리다이렉트가 끝날 때까지 기다렸다가 다시 시도한다.
    """
    last_err: Exception | None = None
    for _ in range(retries):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=120_000)
            return
        except Exception as e:  # noqa: BLE001
            msg = str(e)
            if "interrupted by another navigation" in msg or "NS_BINDING_ABORTED" in msg:
                last_err = e
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=15_000)
                except Exception:  # noqa: BLE001
                    pass
                page.wait_for_timeout(800)
                continue
            raise
    if last_err:
        raise last_err


def wait_bulk_ready(page: Page) -> None:
    try:
        page.wait_for_load_state("domcontentloaded", timeout=15_000)
    except Exception:  # noqa: BLE001
        pass
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    log("  대량수집 화면 로딩 확인 중...")
    url_search_button(page).first.wait_for(state="visible", timeout=60_000)


def reset_to_bulk_menu(page: Page) -> None:
    """0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭"""
    href = page.locator('a[href*="getGoodsNew"]').first
    if href.count() > 0:
        try:
            href.click(timeout=5000)
        except PWTimeout:
            href.evaluate("(node) => node.click()")
        page.wait_for_timeout(800)
        if BULK_PATH in page.url:
            wait_bulk_ready(page)
            return

    page.evaluate(
        """() => {
            const clean = (s) => (s || '').replace(/\\s+/g, '');
            const nodes = Array.from(document.querySelectorAll('a, li, span, td, div, button'));
            const byHref = Array.from(document.querySelectorAll('a[href*="getGoodsNew"]'));
            if (byHref[0]) { byHref[0].click(); return; }
            const top = nodes.find(el => clean(el.textContent) === '상품데이터수집');
            if (top) top.click();
            const sub = nodes.find(el => {
                const t = clean(el.textContent);
                if (t.length > 30) return false;
                return /대량데이터수집|대량수집|상품데이터대량/.test(t);
            });
            if (sub) (sub.closest('a') || sub).click();
        }"""
    )
    page.wait_for_timeout(1000)
    if BULK_PATH not in page.url:
        safe_goto(page, BULK_URL)
    wait_bulk_ready(page)


def process_row(page: Page, row: dict, save_count: int) -> None:
    label = row["label"]
    url = normalize_url(row["url"])
    log(f"--- {row['row']}행 : {label} ---")

    log("0. 초기화 : 상품데이터수집 -> 대량데이터수집")
    reset_to_bulk_menu(page)
    page.wait_for_timeout(500)

    log(f"1. 필드값 입력: {url}")
    target = url_input(page)
    type_into(page, target, url)
    actual = ""
    try:
        actual = target.input_value()
    except Exception:
        pass
    log(f"  입력칸 최종 값: {actual!r}")
    log("1. URL상품검색하기 클릭")
    click_it(url_search_button(page))

    log("1. 팝업창이 없어질 때까지 대기")
    wait_popups_gone(page)
    log("1. 팝업창 닫힘")

    log("2. 검색된 상품 모두저장 클릭")
    click_it(save_all_button(page))
    save_modal(page).wait_for(state="visible", timeout=MODAL_WAIT_SEC * 1000)
    page.wait_for_timeout(400)

    log(f"2. 검색필터명 입력: {label}")
    type_into(page, modal_field(page, FILTER_NAME_LABEL), label)

    count_field = modal_field(page, SAVE_COUNT_LABEL)
    if count_field.count() > 0:
        type_into(page, count_field, str(save_count))
        type_into(page, modal_field(page, FILTER_NAME_LABEL), label)  # 덮어쓰기 방지 재확인

    log("2. 저장하기 버튼 클릭")
    click_it(
        save_modal(page)
        .locator('input[value*="저장하기"]')
        .or_(save_modal(page).locator('button:has-text("저장하기")'))
        .or_(save_modal(page).get_by_text(re.compile("^저장하기$")))
    )

    log("3. 팝업창이 없어질 때까지 대기")
    end = time.time() + MODAL_WAIT_SEC
    closed = False
    while time.time() < end:
        if not save_modal_visible(page):
            wait_popups_gone(page, grace_sec=0.5)  # 남은 팝업이 있으면만 대기
            closed = True
            break
        page.wait_for_timeout(500)
    if not closed:
        raise TimeoutError(f"#{row['row']} 저장 팝업창이 닫히지 않음")
    log("3. 팝업창 닫힘")

    log("4. -> 0. 초기화")


def ensure_ready_page(page: Page) -> None:
    """이미 망고 화면이면 그대로, 아니면 메인화면 진입 → 필요시 로그인 대기 → 0.초기화"""
    if ADMIN_HOST not in page.url or page.url in ("about:blank", ""):
        log("메인화면으로 이동: " + MAIN_URL)
        safe_goto(page, MAIN_URL)

    if "admin_login" in page.url:
        input("로그인이 필요합니다 — 브라우저에서 로그인 후 이 창에서 Enter 를 누르세요...")
        # 로그인 직후 사이트 자체가 리다이렉트 중일 수 있으므로(m_login_ok.php 등)
        # 안정될 때까지 잠깐 기다린 뒤에 필요하면 이동한다.
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
        except Exception:  # noqa: BLE001
            pass
        page.wait_for_timeout(1000)
        if "admin_login" in page.url or ADMIN_HOST not in page.url:
            safe_goto(page, MAIN_URL)

    log("0. 초기화 : 상품데이터수집 -> 대량데이터수집 클릭")
    if BULK_PATH in page.url:
        wait_bulk_ready(page)
    else:
        reset_to_bulk_menu(page)


def main() -> None:
    if len(sys.argv) < 2:
        print("사용법: python collect.py 엑셀파일.xlsx [저장수(기본 3)]")
        sys.exit(1)

    excel_path = sys.argv[1]
    save_count = int(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_SAVE_COUNT

    rows = read_excel(excel_path)
    if not rows:
        print("엑셀에 처리할 행이 없습니다.")
        sys.exit(1)
    print(f"엑셀 {len(rows)}행 로드 완료. 저장수={save_count}")

    with sync_playwright() as p:
        _browser, page = connect_browser(p)
        page.set_default_timeout(120_000)
        ensure_ready_page(page)

        ok = 0
        for row in rows:
            try:
                process_row(page, row, save_count)
                ok += 1
            except Exception as e:  # noqa: BLE001
                log(f"오류: {e}")
                if input("계속 진행할까요? (y/n) ").strip().lower() != "y":
                    break

        print(f"완료: {ok}/{len(rows)}행 처리")
        print("브라우저는 그대로 열어둡니다 (이 창만 닫으면 됩니다).")


if __name__ == "__main__":
    main()
