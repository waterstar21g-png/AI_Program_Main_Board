"""
P1_101 상품수 추출

엑셀(URL 주소 포함)을 읽어 각 URL을 브라우저로 열고,
(1)첫 팝업 닫기 → (2)푸터까지 스크롤 다운 → (3)하단→상단 스크롤하며
카드 이미지를 센 뒤 → (4)그 갯수를 총상품수로 엑셀·로그에 기록한다.

사용법:
    python extract.py 엑셀.xlsx
    python extract.py 엑셀.xlsx --headless
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path, PureWindowsPath
from typing import Callable
from urllib.parse import urlparse

from openpyxl import load_workbook

# Windows cp949 콘솔 안전 출력
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

ProgressFn = Callable[[str, str], None]

DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# ★요건: 팝업 닫기 후 3초 대기 → 상품수 수집
POST_POPUP_WAIT_SEC = 3.0

# 메인보드 직접실행 시 중단 플래그 (P2와 동일 패턴)
STOP_FLAG_PATH = Path(__file__).resolve().parent / ".extract_stop"

URL_HEADER_CANDIDATES = (
    "최종 카테고리 URL주소",
    "최종 카테고리 URL",
    "카테고리 URL",
    "URL주소",
    "URL",
    "url",
)
COUNT_HEADER = "총상품수"
COLLECTIBLE_HEADER = "상품수집가능개수"

# ★요건: 출력 엑셀은 입력과 별도 — 지정 폴더에 파일명+버전으로 생성
OUTPUT_DIR = Path(r"D:\My_Project\AI_Program_Main_Board\P2_INPUT_건수집계")
ROOT_DIR = Path(__file__).resolve().parent.parent
VERSION_FILE = ROOT_DIR / "VERSION.txt"

# 팝업/레이어 닫기 버튼 후보 (순서대로 시도)
POPUP_CLOSE_SELECTORS = (
    'button[aria-label="Close"]',
    'button[aria-label="닫기"]',
    '[aria-label="Close"]',
    '[aria-label="닫기"]',
    'button.close',
    '.btn-close',
    '.layer-popup .close',
    '.popup .close',
    '.modal .close',
    '#popupClose',
    '.popup-close',
    '.btn_close',
    'button:has-text("닫기")',
    'a:has-text("닫기")',
    'button:has-text("닫 기")',
    'button:has-text("Close")',
    'button:has-text("동의하고 닫기")',
    'button:has-text("오늘 하루 보지 않기")',
    'button:has-text("다시 보지 않기")',
    '[class*="close-button"]',
    '[class*="CloseButton"]',
    '[data-qa-action="stay-on-store"]',
    'button:has-text("Accept")',
    'button:has-text("동의")',
    'button:has-text("수락")',
)


@dataclass
class RowJob:
    excel_row: int  # 1-based sheet row
    url: str
    label: str = ""


@dataclass
class RowResult:
    """행별 최종출력용."""

    label: str
    count: int | None  # None = 수집 실패
    url: str


@dataclass
class ExtractResult:
    ok: bool
    excel_path: str  # 출력(저장) 엑셀 경로
    total: int = 0
    updated: int = 0
    failed: int = 0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    rows: list[RowResult] = field(default_factory=list)
    input_excel_path: str = ""


def read_app_version() -> str:
    """프로젝트 VERSION.txt 버전 숫자 (예: 2.0.95)."""
    try:
        text = VERSION_FILE.read_text(encoding="utf-8")
    except OSError:
        return "0.0.0"
    m = re.search(r"(?:버전|version)\s*([0-9]+(?:\.[0-9]+)+)", text, re.I)
    if m:
        return m.group(1)
    m = re.search(r"([0-9]+\.[0-9]+\.[0-9]+)", text)
    return m.group(1) if m else "0.0.0"


def build_output_excel_path(
    input_path: Path,
    *,
    version: str | None = None,
    when: time.struct_time | None = None,
    out_dir: Path | None = None,
) -> Path:
    """입력 파일명 + 버전으로 출력 경로 생성.

    예: sample.xlsx → sample_상품수_v2.0.95_20260812_151543.xlsx
    """
    ver = (version or read_app_version()).strip() or "0.0.0"
    stamp = time.strftime("%Y%m%d_%H%M%S", when or time.localtime())
    raw = str(input_path)
    # Windows 경로(D:\...) 가 Linux 테스트에서도 stem 만 쓰이도록
    if "\\" in raw or (len(raw) >= 2 and raw[1] == ":"):
        win = PureWindowsPath(raw)
        stem = win.stem
        suffix = win.suffix.lower() or ".xlsx"
    else:
        stem = Path(raw).stem
        suffix = Path(raw).suffix.lower() or ".xlsx"
    name = f"{stem}_상품수_v{ver}_{stamp}{suffix}"
    return (out_dir or OUTPUT_DIR) / name


def _log(progress: ProgressFn | None, step: str, message: str) -> None:
    # 보드 서브프로세스 stdout 파싱용 — 항상 출력
    print(f"[{step}] {message}", flush=True)
    if progress:
        progress(step, message)


def format_final_output(label: str, count: int | None, url: str) -> str:
    """실행로그 최종출력 1행 — 상위 최종 카테고리명 · 상품갯수 · url."""
    name = (label or "").strip() or "(이름없음)"
    cnt = "실패" if count is None else str(count)
    return f"상위 최종 카테고리명={name} · 상품갯수={cnt} · url={url}"


def log_final_outputs(
    progress: ProgressFn | None,
    rows: list[RowResult],
    *,
    title: str = "최종출력",
) -> None:
    """수집 결과를 실행로그에 최종출력 형식으로 일괄 표시."""
    _log(progress, "최종", f"===== {title} ({len(rows)}건) =====")
    if not rows:
        _log(progress, "최종", "(출력할 행 없음)")
        return
    for r in rows:
        _log(progress, "최종", format_final_output(r.label, r.count, r.url))


def clear_stop_flag() -> None:
    try:
        STOP_FLAG_PATH.unlink(missing_ok=True)  # type: ignore[call-arg]
    except TypeError:
        if STOP_FLAG_PATH.exists():
            try:
                STOP_FLAG_PATH.unlink()
            except OSError:
                pass
    except OSError:
        pass


def stop_requested() -> bool:
    return STOP_FLAG_PATH.is_file()


def _parse_int_count(raw: str | int | float | None) -> int:
    if raw is None or isinstance(raw, bool):
        return 0
    if isinstance(raw, (int, float)):
        return max(0, int(raw))
    s = re.sub(r"[^\d]", "", str(raw))
    return int(s) if s else 0


# Zara 등 SPA 상품 그리드 카드
PRODUCT_CARD_SELECTORS = (
    "li.product-grid-product",
    "article.product-grid-product",
    ".product-grid-product",
    "[data-productid]",
    "[data-qa-qualifier='product-grid-product']",
    "li[class*='product-grid']",
)


def parse_product_count_from_html(html: str) -> int:
    """페이지 HTML에서 노출 상품수를 파싱한다 (A-RT · Zara · 일반 패턴)."""
    text = html or ""
    if not text:
        return 0

    # A-RT: hidden totalCount / result-cnt
    m = re.search(
        r'name=["\']totalCount["\'][^>]*value=["\']([^"\']*)["\']',
        text,
        re.I,
    )
    if m:
        n = _parse_int_count(m.group(1))
        if n:
            return n
    m = re.search(
        r'value=["\']([^"\']*)["\'][^>]*name=["\']totalCount["\']',
        text,
        re.I,
    )
    if m:
        n = _parse_int_count(m.group(1))
        if n:
            return n
    m = re.search(
        r'class=["\'][^"\']*result-cnt[^"\']*["\'][^>]*>\s*([\d,]+)',
        text,
        re.I,
    )
    if m:
        n = _parse_int_count(m.group(1))
        if n:
            return n

    # JSON / JS / JSON-LD 흔한 키 (Zara·일반)
    for pat in (
        r'"numberOfItems"\s*:\s*(\d+)',
        r'"totalCount"\s*:\s*(\d+)',
        r'"productCount"\s*:\s*(\d+)',
        r'"productsCount"\s*:\s*(\d+)',
        r'"totalProducts"\s*:\s*(\d+)',
        r'"productsTotal"\s*:\s*(\d+)',
        r'"total_count"\s*:\s*(\d+)',
        r'"resultsCount"\s*:\s*(\d+)',
        r'"resultCount"\s*:\s*(\d+)',
        r'"numProducts"\s*:\s*(\d+)',
        r'"productsTotalCount"\s*:\s*(\d+)',
    ):
        m = re.search(pat, text, re.I)
        if m:
            n = _parse_int_count(m.group(1))
            if n:
                return n

    # 화면 표기: "총 930개", "상품 930개", "930 products", "930 results"
    patterns = (
        r"총\s*상품\s*수?\s*[:：]?\s*([\d,]+)",
        r"상품\s*수\s*[:：]?\s*([\d,]+)",
        r"총\s*([\d,]+)\s*개\s*상품",
        r"총\s*([\d,]+)\s*개",
        r"상품\s*([\d,]+)\s*개",
        r"([\d,]+)\s*개\s*상품",
        r"([\d,]+)\s*products?\b",
        r"([\d,]+)\s*results?\b",
        r"([\d,]+)\s*items?\b",
        r"([\d,]+)\s*Artikel\b",
        r"([\d,]+)\s*Produkte?\b",
    )
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            n = _parse_int_count(m.group(1))
            if n:
                return n

    # Zara 등: data-productid / 상품 상세 링크(-p####.html)
    ids = set(re.findall(r'data-productid=["\'](\d+)["\']', text, flags=re.I))
    if ids:
        return len(ids)
    p_links = set(
        re.findall(
            r"https?://[^\"'\s]+?-p\d+\.html",
            text,
            flags=re.I,
        )
    )
    if not p_links:
        p_links = set(re.findall(r"/[^\"'\s]+?-p\d+\.html", text, flags=re.I))
    if p_links:
        return len(p_links)
    return 0


def _total_from_json_obj(data: object) -> int:
    """API/JSON 객체에서 총상품수 후보를 찾는다."""
    if data is None:
        return 0
    if isinstance(data, (int, float)) and not isinstance(data, bool):
        return max(0, int(data))
    if isinstance(data, list):
        # 상품 배열로 보이면 길이
        if data and isinstance(data[0], dict):
            keys = set(data[0].keys())
            if keys & {"id", "productId", "seo", "detail", "name", "price"}:
                return len(data)
        best = 0
        for item in data[:50]:
            best = max(best, _total_from_json_obj(item))
        return best
    if not isinstance(data, dict):
        return 0
    for key in (
        "totalProducts",
        "productsCount",
        "productCount",
        "totalCount",
        "numberOfItems",
        "numProducts",
        "productsTotalCount",
        "resultsCount",
    ):
        if key in data:
            n = _parse_int_count(data.get(key))
            if n:
                return n
    # productGroups / products 배열
    for key in ("products", "productGroups", "items", "results"):
        if key not in data:
            continue
        val = data[key]
        if isinstance(val, list):
            if key == "productGroups":
                ids: set[str] = set()
                for g in val:
                    if not isinstance(g, dict):
                        continue
                    for p in g.get("products") or g.get("elements") or []:
                        if isinstance(p, dict):
                            pid = p.get("id") or p.get("productId")
                            if pid is not None:
                                ids.add(str(pid))
                if ids:
                    return len(ids)
            n = len(val)
            if n and isinstance(val[0], dict):
                return n
    best = 0
    for v in list(data.values())[:40]:
        if isinstance(v, (dict, list)):
            best = max(best, _total_from_json_obj(v))
    return best


def find_header_index(headers: list[str], candidates: tuple[str, ...] | list[str]) -> int | None:
    normalized = [str(h or "").strip() for h in headers]
    lower_map = {h.lower(): i for i, h in enumerate(normalized) if h}
    for cand in candidates:
        c = cand.strip()
        if c in normalized:
            return normalized.index(c)
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    return None


def ensure_column(headers: list[str], ws, header_name: str) -> int:
    """헤더에 열이 없으면 맨 끝에 추가하고 0-based index 반환."""
    idx = find_header_index(headers, (header_name,))
    if idx is not None:
        return idx
    new_idx = len(headers)
    ws.cell(row=1, column=new_idx + 1, value=header_name)
    headers.append(header_name)
    return new_idx


def read_url_jobs(ws) -> tuple[list[str], int, list[RowJob]]:
    """시트에서 헤더·URL열·작업 행 목록을 읽는다."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("엑셀 헤더(1행)가 비어 있습니다.")
    headers = [str(c or "").strip() for c in header_row]
    url_idx = find_header_index(headers, URL_HEADER_CANDIDATES)
    if url_idx is None:
        # 값이 URL처럼 보이는 첫 열 탐색
        for col_i, h in enumerate(headers):
            sample = ws.cell(row=2, column=col_i + 1).value
            if isinstance(sample, str) and sample.strip().lower().startswith("http"):
                url_idx = col_i
                break
    if url_idx is None:
        raise ValueError(
            "URL 열을 찾지 못했습니다. 헤더에 "
            "'최종 카테고리 URL주소' (또는 URL) 열이 필요합니다."
        )

    label_idx = find_header_index(
        headers,
        ("상위 최종 카테고리명", "최종 카테고리명", "카테고리명"),
    )

    jobs: list[RowJob] = []
    for excel_row in range(2, ws.max_row + 1):
        raw_url = ws.cell(row=excel_row, column=url_idx + 1).value
        url = str(raw_url or "").strip()
        # http(s) · file(로컬 검증) 허용
        low = url.lower()
        if not url or not (low.startswith("http://") or low.startswith("https://") or low.startswith("file:")):
            continue
        label = ""
        if label_idx is not None:
            label = str(ws.cell(row=excel_row, column=label_idx + 1).value or "").strip()
        if label == "목차" or label.startswith("목차") or label.upper() == "TOC":
            continue
        jobs.append(RowJob(excel_row=excel_row, url=url, label=label))
    return headers, url_idx, jobs


def dismiss_popups(page) -> int:
    """URL 진입 후 팝업·레이어·추가 창을 닫는다. 닫은 횟수 반환."""
    closed = 0

    # 1) 추가 창(팝업 윈도우) 닫기 — 메인 page 제외
    try:
        main = page
        for p in list(page.context.pages):
            if p is main:
                continue
            try:
                if not p.is_closed():
                    p.close()
                    closed += 1
            except Exception:
                pass
    except Exception:
        pass

    # 2) JS dialog (alert/confirm) — 다음부터 자동 수락
    try:
        page.on("dialog", lambda d: d.dismiss())
    except Exception:
        pass

    # 3) 화면 위 닫기 버튼 클릭
    for sel in POPUP_CLOSE_SELECTORS:
        try:
            loc = page.locator(sel)
            count = loc.count()
        except Exception:
            continue
        for i in range(min(count, 3)):
            try:
                el = loc.nth(i)
                if el.is_visible(timeout=300):
                    el.click(timeout=800, force=True)
                    closed += 1
                    time.sleep(0.2)
            except Exception:
                continue

    # 4) Escape 키로 모달 닫기 시도
    try:
        page.keyboard.press("Escape")
        time.sleep(0.15)
    except Exception:
        pass

    return closed


FOOTER_SELECTORS = (
    "footer",
    "[role='contentinfo']",
    "#footer",
    ".layout-footer",
    ".footer",
    "[data-qa-qualifier='footer']",
    ".page-footer",
)

# ★요건: 상→푸터↓ → 하단→상단↑ 하며 카드 갯수 집계 (최대한 빠르게)
SCROLL_MAX_ROUNDS = 220
SCROLL_PAUSE_SEC = 0.05  # 하위 호환 기본값
SCROLL_DOWN_PAUSE_SEC = 0.04  # 하향(집계 없음) — 최대 속도
SCROLL_UP_PAUSE_SEC = 0.06  # 상향(카드집계) — DOM 반영 최소 대기
SCROLL_STABLE_ROUNDS = 2  # scrollHeight 불변 연속 횟수
FOOTER_STABLE_ROUNDS = 2  # 푸터/하단 도달 연속 확인
TOP_STABLE_ROUNDS = 2


def count_product_cards(page) -> int:
    """현재 DOM에 보이는 상품 카드 개수 (가상스크롤이면 일부만)."""
    best = 0
    for sel in PRODUCT_CARD_SELECTORS:
        try:
            c = page.locator(sel).count()
            if c > best:
                best = c
        except Exception:
            continue
    return best


def collect_visible_card_keys(page) -> set[str]:
    """현재 화면의 상품 카드 고유키 수집 (상향 스크롤 집계용)."""
    try:
        keys = page.evaluate(
            """() => {
              const out = new Set();
              const cardSel = [
                'li.product-grid-product',
                'article.product-grid-product',
                '.product-grid-product',
                '[data-qa-qualifier="product-grid-product"]',
                'li[class*="product-grid"]',
                '[data-productid]',
              ].join(',');
              document.querySelectorAll(cardSel).forEach((card) => {
                const pid = card.getAttribute('data-productid') ||
                  card.getAttribute('data-product-id') || '';
                if (pid) { out.add('pid:' + String(pid)); return; }
                let href = '';
                const a = card.querySelector('a[href*="-p"]');
                if (a) href = a.href || a.getAttribute('href') || '';
                const m = href.match(/-p(\\d+)\\.html/i);
                if (m) { out.add('pid:' + m[1]); return; }
                if (href) { out.add('href:' + String(href).split('?')[0]); return; }
                const img = card.querySelector('img');
                const src = img ? (img.currentSrc || img.src ||
                  img.getAttribute('src') || img.getAttribute('data-src') || '') : '';
                if (src) { out.add('img:' + String(src).split('?')[0]); return; }
              });
              return Array.from(out);
            }"""
        )
        return {str(x) for x in (keys or []) if x}
    except Exception:
        return set()


# 이전 이름 호환
collect_visible_card_image_keys = collect_visible_card_keys


def parse_card_image_keys_from_html(html: str) -> set[str]:
    """단위테스트용 — HTML 조각에서 카드 키를 추출."""
    text = html or ""
    keys: set[str] = set()
    for m in re.finditer(r'data-productid=["\'](\d+)["\']', text, flags=re.I):
        keys.add(f"pid:{m.group(1)}")
    return keys


def _collect_ids_from_json(data: object, out: set[str], depth: int = 0) -> None:
    if depth > 8 or data is None:
        return
    if isinstance(data, dict):
        pid = data.get("id") or data.get("productId") or data.get("product_id")
        if pid is not None and (
            "price" in data
            or "seo" in data
            or "detail" in data
            or "name" in data
            or "brand" in data
            or "sectionName" in data
            or "productType" in data
        ):
            out.add(str(pid))
        for v in data.values():
            if isinstance(v, (dict, list)):
                _collect_ids_from_json(v, out, depth + 1)
    elif isinstance(data, list):
        for item in data[:500]:
            _collect_ids_from_json(item, out, depth + 1)


def is_near_page_bottom(page) -> bool:
    try:
        return bool(
            page.evaluate(
                """() => {
                  const el = document.scrollingElement || document.documentElement;
                  const remain = el.scrollHeight - el.scrollTop - el.clientHeight;
                  return remain < 120;
                }"""
            )
        )
    except Exception:
        return False


def is_near_page_top(page) -> bool:
    try:
        return bool(
            page.evaluate(
                """() => {
                  const el = document.scrollingElement || document.documentElement;
                  return (el.scrollTop || 0) < 80;
                }"""
            )
        )
    except Exception:
        return False


def footer_in_view(page) -> bool:
    """푸터 영역이 뷰포트에 들어왔는지."""
    for sel in FOOTER_SELECTORS:
        try:
            loc = page.locator(sel).first
            if loc.count() == 0:
                continue
            if loc.is_visible(timeout=80):
                box = loc.bounding_box()
                if not box:
                    continue
                vh = (page.viewport_size or {}).get("height") or 900
                if box["y"] < vh + 40:
                    return True
        except Exception:
            continue
    return is_near_page_bottom(page)


def scroll_step_down(page) -> None:
    """아래로 한 걸음 — 하향은 집계 없으므로 큰 폭으로 빠르게."""
    try:
        page.evaluate(
            """() => {
              const el = document.scrollingElement || document.documentElement;
              const step = Math.max(900, Math.floor(window.innerHeight * 0.95));
              window.scrollBy(0, step);
              const remain = el.scrollHeight - el.scrollTop - el.clientHeight;
              if (remain < 220) {
                window.scrollTo(0, el.scrollHeight);
              }
            }"""
        )
    except Exception:
        try:
            page.mouse.wheel(0, 1600)
        except Exception:
            pass


def scroll_step_up(page) -> None:
    """위로 한 걸음 — 카드 집계용, 가능한 한 빠르게."""
    try:
        page.evaluate(
            """() => {
              const step = Math.max(700, Math.floor(window.innerHeight * 0.85));
              window.scrollBy(0, -step);
              const el = document.scrollingElement || document.documentElement;
              if ((el.scrollTop || 0) < 80) {
                window.scrollTo(0, 0);
              }
            }"""
        )
    except Exception:
        try:
            page.mouse.wheel(0, -1400)
        except Exception:
            pass


def _page_scroll_height(page) -> int:
    try:
        return int(
            page.evaluate(
                "() => (document.scrollingElement||document.documentElement).scrollHeight"
            )
            or 0
        )
    except Exception:
        return 0


def scroll_down_to_footer(
    page,
    *,
    progress: ProgressFn | None = None,
    max_rounds: int = SCROLL_MAX_ROUNDS,
    pause: float = SCROLL_DOWN_PAUSE_SEC,
) -> None:
    """스크롤을 위에서 하단 푸터 영역까지 빠르게 내린다. (집계 안 함)"""
    prev_h = -1
    stable = 0
    footer_hits = 0
    # ★요건: 스크롤 상세 로그 출력 안 함 (4단계·주요정보만)
    del progress

    try:
        page.evaluate("() => window.scrollTo(0, 0)")
    except Exception:
        pass
    time.sleep(0.05)

    for i in range(1, max_rounds + 1):
        h = _page_scroll_height(page)
        at_footer = is_near_page_bottom(page) or footer_in_view(page)
        if h > 0 and h == prev_h:
            stable += 1
        else:
            stable = 0
        prev_h = h
        if at_footer:
            footer_hits += 1
        else:
            footer_hits = 0

        if footer_hits >= FOOTER_STABLE_ROUNDS and stable >= SCROLL_STABLE_ROUNDS:
            break

        scroll_step_down(page)
        time.sleep(pause)
        # 팝업 재확인은 드물게만 (속도 저하 방지)
        if i == 1 or i % 25 == 0:
            dismiss_popups(page)
        if stop_requested():
            break

    try:
        page.evaluate(
            "() => window.scrollTo(0, (document.scrollingElement||document.documentElement).scrollHeight)"
        )
    except Exception:
        pass
    time.sleep(0.12)


def scroll_up_count_card_images(
    page,
    *,
    progress: ProgressFn | None = None,
    max_rounds: int = SCROLL_MAX_ROUNDS,
    pause: float = SCROLL_UP_PAUSE_SEC,
) -> int:
    """맨 하단부터 맨 위까지 빠르게 올리며 카드 갯수를 센다 → 상품수."""
    seen: set[str] = set()
    prev_n = -1
    stable = 0
    top_hits = 0
    # ★요건: 스크롤 상세 로그 출력 안 함 (4단계·주요정보만)
    del progress

    try:
        page.evaluate(
            "() => window.scrollTo(0, (document.scrollingElement||document.documentElement).scrollHeight)"
        )
    except Exception:
        pass
    time.sleep(0.08)

    for i in range(1, max_rounds + 1):
        seen |= collect_visible_card_keys(page)
        n = len(seen)
        at_top = is_near_page_top(page)
        if n > 0 and n == prev_n:
            stable += 1
        else:
            stable = 0
        prev_n = n
        if at_top:
            top_hits += 1
        else:
            top_hits = 0

        if top_hits >= TOP_STABLE_ROUNDS and stable >= SCROLL_STABLE_ROUNDS:
            break

        scroll_step_up(page)
        time.sleep(pause)
        if i == 1 or i % 25 == 0:
            dismiss_popups(page)
        if stop_requested():
            break

    try:
        page.evaluate("() => window.scrollTo(0, 0)")
    except Exception:
        pass
    time.sleep(0.1)
    seen |= collect_visible_card_keys(page)
    return len(seen)


def extract_count_from_page(
    page,
    *,
    api_totals: list[int] | None = None,
    api_ids: set[str] | None = None,
    progress: ProgressFn | None = None,
) -> int:
    """상품수 추출 — 상→푸터↓ → 하단→상단↑ 카드갯수.

    1) 첫 팝업 닫기(호출측)
    2) 푸터까지 고속 다운
    3) 하단→상단 고속 올리며 카드 갯수 집계
    4) 그 갯수 = 상품수
    """
    del api_totals, api_ids  # 카드 갯수만 사용

    try:
        page.wait_for_selector(
            ", ".join(PRODUCT_CARD_SELECTORS[:4]),
            timeout=5_000,
        )
    except Exception:
        pass

    _log(progress, "로직", "2) 상단 → 푸터 스크롤")
    scroll_down_to_footer(page, progress=None)

    if stop_requested():
        return len(collect_visible_card_keys(page))

    _log(progress, "로직", "3) 하단 → 상단 스크롤 · 카드 갯수 집계")
    total = scroll_up_count_card_images(page, progress=None)

    _log(progress, "로직", f"4) 상품수={total}")
    return total


def attach_api_total_listener(page) -> tuple[list[int], set[str]]:
    """카테고리/상품 API 응답에서 total 후보 + 상품 ID 누적."""
    found: list[int] = []
    ids: set[str] = set()

    def _on_response(response) -> None:
        try:
            if response.status != 200:
                return
            url = response.url or ""
            low = url.lower()
            if not any(
                k in low
                for k in (
                    "/products",
                    "product",
                    "catalog",
                    "category",
                    "commercial",
                    "ajax=true",
                )
            ):
                return
            ctype = (response.headers or {}).get("content-type", "")
            if "json" not in ctype and "text" not in ctype and "javascript" not in ctype:
                return
            data = response.json()
            n = _total_from_json_obj(data)
            if n > 0:
                found.append(n)
            _collect_ids_from_json(data, ids)
        except Exception:
            return

    page.on("response", _on_response)
    return found, ids


def _default_headless() -> bool:
    # 로컬 Windows는 화면 확인용으로 창 표시, CI/리눅스는 headless
    return os.name != "nt"


def run_extract(
    excel_path: str | Path,
    *,
    progress: ProgressFn | None = None,
    headless: bool | None = None,
    post_popup_wait_sec: float = POST_POPUP_WAIT_SEC,
) -> ExtractResult:
    """엑셀 URL별 상품수 수집 후 지정 폴더에 별도 파일(파일명+버전)로 저장."""
    path = Path(excel_path).expanduser().resolve()
    out_path = build_output_excel_path(path)
    result = ExtractResult(
        ok=False,
        excel_path=str(out_path),
        input_excel_path=str(path),
    )
    if not path.is_file():
        result.errors.append(f"파일 없음: {path}")
        return result
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        result.errors.append("xlsx/xlsm 엑셀 파일만 지원합니다.")
        return result

    use_headless = _default_headless() if headless is None else headless
    clear_stop_flag()
    stopped = False

    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"Playwright 미설치: {e}")
        return result

    try:
        wb = load_workbook(str(path))
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"엑셀 열기 실패: {e}")
        return result

    ws = wb.active
    try:
        headers, _url_idx, jobs = read_url_jobs(ws)
    except Exception as e:  # noqa: BLE001
        result.errors.append(str(e))
        try:
            wb.close()
        except Exception:
            pass
        return result

    count_idx = ensure_column(headers, ws, COUNT_HEADER)
    collectible_idx = find_header_index(headers, (COLLECTIBLE_HEADER,))
    # 상품수집가능개수 열이 있으면 함께 갱신, 없으면 추가하지 않음(기존 엑셀 보존)

    result.total = len(jobs)
    _log(
        progress,
        "준비",
        f"입력: {path.name} · URL {result.total}건 · 출력폴더: {OUTPUT_DIR}",
    )
    if result.total == 0:
        result.errors.append("처리할 URL 행이 없습니다.")
        try:
            wb.close()
        except Exception:
            pass
        return result

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=use_headless)
            # Zara 영문 DE 등 — en-GB 가 차단·팝업에 유리한 경우가 많음
            context = browser.new_context(
                user_agent=DEFAULT_UA,
                viewport={"width": 1440, "height": 900},
                locale="en-GB",
            )
            page = context.new_page()
            api_totals, api_ids = attach_api_total_listener(page)

            for i, job in enumerate(jobs, start=1):
                if stop_requested():
                    stopped = True
                    _log(progress, "중단", "사용자 중단 요청 — 현재까지 UPDATE 저장")
                    break

                label = (job.label or "").strip() or urlparse(job.url).path
                _log(
                    progress,
                    "URL",
                    f"{i}/{result.total} 행{job.excel_row} 열기: {label} · {job.url}",
                )
                api_totals.clear()
                api_ids.clear()
                try:
                    page.goto(job.url, wait_until="domcontentloaded", timeout=90_000)
                except Exception as e:  # noqa: BLE001
                    result.failed += 1
                    result.rows.append(
                        RowResult(label=label, count=None, url=job.url)
                    )
                    _log(progress, "오류", f"행{job.excel_row} 접속 실패: {e}")
                    _log(
                        progress,
                        "최종",
                        format_final_output(label, None, job.url),
                    )
                    continue

                # ★요건1: url 입력 후 첫번째 팝업창 닫기
                closed = dismiss_popups(page)
                _log(
                    progress,
                    "로직",
                    f"1) 첫 팝업 닫기"
                    + (f" ({closed}건)" if closed else " (없음)"),
                )

                # 대기 (상세 로그 없음)
                waited = 0.0
                step = 0.25
                while waited < float(post_popup_wait_sec):
                    if stop_requested():
                        stopped = True
                        break
                    time.sleep(step)
                    waited += step
                if stopped:
                    _log(progress, "중단", "사용자 중단 요청 — 현재까지 UPDATE 저장")
                    break

                # 대기 중 다시 뜬 팝업이 있으면 한 번 더 닫고, 짧게만 안정화
                closed2 = dismiss_popups(page)
                if closed2:
                    time.sleep(0.5)

                count = extract_count_from_page(
                    page,
                    api_totals=api_totals,
                    api_ids=api_ids,
                    progress=progress,
                )
                if count == 0:
                    _log(
                        progress,
                        "경고",
                        f"행{job.excel_row} 상품수 0 — 카드 미검출",
                    )
                ws.cell(row=job.excel_row, column=count_idx + 1, value=count)
                if collectible_idx is not None:
                    ws.cell(
                        row=job.excel_row,
                        column=collectible_idx + 1,
                        value=count,
                    )
                result.updated += 1
                result.rows.append(RowResult(label=label, count=count, url=job.url))
                # ★요건: 실행로그 최종출력 = 상위 최종 카테고리명 · 상품갯수 · url
                _log(
                    progress,
                    "최종",
                    format_final_output(label, count, job.url),
                )

            browser.close()
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"브라우저 실행 실패: {e}")
        try:
            wb.close()
        except Exception:
            pass
        clear_stop_flag()
        return result

    # ★요건: 입력 파일은 유지, 출력은 P2_INPUT_건수집계 에 파일명+버전으로 별도 생성
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(str(out_path))
        wb.close()
        result.excel_path = str(out_path)
    except Exception as e:  # noqa: BLE001
        result.errors.append(f"엑셀 저장 실패: {e} → {out_path}")
        try:
            wb.close()
        except Exception:
            pass
        clear_stop_flag()
        return result

    clear_stop_flag()
    # ★요건: 종료 시 최종출력을 상위 최종 카테고리명 · 상품갯수 · url 로 모아 표시
    log_final_outputs(progress, result.rows)

    if stopped:
        result.ok = result.updated > 0
        result.warnings.append("사용자 중단으로 일부만 저장")
        _log(
            progress,
            "중단",
            f"부분 저장 {result.updated}/{result.total} · 출력: {out_path}",
        )
        return result

    result.ok = result.updated > 0 and not result.errors
    if result.updated == 0 and not result.errors:
        result.errors.append("상품수를 하나도 갱신하지 못했습니다.")
        result.ok = False
    _log(
        progress,
        "완료",
        f"저장 {result.updated}/{result.total} · 실패 {result.failed} · 출력: {out_path}",
    )
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="P1_101 상품수 추출 — 엑셀 URL UPDATE")
    parser.add_argument("excel", help="URL이 담긴 엑셀 파일 경로")
    parser.add_argument(
        "--headless",
        action="store_true",
        help="브라우저를 숨김 모드로 실행",
    )
    parser.add_argument(
        "--headed",
        action="store_true",
        help="브라우저 창을 표시하며 실행",
    )
    parser.add_argument(
        "--wait",
        type=float,
        default=POST_POPUP_WAIT_SEC,
        help=f"팝업 닫기 후 대기 초 (기본 {POST_POPUP_WAIT_SEC:g})",
    )
    args = parser.parse_args(argv)

    headless: bool | None
    if args.headless:
        headless = True
    elif args.headed:
        headless = False
    else:
        headless = None

    result = run_extract(
        args.excel,
        headless=headless,
        post_popup_wait_sec=args.wait,
    )
    if not result.ok:
        for e in result.errors:
            print(f"[오류] {e}", flush=True)
        return 1
    # 보드: 부분 UPDATE(중단)도 exit 0 — 로그의 구분
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
